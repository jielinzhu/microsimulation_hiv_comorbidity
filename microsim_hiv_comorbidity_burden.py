"""
Microsimulation model to forecast multimorbidity in aging people with HIV

Data needed:
- stop.sas7bdat: individual characteristics from the STOP HIV/AIDS cohort
    variables included: moh_id (individual ID), sex_at_birth_dv (sex at birth), DOB (date of birth), 
                        earliest_HIV (earliest date known to be HIV+), FARVDT (first ART date), 
                        end_fu_dt (end of follow-up date), IDU_all (IDU status), 
                        diag_source (source to determine earliest_HIV), PREV_ARV (previous ART experience before the DTP)
- dad.sas7bdat: hospitalization record from the STOP HIV/AIDS cohort
    variables included: moh_id (individual ID), addate (date of admission), DIAG1-DIAG16 (ICD9 code for diagnosis)
                        DIAGX1-DIAGX25 (ICD10 code for diagnosis)
- msp_app.sas7bdat and msp_ffs.pkl: record of Medical Service Plan for BC from the STOP HIV/AIDS cohort
    variables included: moh_id (individual ID), SERVDT (date of service), CLMSPEC (claim specialty), 
                        DIAGCD (ICD9 code for diagnosis), FITM (fee item)
- pnet.pkl: pharmanet record from the STOP HIV/AIDS cohort
    variables included: moh_id (individual ID), DIN_PIN (drug identification number and pseido identification number for drugs),
                        date_of_service (date of service)
- stop_cd4.sas7bdat: record of CD4 cell count from the STOP HIV/AIDS cohort
    variables included: moh_id (individual ID), TESTDATE (date of test), RESULT (CD4)
- virload.sas7bdat: record of viral load test from the STOP HIV/AIDS cohort
    variables included: moh_id (individual ID), COLDATE (date of test), VLOAD (viral load)
- art_records.sas7bdat: record of ART dispensation from the STOP HIV/AIDS cohort
    variables included: moh_id (individual ID), STARTDATE (date to start ART), STOPDATE (date to stop ART),
                        REGIMEN (combination of regimens), NON_BACKNONE_ARV (non-backbone regimen),
                        BACKBONE_ARV (backbone regimen), NON_BACKNONE_CLASSES (class of the third regimen)
- cvd/htn/dm/oa/copd/ckd/cld_cmb/manx/sczo_moh/prsn/cancer_noaids.sas7bdat: date of healthcare visit related to each disease from the STOP HIV/AIDS cohort
    variables included: moh_id (individual ID), dt (date of the visit), source (source of the record)
- earliest_health_record.sas7bdat: record of the earliest healthcare visit from the STOP HIV/AIDS cohort
    variables included: moh_id (individual ID), earliest_record_dt (date of the earliest record)
- dtp.sas7bdat: individual characteristics from the DTP
    variables included: PSEUDO (individual ID), BIRTHDATE (date of birth), FARVDT (first ART date), 
                        LASTCTDT (last contact date), DTHDT (date of death)
- dtp_art_records.sas7bdat: record of ART dispensation from the DTP
    variables included: PSEUDO (individual ID), STARTDATE (date to start ART), STOPDATE (date to stop ART),
                        REGIMEN (combination of regimens), NON_BACKNONE_ARV (non-backbone regimen),
                        BACKBONE_ARV (backbone regimen), NON_BACKNONE_CLASSES (class of the third regimen)
- dtp_virload.sas7bdat: record of viral load test from the DTP
    variables included: PSEUDO (individual ID), COLDATE (date of test), VLOAD (viral load)
- phac_inc_prev.xlsx: HIV incidence and prevalence estimation (with plausible range) between 1975 and 2020 from Public Health Agency of Canada,
                      and the number of PLWH with viral load <200 (column '200') between 2000 and 2022 from the DTP
- table_ipr.xlsx: Incidence-to-unsuppressed ratio between 2000 and 2022 (PMID: 38848736)

To run the microsimulation model:
comment out the rest of runs=.... to run each module
"""

import sys
import time
import pickle
import warnings
import random
import itertools
import copy
import numpy as np
import pandas as pd
import scipy.stats as st
import seaborn as sns
import matplotlib.pyplot as plt
from collections import Counter
from typing import Dict,List 
from datetime import datetime
from itertools import combinations 
from scipy import integrate
from scipy.optimize import minimize
from scipy.stats import uniform
from matplotlib.patches import Patch 
from pyDOE import *
from openpyxl import Workbook, load_workbook
from multiprocessing import Pool
from PIL import Image

def sigmoid(y0, y1, tm, t):
    '''monotonically increasing function that starts at y0, plateaus at y1, and reaches
    the half-point (y0+y1)/2 at tm, as a function of time t'''
    return (y1 - y0)*(1. + np.tanh(2.*(t - tm)/tm))/2. + y0

def func_target_sigmoid_iur_ci(arr_params,data_iur,c,key_col):
    """for uncertainty in iur: fit boundries of time-varying incidence-unsuppressed(200)-ratio to a sigmoid function with fixed tm/t0 from the median estimation"""

    ####initialize parameter values
    iur_y0 = arr_params[0]
    iur_y1 = arr_params[1]
    iur_tm = c.sig_iur_params[-2]
    iur_t0 = c.sig_iur_params[-1]

    ####simulate iur using a sigmoid function
    arr_iur_sim = sigmoid(iur_y0,iur_y1,iur_tm,data_iur['year'].values-iur_t0) #treat t=0 at data_iur['year].values[0]

    return sum((arr_iur_sim/data_iur[key_col].values-1)**2)

def func_target_sigmoid_iur(arr_params,data_iur):
    """fit time-varying incidence-unsuppressed(200)-ratio to a sigmoid function"""

    ####initialize parameter values
    iur_y0 = arr_params[0]
    iur_y1 = arr_params[1]
    iur_tm = arr_params[2]
    iur_t0 = arr_params[-1]

    ####simulate iur using a sigmoid function
    arr_iur_sim = sigmoid(iur_y0,iur_y1,iur_tm,data_iur['year'].values-iur_t0) #treat t=0 at data_iur['year].values[0]

    return sum((arr_iur_sim/data_iur['median'].values-1)**2)

def func_target_sigmoid_rdiag(arr_params,data_diag):
    """fit time-varying diagnosis rate as new diagnosis/undiagnosed PLWH to a sigmoid function"""

    ####initialize parameter values
    rdiag_y0 = arr_params[0]
    rdiag_y1 = arr_params[1]
    rdiag_tm = arr_params[2]
    rdiag_t0 = arr_params[3] 

    ####simulate diagnosis rate using a sigmoid function
    arr_rdiag_sim = sigmoid(rdiag_y0,rdiag_y1,rdiag_tm,data_diag['year'].values-rdiag_t0)

    return sum((arr_rdiag_sim/data_diag['rdiag'].values-1)**2)

def func_target_sigmoid_general(arr_params,data_hist_fit,key_fit):
    """fitting time-varying variable key_fit to a sigmoid function"""

    ####initialize parameter values
    y0 = arr_params[0]
    y1 = arr_params[1]
    tm = arr_params[2]
    t0 = arr_params[3] 

    ####simulate key_fit variable using a sigmoid function
    arr_sim = sigmoid(y0,y1,tm,data_hist_fit['year'].values-t0)

    return sum((arr_sim/data_hist_fit[key_fit].values-1)**2)

def sample_from_prob_matrix(arr_prob):
    """sample from np.range(arr_prob.shape[1]), based on different discrete probability distribution from matrix arr_prob, probability per row
    link:https://stackoverflow.com/questions/40474436/how-to-apply-numpy-random-choice-to-a-matrix-of-probability-values-vectorized-s"""

    cumprob = arr_prob.cumsum(axis=1)
    uni_sample = np.random.rand(cumprob.shape[0], 1)
    choices = (uni_sample < cumprob).argmax(axis=1) #find the index that fit in cumprob

    return choices

def func_sort_byrange(col_value,dic_label):
    """use to sort certain column value into different classes stored in dic_label, class name is the key of dic_label if col_value fall in the range"""
    
    for key_i in list(dic_label):
        if col_value>=dic_label[key_i][0] and col_value<dic_label[key_i][1]:
            sorted_key = key_i
            break
        elif np.isnan(col_value):
            sorted_key = 'missing' #include the situation when value is missing (showing as NaN)
    return sorted_key

def func_apply_key_by_id(row,dic_state_id):
    """used to create new columns to record status of each individual at given time step"""

    list_keys = list(dic_state_id)
    i = 0
    while i<len(list_keys) and row['moh_id'] not in dic_state_id[list_keys[i]]:
        i+=1
    return list_keys[i] if i<len(list_keys) else np.nan #NaN if id not in dic_state_id

def func_apply_spvl_status(row,col_date,data_spvl_rebound,list_columns_interest):
    """use to determine whether the individual in each row achieved viral suppression, with ART failure (AF) or ART interruption (AI) after FARVDT at date col_date, given data_spvl_rebound"""

    if row[col_date]=='T': #generalized for any given step
        arr_days = np.array([(pd.Timestamp(date_i)-pd.Timestamp(col_date)).days for date_i in data_spvl_rebound[list_columns_interest][data_spvl_rebound['moh_id']==row['moh_id']].values[0]]) #count number of days between spvl/rebound date and the time step
        index_min = len(arr_days[arr_days<=0])-1 #spvl/rebound status is determined by the last spvl/rebound date before col_date
        if 'FARVDT'==list_columns_interest[index_min]:
            return row[col_date]  #keep status T if individual initiated ART before col_date but not achieved first suppression
        elif 'rebound' in list_columns_interest[index_min]:
            return 'AF' if data_spvl_rebound[list_columns_interest[index_min][:-2]+'status'][data_spvl_rebound['moh_id']==row['moh_id']].values[0]=='F' else 'AI'
        elif 'spvl' in list_columns_interest[index_min]:
            return 'S'
        else:
            print ('Existing conditions not considered')
    else:
        return row[col_date]

def func_spvl_dt(c,data_record,data_pvl,col_startdt,col_enddt,spvl_i):
    """return to dataframe recording the first date of the first two consecutive pvl<200 between col_startdt and col_enddt,
    only for individuals in data_record, which only inlcudes those with startdt available
    spvl_i represents the time number to achieve spvl
    c.dic_spvl_thrh is used to determine viral suppression information before threshold of detectable pvl=50"""

    #####initialize dataframe to save final outcomes
    data_outcome = pd.DataFrame([])
    data_outcome['moh_id'] = pd.Series(data_record.moh_id)
    assert(data_outcome.shape[0]==data_outcome.index.max()+1),'Inconsistent index numer in comparison to the shape of the dataframe'

    #####find first spvl date for each individual, suppression defined as the first date of two consecutive spvl<200 (PMID:31517267)
    list_dates_spvl_thrh = list(c.dic_spvl_thrh)
    list_dates_spvl_thrh.sort() 
    dic_pvl_after_startdt = {'y':[],'n':[]}
    list_spvl_dt = []
    for i,id_i in enumerate(data_record.moh_id.values):
        t1 = data_record[col_startdt][data_record.moh_id==id_i].values[0]
        t2 = data_record[col_enddt][data_record.moh_id==id_i].values[0]
        data_pvl_i = data_pvl[(data_pvl.moh_id==id_i)&(data_pvl.COLDATE>=t1)&(data_pvl.COLDATE<=t2)].sort_values(by='COLDATE').reset_index(drop=True) 
        if not data_pvl_i.empty:
            dic_pvl_after_startdt['y'].append(id_i) #pvl tested after FARVDT
            if data_pvl_i.shape[0]>=2:
                for index_i,date_i in enumerate(data_pvl_i.COLDATE.values[:-1]): #need at least two pvl records between t1 and t2
                    if pd.Timestamp(date_i)<pd.Timestamp(list_dates_spvl_thrh[0]) and (data_pvl_i['VLOAD'][(data_pvl_i.index>=index_i)&(data_pvl_i.index<=index_i+1)]<c.dic_spvl_thrh[list_dates_spvl_thrh[0]]).all():
                        list_spvl_dt.append(data_pvl_i['COLDATE'][data_pvl_i.index==index_i].values[0])
                        break
                    elif pd.Timestamp(date_i)>=pd.Timestamp(list_dates_spvl_thrh[0]) and pd.Timestamp(date_i)<pd.Timestamp(list_dates_spvl_thrh[1]) and (data_pvl_i['VLOAD'][(data_pvl_i.index>=index_i)&(data_pvl_i.index<=index_i+1)]<c.dic_spvl_thrh[list_dates_spvl_thrh[1]]).all():
                        list_spvl_dt.append(data_pvl_i['COLDATE'][data_pvl_i.index==index_i].values[0])
                        break
                    elif pd.Timestamp(date_i)>=pd.Timestamp(list_dates_spvl_thrh[1]) and (data_pvl_i['VLOAD'][(data_pvl_i.index>=index_i)&(data_pvl_i.index<=index_i+1)]<200).all():
                        list_spvl_dt.append(data_pvl_i['COLDATE'][data_pvl_i.index==index_i].values[0])
                        break
                if len(list_spvl_dt)<i+1:
                    list_spvl_dt.append(np.nan) 
            else:
                list_spvl_dt.append(np.nan) #no viral suppression if only one pvl record in the period
        else:
            dic_pvl_after_startdt['n'].append(id_i) #pvl not tested between FARVDT and end_fu_dt
            list_spvl_dt.append(np.nan)
    data_outcome[str(spvl_i)+'_spvl_dt'] = pd.Series(list_spvl_dt)
    data_outcome['pvl_after_'+col_startdt] = data_outcome.apply(func_apply_key_by_id,args=(dic_pvl_after_startdt,),axis=1)

    return data_outcome

def func_rebound_dt(c,data_record,data_pvl,col_startdt,col_enddt,rebound_i):
    """return to dataframe recording the first date of the first two consecutive pvl>=200 between col_startdt and col_enddt,
    only for individuals in data_record, which only inlcudes those with startdt available
    i represents the time number to viral rebound
    c.dic_spvl_thrh is used to determine viral suppression information before threshold of detectable pvl=50"""

    #####initialize dataframe to save final outcomes
    data_outcome = pd.DataFrame([])
    data_outcome['moh_id'] = pd.Series(data_record.moh_id)
    assert(data_outcome.shape[0]==data_outcome.index.max()+1),'Inconsistent index numer in comparison to the shape of the dataframe'

    #####find first rebound date for each individual, defined as the first date of two consecutive spvl>=200 and >=30 days apart (PMID:27769246)
    list_dates_spvl_thrh = list(c.dic_spvl_thrh)
    list_dates_spvl_thrh.sort() 
    list_rebound_dt = []
    for i,id_i in enumerate(data_record.moh_id.values):
        t1 = data_record[col_startdt][data_record.moh_id==id_i].values[0]
        t2 = data_record[col_enddt][data_record.moh_id==id_i].values[0]
        data_pvl_i = data_pvl[(data_pvl.moh_id==id_i)&(data_pvl.COLDATE>=t1)&(data_pvl.COLDATE<=t2)].sort_values(by='COLDATE').reset_index(drop=True) 
        if not data_pvl_i.empty:
            if data_pvl_i.shape[0]>=2:
                for index_i,date_i in enumerate(data_pvl_i.COLDATE.values[:-1]): #need at least two pvl records between t1 and t2
                    if pd.Timestamp(date_i)<pd.Timestamp(list_dates_spvl_thrh[0]) and (data_pvl_i['VLOAD'][(data_pvl_i.index>=index_i)&(data_pvl_i.index<=index_i+1)]>=c.dic_spvl_thrh[list_dates_spvl_thrh[0]]).all() and (pd.Timestamp(data_pvl_i['COLDATE'][data_pvl_i.index==index_i+1].values[0])-pd.Timestamp(data_pvl_i['COLDATE'][data_pvl_i.index==index_i].values[0])).days>=30:
                        list_rebound_dt.append(data_pvl_i['COLDATE'][data_pvl_i.index==index_i].values[0])
                        break
                    elif pd.Timestamp(date_i)>=pd.Timestamp(list_dates_spvl_thrh[0]) and pd.Timestamp(date_i)<pd.Timestamp(list_dates_spvl_thrh[1]) and (data_pvl_i['VLOAD'][(data_pvl_i.index>=index_i)&(data_pvl_i.index<=index_i+1)]>=c.dic_spvl_thrh[list_dates_spvl_thrh[1]]).all() and (pd.Timestamp(data_pvl_i['COLDATE'][data_pvl_i.index==index_i+1].values[0])-pd.Timestamp(data_pvl_i['COLDATE'][data_pvl_i.index==index_i].values[0])).days>=30:
                        list_rebound_dt.append(data_pvl_i['COLDATE'][data_pvl_i.index==index_i].values[0])
                        break
                    elif pd.Timestamp(date_i)>=pd.Timestamp(list_dates_spvl_thrh[1]) and (data_pvl_i['VLOAD'][(data_pvl_i.index>=index_i)&(data_pvl_i.index<=index_i+1)]>=200).all() and (pd.Timestamp(data_pvl_i['COLDATE'][data_pvl_i.index==index_i+1].values[0])-pd.Timestamp(data_pvl_i['COLDATE'][data_pvl_i.index==index_i].values[0])).days>=30:
                        list_rebound_dt.append(data_pvl_i['COLDATE'][data_pvl_i.index==index_i].values[0])
                        break
                if len(list_rebound_dt)<i+1:
                    list_rebound_dt.append(np.nan) 
            else:
                list_rebound_dt.append(np.nan) #no viral rebound if only one pvl record in the period
        else:
            list_rebound_dt.append(np.nan) #no viral rebound if no further viral load information between t1 and t2
    data_outcome[str(rebound_i)+'_rebound_dt'] = pd.Series(list_rebound_dt)

    return data_outcome

def func_update_comorb_variable(c,data_comorb,data_char,data_hiv,data_pvl,dict_reg,t0):
    """derive timely updated variables as dataframes, used to derive probability of comorbidity incidence
    only need to derive once at each time step for all comorbidities"""

    #####initialize dataframe to save covariates necessary to estimate probability of comorbidity incidence
    data_comorb_var = data_comorb[['moh_id']].copy() 

    #####incorporate behaviour variable if available
    if 'pa' in list(data_char):
        data_comorb_var = pd.merge(data_comorb_var,data_char[['moh_id','pa','smk','alc']].copy(),how='left',on='moh_id')

    #####update data_char/data_comorb/data_hiv/data_pvl/dic_reg to get variables necessary to estimate probability of comorbidity incidence and merge the necessary information to one dataframe
    data_char['age_scale10_dv'] = (t0 - data_char['DOB']).dt.total_seconds()/(60*60*24)/c.def_year/10 #count full time between t0 and DOB instead of rounded days
    data_char['counter'] = (t0 - data_char['baseline_dt']).dt.total_seconds()/(60*60*24)/c.def_year + 1 #time between t0 and beginning of follow-up, initial value as 1
    data_comorb_var = pd.merge(data_comorb_var,data_char[['moh_id','sex_dv','age_bsln_scale10_dv','age_scale10_dv','ncd4_bsln_dv','su_bsln_dv','year_diag_dv','counter']],how='left',on='moh_id') 
    list_prev_comorb = ['manx','prsn','sczo','dm','ckd','htn','copd','cld','cancer','cvd','oa']
    for comorb_i in list_prev_comorb:
        data_comorb[comorb_i+'_dv'] = data_comorb['earliest_'+comorb_i+'_dt'].apply(lambda x: 0 if x>t0 else 1) #derive prevalent comorbidities of interest at t=t0
    list_comorb_dv = [comorb_i+'_dv' for comorb_i in list_prev_comorb]
    data_comorb_var = pd.merge(data_comorb_var,data_comorb[['moh_id']+list_comorb_dv],how='left',on='moh_id') 
    if str(t0)[:10] not in list(data_pvl):
        t1 = t0-pd.to_timedelta(c.def_year*c.dt,unit='d')
        data_comorb_var = pd.merge(data_comorb_var,data_pvl[['moh_id',str(t1)[:10]]],how='left',on='moh_id').rename(columns={str(t1)[:10]:'pvl_dv'}) #data_hiv at t0 with derived viral suppression status over 1yr period
        for k in dict_reg.keys():
            data_comorb_var = pd.merge(data_comorb_var,dict_reg[k][['moh_id',str(t1)[:10]]].copy(),how='left',on='moh_id').rename(columns={str(t1)[:10]:k+'_dv'}) #add information of sproportion of art regimen over all treatment for 1yr period
    else:
        data_comorb_var = pd.merge(data_comorb_var,data_pvl[['moh_id',str(t0)[:10]]],how='left',on='moh_id').rename(columns={str(t0)[:10]:'pvl_dv'}) 
        for k in dict_reg.keys():
            data_comorb_var = pd.merge(data_comorb_var,dict_reg[k][['moh_id',str(t0)[:10]]].copy(),how='left',on='moh_id').rename(columns={str(t0)[:10]:k+'_dv'}) 
    
    return data_comorb_var

def func_comorb_rate2prob(c,dic_prob_art_stat,data_comorb_prob):
    """"estimate comorbidity incidence rate based on art-specified coefficients and convert to probability"""

    #use iteration through keys of dic_prob_comorb_art/noart without explicitly specifying linear relationship to avoid mistakes, time consumptions are similar
    for art_stat in dic_prob_art_stat.keys():
        list_keys = list(dic_prob_art_stat[art_stat])
        list_keys.remove('beta0')
        data_comorb_prob['lnr_'+art_stat] = pd.Series([dic_prob_art_stat[art_stat]['beta0'][0]]*data_comorb_prob.shape[0])
        list_keys_check = []
        for key_i in list_keys:
            list_keys_check.append(key_i)
            data_comorb_prob['lnr_'+art_stat] += dic_prob_art_stat[art_stat][key_i][0]*data_comorb_prob[key_i]
        data_comorb_prob['rate_'+art_stat] = np.exp(data_comorb_prob['lnr_'+art_stat])*c.def_year #include offset (1 year or time to incidence date) 
        data_comorb_prob['prob_'+art_stat] = 1-np.exp(-data_comorb_prob['rate_'+art_stat]*c.dt) #convert the incidence rate to probability by time step

    return data_comorb_prob

def func_prob_update_cvd(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without cvd  to with cvd based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without cvd at t0"""

    #####initialize dataframe to save transition probability from without cvd to without cvd and with cvd
    data_prob = data_comorb[['moh_id']][data_comorb['earliest_cvd_dt']>t0].reset_index(drop=True) #only consider those without cvd at t0

    #####incorpoarate data_prob and data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without cvd at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_cvd_art,
        'noart':c.dic_prob_cvd_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #rate-converted probability based on art status instead of rates
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']
    
    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['cvd'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['cvd'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['cvd'][0][0] if x==1 else (c.dic_coeff_smk['cvd'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['cvd']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] 
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_prob['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['cvd']*data_prob['prob_w_comorb']
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_htn(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without htn to with htn based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without htn at t0"""

    #####initialize dataframe to save transition probability from without htn to without htn and with htn
    data_prob = data_comorb[['moh_id']][data_comorb['earliest_htn_dt']>t0].reset_index(drop=True) #only consider those without htn at t0

    #####incorpoarate data_cvd and data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without cvd at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_htn_art,
        'noart':c.dic_prob_htn_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #rate-converted probability based on art status instead of rates
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']
    
    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['htn'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['htn'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['htn'][0][0] if x==1 else (c.dic_coeff_smk['htn'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['htn']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] 
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_prob['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['htn']*data_prob['prob_w_comorb']
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_dm(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without dm to with dm based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without dm at t0"""

    #####initialize dataframe to save transition probability from without dm to without dm and with dm
    data_prob = data_comorb[['moh_id']][data_comorb['earliest_dm_dt']>t0].reset_index(drop=True) #only consider those without dm at t0, rename the dataframe as data_prob for simplification purpose for other diseases

    #####incorpoarate data_cvd and data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without cvd at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_dm_art,
        'noart':c.dic_prob_dm_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #rate-converted probability based on art status instead of rates
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']
    
    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['dm'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['dm'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['dm'][0][0] if x==1 else (c.dic_coeff_smk['dm'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['dm']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] 
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_prob['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['dm']*data_prob['prob_w_comorb']
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_oa(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without oa to with oa based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without oa at t0"""

    #####initialize dataframe to save transition probability from without oa to without oa and with oa
    data_prob = data_comorb[['moh_id']][data_comorb['earliest_oa_dt']>t0].reset_index(drop=True) #only consider those without dm at t0, rename the dataframe as data_prob for simplification purpose for other diseases
    
    #####incorpoarate data_prob and data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without cvd at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_oa_art,
        'noart':c.dic_prob_oa_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #rate-converted probability based on art status instead of rates
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['oa'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['oa'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['oa'][0][0] if x==1 else (c.dic_coeff_smk['oa'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['oa']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] 
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_prob['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['oa']*data_prob['prob_w_comorb']
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_copd(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without copd to with copd based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without copd at t0"""

    #####initialize dataframe to save transition probability from without copd to with copd
    data_copd = data_comorb[['moh_id']][data_comorb['earliest_copd_dt']>t0].reset_index(drop=True) #only consider those without cvd at t0
    
    #####incorpoarate data_copd and data_comorb_var
    data_copd = pd.merge(data_copd,data_comorb_var,how='left',on='moh_id') #only update the information for those without copd at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    data_hiv_copd = data_hiv[data_hiv['moh_id'].isin(data_copd['moh_id'])].reset_index(drop=True)
    arr_id_art = data_hiv_copd['moh_id'][(data_hiv_copd['moh_id'].isin(data_hiv_copd['moh_id']))&(data_hiv_copd[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv_copd['moh_id'][data_hiv_copd[str(t0)[:10]].isin(['A','U'])].values 
    dic_prob = {
        'art':c.dic_prob_copd_art,
        'noart':c.dic_prob_copd_noart,
    }
    data_copd = func_comorb_rate2prob(c,dic_prob,data_copd) 
    data_copd['prob_w_comorb'] = data_copd[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #rate-converted probability based on art status instead of rates
    assert data_copd[pd.isnull(data_copd['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_copd['prob_wo_comorb'] = 1-data_copd['prob_w_comorb']

    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_copd['coeff_pa'] = data_copd['pa'].apply(lambda x: c.dic_coeff_pa['copd'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_copd['coeff_alc'] = data_copd['alc'].apply(lambda x: c.dic_coeff_alc['copd'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_copd['coeff_smk'] = data_copd['smk'].apply(lambda x: c.dic_coeff_smk['copd'][0][0] if x==1 else (c.dic_coeff_smk['copd'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_copd['prob_w_comorb'] = c.dic_prob_comorb_coeff['copd']*data_copd['prob_w_comorb']*data_copd['coeff_pa']*data_copd['coeff_alc']*data_copd['coeff_smk'] 
        data_copd['prob_wo_comorb'] = 1-data_copd['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_copd['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['copd']*data_copd['prob_w_comorb']
    data_copd['prob_wo_comorb'] = 1-data_copd['prob_w_comorb']

    return data_copd[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_ckd(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without ckd to with ckd based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without ckd at t0"""

    #####initialize dataframe to save transition probability from without ckd to without ckd and with ckd
    data_prob = data_comorb[['moh_id']][data_comorb['earliest_ckd_dt']>t0].reset_index(drop=True) #only consider those without dm at t0, rename the dataframe as data_prob for simplification purpose for other diseases

    #####incorpoarate data_cvd and data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without cvd at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_ckd_art,
        'noart':c.dic_prob_ckd_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #rate-converted probability based on art status instead of rates
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']
    
    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['ckd'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['ckd'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['ckd'][0][0] if x==1 else (c.dic_coeff_smk['ckd'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['ckd']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] 
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_prob['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['ckd']*data_prob['prob_w_comorb']
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_cld(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without cld to with cld based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without cld at t0"""

    #####initialize dataframe to save transition probability from without cld to without cld and with cld
    data_prob = data_comorb[['moh_id']][data_comorb['earliest_cld_dt']>t0].reset_index(drop=True) #only consider those without dm at t0, rename the dataframe as data_prob for simplification purpose for other diseases

    #####incorpoarate data_cvd and data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without cvd at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_cld_art,
        'noart':c.dic_prob_cld_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #rate-converted probability based on art status instead of rates
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']
    
    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['cld'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['cld'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['cld'][0][0] if x==1 else (c.dic_coeff_smk['cld'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['cld']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] 
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_prob['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['cld']*data_prob['prob_w_comorb']
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_cancer(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without cancer to with cancer based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without cancer at t0"""

    #####initialize dataframe to save transition probability from without cancer to without cancer and with cancer
    data_prob = data_comorb[['moh_id']][data_comorb['earliest_cancer_dt']>t0].reset_index(drop=True) #only consider those without dm at t0, rename the dataframe as data_prob for simplification purpose for other diseases

    #####incorpoarate data_prob and data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without cancer at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_cancer_art,
        'noart':c.dic_prob_cancer_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #rate-converted probability based on art status instead of rates
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['cancer'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['cancer'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['cancer'][0][0] if x==1 else (c.dic_coeff_smk['cancer'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['cancer']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] 
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_prob['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['cancer']*data_prob['prob_w_comorb']
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_manx(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without manx to with manx based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without manx at t0"""

    #####initialize dataframe to save transition probability from without manx to without manx and with manx
    data_prob = data_comorb[['moh_id']][data_comorb['earliest_manx_dt']>t0].reset_index(drop=True) #only consider those without dm at t0, rename the dataframe as data_prob for simplification purpose for other diseases

    #####incorpoarate data_cvd and data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without manx at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_manx_art,
        'noart':c.dic_prob_manx_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #rate-converted probability based on art status instead of rates
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']
    
    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['manx'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['manx'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['manx'][0][0] if x==1 else (c.dic_coeff_smk['manx'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['manx']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] 
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_prob['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['manx']*data_prob['prob_w_comorb']
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_sczo(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without sczo to with sczo based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without sczo at t0"""

    #####initialize dataframe to save transition probability from without sczo to without sczo and with sczo
    data_prob = data_comorb[['moh_id']][data_comorb['earliest_sczo_dt']>t0].reset_index(drop=True) #only consider those without dm at t0, rename the dataframe as data_prob for simplification purpose for other diseases

    #####incorpoarate data_prob and data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without sczo at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_sczo_art,
        'noart':c.dic_prob_sczo_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #rate-converted probability based on art status instead of rates
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']
    
    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['sczo'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['sczo'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['sczo'][0][0] if x==1 else (c.dic_coeff_smk['sczo'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['sczo']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] 
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_prob['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['sczo']*data_prob['prob_w_comorb']
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_prsn(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update transition probability from without prsn to with prsn based on charactersitics, HIV status and other comorbidity status at t0,
    only take into account those without prsn at t0"""

    #####initialize dataframe to save transition probability from without prsn to without prsn and with prsn
    data_prob = data_comorb[['moh_id']][data_comorb['earliest_prsn_dt']>t0].reset_index(drop=True) #only consider those without dm at t0, rename the dataframe as data_prob for simplification purpose for other diseases

    #####incorpoarate data_prob and data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without prsn at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_prsn_art,
        'noart':c.dic_prob_prsn_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: x['prob_art'] if x['moh_id'] in arr_id_art else (x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #choose rate-converted probability based on art status instead of rates
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']
    
    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['prsn'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['prsn'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['prsn'][0][0] if x==1 else (c.dic_coeff_smk['prsn'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['prsn']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] 
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    ####introduce dic_prob_comorb_sens_coeff to adjust incidence probability for sensitivity analysis
    data_prob['prob_w_comorb'] = c.dic_prob_comorb_sens_coeff['prsn']*data_prob['prob_w_comorb']
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_prob_update_mortality(c,data_comorb,data_hiv,data_comorb_var,t0):
    """update probability of death based on charactersitics, HIV status and other comorbidity status at t0"""

    #####initialize dataframe to save probability of death
    data_prob = data_comorb[['moh_id']].copy() #need to include all alive ones, determine alive status in mp_microsim_comorb_inc...

    #####incorpoarate data_comorb_var
    data_prob = pd.merge(data_prob,data_comorb_var,how='left',on='moh_id') #only update the information for those without cancer at t0

    #####implement the probability based on statistical modeling for participants with ART and without ART initiation separately
    arr_id_art = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(c.arr_state_art))].values #introduce arr_state_art to specify states with/without ART initiation
    arr_id_noart = data_hiv['moh_id'][(data_hiv['moh_id'].isin(data_prob['moh_id']))&(data_hiv[str(t0)[:10]].isin(['A','U']))].values 
    dic_prob = {
        'art':c.dic_prob_dead_art,
        'noart':c.dic_prob_dead_noart,
    }
    data_prob = func_comorb_rate2prob(c,dic_prob,data_prob) 
    data_prob['prob_w_comorb'] = data_prob[['moh_id','prob_art','prob_noart']].apply(lambda x: c.coeff_prob_dead_ltfu_art*x['prob_art'] if x['moh_id'] in arr_id_art else (c.coeff_prob_dead_ltfu_noart*x['prob_noart'] if x['moh_id'] in arr_id_noart else np.nan),axis=1) #separate coeff_prob_dead_ltfu for art and noart to improve fitting for deaths
    assert data_prob[pd.isnull(data_prob['prob_w_comorb'])].shape[0]==0, 'Missing probability for some individuals'
    data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']
    
    #####introduce association between comorbidity and heath behavior parameters
    if 'pa' in list(data_comorb_var):
        data_prob['coeff_pa'] = data_prob['pa'].apply(lambda x: c.dic_coeff_pa['mortality'][0] if x==1 else 1.) #increased risk if physical inactive (pa=1)
        data_prob['coeff_alc'] = data_prob['alc'].apply(lambda x: c.dic_coeff_alc['mortality'][0] if x==1 else 1.) #increased risk if heavy drinker (alc=1)
        data_prob['coeff_smk'] = data_prob['smk'].apply(lambda x: c.dic_coeff_smk['mortality'][0][0] if x==1 else (c.dic_coeff_smk['mortality'][1][0] if x==2 else 1.)) #increased risk if current smoke (smk=1) or ever smoke (smk=2)
        data_prob['prob_w_comorb'] = c.dic_prob_comorb_coeff['dead']*data_prob['prob_w_comorb']*data_prob['coeff_pa']*data_prob['coeff_alc']*data_prob['coeff_smk'] #introduce dic_prob_comorb_coeff to adjust probability by comorbidity name
        data_prob['prob_wo_comorb'] = 1-data_prob['prob_w_comorb']

    return data_prob[['moh_id','prob_w_comorb','prob_wo_comorb']]

def func_art_logit2prob(dic_prob_stat,data_art_prob,rand):
    """estimate probability from diagnosis to ART initiation based on logistic regression, stratified by year of diagnosis
    rand=0/1 indicates whether to include random term for each individual (0=no)"""

    for stat_i in dic_prob_stat.keys():
        list_keys = list(dic_prob_stat[stat_i])
        list_keys.remove('beta0')
        data_art_prob['logit_'+stat_i] = pd.Series([dic_prob_stat[stat_i]['beta0'][0]]*data_art_prob.shape[0])
        list_keys_check = []
        for key_i in list_keys:
            list_keys_check.append(key_i)
            data_art_prob['logit_'+stat_i] += dic_prob_stat[stat_i][key_i][0]*data_art_prob[key_i]

        ####additional term based on whether to include random term for each individual
        if rand==1:
            data_art_prob['logit_'+stat_i] += data_art_prob['blup'] #blup term was generated based on diagnosis year in a2t_cat
        data_art_prob['prob_'+stat_i] = np.exp(data_art_prob['logit_'+stat_i])/(1+np.exp(data_art_prob['logit_'+stat_i]))
    
    return data_art_prob

def func_prob_update_a2t(c,data_char,t0,rand):
    """Update probability from diagnosis to ART initiation based on characteristics, t0 starts from 2008-01-01 as the beginning of each follow-up interval
    rand=0/1 indicates whether to include the random term for each individual"""

    #####update time-varying characteristics used to estimate probability a2t
    data_char['counter'] = np.floor((t0 - data_char['baseline_dt']).dt.total_seconds()/(60*60*24)/(c.def_year/2)) + 1 #only consider integer as counter

    #####initialize dataframe to save probability of a2t
    data_prob = data_char[['moh_id']].copy() 

    #####incorpoarate data_prob and data_char
    data_prob = pd.merge(data_prob,data_char,how='left',on='moh_id')

    #####estimate probability using different models based on year of diagnosis separately
    dic_prob = {
        'bf08':c.dic_prob_a2t_bf08,
        '0811':c.dic_prob_a2t_0811,
        'sin12':c.dic_prob_a2t_sin12,
    }
    data_prob = func_art_logit2prob(dic_prob,data_prob,rand) #0 indicate no blup/random term
    data_prob['prob_a2t'] = data_prob[['a2t_cat','prob_bf08','prob_0811','prob_sin12']].apply(lambda x: min(x['prob_'+x['a2t_cat']]*c.dic_coeff_adj_a2t[x['a2t_cat']],1.),axis=1) #introduce coeffecients to adjust transition probability to fit to historical values, be aware that it's likely to make the probability>1 while the coefficient increases
    data_prob['prob_a2a'] = 1-data_prob[['prob_a2t']]
    
    return data_prob[['moh_id','prob_a2t','prob_a2a']]

def func_art_multinomial2prob(dic_prob_off,dic_prob_on,data_art_prob,rand):
    """estimate probabilities for transitions between suppressed and unsuppressed (U_off/on) based on multinomial model"""

    ####update logit function for probabilities
    list_keys = list(dic_prob_off)
    list_keys.remove('beta0')
    data_art_prob['logit_off'] = pd.Series([dic_prob_off['beta0'][0]]*data_art_prob.shape[0])
    data_art_prob['logit_on'] = pd.Series([dic_prob_on['beta0'][0]]*data_art_prob.shape[0])
    list_keys_check = []
    for key_i in list_keys:
        list_keys_check.append(key_i)
        data_art_prob['logit_off'] += dic_prob_off[key_i][0]*data_art_prob[key_i]
        data_art_prob['logit_on'] += dic_prob_on[key_i][0]*data_art_prob[key_i]

    ####additional term based on whether to include random term for each individual
    if rand==1:
        data_art_prob['logit_off'] += data_art_prob['blup_off'] 
        data_art_prob['logit_on'] += data_art_prob['blup_on'] 

    ####estimate probabilities
    data_art_prob['prob_s'] = 1/(1+np.exp(data_art_prob['logit_off'])+np.exp(data_art_prob['logit_on']))
    data_art_prob['prob_off'] = data_art_prob['prob_s']*np.exp(data_art_prob['logit_off'])
    data_art_prob['prob_on'] = data_art_prob['prob_s']*np.exp(data_art_prob['logit_on'])

    return data_art_prob

def func_prob_update_trans_art(c,data_art,data_char,t0,rand):
    """Update probability from diagnosis to ART initiation based on characteristics, t0 starts from 2008-01-01 as the beginning of each follow-up interval
    rand=0/1 indicates whether to include the random term for each individual"""

    #####update time-varying characteristics 
    data_char_var = data_char.copy() #create a copy of data_char first so that u_off/on_dv can be updated at each step using pd.merge
    data_char_var['art_counter'] = np.floor((t0 - data_char_var['art_baseline_dt']).dt.total_seconds()/(60*60*24)/(c.def_year/2)) + 1 #art_counter is different from counter based on baseline_dt
    data_char_var = data_char_var.merge(data_art[['moh_id',str(t0)[:10]]],how='left',on='moh_id').rename(columns={str(t0)[:10]:'state'})
    data_char_var['u_off_dv'] = data_char_var['state'].apply(lambda x: 1 if x=='U_off' else 0) 
    data_char_var['u_on_dv'] = data_char_var['state'].apply(lambda x: 1 if x=='U_on' else 0)
    
    #####initialize dataframe to save probability of transitions after ART initiation
    data_prob = data_art[['moh_id']].copy()
    data_prob = pd.merge(data_prob,data_char_var,how='left',on='moh_id')

    ####update art transition probability
    data_prob = func_art_multinomial2prob(c.dic_prob_u_off,c.dic_prob_u_on,data_prob,rand)

    return data_prob #keep all data_char information including u_on/off_dv for further probability adjustment

def func_prob_adj_trans_art(c,data_prob_art):
    """Adjust probabilities of ART transitions based on 6 coefficients in c.dic_coeff_prob_art"""

    #####use coefficients to adjust each probability
    data_prob_art['sum'] = data_prob_art[['prob_s','prob_off','prob_on']].sum(axis=1)
    data_prob_art['prob_s_adj'] = data_prob_art[['prob_s','u_on_dv','u_off_dv']].apply(lambda x: x['prob_s']*c.dic_coeff_prob_art['off_s'] if x['u_off_dv']==1 else x['prob_s'],axis=1)
    data_prob_art['prob_off_adj'] = data_prob_art[['prob_off','u_on_dv','u_off_dv']].apply(lambda x: x['prob_off']*c.dic_coeff_prob_art['on_off'] if x['u_on_dv']==1 else (x['prob_off']*c.dic_coeff_prob_art['s_off'] if x['u_off_dv']==0 else x['prob_off']),axis=1)
    data_prob_art['prob_on_adj'] = data_prob_art[['prob_on','u_on_dv','u_off_dv']].apply(lambda x: x['prob_on']*c.dic_coeff_prob_art['on_on'] if x['u_on_dv']==1 else (x['prob_on']*c.dic_coeff_prob_art['off_on'] if x['u_off_dv']==1 else x['prob_on']*c.dic_coeff_prob_art['s_on']),axis=1)

    #####make the sum of probabilities as one
    data_prob_art['prob_off_adj2'] = data_prob_art[['prob_s_adj','prob_off_adj','prob_on_adj','u_on_dv','u_off_dv']].apply(lambda x: x['prob_off_adj']/(x['prob_off_adj']+x['prob_on_adj']) if x['u_off_dv']!=1 and x['prob_off_adj']+x['prob_on_adj']>=1 else x['prob_off_adj'],axis=1)
    data_prob_art['prob_on_adj2'] = data_prob_art[['prob_s_adj','prob_off_adj','prob_on_adj','u_on_dv','u_off_dv']].apply(lambda x: x['prob_on_adj']/(x['prob_off_adj']+x['prob_on_adj']) if x['u_off_dv']!=1 and x['prob_off_adj']+x['prob_on_adj']>=1 else (x['prob_on_adj']/(x['prob_s_adj']+x['prob_on_adj']) if x['u_off_dv']==1 and x['prob_s_adj']+x['prob_on_adj']>=1 else x['prob_on_adj']),axis=1)
    data_prob_art['prob_s_adj2'] = data_prob_art[['prob_s_adj','prob_off_adj','prob_on_adj','u_on_dv','u_off_dv']].apply(lambda x: x['prob_s_adj']/(x['prob_s_adj']+x['prob_on_adj']) if x['u_off_dv']==1 and x['prob_s_adj']+x['prob_on_adj']>=1 else x['prob_s_adj'],axis=1)
    data_prob_art['prob_s_adj3'] = data_prob_art[['prob_s_adj2','prob_off_adj2','prob_on_adj2','u_on_dv','u_off_dv']].apply(lambda x: max(1.-x['prob_off_adj2']-x['prob_on_adj2'],0.) if x['u_off_dv']!=1 else x['prob_s_adj2'],axis=1)
    data_prob_art['prob_off_adj3'] = data_prob_art[['prob_s_adj2','prob_off_adj2','prob_on_adj2','u_on_dv','u_off_dv']].apply(lambda x: max(1.-x['prob_s_adj2']-x['prob_on_adj2'],0.) if x['u_off_dv']==1 else x['prob_off_adj2'],axis=1)
    data_prob_art['prob_on_adj3'] = data_prob_art['prob_on_adj2'].copy()
    data_prob_art['sum_adj3'] = data_prob_art[['prob_s_adj3','prob_off_adj3','prob_on_adj3']].sum(axis=1)
    
    return data_prob_art

def func_micro_sim_outcomes(c,dict_cohort):
    """return to a dictionary with dataframes of annual outcomes for final results, such as age distribution among those diagnosed and distribution of comorbidity burden"""

    #####load simulated information of the cohort
    data_char_sim = dict_cohort['char'].copy()
    data_hiv_sim = dict_cohort['hiv'].copy()
    data_comorb_sim = dict_cohort['comorb'].copy()
    data_new_sim = dict_cohort['new'].copy() #if using mp_microsim_all_rand

    #####create time series and corresponding annual markers
    arr_step = np.arange(c.n_step+1) #start from step 0 and end at n_step, every two steps to get the end of the year
    arr_date = [c.t0+pd.to_timedelta(c.def_year*c.dt*step_i,unit='d') for step_i in arr_step] #keep time format for comparison purpose
    annual_date = arr_date[0::2]
    arr_year = np.arange(c.t0.year-1,c.t_end.year)
    assert (len(annual_date)==len(arr_year)), 'Inconsistent date and year lists'

    #####create dataframes for outcomes
    data_plwh_cas = pd.DataFrame([],columns=['date','year','sex','age']+['new_inc','new_diag','new_art','A','S','U_on','U_off','undiag','dead']) #incoporate age and sex information for each state
    data_comorb_inc = pd.DataFrame([],columns=['date','year','sex','age']+c.list_comorb) #among all diagnosed PLWH
    data_comorb_prev = pd.DataFrame([],columns=['date','year','sex','age']+c.list_comorb+['0_b','1_b','2_b','>=3_b','0_p','1_p','2_p','>=3_p','0_m','1_m','2_m','3_m'])
    c.list_phy_comorb = c.list_comorb[:-3]
    c.list_mental_comorb = c.list_comorb[-3:]
    data_comb1_both = pd.DataFrame([],columns=['date','year','n_dplwh','c','n_c','rank_both','rank_phys']) #dataframe to rank all 11 comorbidities 
    data_comb2_both = pd.DataFrame([],columns=['date','year','n_dplwh','cc','n_cc','rank_both','rank_phys']) #each row refers to one specific combination of 2 comorbidities
    data_comb1_both_agesex = pd.DataFrame([],columns=['date','year','sex','age','n_dplwh','c','n_c','rank_both','rank_phys']) #for age-sex specific subgroup
    data_comb2_both_agesex = pd.DataFrame([],columns=['date','year','sex','age','n_dplwh','cc','n_cc','rank_both','rank_phys'])
    dict_comb2_both_agesex = {sex_i:{agecat_j:pd.DataFrame([],columns=['year','n_dplwh','cc','n_cc','rank_both','rank_phys']) for agecat_j in c.dic_coprev_age_cat1} for sex_i in ['M','F']}
    
    #####initialize all combinations of two comorbidities
    list_comb_cc = list(combinations(c.list_comorb,2))

    #####estimates stratified by age and sex by the end of each year
    for i,date_i in enumerate(annual_date):

        #####focus on those alive by date_i
        data_char_sim_i = data_char_sim[data_char_sim['DOB']<=date_i].reset_index(drop=True) #focus on those born before date_i
        data_char_sim_i['age'] = (date_i-data_char_sim_i['DOB']).dt.total_seconds()/(60*60*24)/c.def_year
        data_char_sim_i['age_cat_i'] = data_char_sim_i['age'].apply(func_sort_byrange,args=(c.dic_age_cat,))

        ####comorbidity status
        for comorb_m in c.list_comorb:
            data_comorb_sim[comorb_m] = data_comorb_sim['earliest_'+comorb_m+'_dt'].apply(lambda x:0 if x>date_i else 1)

        ####derive outcomes by age and sex
        for j,sex_j in enumerate(['M','F']):
            for k,age_k in enumerate(c.dic_age_cat.keys()):

                ####only consider those alive by date_i, stratify by age and sex
                arr_id_jk = data_char_sim_i['moh_id_sim'][(data_char_sim_i['sex_at_birth_dv']==sex_j)&(data_char_sim_i['age_cat_i']==age_k)].values

                ####derive the number of PLWH in each state by age and sex
                dict_plwh_cas = {'date':str(date_i)[:10], 'year':arr_year[i], 'sex':sex_j, 'age':age_k}
                for stat_m in ['A','S','U_on','U_off']:
                    dict_plwh_cas[stat_m] = data_hiv_sim[(data_hiv_sim['moh_id_sim'].isin(arr_id_jk))&(data_hiv_sim[str(date_i)[:10]]==stat_m)].shape[0]
                if i>0:
                    t0 = arr_date[i*2-2]
                    t1 = arr_date[i*2-1]
                    t2 = arr_date[i*2]
                    new_diag1 = data_hiv_sim[(data_hiv_sim['moh_id_sim'].isin(arr_id_jk))&(data_hiv_sim[str(t0)[:10]]=='U')&(data_hiv_sim[str(t1)[:10]].isin(['A','S','U_on','U_off']))].shape[0] 
                    new_diag2 = data_hiv_sim[(data_hiv_sim['moh_id_sim'].isin(arr_id_jk))&(data_hiv_sim[str(t1)[:10]]=='U')&(data_hiv_sim[str(t2)[:10]].isin(['A','S','U_on','U_off']))].shape[0]
                    dict_plwh_cas['new_diag'] = new_diag1+new_diag2
                    new_art = data_char_sim_i[(data_char_sim_i['moh_id_sim'].isin(arr_id_jk))&(data_char_sim_i['FARVDT_sim']>=annual_date[i-1])&(data_char_sim_i['FARVDT_sim']<date_i)].shape[0]
                    dict_plwh_cas['new_art'] = new_art
                    dead1 = data_hiv_sim[(data_hiv_sim['moh_id_sim'].isin(arr_id_jk))&(data_hiv_sim[str(t0)[:10]]!='D')&(data_hiv_sim[str(t1)[:10]]=='D')].shape[0]
                    dead2 = data_hiv_sim[(data_hiv_sim['moh_id_sim'].isin(arr_id_jk))&(data_hiv_sim[str(t1)[:10]]!='D')&(data_hiv_sim[str(t2)[:10]]=='D')].shape[0]
                    dict_plwh_cas['dead'] = dead1+dead2
                data_plwh_cas = pd.concat([data_plwh_cas,pd.DataFrame(dict_plwh_cas,index=[0])],ignore_index=True)

                ####derive the number of comorb incident cases by age and sex
                dict_comorb_inc = {'date':str(date_i)[:10], 'year':arr_year[i], 'sex':sex_j, 'age':age_k} 
                if i>0:
                    for comorb_m in c.list_comorb:
                        dict_comorb_inc[comorb_m] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_jk))&(data_comorb_sim['earliest_'+comorb_m+'_dt']>data_comorb_sim['baseline_dt'])&(data_comorb_sim['earliest_'+comorb_m+'_dt']>annual_date[i-1])&(data_comorb_sim['earliest_'+comorb_m+'_dt']<=annual_date[i])].shape[0] #incidence as diagnosed after baseline_dt to distinguish from the prevalent cases
                data_comorb_inc = pd.concat([data_comorb_inc,pd.DataFrame(dict_comorb_inc,index=[0])],ignore_index=True)

                ####derive multicomorbidity status of PLWH, combined physical and mental and separated
                dict_comorb_prev = {'date':str(date_i)[:10],'year':arr_year[i], 'sex':sex_j, 'age':age_k}
                arr_id_dplwh_jk = data_hiv_sim['moh_id_sim'][(data_hiv_sim['moh_id_sim'].isin(arr_id_jk))&(data_hiv_sim[str(date_i)[:10]].isin(['A','S','U_on','U_off']))].values
                for comorb_m in c.list_comorb:
                    dict_comorb_prev[comorb_m] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim[comorb_m]==1)].shape[0]
                data_comorb_sim['n_both'] = data_comorb_sim[c.list_comorb].sum(axis=1)
                dict_comorb_prev['0_b'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_both']==0)].shape[0]
                dict_comorb_prev['1_b'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_both']==1)].shape[0]
                dict_comorb_prev['2_b'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_both']==2)].shape[0]
                dict_comorb_prev['>=3_b'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_both']>=3)].shape[0]
                data_comorb_sim['n_phys'] = data_comorb_sim[c.list_phy_comorb].sum(axis=1)
                dict_comorb_prev['0_p'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_phys']==0)].shape[0]
                dict_comorb_prev['1_p'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_phys']==1)].shape[0]
                dict_comorb_prev['2_p'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_phys']==2)].shape[0]
                dict_comorb_prev['>=3_p'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_phys']>=3)].shape[0]
                data_comorb_sim['n_mental'] = data_comorb_sim[c.list_mental_comorb].sum(axis=1)
                dict_comorb_prev['0_m'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_mental']==0)].shape[0]
                dict_comorb_prev['1_m'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_mental']==1)].shape[0]
                dict_comorb_prev['2_m'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_mental']==2)].shape[0]
                dict_comorb_prev['3_m'] = data_comorb_sim[(data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk))&(data_comorb_sim['n_mental']>=3)].shape[0] 
                data_comorb_prev = pd.concat([data_comorb_prev,pd.DataFrame(dict_comorb_prev,index=[0])],ignore_index=True)

        #####add overall new_inc and undiag which cannot be stratified by age and sex
        dict_plwh_cas = {'date':str(date_i)[:10], 'year':arr_year[i], 'sex':'all', 'age':'all'}
        for stat_m in ['A','S','U_on','U_off']:
            dict_plwh_cas[stat_m] = data_hiv_sim[(data_hiv_sim[str(date_i)[:10]]==stat_m)].shape[0]
        dict_plwh_cas['undiag'] = data_new_sim['undiag'][data_new_sim['date']==str(date_i)[:10]].values[0]
        if i>0:
            t0 = arr_date[i*2-2]
            t1 = arr_date[i*2-1]
            t2 = arr_date[i*2]
            dict_plwh_cas['new_inc'] = data_new_sim['new_inc'][data_new_sim['date'].isin([str(t1)[:10],str(t2)[:10]])].sum()
            for stat_m in ['new_diag','new_art','dead']:
                dict_plwh_cas[stat_m] = data_plwh_cas[stat_m][(data_plwh_cas['year']==arr_year[i])].sum()
        data_plwh_cas = pd.concat([data_plwh_cas,pd.DataFrame(dict_plwh_cas,index=[0])],ignore_index=True)  

        ####add ranking and prevalence of combination of two comorbidities
        arr_id_dplwh = data_hiv_sim['moh_id_sim'][data_hiv_sim[str(date_i)[:10]].isin(['A','S','U_on','U_off'])].values
        data_comorb_cc = data_comorb_sim[data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh)].reset_index(drop=True) 
        n_dplwh = data_plwh_cas[['A','S','U_on','U_off']][(data_plwh_cas['date']==str(date_i)[:10])&(data_plwh_cas['sex']=='all')].sum(axis=1).values[0] #add number of diagnosed PLWH on date_i for percentage estimation
        data_comb1_both_i = pd.DataFrame([],columns=['date','year','n_dplwh','c','n_c'])
        dict_comb1_both_i = {'date':str(date_i)[:10],'year':arr_year[i],'n_dplwh':n_dplwh}
        for c_j in c.list_comorb:
            dict_comb1_both_i['c'] = c_j
            dict_comb1_both_i['n_c'] = data_comorb_cc[(data_comorb_cc['moh_id_sim'].isin(arr_id_dplwh))&(data_comorb_cc[c_j]==1)].shape[0]
            data_comb1_both_i = pd.concat([data_comb1_both_i,pd.DataFrame(dict_comb1_both_i,index=[0])],ignore_index=True)
        data_comb1_both_i['rank_both'] = data_comb1_both_i['n_c'].rank(method='min',ascending=False)
        data_comb1_phys_i = data_comb1_both_i[['c','n_c']][data_comb1_both_i['c'].isin(['manx','prsn','sczo'])==False].reset_index(drop=True)
        data_comb1_phys_i['rank_phys'] = data_comb1_phys_i['n_c'].rank(method='min',ascending=False)
        data_comb1_both_i = data_comb1_both_i.merge(data_comb1_phys_i[['c','rank_phys']],how='left',on='c')
        data_comb1_both = pd.concat([data_comb1_both,data_comb1_both_i],ignore_index=True)
        data_comb2_both_i = pd.DataFrame([],columns=['date','year','n_dplwh','cc','n_cc'])
        dict_comb2_both_i = {'date':str(date_i)[:10],'year':arr_year[i],'n_dplwh':n_dplwh}
        for cc_j in list_comb_cc:
            dict_comb2_both_i['cc'] = cc_j[0]+'|'+cc_j[1]
            dict_comb2_both_i['n_cc'] = data_comorb_cc[(data_comorb_cc['moh_id_sim'].isin(arr_id_dplwh))&(data_comorb_cc[cc_j[0]]==1)&(data_comorb_cc[cc_j[1]]==1)].shape[0]
            data_comb2_both_i = pd.concat([data_comb2_both_i,pd.DataFrame(dict_comb2_both_i,index=[0])],ignore_index=True)
        data_comb2_both_i['rank_both'] = data_comb2_both_i['n_cc'].rank(method='min',ascending=False)
        data_comb2_phys_i = data_comb2_both_i[['cc','n_cc']][data_comb2_both_i['cc'].str.contains('manx|prsn|sczo')==False].reset_index(drop=True)
        data_comb2_phys_i['rank_phys'] = data_comb2_phys_i['n_cc'].rank(method='min',ascending=False)
        data_comb2_both_i = data_comb2_both_i.merge(data_comb2_phys_i[['cc','rank_phys']],how='left',on='cc')
        data_comb2_both = pd.concat([data_comb2_both,data_comb2_both_i],ignore_index=True)

        ####add ranking and prevalence of combination of two comorbidities by age and sex category
        data_char_sim_i['agecat_i'] = data_char_sim_i['age'].apply(func_sort_byrange,args=(c.dic_coprev_age_cat1,)) #focus on <50,50-60,60-70 and >=70 subgroup
        for sex_j in ['M','F']:
            for age_k in c.dic_coprev_age_cat1.keys():
                arr_id_age_k= data_char_sim_i['moh_id_sim'][(data_char_sim_i['sex_at_birth_dv']==sex_j)&(data_char_sim_i['agecat_i']==age_k)].values
                arr_id_dplwh_jk = data_hiv_sim['moh_id_sim'][(data_hiv_sim[str(date_i)[:10]].isin(['A','S','U_on','U_off']))&(data_hiv_sim['moh_id_sim'].isin(arr_id_age_k))].values
                n_dplwh = len(arr_id_dplwh_jk)
                data_comorb_sim_jk = data_comorb_sim[data_comorb_sim['moh_id_sim'].isin(arr_id_dplwh_jk)].reset_index(drop=True)
                data_comb1_both_jk = pd.DataFrame([],columns=['date','year','sex','age','n_dplwh','c','n_c'])
                dict_comb1_both_jk = {'date':str(date_i)[:10],'year':arr_year[i],'sex':sex_j,'age':age_k,'n_dplwh':n_dplwh}
                for c_j in c.list_comorb:
                    dict_comb1_both_jk['c'] = c_j
                    dict_comb1_both_jk['n_c'] = data_comorb_sim_jk[(data_comorb_sim_jk[c_j]==1)].shape[0] 
                    data_comb1_both_jk = pd.concat([data_comb1_both_jk,pd.DataFrame(dict_comb1_both_jk,index=[0])],ignore_index=True)
                data_comb1_both_jk['rank_both'] = data_comb1_both_jk['n_c'].rank(method='min',ascending=False)
                data_comb1_phys_jk = data_comb1_both_jk[['c','n_c']][data_comb1_both_jk['c'].isin(['manx','prsn','sczo'])==False].reset_index(drop=True)
                data_comb1_phys_jk['rank_phys'] = data_comb1_phys_jk['n_c'].rank(method='min',ascending=False)
                data_comb1_both_jk = data_comb1_both_jk.merge(data_comb1_phys_jk[['c','rank_phys']],how='left',on='c')
                data_comb1_both_agesex = pd.concat([data_comb1_both_agesex,data_comb1_both_jk],ignore_index=True)
                data_comb2_both_jk = pd.DataFrame([],columns=['date','year','sex','age','n_dplwh','cc','n_cc'])
                dict_comb2_both_jk = {'date':str(date_i)[:10],'year':arr_year[i],'sex':sex_j,'age':age_k,'n_dplwh':n_dplwh}
                for cc_j in list_comb_cc:
                    dict_comb2_both_jk['cc'] = cc_j[0]+'|'+cc_j[1]
                    dict_comb2_both_jk['n_cc'] = data_comorb_sim_jk[(data_comorb_sim_jk[cc_j[0]]==1)&(data_comorb_sim_jk[cc_j[1]]==1)].shape[0]
                    data_comb2_both_jk = pd.concat([data_comb2_both_jk,pd.DataFrame(dict_comb2_both_jk,index=[0])],ignore_index=True)
                data_comb2_both_jk['rank_both'] = data_comb2_both_jk['n_cc'].rank(method='min',ascending=False)
                data_comb2_phys_jk = data_comb2_both_jk[['cc','n_cc']][data_comb2_both_jk['cc'].str.contains('manx|prsn|sczo')==False].reset_index(drop=True)
                data_comb2_phys_jk['rank_phys'] = data_comb2_phys_jk['n_cc'].rank(method='min',ascending=False)
                data_comb2_both_jk = data_comb2_both_jk.merge(data_comb2_phys_jk[['cc','rank_phys']],how='left',on='cc')
                data_comb2_both_agesex = pd.concat([data_comb2_both_agesex,data_comb2_both_jk],ignore_index=True)
                dict_comb2_both_agesex[sex_j][age_k] = pd.concat([dict_comb2_both_agesex[sex_j][age_k],data_comb2_both_jk.drop(columns=['date','sex','age'])],ignore_index=True)

    #####build dictionary to save final outcomes
    dict_outcomes = {
        'hiv':data_plwh_cas, #hiv-related outcomes
        'comorb_inc':data_comorb_inc,
        'comorb_prev':data_comorb_prev, #put prev of phys and mental and both together
        'comb1_both':data_comb1_both, #add comorbidity prevalence with ranking
        'comb2_both':data_comb2_both, #add top 3 prevalent combination of two comorbidities
        'comb1_both_agesex':data_comb1_both_agesex.drop(columns=['date']), #data_comb1_both_agesex,
        'comb2_both_agesex':dict_comb2_both_agesex, #data_comb2_both_agesex.drop.drop(columns=['date']),
    }

    return dict_outcomes 

def func_micro_outcomes_target(c,dict_cohort):
    """return to a dictionary with dataframes of annual outcomes for calibration"""

    #####load simulated information of the cohort
    data_char_sim = dict_cohort['char'].copy()
    data_hiv_sim = dict_cohort['hiv'].copy()
    data_comorb_sim = dict_cohort['comorb'].copy()

    #####create time series and corresponding annual markers
    arr_step = np.arange(c.n_step+1) #start from step 0 and end at n_step, every two steps to get the end of the year
    arr_date = [c.t0+pd.to_timedelta(c.def_year*c.dt*step_i,unit='d') for step_i in arr_step] 
    annual_date = arr_date[0::2]
    arr_year = np.arange(c.t0.year-1,c.t_end.year)
    assert (len(annual_date)==len(arr_year)), 'Inconsistent date and year lists'

    #####create dataframes for outcomes
    data_plwh_cas = pd.DataFrame([],columns=['date','year']+['new_diag','new_art','A','S','U_on','U_off','S2Uoff','Uoff2S','S2Uon','Uon2S','Uon2Uoff','Uoff2Uon','dead','A2D','S2D','Uon2D','Uoff2D','A2S','A2Uon','A2Uoff']) 
    data_comorb_inc = pd.DataFrame([],columns=['date','year']+c.list_comorb)
    data_comorb_check = data_comorb_sim[data_comorb_sim['micro_baseline_comorb']=='known'].reset_index(drop=True) 
    for i,date_i in enumerate(annual_date):

        ####create dataframe related to HIV cascade
        dict_plwh_cas = {'date':str(date_i)[:10],'year':arr_year[i]}
        for stat_j in ['A','S','U_on','U_off']:
            dict_plwh_cas[stat_j] = data_hiv_sim[data_hiv_sim[str(date_i)[:10]]==stat_j].shape[0]
        if i>0:
            t0 = arr_date[i*2-2]
            t1 = arr_date[i*2-1]
            t2 = arr_date[i*2]
            new_diag1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U')&(data_hiv_sim[str(t1)[:10]].isin(['A','S','U_on','U_off']))].shape[0] #D is not included as possible deaths happened before ART diagnosis if undiagnosed population was set at the beginning
            new_diag2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U')&(data_hiv_sim[str(t2)[:10]].isin(['A','S','U_on','U_off']))].shape[0]
            dict_plwh_cas['new_diag'] = new_diag1+new_diag2
            new_art = data_char_sim[(data_char_sim['FARVDT_sim']>=annual_date[i-1])&(data_char_sim['FARVDT_sim']<date_i)].shape[0]
            dict_plwh_cas['new_art'] = new_art
            for diag_cat in ['bf08','0811','sin12']: #separate art initiation by a2t_cat/earliest_HIV so that dic_coeff_adj_a2t can be adjusted properly
                arr_id_cat = data_char_sim['moh_id_sim'][data_char_sim['a2t_cat']==diag_cat].values
                new_art_cat = data_char_sim[(data_char_sim['moh_id_sim'].isin(arr_id_cat))&(data_char_sim['FARVDT_sim']>=annual_date[i-1])&(data_char_sim['FARVDT_sim']<date_i)].shape[0] #use simulated FARVDT_sim to count new ART initiation similar to hist_target, which is higher than estimated transitions as FARVDT_sim was determined by baseline_dt and counter, similar to simulating a2t alone
                dict_plwh_cas['new_art '+diag_cat] = new_art_cat
            dead1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]!='D')&(data_hiv_sim[str(t1)[:10]]=='D')].shape[0]
            dead2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]!='D')&(data_hiv_sim[str(t2)[:10]]=='D')].shape[0]
            dict_plwh_cas['dead'] = dead1+dead2
            n_s2off1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='S')&(data_hiv_sim[str(t1)[:10]]=='U_off')].shape[0]
            n_s2off2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='S')&(data_hiv_sim[str(t2)[:10]]=='U_off')].shape[0]
            n_off2s1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_off')&(data_hiv_sim[str(t1)[:10]]=='S')].shape[0]
            n_off2s2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_off')&(data_hiv_sim[str(t2)[:10]]=='S')].shape[0]
            n_s2on1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='S')&(data_hiv_sim[str(t1)[:10]]=='U_on')].shape[0]
            n_s2on2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='S')&(data_hiv_sim[str(t2)[:10]]=='U_on')].shape[0]
            n_on2s1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_on')&(data_hiv_sim[str(t1)[:10]]=='S')].shape[0]
            n_on2s2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_on')&(data_hiv_sim[str(t2)[:10]]=='S')].shape[0]
            n_on2off1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_on')&(data_hiv_sim[str(t1)[:10]]=='U_off')].shape[0]
            n_on2off2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_on')&(data_hiv_sim[str(t2)[:10]]=='U_off')].shape[0]
            n_off2on1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_off')&(data_hiv_sim[str(t1)[:10]]=='U_on')].shape[0]
            n_off2on2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_off')&(data_hiv_sim[str(t2)[:10]]=='U_on')].shape[0]
            dict_plwh_cas['S2Uoff'] = n_s2off1+n_s2off2
            dict_plwh_cas['Uoff2S'] = n_off2s1+n_off2s2
            dict_plwh_cas['S2Uon'] = n_s2on1+n_s2on2
            dict_plwh_cas['Uon2S'] = n_on2s1+n_on2s2
            dict_plwh_cas['Uon2Uoff'] = n_on2off1+n_on2off2
            dict_plwh_cas['Uoff2Uon'] = n_off2on1+n_off2on2
            n_a2d1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='A')&(data_hiv_sim[str(t1)[:10]]=='D')].shape[0]
            n_a2d2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='A')&(data_hiv_sim[str(t2)[:10]]=='D')].shape[0]
            n_s2d1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='S')&(data_hiv_sim[str(t1)[:10]]=='D')].shape[0]
            n_s2d2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='S')&(data_hiv_sim[str(t2)[:10]]=='D')].shape[0]
            n_on2d1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_on')&(data_hiv_sim[str(t1)[:10]]=='D')].shape[0]
            n_on2d2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_on')&(data_hiv_sim[str(t2)[:10]]=='D')].shape[0]
            n_off2d1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_off')&(data_hiv_sim[str(t1)[:10]]=='D')].shape[0]
            n_off2d2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_off')&(data_hiv_sim[str(t2)[:10]]=='D')].shape[0]
            dict_plwh_cas['A2D'] = n_a2d1+n_a2d2
            dict_plwh_cas['S2D'] = n_s2d1+n_s2d2
            dict_plwh_cas['Uon2D'] = n_on2d1+n_on2d2
            dict_plwh_cas['Uoff2D'] = n_off2d1+n_off2d2
            n_a2s1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='A')&(data_hiv_sim[str(t1)[:10]]=='S')].shape[0]
            n_a2s2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='A')&(data_hiv_sim[str(t2)[:10]]=='S')].shape[0]
            n_a2on1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='A')&(data_hiv_sim[str(t1)[:10]]=='U_on')].shape[0]
            n_a2on2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='A')&(data_hiv_sim[str(t2)[:10]]=='U_on')].shape[0]
            n_a2off1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='A')&(data_hiv_sim[str(t1)[:10]]=='U_off')].shape[0]
            n_a2off2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='A')&(data_hiv_sim[str(t2)[:10]]=='U_off')].shape[0]
            dict_plwh_cas['A2S'] = n_a2s1+n_a2s2
            dict_plwh_cas['A2Uon'] = n_a2on1+n_a2on2
            dict_plwh_cas['A2Uoff'] = n_a2off1+n_a2off2
        data_plwh_cas = pd.concat([data_plwh_cas,pd.DataFrame(dict_plwh_cas,index=[0])],ignore_index=True)

        ####create dataframe for comorbidity incidence
        dict_comorb_inc = {'date':date_i,'year':arr_year[i]} #keep full date information which is related to derivation of incidence
        if i>0:
            for comorb_j in c.list_comorb:
                dict_comorb_inc[comorb_j] = data_comorb_check[(data_comorb_check['earliest_'+comorb_j+'_dt']>data_comorb_check['baseline_dt'])&(data_comorb_check['earliest_'+comorb_j+'_dt']>annual_date[i-1])&(data_comorb_check['earliest_'+comorb_j+'_dt']<=annual_date[i])].shape[0] #incidence as diagnosed after baseline_dt to distinguish from the prevalent cases
        data_comorb_inc = pd.concat([data_comorb_inc,pd.DataFrame(dict_comorb_inc,index=[0])],ignore_index=True)

    dict_outcomes = {
        'hiv':data_plwh_cas,
        'comorb':data_comorb_inc,
    }
    
    return dict_outcomes

def func_micro_outcomes_target_ci(c,list_outcomes):
    """return to a dataframes with summary of each outcomes and corresponding credible interval"""

    ####separate hiv and comorb results
    list_hiv = [result['hiv'] for result in list_outcomes]
    list_comorb = [result['comorb'] for result in list_outcomes]
    arr_year = list_hiv[0]['year'].values

    ####initialize dictionary of dataframes for final outcomes
    dict_summary = {}

    ####summarize hiv-related outcomes/cum. outcomes by mean, median and 2.5/97.5 percentiles
    data_summary = list_outcomes[0]['hiv'][['date','year']].copy()
    for key_i in ['A','S','U_on','U_off']:
        arr_sim = np.array([data_k[key_i].values for data_k in list_hiv])
        data_summary[key_i+' avg'] = np.mean(arr_sim,axis=0)
        data_summary[key_i+ ' med'] = np.percentile(arr_sim,50,axis=0)
        data_summary[key_i+' lb'] = np.percentile(arr_sim,2.5,axis=0)
        data_summary[key_i+' ub'] = np.percentile(arr_sim,97.5,axis=0)
    dict_key_cas = {
        'diag':['A','S','U_on','U_off'],
        'art_ever':['S','U_on','U_off'],
        'art':['S','U_on'],
    }
    for key_i in dict_key_cas.keys():
        arr_sim = np.array([data_k[dict_key_cas[key_i]].sum(axis=1).values for data_k in list_hiv])
        data_summary[key_i+' avg'] = np.mean(arr_sim,axis=0)
        data_summary[key_i+ ' med'] = np.percentile(arr_sim,50,axis=0)
        data_summary[key_i+' lb'] = np.percentile(arr_sim,2.5,axis=0)
        data_summary[key_i+' ub'] = np.percentile(arr_sim,97.5,axis=0)
    for key_i in ['new_art','new_art bf08','new_art 0811','new_art sin12','dead','S2Uoff','Uoff2S','S2Uon','Uon2S','Uon2Uoff','Uoff2Uon','A2D','S2D','Uon2D','Uoff2D','A2S','A2Uon','A2Uoff']: 
        data_summary_i = pd.DataFrame([]) 
        arr_sim = np.array([data_k[key_i].values for data_k in list_hiv])
        data_summary_i[key_i+' avg'] = np.mean(arr_sim,axis=0)
        data_summary_i[key_i+ ' med'] = np.percentile(arr_sim,50,axis=0)
        data_summary_i[key_i+' lb'] = np.percentile(arr_sim,2.5,axis=0)
        data_summary_i[key_i+' ub'] = np.percentile(arr_sim,97.5,axis=0)

        arr_cum_sim = np.array([data_k[key_i][data_k['year']>arr_year[0]].cumsum() for data_k in list_hiv])
        arr_cum_sim = np.hstack(((np.ones(arr_cum_sim.shape[0])*np.nan).reshape(-1,1),arr_cum_sim))
        data_summary_i['cum '+key_i+' avg'] = np.mean(arr_cum_sim,axis=0)
        data_summary_i['cum '+key_i+ ' med'] = np.percentile(arr_cum_sim,50,axis=0)
        data_summary_i['cum '+key_i+ ' lb'] = np.percentile(arr_cum_sim,2.5,axis=0)
        data_summary_i['cum '+key_i+ ' ub'] = np.percentile(arr_cum_sim,97.5,axis=0)
        data_summary = pd.concat([data_summary,data_summary_i],axis=1)
    dict_summary['hiv'] = data_summary.copy()

    ####summarize hiv-related outcomes/cum. outcomes by mean, median and 2.5/97.5 percentiles
    data_summary = list_outcomes[0]['hiv'][['date','year']].copy()
    for key_i in c.list_comorb:
        arr_sim = np.array([data_k[key_i].values for data_k in list_comorb])
        data_summary[key_i+' avg'] = np.mean(arr_sim,axis=0)
        data_summary[key_i+ ' med'] = np.percentile(arr_sim,50,axis=0)
        data_summary[key_i+' lb'] = np.percentile(arr_sim,2.5,axis=0)
        data_summary[key_i+' ub'] = np.percentile(arr_sim,97.5,axis=0)

        arr_cum_sim = np.array([data_k[key_i].cumsum() for data_k in list_comorb]) #.cumsum() can avoid NaN at the beginning
        data_summary['cum '+key_i+' avg'] = np.mean(arr_cum_sim,axis=0)
        data_summary['cum '+key_i+ ' med'] = np.percentile(arr_cum_sim,50,axis=0)
        data_summary['cum '+key_i+ ' lb'] = np.percentile(arr_cum_sim,2.5,axis=0)
        data_summary['cum '+key_i+ ' ub'] = np.percentile(arr_cum_sim,97.5,axis=0)
    dict_summary['comorb'] = data_summary.copy()

    return dict_summary

def func_micro_1outcome_ci(c,list_outcomes,key_outcome,arg_inc):
    """return to dataframe containing average/median of the specific outcome (key_outcome) and credible interval
    arg_inc=inc/prev indicates whether including cumulative annual estimates """

    arr_year = list_outcomes[0]['year'].values
    data_summary = list_outcomes[0][['year']].copy()
    assert (arg_inc in ['inc','prev']), "Double check the argument for whether to include cumulative annual estimates"
    arr_sim = np.array([data_k[key_outcome].values for data_k in list_outcomes])
    data_summary[key_outcome+' avg'] = np.mean(arr_sim,axis=0)
    data_summary[key_outcome+ ' med'] = np.percentile(arr_sim,50,axis=0)
    data_summary[key_outcome+' lb'] = np.percentile(arr_sim,2.5,axis=0)
    data_summary[key_outcome+' ub'] = np.percentile(arr_sim,97.5,axis=0)
    if arg_inc=='inc':
        arr_cum_sim = np.array([data_k[key_outcome][data_k['year']>arr_year[0]].cumsum() for data_k in list_outcomes])
        arr_cum_sim = np.hstack(((np.ones(arr_cum_sim.shape[0])*np.nan).reshape(-1,1),arr_cum_sim))
        data_summary['cum '+key_outcome+' avg'] = np.mean(arr_cum_sim,axis=0)
        data_summary['cum '+key_outcome+ ' med'] = np.percentile(arr_cum_sim,50,axis=0)
        data_summary['cum '+key_outcome+ ' lb'] = np.percentile(arr_cum_sim,2.5,axis=0)
        data_summary['cum '+key_outcome+ ' ub'] = np.percentile(arr_cum_sim,97.5,axis=0)

    return data_summary

def func_micro_outcomes_hiv_ci(c,list_hiv):
    """return to dataframe summarizing credible interval for HIV-related outcomes for final simulation
    list_hiv should have one row for each year, for certain age-sex subgroup or the whole population"""

    ####separate hiv and comorb results
    arr_year = list_hiv[0]['year'].values

    ####summarize hiv-related outcomes/cum. outcomes by mean, median and 2.5/97.5 percentiles
    data_summary = list_hiv[0][['date','year']].copy()
    for key_i in ['A','S','U_on','U_off','undiag']:
        arr_sim = np.array([data_k[key_i].values for data_k in list_hiv])
        data_summary[key_i+' avg'] = np.mean(arr_sim,axis=0)
        data_summary[key_i+ ' med'] = np.percentile(arr_sim,50,axis=0)
        data_summary[key_i+' lb'] = np.percentile(arr_sim,2.5,axis=0)
        data_summary[key_i+' ub'] = np.percentile(arr_sim,97.5,axis=0)
    dict_key_cas = {
        'diag':['A','S','U_on','U_off'],
        'art_ever':['S','U_on','U_off'],
        'art':['S','U_on'],
    }
    for key_i in dict_key_cas.keys():
        arr_sim = np.array([data_k[dict_key_cas[key_i]].sum(axis=1).values for data_k in list_hiv])
        data_summary[key_i+' avg'] = np.mean(arr_sim,axis=0)
        data_summary[key_i+ ' med'] = np.percentile(arr_sim,50,axis=0)
        data_summary[key_i+' lb'] = np.percentile(arr_sim,2.5,axis=0)
        data_summary[key_i+' ub'] = np.percentile(arr_sim,97.5,axis=0)
    for key_i in ['new_inc','new_diag','new_art','dead',]: 
        data_summary_i = pd.DataFrame([]) 
        
        arr_sim = np.array([data_k[key_i].values for data_k in list_hiv])
        data_summary_i[key_i+' avg'] = np.mean(arr_sim,axis=0)
        data_summary_i[key_i+ ' med'] = np.percentile(arr_sim,50,axis=0)
        data_summary_i[key_i+' lb'] = np.percentile(arr_sim,2.5,axis=0)
        data_summary_i[key_i+' ub'] = np.percentile(arr_sim,97.5,axis=0)

        arr_cum_sim = np.array([data_k[key_i][(data_k['year']>arr_year[0])&(data_k['sex']=='all')].cumsum() for data_k in list_hiv])
        arr_cum_sim = np.hstack(((np.ones(arr_cum_sim.shape[0])*np.nan).reshape(-1,1),arr_cum_sim))
        data_summary_i['cum '+key_i+' avg'] = np.mean(arr_cum_sim,axis=0)
        data_summary_i['cum '+key_i+ ' med'] = np.percentile(arr_cum_sim,50,axis=0)
        data_summary_i['cum '+key_i+ ' lb'] = np.percentile(arr_cum_sim,2.5,axis=0)
        data_summary_i['cum '+key_i+ ' ub'] = np.percentile(arr_cum_sim,97.5,axis=0)
        data_summary = pd.concat([data_summary,data_summary_i],axis=1)

    return data_summary

def func_micro_outcomes_age_ci(c,list_hiv_all):
    """return to dataframe summarizing age distribution of PLWH in certain subgroup"""

    ####separate hiv and comorb results
    arr_date = list_hiv_all[0]['date'].unique()
    arr_year = list_hiv_all[0]['year'].unique()

    ####create dataframe for age distribution among PLWH on ART
    data_art_age = pd.DataFrame.from_dict({'date':arr_date,'year':arr_year})
    arr_sim_all = np.array([data_k[['S','U_on']][(data_k['age']=='all')].sum(axis=1).values for data_k in list_hiv_all],dtype='f') #age=all and sex=all simultaneously
    for age_i in c.dic_age_cat.keys():
        list_sim = [data_k[['S','U_on']][(data_k['age']==age_i)].sum(axis=1).values for data_k in list_hiv_all]
        arr_sim = np.array([np.add.reduceat(sim_i, np.arange(0, len(sim_i), 2)) for sim_i in list_sim],dtype='f')
        arr_p_sim = arr_sim/arr_sim_all
        data_art_age[age_i+' med'] = np.percentile(arr_p_sim,50,axis=0)
        data_art_age[age_i+' lb'] = np.percentile(arr_p_sim,2.5,axis=0)
        data_art_age[age_i+' ub'] = np.percentile(arr_p_sim,97.5,axis=0)

    return data_art_age

def func_hist_cali_target_dtp(c,data_hiv_dtp):
    """return to a dataframe for hiv validation target using DTP data"""

    #####create time series and corresponding annual markers
    c.t0_dtp = pd.Timestamp('2008-01-01') 
    c.t_end_dtp = pd.Timestamp('2020-01-01') 
    c.n_step_dtp = int((c.t_end_dtp.year-c.t0_dtp.year)/c.dt)
    arr_step = np.arange(c.n_step_dtp+1) 
    arr_date = [c.t0_dtp+pd.to_timedelta(c.def_year*c.dt*step_i,unit='d') for step_i in arr_step] 
    annual_date = arr_date[0::2]
    annual_date[-1] = pd.Timestamp('2019-12-31') #reset the last day as the cutoff date of DTP as no data available beyond it
    arr_year = np.arange(c.t0_dtp.year-1,c.t_end_dtp.year)

    #####create dataframe for outcomes
    data_plwh_cas = pd.DataFrame([],columns=['date','year']+['new_art','S','U_on','U_off','dead'])
    data_art_age = pd.DataFrame([],columns=['date','year']+list(c.dic_age_cat)) 
    for i,date_i in enumerate(annual_date):

        ####create dataframe related to HIV cascade
        dict_plwh_cas = {'date':str(date_i)[:10],'year':arr_year[i]}
        for stat_j in ['S','U_on','U_off']:
            dict_plwh_cas[stat_j] = data_hiv_dtp[data_hiv_dtp[str(date_i)[:10]]==stat_j].shape[0]
        dict_plwh_cas['art_ever'] = dict_plwh_cas['S']+dict_plwh_cas['U_on']+dict_plwh_cas['U_off']
        dict_plwh_cas['art'] = dict_plwh_cas['S']+dict_plwh_cas['U_on']
        if i>0:
            t0 = arr_date[i*2-2]
            t1 = arr_date[i*2-1]
            t2 = arr_date[i*2]
            new_art = data_hiv_dtp[(data_hiv_dtp['FARVDT']>=annual_date[i-1])&(data_hiv_dtp['FARVDT']<date_i)].shape[0] 
            dead = data_hiv_dtp[(data_hiv_dtp['end_fu_dt']>=annual_date[i-1])&(data_hiv_dtp['end_fu_dt']<date_i)].shape[0]
            dict_plwh_cas['new_art'] = new_art 
            dict_plwh_cas['dead'] = dead 
        data_plwh_cas = pd.concat([data_plwh_cas,pd.DataFrame(dict_plwh_cas,index=[0])],ignore_index=True)

        ####estimate age distribution among PLWH on ART on date_i
        data_hiv_dtp['age_t'] = (date_i-data_hiv_dtp['BIRTHDATE']).dt.total_seconds()/(60*60*24)/c.def_year
        dict_art_age = {'date':str(date_i)[:10],'year':arr_year[i]}
        for age_j in c.dic_age_cat.keys():
            dict_art_age[age_j] = data_hiv_dtp[(data_hiv_dtp[str(date_i)[:10]].isin(['S','U_on']))&(data_hiv_dtp['age_t']>=c.dic_age_cat[age_j][0])&(data_hiv_dtp['age_t']<c.dic_age_cat[age_j][1])].shape[0]/dict_plwh_cas['art']
        data_art_age = pd.concat([data_art_age,pd.DataFrame(dict_art_age,index=[0])],ignore_index=True)

    return data_plwh_cas,data_art_age

def func_hist_cali_target(c,dict_data):
    """return to a dictionary of dataframes for hiv and comorbidity calibration target"""

    #####load historical information of the cohort
    data_char_hist = dict_data['char'].copy()
    data_hiv_sim = dict_data['hiv'].copy()
    data_comorb_sim = dict_data['comorb'].copy()

    #####create time series and corresponding annual markers
    c.t0_cali = pd.Timestamp('2008-01-01') 
    c.t_end_cali = pd.Timestamp('2017-01-01')
    c.n_step_cali = int((c.t_end_cali.year-c.t0_cali.year)/c.dt)
    arr_step = np.arange(c.n_step_cali+1) #start from step 0 and end at n_step, every two steps to get the end of the year
    arr_date = [c.t0_cali+pd.to_timedelta(c.def_year*c.dt*step_i,unit='d') for step_i in arr_step] 
    annual_date = arr_date[0::2]
    arr_year = np.arange(c.t0_cali.year-1,c.t_end_cali.year)
    assert (len(annual_date)==len(arr_year)), 'Inconsistent date and year lists'

    #####create dataframes for outcomes
    data_plwh_cas = pd.DataFrame([],columns=['date','year']+['new_diag','new_art','A','S','U_on','U_off','S2Uoff','Uoff2S','S2Uon','Uon2S','Uon2Uoff','Uoff2Uon','dead']) #initialize the dataframe related to HIV cascade of care
    data_comorb_inc = pd.DataFrame([],columns=['date','year']+c.list_comorb)
    data_comorb_check = data_comorb_sim[data_comorb_sim['micro_baseline_comorb']=='known'].reset_index(drop=True) 
    for i,date_i in enumerate(annual_date):

        ####create dataframe related to HIV cascade
        dict_plwh_cas = {'date':str(date_i)[:10],'year':arr_year[i]}
        for stat_j in ['A','S','U_on','U_off']:
            dict_plwh_cas[stat_j] = data_hiv_sim[data_hiv_sim[str(date_i)[:10]]==stat_j].shape[0]
        dict_plwh_cas['diag'] = dict_plwh_cas['A']+dict_plwh_cas['S']+dict_plwh_cas['U_on']+dict_plwh_cas['U_off']
        dict_plwh_cas['art_ever'] = dict_plwh_cas['S']+dict_plwh_cas['U_on']+dict_plwh_cas['U_off']
        dict_plwh_cas['art'] = dict_plwh_cas['S']+dict_plwh_cas['U_on']
        if i>0:
            t0 = arr_date[i*2-2]
            t1 = arr_date[i*2-1]
            t2 = arr_date[i*2]
            new_diag = data_char_hist[(data_char_hist['earliest_HIV']>=annual_date[i-1])&(data_char_hist['earliest_HIV']<date_i)].shape[0] #use earliest_HIV instead of transition from NaN to A/S/U_on/U_off to also include those dead after diagnosis
            new_art = data_char_hist[(data_char_hist['FARVDT']>=annual_date[i-1])&(data_char_hist['FARVDT']<date_i)].shape[0] #use FARVDT instead of transition from A to S/U_on/U_off which ignores the transition from NaN to S/U_on/U_off/D when FARVDT=earliest_HIV
            dead = data_char_hist[(data_char_hist['end_fu_dt']>=annual_date[i-1])&(data_char_hist['end_fu_dt']<date_i)].shape[0]
            dict_plwh_cas['new_diag'] = new_diag 
            dict_plwh_cas['new_art'] = new_art 
            dict_plwh_cas['dead'] = dead 
            for diag_cat in ['bf08','0811','sin12']:
                arr_id_cat = data_char_hist['moh_id'][data_char_hist['a2t_cat']==diag_cat].values
                new_art_cat = data_char_hist[(data_char_hist['moh_id'].isin(arr_id_cat))&(data_char_hist['FARVDT']>=annual_date[i-1])&(data_char_hist['FARVDT']<date_i)].shape[0]
                dict_plwh_cas['new_art '+diag_cat] = new_art_cat
            n_s2off1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='S')&(data_hiv_sim[str(t1)[:10]]=='U_off')].shape[0]
            n_s2off2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='S')&(data_hiv_sim[str(t2)[:10]]=='U_off')].shape[0]
            n_off2s1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_off')&(data_hiv_sim[str(t1)[:10]]=='S')].shape[0]
            n_off2s2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_off')&(data_hiv_sim[str(t2)[:10]]=='S')].shape[0]
            n_s2on1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='S')&(data_hiv_sim[str(t1)[:10]]=='U_on')].shape[0]
            n_s2on2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='S')&(data_hiv_sim[str(t2)[:10]]=='U_on')].shape[0]
            n_on2s1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_on')&(data_hiv_sim[str(t1)[:10]]=='S')].shape[0]
            n_on2s2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_on')&(data_hiv_sim[str(t2)[:10]]=='S')].shape[0]
            n_on2off1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_on')&(data_hiv_sim[str(t1)[:10]]=='U_off')].shape[0]
            n_on2off2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_on')&(data_hiv_sim[str(t2)[:10]]=='U_off')].shape[0]
            n_off2on1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_off')&(data_hiv_sim[str(t1)[:10]]=='U_on')].shape[0]
            n_off2on2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_off')&(data_hiv_sim[str(t2)[:10]]=='U_on')].shape[0]
            dict_plwh_cas['S2Uoff'] = n_s2off1+n_s2off2
            dict_plwh_cas['Uoff2S'] = n_off2s1+n_off2s2
            dict_plwh_cas['S2Uon'] = n_s2on1+n_s2on2
            dict_plwh_cas['Uon2S'] = n_on2s1+n_on2s2
            dict_plwh_cas['Uon2Uoff'] = n_on2off1+n_on2off2
            dict_plwh_cas['Uoff2Uon'] = n_off2on1+n_off2on2
            n_a2d1 = data_hiv_sim[((data_hiv_sim[str(t0)[:10]]=='A')|(pd.isnull(data_hiv_sim[str(t0)[:10]])))&(data_hiv_sim[str(t1)[:10]]=='D')].shape[0] #include those died within 6 months of diagnosis
            n_a2d2 = data_hiv_sim[((data_hiv_sim[str(t1)[:10]]=='A')|(pd.isnull(data_hiv_sim[str(t1)[:10]])))&(data_hiv_sim[str(t2)[:10]]=='D')].shape[0]
            n_s2d1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='S')&(data_hiv_sim[str(t1)[:10]]=='D')].shape[0]
            n_s2d2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='S')&(data_hiv_sim[str(t2)[:10]]=='D')].shape[0]
            n_on2d1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_on')&(data_hiv_sim[str(t1)[:10]]=='D')].shape[0]
            n_on2d2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_on')&(data_hiv_sim[str(t2)[:10]]=='D')].shape[0]
            n_off2d1 = data_hiv_sim[(data_hiv_sim[str(t0)[:10]]=='U_off')&(data_hiv_sim[str(t1)[:10]]=='D')].shape[0]
            n_off2d2 = data_hiv_sim[(data_hiv_sim[str(t1)[:10]]=='U_off')&(data_hiv_sim[str(t2)[:10]]=='D')].shape[0]
            dict_plwh_cas['A2D'] = n_a2d1+n_a2d2
            dict_plwh_cas['S2D'] = n_s2d1+n_s2d2
            dict_plwh_cas['Uon2D'] = n_on2d1+n_on2d2
            dict_plwh_cas['Uoff2D'] = n_off2d1+n_off2d2
            n_a2s1 = data_hiv_sim[((data_hiv_sim[str(t0)[:10]]=='A')|(pd.isnull(data_hiv_sim[str(t0)[:10]])))&(data_hiv_sim[str(t1)[:10]]=='S')].shape[0]
            n_a2s2 = data_hiv_sim[((data_hiv_sim[str(t1)[:10]]=='A')|(pd.isnull(data_hiv_sim[str(t1)[:10]])))&(data_hiv_sim[str(t2)[:10]]=='S')].shape[0]
            n_a2on1 = data_hiv_sim[((data_hiv_sim[str(t0)[:10]]=='A')|(pd.isnull(data_hiv_sim[str(t0)[:10]])))&(data_hiv_sim[str(t1)[:10]]=='U_on')].shape[0]
            n_a2on2 = data_hiv_sim[((data_hiv_sim[str(t1)[:10]]=='A')|(pd.isnull(data_hiv_sim[str(t1)[:10]])))&(data_hiv_sim[str(t2)[:10]]=='U_on')].shape[0]
            n_a2off1 = data_hiv_sim[((data_hiv_sim[str(t0)[:10]]=='A')|(pd.isnull(data_hiv_sim[str(t0)[:10]])))&(data_hiv_sim[str(t1)[:10]]=='U_off')].shape[0]
            n_a2off2 = data_hiv_sim[((data_hiv_sim[str(t1)[:10]]=='A')|(pd.isnull(data_hiv_sim[str(t1)[:10]])))&(data_hiv_sim[str(t2)[:10]]=='U_off')].shape[0]
            dict_plwh_cas['A2S'] = n_a2s1+n_a2s2
            dict_plwh_cas['A2Uon'] = n_a2on1+n_a2on2
            dict_plwh_cas['A2Uoff'] = n_a2off1+n_a2off2
        data_plwh_cas = pd.concat([data_plwh_cas,pd.DataFrame(dict_plwh_cas,index=[0])],ignore_index=True)

        ####create dataframe for comorbidity incidence
        dict_comorb_inc = {'date':date_i,'year':arr_year[i]} #keep full date information which is related to derivation of incidence
        if i>0:
            for comorb_j in c.list_comorb:
                dict_comorb_inc[comorb_j] = data_comorb_check[(data_comorb_check['earliest_'+comorb_j+'_dt']>data_comorb_check['baseline_dt'])&(data_comorb_check['earliest_'+comorb_j+'_dt']>annual_date[i-1])&(data_comorb_check['earliest_'+comorb_j+'_dt']<=annual_date[i])].shape[0] #incidence as diagnosed after baseline_dt to distinguish from the prevalent cases
        data_comorb_inc = pd.concat([data_comorb_inc,pd.DataFrame(dict_comorb_inc,index=[0])],ignore_index=True)

    dict_outcomes = {
        'hiv':data_plwh_cas,
        'comorb':data_comorb_inc,
    }
    
    return dict_outcomes

def func_reg_add_new(c,dict_reg_status,t_start,t_stop,arr_id_trans):
    """return to dict_reg_status with t_stop modified from nan/0 to sampled regimens, representing changes for ART initiation and from U_off to S/U_on"""

    data_trans_reg_sim = pd.DataFrame({'moh_id':arr_id_trans})
    data_trans_reg_sim['bb'] = random.choices(c.list_bb,weights=[c.reg_dist['init']['percent'][(c.reg_dist['init']['date']==c.init_dist_dt)&(c.reg_dist['init']['class_bb']==bb_i)].sum() for bb_i in c.list_bb],k=data_trans_reg_sim.shape[0])
    data_trans_reg_sim['bb'] = data_trans_reg_sim['bb'].replace({'tdf':1,'other':0})
    dict_reg_status['tdf'].loc[dict_reg_status['tdf']['moh_id_sim'].isin(arr_id_trans),str(t_stop)[:10]] = data_trans_reg_sim['bb'][data_trans_reg_sim['moh_id'].isin(arr_id_trans)].values
    data_trans_reg_sim['nbb'] = random.choices(c.list_nbb,weights=[c.reg_dist['init']['percent'][(c.reg_dist['init']['date']==c.init_dist_dt)&(c.reg_dist['init']['class_nbb']==nbb_i)].sum() for nbb_i in c.list_nbb],k=data_trans_reg_sim.shape[0])
    for reg_i in c.list_nbb:
        arr_id_reg = data_trans_reg_sim['moh_id'][data_trans_reg_sim['nbb']==reg_i].values
        if reg_i!='super':
            dict_reg_status[reg_i].loc[dict_reg_status[reg_i]['moh_id_sim'].isin(arr_id_reg),str(t_stop)[:10]] = 1.
        else:
            data_trans_reg_sim.loc[data_trans_reg_sim['moh_id'].isin(arr_id_reg),'nbb'] = random.choices(c.super_state,weights=c.super_dist,k=len(arr_id_reg))
            for reg_k in c.list_nbb[:-1]:
                arr_id_super_reg = data_trans_reg_sim['moh_id'][(data_trans_reg_sim['moh_id'].isin(arr_id_reg))&(data_trans_reg_sim['nbb'].str.contains(reg_k))].values 
                dict_reg_status[reg_k].loc[dict_reg_status[reg_k]['moh_id_sim'].isin(arr_id_super_reg),str(t_stop)[:10]] = 1.
            
    return dict_reg_status

def mp_microsim_all_test(c,dict_hist_data,seed_i):
    """run microsimulation step by step for mortality, aware to treatment, ART interruption and comorbidity incidence for calibration"""

    #####initialize random generator for each simulation of the cohort
    random.seed(seed_i)
    np.random.seed(seed_i)

    #####initialize cohort on c.t0
    dict_cohort_sim = func_init_pop(c,dict_hist_data)
    
    #####rename dataframe for simplification purpose
    data_char_sim = dict_cohort_sim['char'].assign(moh_id_sim=dict_cohort_sim['char'].index) 
    data_char_sim['FARVDT_sim'] = data_char_sim['FARVDT'].apply(lambda x:x if x<=c.t0 else np.nan) #FARVDT_sim should be simulated for new ART initiation after t0 which is the beginning of the microsimulation
    data_char_sim['art_baseline_dt'] = data_char_sim['art_baseline_dt'].apply(lambda x: x if x<=c.study_dt else np.nan) #remove art_baseline_dt for those with simulated FARVDT
    data_char_sim['end_fu_dt_sim'] = c.dummy_end_dt #add simulated end_fu_dt
    data_char_sim['blup'] = data_char_sim['a2t_cat'].apply(lambda x: np.random.normal(0.,np.sqrt(c.dic_random_a2t[x][0]))) #random term sampled from a normal distribution with mean=0 and variance provided
    data_char_sim['blup_off'] = np.random.normal(0.,np.sqrt(c.dic_random_t['u_off'][0]),size=data_char_sim.shape[0])
    data_char_sim['blup_on'] = np.random.normal(0.,np.sqrt(c.dic_random_t['u_on'][0]),size=data_char_sim.shape[0])
    data_hiv_sim = dict_cohort_sim['hiv'].assign(moh_id_sim=dict_cohort_sim['hiv'].index)
    data_comorb_sim = dict_cohort_sim['comorb'].assign(moh_id_sim=dict_cohort_sim['comorb'].index) #be aware that the specific comorbidity information was replaced by dummy dates if earliest_comorb_dt>t0
    data_pvl_sim = dict_cohort_sim['pvl'].assign(moh_id_sim=dict_cohort_sim['pvl'].index)
    dict_reg_status_sim = {k:v.assign(moh_id_sim=v.index) for k,v in dict_cohort_sim['reg_status'].items()}
    dict_reg_prop_sim = {k:v.assign(moh_id_sim=v.index) for k,v in dict_cohort_sim['reg_prop'].items()}

    ######estimate probability and run randomization for mortality, a2t, trans_art and comorbidities
    for i in range(1,c.n_step+1):
        
        ####define the beginning and the end of the time step
        t_start = c.t0+pd.to_timedelta(c.def_year*c.dt*(i-1),unit='d')
        t_stop = t_start+pd.to_timedelta(c.def_year*c.dt,unit='d')

        ####use piece-wise dic_coeff_adj_a2t before and since 2012 and use different split time for different HIV diagnosis period
        if t_start<c.dt_split_prob_a2t_bf08:
            c.dic_coeff_adj_a2t['bf08'] = c.dic_coeff_adj_a2t_pw['bf08'][0]
        else:
            c.dic_coeff_adj_a2t['bf08'] = c.dic_coeff_adj_a2t_pw['bf08'][1]
        if t_start<c.dt_split_prob_a2t_0811:
            c.dic_coeff_adj_a2t['0811'] = c.dic_coeff_adj_a2t_pw['0811'][0]
        else:
            c.dic_coeff_adj_a2t['0811'] = c.dic_coeff_adj_a2t_pw['0811'][1]
        if t_start<c.dt_split_prob_a2t_sin12:
            c.dic_coeff_adj_a2t['sin12'] = c.dic_coeff_adj_a2t_pw['sin12'][0]
        else:
            c.dic_coeff_adj_a2t['sin12'] = c.dic_coeff_adj_a2t_pw['sin12'][1]
        
        ####use piece-wise coeff_prob_dead_ltfu_noart to account for elevated overdose risk (most likely) since 2014 among those not on ART
        if t_start<c.dt_split_prob_dead_noart: #use date instead of year for split year other than 2012
            c.coeff_prob_dead_ltfu_noart = c.coeff_prob_dead_ltfu_noart_pw[0]
        else:
            c.coeff_prob_dead_ltfu_noart = c.coeff_prob_dead_ltfu_noart_pw[1]

        ####initialize columns for t_stop for simulated results
        data_hiv_sim[str(t_stop)[:10]] = data_hiv_sim[str(t_start)[:10]].copy() #initialize new hiv status by t_stop, 'D' from previous step will be copied and no longer requires further modification
        for key_i in dict_reg_status_sim.keys():
                dict_reg_status_sim[key_i][str(t_stop)[:10]] = dict_reg_status_sim[key_i][str(t_start)[:10]].copy() #regimen type will be determined every 6 months
        if i==1: #keep pvl/reg_prop the same as t_start for i=1, the rest can be derived for i>=2
            data_pvl_sim[str(t_stop)[:10]] = data_pvl_sim[str(t_start)[:10]].copy() #pvl status will be updated after one year
            for key_i in dict_reg_prop_sim.keys():
                dict_reg_prop_sim[key_i][str(t_stop)[:10]] = dict_reg_prop_sim[key_i][str(t_start)[:10]].copy() #regimen status will be updated after one year

        ####only consider the probabilities among those alive at t_start
        arr_id_plwh = data_hiv_sim['moh_id_sim'][data_hiv_sim[str(t_start)[:10]].isin(c.arr_state_alive)].values #introduce arr_state_alive so that PLWH can be selected at each step if HIV status is updated
        data_char_plwh = data_char_sim[data_char_sim['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'})
        data_hiv_plwh = data_hiv_sim[data_hiv_sim['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'})
        data_comorb_plwh = data_comorb_sim[data_comorb_sim['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'})
        data_pvl_plwh = data_pvl_sim[data_pvl_sim['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'})
        dict_reg_status_plwh = {k:dict_reg_status_sim[k][dict_reg_status_sim[k]['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'}) for k in dict_reg_status_sim.keys()}
        dict_reg_prop_plwh = {k:dict_reg_prop_sim[k][dict_reg_prop_sim[k]['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'}) for k in dict_reg_prop_sim.keys()}

        ####update covariants for mortality and other comorbidities based on simulated cohort on t_start
        data_comorb_var = func_update_comorb_variable(c,data_comorb_plwh,data_char_plwh,data_hiv_plwh,data_pvl_plwh,dict_reg_prop_plwh,t_start) 

        ####determine the probability of each comorbidity and simulate comorbidity status, which is independent from mortality and ART module
        for comorb in c.list_comorb:
            data_prob_comorb = c.dic_prob_comorb[comorb](c,data_comorb_plwh,data_hiv_plwh,data_comorb_var,t_start) 
            assert (data_prob_comorb[pd.isnull(data_prob_comorb['prob_wo_comorb'])].shape[0]==0), 'Invalid probability estimation for comorbidities'
            data_trans_comorb = pd.DataFrame({'moh_id':data_prob_comorb['moh_id'].astype(int),'0':np.zeros(data_prob_comorb.shape[0],dtype=int)})
            data_trans_comorb['1'] = pd.Series(sample_from_prob_matrix(data_prob_comorb[['prob_wo_comorb','prob_w_comorb']].values))
            arr_id_comorb_inc = data_trans_comorb['moh_id'][data_trans_comorb['1']==1].values
            data_comorb_sim.loc[data_comorb_sim['moh_id_sim'].isin(arr_id_comorb_inc),'earliest_'+comorb+'_dt'] = t_stop #set diagnosis date of comorbidity as t_stop for newly simulated incident cases in the full dataset, not for plwh alive at t_start

        ####determine the probability of the mortality for each individual and simulate mortality status
        data_prob_dead = func_prob_update_mortality(c,data_comorb_plwh,data_hiv_plwh,data_comorb_var,t_start) 
        assert (data_prob_dead[pd.isnull(data_prob_dead['prob_wo_comorb'])].shape[0]==0), 'Invalid probability estimation for mortality'
        data_trans_dead = pd.DataFrame({'moh_id':data_prob_dead['moh_id'].astype(int),'0':np.zeros(data_prob_dead.shape[0],dtype=int)})
        data_trans_dead['1'] = pd.Series(sample_from_prob_matrix(data_prob_dead[['prob_wo_comorb','prob_w_comorb']].values))
        arr_id_new_dead = data_trans_dead['moh_id'][data_trans_dead['1']==1].values
        data_char_sim.loc[data_comorb_sim['moh_id_sim'].isin(arr_id_new_dead),'end_fu_dt_sim'] = t_stop #convert timestamp to datetime to be consistent with dummy_end_date
        data_char_sim['end_fu_dt_sim'] = pd.to_datetime(data_char_sim['end_fu_dt_sim'].dt.date) #only keep the date information using datetime format
        data_hiv_plwh.loc[data_hiv_plwh['moh_id'].isin(arr_id_new_dead),str(t_stop)[:10]] = 'D' #modify data_plwh for simulation later to exclude the dead
        data_hiv_sim.loc[data_hiv_sim['moh_id_sim'].isin(arr_id_new_dead),str(t_stop)[:10]] = 'D'

        ####for those without ART initiation and alive by t_stop, determine status for ART initiation
        arr_id_plwh_A = data_hiv_plwh['moh_id'][(data_hiv_plwh[str(t_start)[:10]]=='A')&(data_hiv_plwh[str(t_stop)[:10]]!='D')].values
        data_char_A = data_char_plwh[data_char_plwh['moh_id'].isin(arr_id_plwh_A)].sort_values(by='moh_id').reset_index(drop=True)
        data_prob_a2t = func_prob_update_a2t(c,data_char_A,t_start,c.rand_blup_a2t) 
        assert (data_prob_a2t[pd.isnull(data_prob_a2t['prob_a2t'])].shape[0]==0), 'Invalid probability estimation for transition A2T'
        data_trans_a2t = pd.DataFrame({'moh_id':data_prob_a2t['moh_id'].astype(int),'0':np.zeros(data_prob_a2t.shape[0],dtype=int)})
        data_trans_a2t['1'] = pd.Series(sample_from_prob_matrix(data_prob_a2t[['prob_a2a','prob_a2t']].values)) 
        data_trans_a2t = data_trans_a2t.merge(data_char_A[['moh_id','baseline_dt','counter']].copy(),how='left',on='moh_id')
        arr_id_a2t = data_trans_a2t['moh_id'][(data_trans_a2t['0']==0)&(data_trans_a2t['1']==1)].values
        data_trans_a2t['t_sim'] = data_trans_a2t['moh_id'].apply(lambda x: np.random.uniform(0,1,1)[0] if x in arr_id_a2t else np.nan) #assume ART initiation between t_start and t_stop instead of baseline+counter*dt*def_year for simplification
        data_trans_a2t['FARVDT_sim'] = (data_trans_a2t['baseline_dt']+pd.to_timedelta(c.def_year*c.dt*((data_trans_a2t['counter']-1)+data_trans_a2t['t_sim']),unit='d')).dt.round('d') #FARVDT_sim should be updated based on counter and baseline_dt instead of t_start and t_stop
        data_char_sim.loc[data_char_sim['moh_id_sim'].isin(arr_id_a2t),'FARVDT_sim'] = data_trans_a2t['FARVDT_sim'][data_trans_a2t['moh_id'].isin(arr_id_a2t)].values 
        data_char_sim.loc[data_char_sim['moh_id_sim'].isin(arr_id_a2t),'art_baseline_dt'] = data_char_sim['FARVDT_sim'][data_char_sim['moh_id_sim'].isin(arr_id_a2t)]
        data_char_sim['age_art_bsln_scale10_dv'] = (data_char_sim['art_baseline_dt']-data_char_sim['DOB']).dt.total_seconds()/(60*60*24*c.def_year)/10 #update age at art_baseline after data update
        data_hiv_sim.loc[data_hiv_sim['moh_id_sim'].isin(arr_id_a2t),str(t_stop)[:10]] = random.choices(c.arr_state_art,weights=[c.prob_A2S,c.prob_A2Uon,1-c.prob_A2S-c.prob_A2Uon],k=len(arr_id_a2t))
        
        ####for those with ART initiation and alive by t_stop, determine status among S, U_on and U_off
        arr_id_plwh_T = data_hiv_plwh['moh_id'][(data_hiv_plwh[str(t_start)[:10]].isin(['S','U_off','U_on']))&(data_hiv_plwh[str(t_stop)[:10]]!='D')].values
        data_char_T = data_char_plwh[data_char_plwh['moh_id'].isin(arr_id_plwh_T)].sort_values(by='moh_id').reset_index(drop=True)
        data_comorb_T = data_comorb_var[['moh_id','manx_dv','prsn_dv','sczo_dv']][data_comorb_var['moh_id'].isin(arr_id_plwh_T)].sort_values(by='moh_id').reset_index(drop=True) 
        data_char_T = data_char_T.merge(data_comorb_T,how='left',on='moh_id')
        data_hiv_T = data_hiv_plwh[data_hiv_plwh['moh_id'].isin(arr_id_plwh_T)].sort_values(by='moh_id').reset_index(drop=True)
        data_prob_trans_art = func_prob_update_trans_art(c,data_hiv_T,data_char_T,t_start,c.rand_blup_art) #add rand_art=0/1 representing whether to use random terms in probability estimation
        data_prob_trans_art = func_prob_adj_trans_art(c,data_prob_trans_art) #adjust probabilities of ART transitions using 6 coefficients
        data_prob_trans_art = data_prob_trans_art[['moh_id','prob_s_adj3','prob_off_adj3','prob_on_adj3']].copy().rename(columns={'prob_s_adj3':'prob_s','prob_off_adj3':'prob_off','prob_on_adj3':'prob_on'})
        data_trans_art = data_hiv_T[['moh_id',str(t_start)[:10]]].copy().rename(columns={str(t_start)[:10]:'0'}) #keep state at t_start to count transitions from different states
        data_trans_art['1'] = pd.Series(sample_from_prob_matrix(data_prob_trans_art[['prob_off','prob_on','prob_s']].values)) 
        data_trans_art = data_trans_art.replace({'1':{0:'U_off',1:'U_on',2:'S'}}) #replace sampled number back to strings that represent suppression/unsuppressed(off ART/on ART)
        arr_id_trans_art = data_trans_art['moh_id'][(data_trans_art['0']!=data_trans_art['1'])].values
        data_hiv_sim.loc[data_hiv_sim['moh_id_sim'].isin(arr_id_trans_art),str(t_stop)[:10]] = data_trans_art['1'][data_trans_art['moh_id'].isin(arr_id_trans_art)].values

        ####for those newly initiated ART by t_stop, assign regimen type at t_stop
        c.init_dist_dt = c.reg_dist['init']['date'][c.reg_dist['init']['date']>=str(t_stop)[:10]].min() #distribution among new enrollments started from 2008-12-31
        c.super_dist_dt = c.reg_dist['super']['date'][c.reg_dist['super']['date']<str(t_stop)[:10]].max() #use the latest distribution before t_stop
        c.super_dist = [c.reg_dist['super']['percent'][(c.reg_dist['super']['date']==c.super_dist_dt)&(c.reg_dist['super']['class_nbb_super']==state_i)].values[0] for state_i in c.super_state]
        dict_reg_status_sim = func_reg_add_new(c,dict_reg_status_sim,t_start,t_stop,arr_id_a2t)
        
        ####for those with ART initiation and alive by t_stop, assign regimen type at t_stop
        arr_id_stay = data_hiv_sim['moh_id_sim'][(data_hiv_sim[str(t_start)[:10]].isin(['S','U_on']))&(data_hiv_sim[str(t_stop)[:10]].isin(['S','U_on']))].values #new regimen type will be added in dict_reg_status_sim
        dict_bb_state = {'tdf':['tdf','other'], 'other':['other','tdf']}
        dict_bb_p = {}
        for bb_j in c.list_bb:
            if str(t_stop)[:10]<str(c.dic_reg_trans_p[bb_j][0])+'-01-01':
                dict_bb_p[bb_j] = c.dic_reg_trans_p[bb_j][1]
            else:
                dict_bb_p[bb_j] = c.dic_reg_trans_p[bb_j][2]
        data_bb_sim = pd.DataFrame({'moh_id':arr_id_stay,'bb_old':dict_reg_status_plwh['tdf'][str(t_start)[:10]][dict_reg_status_plwh['tdf']['moh_id'].isin(arr_id_stay)].values})
        data_bb_sim['bb_old'] = data_bb_sim['bb_old'].replace({1:'tdf',0:'other'})
        data_bb_sim['bb'] = data_bb_sim['bb_old'].apply(lambda x:random.choices(dict_bb_state[x],weights=dict_bb_p[x],k=1)[0])
        data_bb_sim[['bb_old','bb']] = data_bb_sim[['bb_old','bb']].replace({'tdf':1,'other':0})
        arr_id_tdf_change = data_bb_sim['moh_id'][data_bb_sim['bb_old']!=data_bb_sim['bb']].values #only change values for a small proportion of PLWH
        dict_reg_status_sim['tdf'].loc[dict_reg_status_sim['tdf']['moh_id_sim'].isin(arr_id_tdf_change),str(t_stop)[:10]] = data_bb_sim['bb'][data_bb_sim['moh_id'].isin(arr_id_tdf_change)].values
        dict_nbb_state = {}
        dict_nbb_p = {}
        for nbb_j in c.list_nbb:
            list_nbb_rest = c.list_nbb[:]
            list_nbb_rest.remove(nbb_j)
            list_nbb_trans = list_nbb_rest[:]
            list_nbb_trans.insert(0,nbb_j)
            dict_nbb_state[nbb_j] = list_nbb_trans
            if str(t_stop)[:10]<str(c.dic_reg_trans_p[nbb_j][0])+'-01-01':
                dict_nbb_p[nbb_j] = c.dic_reg_trans_p[nbb_j][1]
            else:
                dict_nbb_p[nbb_j] = c.dic_reg_trans_p[nbb_j][2]
        dict_id_nbb = {}
        for reg_i in c.list_nbb[:-1]: #find individuals with different regimen type
            dict_id_nbb[reg_i] = dict_reg_status_plwh[reg_i]['moh_id'][(dict_reg_status_plwh[reg_i][str(t_start)[:10]]==1)&(dict_reg_status_plwh[reg_i]['moh_id'].isin(arr_id_stay))].values
        dict_id_nbb['super'] = np.array(list((set(dict_id_nbb['nnrti'])&set(dict_id_nbb['pi']))|(set(dict_id_nbb['nnrti'])&set(dict_id_nbb['insti']))|(set(dict_id_nbb['insti'])&set(dict_id_nbb['pi'])))) #set operation works to determine 'super' as dict_reg_status indicates the regimen type on t_start/t_stop instead of during the period
        dict_id_nbb['super'].sort() #sort in ascending order to be consistent with id in dataframe
        for reg_i in c.list_nbb[:-1]: #super will be considered separately
            arr_id_reg = np.array(list(set(dict_id_nbb[reg_i])-set(dict_id_nbb['super']))) #simulate transitions for each regimen type
            arr_id_reg.sort() #sort in ascending order to be consistent with id in dataframe
            data_nbb_sim = pd.DataFrame({'moh_id':arr_id_reg,'nbb_old':reg_i}) 
            data_nbb_sim['nbb'] = data_nbb_sim['nbb_old'].apply(lambda x:random.choices(dict_nbb_state[x],weights=dict_nbb_p[x],k=1)[0])
            for reg_j in dict_nbb_state[reg_i][1:]: #loop except reg_i itself
                arr_id_reg_change = data_nbb_sim['moh_id'][(data_nbb_sim['nbb']==reg_j)].values
                if reg_j!='super':
                    dict_reg_status_sim[reg_i].loc[dict_reg_status_sim[reg_i]['moh_id_sim'].isin(arr_id_reg_change),str(t_stop)[:10]] = 0. #indicate regimen type changed from reg_i to reg_j
                    dict_reg_status_sim[reg_j].loc[dict_reg_status_sim[reg_j]['moh_id_sim'].isin(arr_id_reg_change),str(t_stop)[:10]] = 1. 
                else:
                    data_nbb_sim.loc[data_nbb_sim['moh_id'].isin(arr_id_reg_change),'nbb'] = random.choices(c.super_state,weights=c.super_dist,k=len(arr_id_reg_change))
                    for reg_k in c.list_nbb[:-1]:
                        if reg_k==reg_i:
                            arr_id_super_reg2 = data_nbb_sim['moh_id'][(data_nbb_sim['moh_id'].isin(arr_id_reg_change))&(data_nbb_sim['nbb'].str.contains(reg_k)==False)].values #for those super does not contain reg_k, change t_stop=0
                            dict_reg_status_sim[reg_i].loc[dict_reg_status_sim[reg_i]['moh_id_sim'].isin(arr_id_super_reg2),str(t_stop)[:10]] = 0.
                        arr_id_super_reg = data_nbb_sim['moh_id'][(data_nbb_sim['moh_id'].isin(arr_id_reg_change))&(data_nbb_sim['nbb'].str.contains(reg_k))].values #id_super_reg can be changed for reg_i with super=nnrti_pi separately, both lead to dict_reg_status_sim[reg_i]=0
                        dict_reg_status_sim[reg_k].loc[dict_reg_status_sim[reg_k]['moh_id_sim'].isin(arr_id_super_reg),str(t_stop)[:10]] = 1.
        arr_id_super = dict_id_nbb['super'] #already sorted ascending
        data_nbb_sim = pd.DataFrame({'moh_id':arr_id_super,'nbb_old':'super'}) #can be simplified as arr_id_reg with reg_i=1
        data_nbb_sim['nbb'] = data_nbb_sim['nbb_old'].apply(lambda x:random.choices(dict_nbb_state[x],weights=dict_nbb_p[x],k=1)[0])
        arr_id_super_change = data_nbb_sim['moh_id'][(data_nbb_sim['nbb']!='super')].values
        data_super_sim = data_nbb_sim[data_nbb_sim['moh_id'].isin(arr_id_super_change)].reset_index(drop=True)
        data_super_sim['super_old'] = data_super_sim['moh_id'].apply(lambda x: '_'.join([reg_m for reg_m in ['nnrti','pi','insti'] if x in dict_id_nbb[reg_m]]))
        for reg_m in ['nnrti','pi','insti']:
            arr_id_reg_to1 = data_super_sim['moh_id'][(data_super_sim['nbb']==reg_m)&(data_super_sim['super_old'].str.contains(reg_m)==False)].values
            dict_reg_status_sim[reg_m].loc[dict_reg_status_sim[reg_m]['moh_id_sim'].isin(arr_id_reg_to1),str(t_stop)[:10]] = 1.
            arr_id_reg_to0 = data_super_sim['moh_id'][(data_super_sim['nbb']!=reg_m)&(data_super_sim['super_old'].str.contains(reg_m))].values
            dict_reg_status_sim[reg_m].loc[dict_reg_status_sim[reg_m]['moh_id_sim'].isin(arr_id_reg_to0),str(t_stop)[:10]] = 0.

        ####for those transitted from U_off to S/U_on by t_stop, assign regimen type at t_stop
        arr_id_trans = data_hiv_sim['moh_id_sim'][(data_hiv_sim[str(t_start)[:10]]=='U_off')&(data_hiv_sim[str(t_stop)[:10]].isin(['S','U_on']))].values
        dict_reg_status_sim = func_reg_add_new(c,dict_reg_status_sim,t_start,t_stop,arr_id_trans)

        ####for those transitted from S/U-on to U_off by t_stop, assign 0 at t_stop for each regimen dataframe
        arr_id_off = data_hiv_sim['moh_id_sim'][(data_hiv_sim[str(t_stop)[:10]]=='U_off')].values #no need to consider status at t_start, including U_off at t_start
        for reg_i in dict_reg_status_sim.keys():
            dict_reg_status_sim[reg_i].loc[dict_reg_status_sim[reg_i]['moh_id_sim'].isin(arr_id_off),str(t_stop)[:10]] = 0.

        ####update data_pvl_sim and dict_reg_prop_sim every 1 year based on data_hiv_sim and dict_reg_status 
        if i>=2: #look at 1-year period (two time steps) before t_stop to determine the status at t_stop
            t_prev1yr = c.t0+pd.to_timedelta(c.def_year*c.dt*(i-2),unit='d') #define the time step one year agao  
            list_dt = [str(t_prev1yr)[:10],str(t_start)[:10],str(t_stop)[:10]]
            data_hiv_sim['comb'] = data_hiv_sim[list_dt].apply(lambda x: "".join(list(x)),axis=1)
            data_hiv_sim['pvl'] = data_hiv_sim['comb'].apply(lambda x: 0 if x=='SSS' else 1) #0 indicates suppression
            data_pvl_sim[str(t_stop)[:10]] = data_hiv_sim['pvl'].values #order in moh_id_sim should be the same for data_hiv_sim and data_pvl_sim
            for k in dict_reg_prop_sim.keys():
                dict_reg_status_sim[k][list_dt] = dict_reg_status_sim[k][list_dt].fillna(4) #replace NaN by 4 to locate the right proportion in c.dic_reg_val
                dict_reg_status_sim[k]['comb'] = dict_reg_status_sim[k][list_dt].apply(lambda x: "".join([str(int(ele)) for ele in x]),axis=1)
                dict_reg_status_sim[k]['prop'] = dict_reg_status_sim[k]['comb'].apply(lambda x: c.dic_reg_prop[x]) #will not fully executed if x outside of keys of dic_reg_val
                dict_reg_prop_sim[k][str(t_stop)[:10]] = dict_reg_status_sim[k]['prop'].values #order of dict_reg_status/prop shoud be the same
                dict_reg_status_sim[k][list_dt] = dict_reg_status_sim[k][list_dt].replace({4:np.nan}) #change no ART initiation back to NaN

        ####create dictionary to save characteristics for new diagnosis at each time step and attach new diagnosis to plwh for micosimulaiton at next step
        dict_new_sim = func_init_new(c,dict_hist_data,t_start,t_stop)
        data_char_new = dict_new_sim['char'].copy()
        data_char_new['FARVDT_sim'] = pd.to_datetime(np.nan) #make type=datetime64 to be consistent with FARVDT_sim in data_char_sim
        data_char_new['art_baseline_dt'] = data_char_new['FARVDT_sim'].copy() #similar to FARVDT_sim
        data_char_new['age_art_bsln_scale10_dv'] = np.nan #change age_art_bsln_scale10_dv to be nan for new diagnosis
        data_char_new['end_fu_dt_sim'] = c.dummy_end_dt #add simulated end_fu_dt
        data_char_new['blup'] = data_char_new['a2t_cat'].apply(lambda x: np.random.normal(0.,np.sqrt(c.dic_random_a2t[x][0]))) #random term sampled from a normal distribution with mean=0 and variance provided
        data_char_new['blup_off'] = np.random.normal(0.,np.sqrt(c.dic_random_t['u_off'][0]),size=data_char_new.shape[0])
        data_char_new['blup_on'] = np.random.normal(0.,np.sqrt(c.dic_random_t['u_on'][0]),size=data_char_new.shape[0])
        data_char_sim = pd.concat([data_char_sim,data_char_new],ignore_index=True) 
        data_char_sim = data_char_sim.assign(moh_id_sim=data_char_sim.index)
        data_hiv_sim = pd.concat([data_hiv_sim,dict_new_sim['hiv']],ignore_index=True)
        data_hiv_sim = data_hiv_sim.assign(moh_id_sim=data_hiv_sim.index)
        data_comorb_sim = pd.concat([data_comorb_sim,dict_new_sim['comorb']],ignore_index=True)
        data_comorb_sim = data_comorb_sim.assign(moh_id_sim=data_comorb_sim.index)
        data_pvl_sim = pd.concat([data_pvl_sim,dict_new_sim['pvl']],ignore_index=True)
        data_pvl_sim = data_pvl_sim.assign(moh_id_sim=data_pvl_sim.index)
        dict_reg_status_sim = {k:pd.concat([v,dict_new_sim['reg_status'][k]],ignore_index=True) for k,v in dict_reg_status_sim.items()} 
        dict_reg_status_sim = {k:v.assign(moh_id_sim=v.index) for k,v in dict_reg_status_sim.items()} 
        dict_reg_prop_sim = {k:pd.concat([v,dict_new_sim['reg_prop'][k]],ignore_index=True) for k,v in dict_reg_prop_sim.items()}
        dict_reg_prop_sim = {k:v.assign(moh_id_sim=v.index) for k,v in dict_reg_prop_sim.items()} 

    #####create dictionary to save dataframes used for outcome derivation
    dict_data_sim = {
        'char':data_char_sim.copy(),
        'hiv':data_hiv_sim.copy(), 
        'comorb':data_comorb_sim.copy(), #pvl,reg_status/prop were used for probability estimation, not final outcomes
        }

    #return dict_outcomes
    return dict_data_sim #show the simulated cohort information

def mp_microsim_outcomes_cali(c,dict_hist_data,seed_i):
    """return to derived outcomes for calibration"""

    dict_micro_results = mp_microsim_all_test(c,dict_hist_data,seed_i)
    dict_outcomes = func_micro_outcomes_target(c,dict_micro_results)

    return dict_outcomes

def mp_microsim_all_rand(c,dict_hist_data,seed_i):
    """run microsimulation step by step for mortality, aware to treatment, ART interruption and comorbidity incidence, with randomized new diagnosis/migrants"""

    #####initialize random generator for each simulation of the cohort
    random.seed(seed_i)
    np.random.seed(seed_i)

    #####initialize cohort on c.t0
    dict_cohort_sim = func_init_pop(c,dict_hist_data)
    
    #####rename dataframe for simplification purpose
    data_char_sim = dict_cohort_sim['char'].assign(moh_id_sim=dict_cohort_sim['char'].index) 
    data_char_sim['FARVDT_sim'] = data_char_sim['FARVDT'].apply(lambda x:x if x<=c.t0 else np.nan) #FARVDT_sim should be simulated for new ART initiation after t0 which is the beginning of the microsimulation
    data_char_sim['art_baseline_dt'] = data_char_sim['art_baseline_dt'].apply(lambda x: x if x<=c.study_dt else np.nan) #remove art_baseline_dt for those with simulated FARVDT
    data_char_sim['end_fu_dt_sim'] = c.dummy_end_dt #add simulated end_fu_dt
    data_char_sim['blup'] = data_char_sim['a2t_cat'].apply(lambda x: np.random.normal(0.,np.sqrt(c.dic_random_a2t[x][0]))) #random term sampled from a normal distribution with mean=0 and variance provided
    data_char_sim['blup_off'] = np.random.normal(0.,np.sqrt(c.dic_random_t['u_off'][0]),size=data_char_sim.shape[0])
    data_char_sim['blup_on'] = np.random.normal(0.,np.sqrt(c.dic_random_t['u_on'][0]),size=data_char_sim.shape[0])
    data_hiv_sim = dict_cohort_sim['hiv'].assign(moh_id_sim=dict_cohort_sim['hiv'].index)
    data_comorb_sim = dict_cohort_sim['comorb'].assign(moh_id_sim=dict_cohort_sim['comorb'].index) #be aware that the specific comorbidity information was replaced by dummy dates if earliest_comorb_dt>t0
    data_pvl_sim = dict_cohort_sim['pvl'].assign(moh_id_sim=dict_cohort_sim['pvl'].index)
    dict_reg_status_sim = {k:v.assign(moh_id_sim=v.index) for k,v in dict_cohort_sim['reg_status'].items()}
    dict_reg_prop_sim = {k:v.assign(moh_id_sim=v.index) for k,v in dict_cohort_sim['reg_prop'].items()}

    #####initialize prevalence and undiagnosed PLWH for estimation of new infection and updated undiagnosed PLWH, using dataframe for simulation over time, introduce uncertainty regarding HIV prevalence
    plwh_rand = np.random.randint(2)
    if plwh_rand==0:
        n_plwh = np.random.uniform(8372.,9381.)
    else:
        n_plwh = np.random.uniform(9381.,10182.)
    n_unsupp_diag = data_hiv_sim[data_hiv_sim[str(c.t0)[:10]].isin(['A','U_off','U_on'])].shape[0]
    n_spvl = data_hiv_sim[data_hiv_sim[str(c.t0)[:10]]=='S'].shape[0]
    n_undiag = n_plwh-n_unsupp_diag-n_spvl
    data_newmicro = pd.DataFrame([],columns=['date','year','new_inc','new_dead_undiag','new_diag','new_migr','undiag','unsupp_diag','spvl','plwh']) #derive new infection, new_diag+new_migr as new participants for microsimulation
    dict_newmicro = {
        'date':str(c.t0)[:10], 'year':c.year0, 
        'plwh':n_plwh, 'spvl':n_spvl, 'unsupp_diag':n_unsupp_diag, 'undiag':n_undiag,
    }
    data_newmicro = pd.concat([data_newmicro,pd.DataFrame(dict_newmicro,index=[0])],ignore_index=True)

    ######estimate probability and run randomization for mortality, a2t, trans_art and comorbidities
    for i in range(1,c.n_step+1):
        
        ####define the beginning and the end of the time step
        t_start = c.t0+pd.to_timedelta(c.def_year*c.dt*(i-1),unit='d')
        t_stop = t_start+pd.to_timedelta(c.def_year*c.dt,unit='d')

        ####use piece-wise dic_coeff_adj_a2t before and since 2012 and use different split time for different HIV diagnosis period
        if t_start<c.dt_split_prob_a2t_bf08:
            c.dic_coeff_adj_a2t['bf08'] = c.dic_coeff_adj_a2t_pw['bf08'][0]
        else:
            c.dic_coeff_adj_a2t['bf08'] = c.dic_coeff_adj_a2t_pw['bf08'][1]
        if t_start<c.dt_split_prob_a2t_0811:
            c.dic_coeff_adj_a2t['0811'] = c.dic_coeff_adj_a2t_pw['0811'][0]
        else:
            c.dic_coeff_adj_a2t['0811'] = c.dic_coeff_adj_a2t_pw['0811'][1]
        if t_start<c.dt_split_prob_a2t_sin12:
            c.dic_coeff_adj_a2t['sin12'] = c.dic_coeff_adj_a2t_pw['sin12'][0]
        else:
            c.dic_coeff_adj_a2t['sin12'] = c.dic_coeff_adj_a2t_pw['sin12'][1]
        
        ####use piece-wise coeff_prob_dead_ltfu_noart to account for elevated overdose risk (most likely) since 2014 among those not on ART
        if t_start<c.dt_split_prob_dead_noart: 
            c.coeff_prob_dead_ltfu_noart = c.coeff_prob_dead_ltfu_noart_pw[0]
        else:
            c.coeff_prob_dead_ltfu_noart = c.coeff_prob_dead_ltfu_noart_pw[1]

        ####initialize columns for t_stop for simulated results
        data_hiv_sim[str(t_stop)[:10]] = data_hiv_sim[str(t_start)[:10]].copy() #initialize new hiv status by t_stop, 'D' from previous step will be copied and no longer requires further modification
        for key_i in dict_reg_status_sim.keys():
                dict_reg_status_sim[key_i][str(t_stop)[:10]] = dict_reg_status_sim[key_i][str(t_start)[:10]].copy() #regimen type will be determined every 6 months
        if i==1: #keep pvl/reg_prop the same as t_start for i=1, the rest can be derived for i>=2
            data_pvl_sim[str(t_stop)[:10]] = data_pvl_sim[str(t_start)[:10]].copy() #pvl status will be updated after one year
            for key_i in dict_reg_prop_sim.keys():
                dict_reg_prop_sim[key_i][str(t_stop)[:10]] = dict_reg_prop_sim[key_i][str(t_start)[:10]].copy() #regimen status will be updated after one year

        ####only consider the probabilities among those alive at t_start
        arr_id_plwh = data_hiv_sim['moh_id_sim'][data_hiv_sim[str(t_start)[:10]].isin(c.arr_state_alive)].values #introduce arr_state_alive so that PLWH can be selected at each step if HIV status is updated
        data_char_plwh = data_char_sim[data_char_sim['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'})
        data_hiv_plwh = data_hiv_sim[data_hiv_sim['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'})
        data_comorb_plwh = data_comorb_sim[data_comorb_sim['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'})
        data_pvl_plwh = data_pvl_sim[data_pvl_sim['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'})
        dict_reg_status_plwh = {k:dict_reg_status_sim[k][dict_reg_status_sim[k]['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'}) for k in dict_reg_status_sim.keys()}
        dict_reg_prop_plwh = {k:dict_reg_prop_sim[k][dict_reg_prop_sim[k]['moh_id_sim'].isin(arr_id_plwh)].sort_values(by='moh_id_sim').reset_index(drop=True).drop(columns=['moh_id']).rename(columns={'moh_id_sim':'moh_id'}) for k in dict_reg_prop_sim.keys()}

        ####update covariants for mortality and other comorbidities based on simulated cohort on t_start
        data_comorb_var = func_update_comorb_variable(c,data_comorb_plwh,data_char_plwh,data_hiv_plwh,data_pvl_plwh,dict_reg_prop_plwh,t_start) #separate derivation of covariates to estimate probability of comorbidity incidence

        ####determine the probability of each comorbidity and simulate comorbidity status, which is independent from mortality and ART module
        for comorb in c.list_comorb:
            data_prob_comorb = c.dic_prob_comorb[comorb](c,data_comorb_plwh,data_hiv_plwh,data_comorb_var,t_start) #only use data_char_plwh,data_pvl_plwh and dict_reg_plwh in data_comorb_var
            assert (data_prob_comorb[pd.isnull(data_prob_comorb['prob_wo_comorb'])].shape[0]==0), 'Invalid probability estimation for comorbidities'
            data_trans_comorb = pd.DataFrame({'moh_id':data_prob_comorb['moh_id'].astype(int),'0':np.zeros(data_prob_comorb.shape[0],dtype=int)})
            data_trans_comorb['1'] = pd.Series(sample_from_prob_matrix(data_prob_comorb[['prob_wo_comorb','prob_w_comorb']].values))
            arr_id_comorb_inc = data_trans_comorb['moh_id'][data_trans_comorb['1']==1].values
            data_comorb_sim.loc[data_comorb_sim['moh_id_sim'].isin(arr_id_comorb_inc),'earliest_'+comorb+'_dt'] = t_stop #set diagnosis date of comorbidity as t_stop for newly simulated incident cases in the full dataset, not for plwh alive at t_start
            
        ####determine the probability of the mortality for each individual and simulate mortality status
        data_prob_dead = func_prob_update_mortality(c,data_comorb_plwh,data_hiv_plwh,data_comorb_var,t_start) 
        assert (data_prob_dead[pd.isnull(data_prob_dead['prob_wo_comorb'])].shape[0]==0), 'Invalid probability estimation for mortality'
        data_trans_dead = pd.DataFrame({'moh_id':data_prob_dead['moh_id'].astype(int),'0':np.zeros(data_prob_dead.shape[0],dtype=int)})
        data_trans_dead['1'] = pd.Series(sample_from_prob_matrix(data_prob_dead[['prob_wo_comorb','prob_w_comorb']].values))
        arr_id_new_dead = data_trans_dead['moh_id'][data_trans_dead['1']==1].values
        data_char_sim.loc[data_comorb_sim['moh_id_sim'].isin(arr_id_new_dead),'end_fu_dt_sim'] = t_stop #convert timestamp to datetime to be consistent with dummy_end_date
        data_char_sim['end_fu_dt_sim'] = pd.to_datetime(data_char_sim['end_fu_dt_sim'].dt.date) #only keep the date information using datetime format
        data_hiv_plwh.loc[data_hiv_plwh['moh_id'].isin(arr_id_new_dead),str(t_stop)[:10]] = 'D' #modify data_plwh for simulation later to exclude the dead
        data_hiv_sim.loc[data_hiv_sim['moh_id_sim'].isin(arr_id_new_dead),str(t_stop)[:10]] = 'D'

        ####for those without ART initiation and alive by t_stop, determine status for ART initiation
        arr_id_plwh_A = data_hiv_plwh['moh_id'][(data_hiv_plwh[str(t_start)[:10]]=='A')&(data_hiv_plwh[str(t_stop)[:10]]!='D')].values
        data_char_A = data_char_plwh[data_char_plwh['moh_id'].isin(arr_id_plwh_A)].sort_values(by='moh_id').reset_index(drop=True)
        data_prob_a2t = func_prob_update_a2t(c,data_char_A,t_start,c.rand_blup_a2t) 
        assert (data_prob_a2t[pd.isnull(data_prob_a2t['prob_a2t'])].shape[0]==0), 'Invalid probability estimation for transition A2T'
        data_trans_a2t = pd.DataFrame({'moh_id':data_prob_a2t['moh_id'].astype(int),'0':np.zeros(data_prob_a2t.shape[0],dtype=int)})
        data_trans_a2t['1'] = pd.Series(sample_from_prob_matrix(data_prob_a2t[['prob_a2a','prob_a2t']].values)) #aware that probability for 0 should be the first one in the list
        data_trans_a2t = data_trans_a2t.merge(data_char_A[['moh_id','baseline_dt','counter']].copy(),how='left',on='moh_id')
        arr_id_a2t = data_trans_a2t['moh_id'][(data_trans_a2t['0']==0)&(data_trans_a2t['1']==1)].values
        data_trans_a2t['t_sim'] = data_trans_a2t['moh_id'].apply(lambda x: np.random.uniform(0,1,1)[0] if x in arr_id_a2t else np.nan) #assume ART initiation between t_start and t_stop instead of baseline+counter*dt*def_year for simplification
        data_trans_a2t['FARVDT_sim'] = (data_trans_a2t['baseline_dt']+pd.to_timedelta(c.def_year*c.dt*((data_trans_a2t['counter']-1)+data_trans_a2t['t_sim']),unit='d')).dt.round('d') #FARVDT_sim should be updated based on counter and baseline_dt instead of t_start and t_stop
        data_char_sim.loc[data_char_sim['moh_id_sim'].isin(arr_id_a2t),'FARVDT_sim'] = data_trans_a2t['FARVDT_sim'][data_trans_a2t['moh_id'].isin(arr_id_a2t)].values 
        data_char_sim.loc[data_char_sim['moh_id_sim'].isin(arr_id_a2t),'art_baseline_dt'] = data_char_sim['FARVDT_sim'][data_char_sim['moh_id_sim'].isin(arr_id_a2t)]
        data_char_sim['age_art_bsln_scale10_dv'] = (data_char_sim['art_baseline_dt']-data_char_sim['DOB']).dt.total_seconds()/(60*60*24*c.def_year)/10 #update age at art_baseline after data update
        data_hiv_sim.loc[data_hiv_sim['moh_id_sim'].isin(arr_id_a2t),str(t_stop)[:10]] = random.choices(c.arr_state_art,weights=[c.prob_A2S,c.prob_A2Uon,1-c.prob_A2S-c.prob_A2Uon],k=len(arr_id_a2t))

        ####for those with ART initiation and alive by t_stop, determine status among S, U_on and U_off
        arr_id_plwh_T = data_hiv_plwh['moh_id'][(data_hiv_plwh[str(t_start)[:10]].isin(['S','U_off','U_on']))&(data_hiv_plwh[str(t_stop)[:10]]!='D')].values
        data_char_T = data_char_plwh[data_char_plwh['moh_id'].isin(arr_id_plwh_T)].sort_values(by='moh_id').reset_index(drop=True)
        data_comorb_T = data_comorb_var[['moh_id','manx_dv','prsn_dv','sczo_dv']][data_comorb_var['moh_id'].isin(arr_id_plwh_T)].sort_values(by='moh_id').reset_index(drop=True) 
        data_char_T = data_char_T.merge(data_comorb_T,how='left',on='moh_id')
        data_hiv_T = data_hiv_plwh[data_hiv_plwh['moh_id'].isin(arr_id_plwh_T)].sort_values(by='moh_id').reset_index(drop=True)
        data_prob_trans_art = func_prob_update_trans_art(c,data_hiv_T,data_char_T,t_start,c.rand_blup_art) #add rand_art=0/1 representing whether to use random terms in probability estimation
        data_prob_trans_art = func_prob_adj_trans_art(c,data_prob_trans_art) 
        data_prob_trans_art = data_prob_trans_art[['moh_id','prob_s_adj3','prob_off_adj3','prob_on_adj3']].copy().rename(columns={'prob_s_adj3':'prob_s','prob_off_adj3':'prob_off','prob_on_adj3':'prob_on'})
        data_trans_art = data_hiv_T[['moh_id',str(t_start)[:10]]].copy().rename(columns={str(t_start)[:10]:'0'}) #keep state at t_start to count transitions from different states
        data_trans_art['1'] = pd.Series(sample_from_prob_matrix(data_prob_trans_art[['prob_off','prob_on','prob_s']].values)) 
        data_trans_art = data_trans_art.replace({'1':{0:'U_off',1:'U_on',2:'S'}}) #replace sampled number back to strings that represent suppression/unsuppressed(off ART/on ART)
        arr_id_trans_art = data_trans_art['moh_id'][(data_trans_art['0']!=data_trans_art['1'])].values
        data_hiv_sim.loc[data_hiv_sim['moh_id_sim'].isin(arr_id_trans_art),str(t_stop)[:10]] = data_trans_art['1'][data_trans_art['moh_id'].isin(arr_id_trans_art)].values

        ####for those newly initiated ART by t_stop, assign regimen type at t_stop
        c.init_dist_dt = c.reg_dist['init']['date'][c.reg_dist['init']['date']>=str(t_stop)[:10]].min() #distribution among new enrollments started from 2008-12-31
        c.super_dist_dt = c.reg_dist['super']['date'][c.reg_dist['super']['date']<str(t_stop)[:10]].max() #use the latest distribution before t_stop
        c.super_dist = [c.reg_dist['super']['percent'][(c.reg_dist['super']['date']==c.super_dist_dt)&(c.reg_dist['super']['class_nbb_super']==state_i)].values[0] for state_i in c.super_state]
        dict_reg_status_sim = func_reg_add_new(c,dict_reg_status_sim,t_start,t_stop,arr_id_a2t)
        
        ####for those with ART initiation and alive by t_stop, assign regimen type at t_stop
        arr_id_stay = data_hiv_sim['moh_id_sim'][(data_hiv_sim[str(t_start)[:10]].isin(['S','U_on']))&(data_hiv_sim[str(t_stop)[:10]].isin(['S','U_on']))].values #new regimen type will be added in dict_reg_status_sim
        dict_bb_state = {'tdf':['tdf','other'], 'other':['other','tdf']}
        dict_bb_p = {}
        for bb_j in c.list_bb:
            if str(t_stop)[:10]<str(c.dic_reg_trans_p[bb_j][0])+'-01-01':
                dict_bb_p[bb_j] = c.dic_reg_trans_p[bb_j][1]
            else:
                dict_bb_p[bb_j] = c.dic_reg_trans_p[bb_j][2]
        data_bb_sim = pd.DataFrame({'moh_id':arr_id_stay,'bb_old':dict_reg_status_plwh['tdf'][str(t_start)[:10]][dict_reg_status_plwh['tdf']['moh_id'].isin(arr_id_stay)].values})
        data_bb_sim['bb_old'] = data_bb_sim['bb_old'].replace({1:'tdf',0:'other'})
        data_bb_sim['bb'] = data_bb_sim['bb_old'].apply(lambda x:random.choices(dict_bb_state[x],weights=dict_bb_p[x],k=1)[0])
        data_bb_sim[['bb_old','bb']] = data_bb_sim[['bb_old','bb']].replace({'tdf':1,'other':0})
        arr_id_tdf_change = data_bb_sim['moh_id'][data_bb_sim['bb_old']!=data_bb_sim['bb']].values #only change values for a small proportion of PLWH
        dict_reg_status_sim['tdf'].loc[dict_reg_status_sim['tdf']['moh_id_sim'].isin(arr_id_tdf_change),str(t_stop)[:10]] = data_bb_sim['bb'][data_bb_sim['moh_id'].isin(arr_id_tdf_change)].values
        dict_nbb_state = {}
        dict_nbb_p = {}
        for nbb_j in c.list_nbb:
            list_nbb_rest = c.list_nbb[:]
            list_nbb_rest.remove(nbb_j)
            list_nbb_trans = list_nbb_rest[:]
            list_nbb_trans.insert(0,nbb_j)
            dict_nbb_state[nbb_j] = list_nbb_trans
            if str(t_stop)[:10]<str(c.dic_reg_trans_p[nbb_j][0])+'-01-01':
                dict_nbb_p[nbb_j] = c.dic_reg_trans_p[nbb_j][1]
            else:
                dict_nbb_p[nbb_j] = c.dic_reg_trans_p[nbb_j][2]
        dict_id_nbb = {}
        for reg_i in c.list_nbb[:-1]: #find individuals with different regimen type
            dict_id_nbb[reg_i] = dict_reg_status_plwh[reg_i]['moh_id'][(dict_reg_status_plwh[reg_i][str(t_start)[:10]]==1)&(dict_reg_status_plwh[reg_i]['moh_id'].isin(arr_id_stay))].values
        dict_id_nbb['super'] = np.array(list((set(dict_id_nbb['nnrti'])&set(dict_id_nbb['pi']))|(set(dict_id_nbb['nnrti'])&set(dict_id_nbb['insti']))|(set(dict_id_nbb['insti'])&set(dict_id_nbb['pi'])))) #set operation works to determine 'super' as dict_reg_status indicates the regimen type on t_start/t_stop instead of during the period
        dict_id_nbb['super'].sort() #sort in ascending order to be consistent with id in dataframe
        for reg_i in c.list_nbb[:-1]: #super will be considered separately
            arr_id_reg = np.array(list(set(dict_id_nbb[reg_i])-set(dict_id_nbb['super']))) #simulate transitions for each regimen type
            arr_id_reg.sort() #sort in ascending order to be consistent with id in dataframe
            data_nbb_sim = pd.DataFrame({'moh_id':arr_id_reg,'nbb_old':reg_i}) #can be simplified as arr_id_reg with reg_i=1
            data_nbb_sim['nbb'] = data_nbb_sim['nbb_old'].apply(lambda x:random.choices(dict_nbb_state[x],weights=dict_nbb_p[x],k=1)[0])
            for reg_j in dict_nbb_state[reg_i][1:]: #loop except reg_i itself
                arr_id_reg_change = data_nbb_sim['moh_id'][(data_nbb_sim['nbb']==reg_j)].values
                if reg_j!='super':
                    dict_reg_status_sim[reg_i].loc[dict_reg_status_sim[reg_i]['moh_id_sim'].isin(arr_id_reg_change),str(t_stop)[:10]] = 0. #indicate regimen type changed from reg_i to reg_j
                    dict_reg_status_sim[reg_j].loc[dict_reg_status_sim[reg_j]['moh_id_sim'].isin(arr_id_reg_change),str(t_stop)[:10]] = 1. 
                else:
                    data_nbb_sim.loc[data_nbb_sim['moh_id'].isin(arr_id_reg_change),'nbb'] = random.choices(c.super_state,weights=c.super_dist,k=len(arr_id_reg_change))
                    for reg_k in c.list_nbb[:-1]:
                        if reg_k==reg_i:
                            arr_id_super_reg2 = data_nbb_sim['moh_id'][(data_nbb_sim['moh_id'].isin(arr_id_reg_change))&(data_nbb_sim['nbb'].str.contains(reg_k)==False)].values #for those super does not contain reg_k, change t_stop=0
                            dict_reg_status_sim[reg_i].loc[dict_reg_status_sim[reg_i]['moh_id_sim'].isin(arr_id_super_reg2),str(t_stop)[:10]] = 0.
                            #print (reg_i,reg_k,len(arr_id_super_reg2))
                        arr_id_super_reg = data_nbb_sim['moh_id'][(data_nbb_sim['moh_id'].isin(arr_id_reg_change))&(data_nbb_sim['nbb'].str.contains(reg_k))].values #id_super_reg can be changed for reg_i with super=nnrti_pi separately, both lead to dict_reg_status_sim[reg_i]=0
                        dict_reg_status_sim[reg_k].loc[dict_reg_status_sim[reg_k]['moh_id_sim'].isin(arr_id_super_reg),str(t_stop)[:10]] = 1.
        arr_id_super = dict_id_nbb['super'] #already sorted ascending
        data_nbb_sim = pd.DataFrame({'moh_id':arr_id_super,'nbb_old':'super'}) #can be simplified as arr_id_reg with reg_i=1
        data_nbb_sim['nbb'] = data_nbb_sim['nbb_old'].apply(lambda x:random.choices(dict_nbb_state[x],weights=dict_nbb_p[x],k=1)[0])
        arr_id_super_change = data_nbb_sim['moh_id'][(data_nbb_sim['nbb']!='super')].values
        data_super_sim = data_nbb_sim[data_nbb_sim['moh_id'].isin(arr_id_super_change)].reset_index(drop=True)
        data_super_sim['super_old'] = data_super_sim['moh_id'].apply(lambda x: '_'.join([reg_m for reg_m in ['nnrti','pi','insti'] if x in dict_id_nbb[reg_m]]))
        for reg_m in ['nnrti','pi','insti']:
            arr_id_reg_to1 = data_super_sim['moh_id'][(data_super_sim['nbb']==reg_m)&(data_super_sim['super_old'].str.contains(reg_m)==False)].values
            dict_reg_status_sim[reg_m].loc[dict_reg_status_sim[reg_m]['moh_id_sim'].isin(arr_id_reg_to1),str(t_stop)[:10]] = 1.
            arr_id_reg_to0 = data_super_sim['moh_id'][(data_super_sim['nbb']!=reg_m)&(data_super_sim['super_old'].str.contains(reg_m))].values
            dict_reg_status_sim[reg_m].loc[dict_reg_status_sim[reg_m]['moh_id_sim'].isin(arr_id_reg_to0),str(t_stop)[:10]] = 0.
        
        ####for those transitted from U_off to S/U_on by t_stop, assign regimen type at t_stop
        arr_id_trans = data_hiv_sim['moh_id_sim'][(data_hiv_sim[str(t_start)[:10]]=='U_off')&(data_hiv_sim[str(t_stop)[:10]].isin(['S','U_on']))].values
        dict_reg_status_sim = func_reg_add_new(c,dict_reg_status_sim,t_start,t_stop,arr_id_trans)

        ####for those transitted from S/U-on to U_off by t_stop, assign 0 at t_stop for each regimen dataframe
        arr_id_off = data_hiv_sim['moh_id_sim'][(data_hiv_sim[str(t_stop)[:10]]=='U_off')].values #no need to consider status at t_start, including U_off at t_start
        for reg_i in dict_reg_status_sim.keys():
            dict_reg_status_sim[reg_i].loc[dict_reg_status_sim[reg_i]['moh_id_sim'].isin(arr_id_off),str(t_stop)[:10]] = 0.

        ####update data_pvl_sim and dict_reg_prop_sim every 1 year based on data_hiv_sim and dict_reg_status 
        if i>=2: #look at 1-year period (two time steps) before t_stop to determine the status at t_stop
            t_prev1yr = c.t0+pd.to_timedelta(c.def_year*c.dt*(i-2),unit='d') #define the time step one year agao  
            list_dt = [str(t_prev1yr)[:10],str(t_start)[:10],str(t_stop)[:10]]
            data_hiv_sim['comb'] = data_hiv_sim[list_dt].apply(lambda x: "".join(list(x)),axis=1)
            data_hiv_sim['pvl'] = data_hiv_sim['comb'].apply(lambda x: 0 if x=='SSS' else 1) #0 indicates suppression
            data_pvl_sim[str(t_stop)[:10]] = data_hiv_sim['pvl'].values #order in moh_id_sim should be the same for data_hiv_sim and data_pvl_sim
            for k in dict_reg_prop_sim.keys():
                dict_reg_status_sim[k][list_dt] = dict_reg_status_sim[k][list_dt].fillna(4) #replace NaN by 4 to locate the right proportion in c.dic_reg_val
                dict_reg_status_sim[k]['comb'] = dict_reg_status_sim[k][list_dt].apply(lambda x: "".join([str(int(ele)) for ele in x]),axis=1)
                dict_reg_status_sim[k]['prop'] = dict_reg_status_sim[k]['comb'].apply(lambda x: c.dic_reg_prop[x]) #will not fully executed if x outside of keys of dic_reg_val
                dict_reg_prop_sim[k][str(t_stop)[:10]] = dict_reg_status_sim[k]['prop'].values #order of dict_reg_status/prop shoud be the same
                dict_reg_status_sim[k][list_dt] = dict_reg_status_sim[k][list_dt].replace({4:np.nan}) #change no ART initiation back to NaN

        ####introduce new diagnosis based on updated undiagnosed PLWH with new infections
        year_start = c.year0+c.dt*(i-1)
        year_stop = c.year0+c.dt*i
        prob_d = 1-np.exp(-c.d_U*c.dt) #estimate probability of dead among undiagnosed PLWH in 6 months and randomize the number of undiagnosed PLWH alive
        n_dead_undiag = np.random.binomial(n_undiag,prob_d) #each individual survival same as success in one bernoulli trial
        r_diag = sigmoid(c.sig_rdiag_params[0],c.sig_rdiag_params[1],c.sig_rdiag_params[2],year_start-c.sig_rdiag_params[3]) #among those alive by t_stop, determine probability of diagnosis over 6 months and randomize the number of new diagnoses
        prob_diag = c.coeff_rdiag*(1-np.sqrt(1-r_diag)) #treat r_diag as probability of diagnosis instead of rates, as rates cannot be estimated due to unknown follow-up time for undiagnosed PLWH, coeff_rdiag used to fit to historical cum. new diagnosis 
        n_newdiag = np.random.binomial(n_undiag-n_dead_undiag,prob_diag) #based on n_undiag and n_unsupp_diag at t_start, estimate the number of new HIV cases 
        iur200 = sigmoid(c.sig_iur_params[0],c.sig_iur_params[1],c.sig_iur_params[2],year_start-c.sig_iur_params[3])
        n_newinc = iur200*(n_undiag+n_unsupp_diag)*c.dt
        n_newmigr = sigmoid(c.sig_migr_params[0],c.sig_migr_params[1],c.sig_migr_params[2],year_start-c.sig_migr_params[3])*c.dt #simulate the number of new migrants aside from new diagnosis from local undiagnosed population using a sigmoid function
        n_undiag = n_undiag - n_dead_undiag - n_newdiag + n_newinc

        ####create dictionary for characteristics for new diagnosis and migrants (new to microsimulation in STOP) and update dataframes for comorb/hiv/pvl/reg_status/reg_prop 
        dict_new_sim = func_init_new_rand(c,dict_hist_data,t_start,t_stop,int(np.round(n_newdiag+n_newmigr)))
        data_char_new = dict_new_sim['char'].copy()
        data_char_new['FARVDT_sim'] = pd.to_datetime(np.nan) #make type=datetime64 to be consistent with FARVDT_sim in data_char_sim
        data_char_new['art_baseline_dt'] = data_char_new['FARVDT_sim'].copy() 
        data_char_new['age_art_bsln_scale10_dv'] = np.nan #change age_art_bsln_scale10_dv to be nan for new diagnosis
        data_char_new['end_fu_dt_sim'] = c.dummy_end_dt #add simulated end_fu_dt
        data_char_new['blup'] = data_char_new['a2t_cat'].apply(lambda x: np.random.normal(0.,np.sqrt(c.dic_random_a2t[x][0]))) 
        data_char_new['blup_off'] = np.random.normal(0.,np.sqrt(c.dic_random_t['u_off'][0]),size=data_char_new.shape[0])
        data_char_new['blup_on'] = np.random.normal(0.,np.sqrt(c.dic_random_t['u_on'][0]),size=data_char_new.shape[0])
        data_char_sim = pd.concat([data_char_sim,data_char_new],ignore_index=True) 
        data_char_sim = data_char_sim.assign(moh_id_sim=data_char_sim.index)
        data_hiv_sim = pd.concat([data_hiv_sim,dict_new_sim['hiv']],ignore_index=True)
        data_hiv_sim = data_hiv_sim.assign(moh_id_sim=data_hiv_sim.index)
        data_comorb_sim = pd.concat([data_comorb_sim,dict_new_sim['comorb']],ignore_index=True)
        data_comorb_sim = data_comorb_sim.assign(moh_id_sim=data_comorb_sim.index)
        data_pvl_sim = pd.concat([data_pvl_sim,dict_new_sim['pvl']],ignore_index=True)
        data_pvl_sim = data_pvl_sim.assign(moh_id_sim=data_pvl_sim.index)
        dict_reg_status_sim = {k:pd.concat([v,dict_new_sim['reg_status'][k]],ignore_index=True) for k,v in dict_reg_status_sim.items()} 
        dict_reg_status_sim = {k:v.assign(moh_id_sim=v.index) for k,v in dict_reg_status_sim.items()} 
        dict_reg_prop_sim = {k:pd.concat([v,dict_new_sim['reg_prop'][k]],ignore_index=True) for k,v in dict_reg_prop_sim.items()}
        dict_reg_prop_sim = {k:v.assign(moh_id_sim=v.index) for k,v in dict_reg_prop_sim.items()} 
        
        ####update dataframe for new participants for microsimulation over time
        n_unsupp_diag = data_hiv_sim[data_hiv_sim[str(t_stop)[:10]].isin(['A','U_on','U_off'])].shape[0]
        n_spvl = data_hiv_sim[data_hiv_sim[str(t_stop)[:10]]=='S'].shape[0]
        n_plwh = n_undiag+n_unsupp_diag+n_spvl
        dict_newmicro = {
            'date':str(t_stop)[:10], 'year':year_stop,
            'new_inc':n_newinc, 'new_diag':n_newdiag, 'new_dead_undiag':n_dead_undiag, 'new_migr':n_newmigr,
            'plwh':n_plwh, 'spvl':n_spvl, 'unsupp_diag':n_unsupp_diag, 'undiag':n_undiag,
        }
        data_newmicro = pd.concat([data_newmicro,pd.DataFrame(dict_newmicro,index=[0])],ignore_index=True)
 
    #####create dictionary to save dataframes used for outcome derivation
    dict_data_sim = {
        'char':data_char_sim.copy(),
        'hiv':data_hiv_sim.copy(), 
        'comorb':data_comorb_sim.copy(), #pvl,reg_status/prop were used for probability estimation, not final outcomes
        'new':data_newmicro.copy(), #information of plwh, undiag, new_inc, new_diag and new_migr outputed as outcomes just in case
        }

    #return dict_outcomes
    return dict_data_sim #show the simulated cohort information

def mp_microsim_outcomes_sim(c,dict_hist_data,seed_i):
    """return to derived outcomes for final simulations with randomized new diagnosis/migrants"""

    dict_micro_results = mp_microsim_all_rand(c,dict_hist_data,seed_i)
    dict_outcomes = func_micro_sim_outcomes(c,dict_micro_results) #derived outcomes for final simulation

    return dict_outcomes

def mp_microsim_inc_undiag_test(c,data_hist,seed_i):
    """return to dataframe of new incidence, new diagnosis, deaths among undiagnosed PLWH, and undiagnosed PLWH over time;
    separate new migrants from new diagnosis from local undiagnosed population to better fit PHAC's records"""

    #####initialize random generator for each simulation of the cohort
    random.seed(seed_i)
    np.random.seed(seed_i)

    #####introduce initial HIV prevalence and diagnosed PLWH in year0 and saved in the dataframe for final outcome
    n_plwh = c.init_prev
    n_unsupp_diag = data_hist['unsupp_diag'][data_hist['year']==c.year0].values[0]
    n_spvl = data_hist['spvl'][data_hist['year']==c.year0].values[0]
    n_undiag = n_plwh-n_unsupp_diag-n_spvl
    data_outcomes = pd.DataFrame([],columns=['date','year','new_inc','new_dead_undiag','new_diag','new_migr','undiag','unsupp_diag','spvl','plwh'])
    dict_outcomes = {
        'date':c.t0, 'year':c.year0, 
        'plwh':n_plwh, 'spvl':n_spvl, 'unsupp_diag':n_unsupp_diag, 'undiag':n_undiag,
    }
    data_outcomes = pd.concat([data_outcomes,pd.DataFrame(dict_outcomes,index=[0])],ignore_index=True)

    #####estimate probability of death and diagnosis every 6 months and randomization for new incidence, new diagnosis and new deaths among undiagnosed PLWH
    for i in range(1,c.n_step+1):
        
        ####define the beginning and the end of the time step
        t_start = c.t0+pd.to_timedelta(c.def_year*c.dt*(i-1),unit='d')
        t_stop = t_start+pd.to_timedelta(c.def_year*c.dt,unit='d')
        year_start = c.year0+c.dt*(i-1)
        year_stop = year_start+c.dt

        ####estimate probability of dead among undiagnosed PLWH in 6 months and randomize the number of undiagnosed PLWH alive
        prob_d = 1-np.exp(-c.d_U*c.dt)
        n_dead_undiag = np.random.binomial(n_undiag,prob_d) #each individual survival same as success in one bernoulli trial

        ####among those alive by t_stop, determine probability of diagnosis over 6 months and randomize the number of new diagnoses
        r_diag = sigmoid(c.sig_rdiag_params[0],c.sig_rdiag_params[1],c.sig_rdiag_params[2],year_start-c.sig_rdiag_params[3])
        prob_diag = c.coeff_rdiag*(1-np.sqrt(1-r_diag))
        n_newdiag = np.random.binomial(n_undiag-n_dead_undiag,prob_diag)

        ####based on n_undiag and n_unsupp_diag at t_start, estimate the number of new HIV cases based on IUR200
        iur200 = sigmoid(c.sig_iur_params[0],c.sig_iur_params[1],c.sig_iur_params[2],year_start-c.sig_iur_params[3])
        n_newinc = iur200*(n_undiag+n_unsupp_diag)*c.dt

        ####simulate the number of new migrants aside from new diagnosis from local undiagnosed population using a sigmoid function
        n_newmigr = sigmoid(c.sig_migr_params[0],c.sig_migr_params[1],c.sig_migr_params[2],year_start-c.sig_migr_params[3])*c.dt

        ####update the number of undiagnosed PLWH and PLWH based on estimates above and historical unsupp_diag and spvl at t_stop
        n_unsupp_diag = data_hist['unsupp_diag'][data_hist['year']==year_stop].values[0]
        n_spvl = data_hist['spvl'][data_hist['year']==year_stop].values[0]
        n_undiag = n_undiag - n_dead_undiag - n_newdiag + n_newinc
        n_plwh = n_undiag+n_unsupp_diag+n_spvl

        ####save results to dataframe
        dict_outcomes = {
            'date':t_stop, 'year':year_stop,
            'new_inc':n_newinc, 'new_diag':n_newdiag, 'new_dead_undiag':n_dead_undiag, 'new_migr':n_newmigr,
            'plwh':n_plwh, 'spvl':n_spvl, 'unsupp_diag':n_unsupp_diag, 'undiag':n_undiag,
        }
        data_outcomes = pd.concat([data_outcomes,pd.DataFrame(dict_outcomes,index=[0])],ignore_index=True)

    #####derive annual outcomes based on results every 6 months
    data_annual = data_outcomes[['date','year','undiag','unsupp_diag','spvl','plwh']][data_outcomes.index.isin(data_outcomes.index.values[::2])].reset_index(drop=True)
    for key_i in ['new_inc','new_diag','new_dead_undiag','new_migr']:
        arr_sim = data_outcomes[key_i].values[1:] #keep non-nan values
        annual_sim = [sum(arr_sim[i:i+2]) for i in range(0,len(arr_sim),2)]
        annual_sim.insert(0,np.nan)
        data_annual[key_i] = pd.Series(annual_sim)

    #return data_outcomes
    return data_annual #focus on annual results

def func_cali_target_micro_rdiag(arr_params,data_for_diag,data_hist,c):
    """estimate sum of squared residual between simulated average based on microsimulation and historical new diagnoses
    data_for_diag has information on historical spvl and unsuppressed PLWH, used to simulated new diagnoses
    data_hist has information on historical new diagnoses, used for residual calculation"""

    ####update the coefficient to adjust probability of diagnosis
    c.coeff_rdiag = arr_params[0]

    ####run microsimulation based on historical records for new diagnoses
    pool = Pool(processes=c.n_pool,maxtasksperchild=1) 
    results = [pool.apply_async(mp_microsim_inc_undiag_test,args=(c,data_for_diag,c.arr_seed[i])) for i in range(c.num_sim)]
    list_results = [x.get() for x in results] 

    ####estimate number of new STOP participants and estimate residuals based on historical records
    data_sim_diag = func_micro_1outcome_ci(c,list_results,'new_diag','inc')
    data_sim_migr = func_micro_1outcome_ci(c,list_results,'new_migr','inc')
    arr_newdiag = data_sim_diag['new_diag avg'][data_sim_diag['year']>c.year0].values+data_sim_migr['new_migr avg'][data_sim_migr['year']>c.year0].values

    ####estimate residual based on historical data
    data_hist = data_hist[data_hist['year'].isin(data_sim_diag['year'].values[1:])].reset_index(drop=True) 
    res_check = (np.cumsum(arr_newdiag)-data_hist['new_diag'].cumsum().values)**2
    res_final = res_check[-1] 

    return res_final #consider cumulative new diagnoses in 2016

def func_cali_target_micro_all(arr_params,dict_hist_target,dict_data,c,arg_cali):
    """estimate sum of squared residual between simulated average based on microsimulation and historical records
    arr_params includes 21 values to adjust probabilities for comorbidities (12, including mortality), ART initiation (3) and ART churn (6)
    dict_data contains characteristics and HIV status inforamtion for microsimulation
    arg_cali=all/comorb etc to represent calibration on different compartments"""

    ####separate dataframe of hiv and comorb for historical data
    data_hiv_hist = dict_hist_target['hiv']
    data_comorb_hist = dict_hist_target['comorb']

    ####update coefficients to adjust probabilities
    if arg_cali=='all':
        assert (len(arr_params)==21), 'Double check number of coefficients adjusted for calibration on HIV and comorbidities'
        c.dic_prob_comorb_coeff = {k:arr_params[i] for i,k in enumerate(list(c.dic_prob_comorb_coeff))} 
        c.dic_coeff_adj_a2t = {k:arr_params[len(list(c.dic_prob_comorb_coeff))+i] for i,k in enumerate(list(c.dic_coeff_adj_a2t))} 
        c.dic_coeff_prob_art = {k:arr_params[len(list(c.dic_prob_comorb_coeff))+len(list(c.dic_coeff_adj_a2t))+i] for i,k in enumerate(list(c.dic_coeff_prob_art))}
    elif arg_cali=='comorb':
        assert (len(arr_params)==11), 'Double check number of coefficients adjusted for calibration on comorbidities'
        for i,comorb_i in enumerate(c.list_comorb):
            c.dic_prob_comorb_coeff[comorb_i] = arr_params[i]
    elif arg_cali in c.list_comorb:
        assert (len(arr_params)==1), 'Double check number of coefficients adjusted for calibration on one comorbidity'
        c.dic_prob_comorb_coeff[arg_cali] = arr_params[0]
    elif arg_cali in ['new_art bf08','new_art 0811','new_art sin12']:

        ####introduce piece-wise probability adjustment before and since 2012
        assert (len(arr_params)==2), 'Double check number of coefficients adjusted for ART initiation based on diagnosis year'
        c.dic_coeff_adj_a2t_pw[arg_cali[8:]] = np.copy(arr_params)

    elif arg_cali=='new_art': #adjusting three ART coefficients together doesn't work when coeff_bf08 jumped from 0.3 to -0.1 with initial guess=1
        assert (len(arr_params)==6), 'Double check number of coefficients adjusted for ART initiation'
        c.dic_coeff_adj_a2t_pw['bf08'] = arr_params[:2]
        c.dic_coeff_adj_a2t_pw['0811'] = arr_params[2:4]
        c.dic_coeff_adj_a2t_pw['sin12'] = arr_params[4:] 

    elif arg_cali=='trans_art':
        assert (len(arr_params)==6), 'Double check number of coefficients adjusted for transitions after ART initiation'
        c.dic_coeff_prob_art = {k:arr_params[i] for i,k in enumerate(list(c.dic_coeff_prob_art))}
    elif arg_cali=='dead': #only adjust deaths

        ####introduce coeff_prob_dead_ltfu_noart_pw to improve fitting on A2D
        assert (len(arr_params)==4), 'Double check number of coefficients adjusted for mortality from different states'
        c.coeff_prob_dead_ltfu_art = arr_params[0] #introduce coeff_prob_dead_ltfu_art/noart to improve fitting to deaths from different states
        c.coeff_prob_dead_ltfu_noart_pw = np.array([arr_params[1],arr_params[2]])
        c.dic_prob_dead_art['pvl_dv'][0] = arr_params[3]

    #####initialize multiprocessing
    pool = Pool(processes=c.n_pool,maxtasksperchild=1) 

    #####run microsimulation using multiprocessing to get derived outcomes for calibration
    results = [pool.apply_async(mp_microsim_outcomes_cali,args=(c,dict_data,c.arr_seed[i])) for i in range(c.num_sim)]
    list_results = [x.get() for x in results] 
    list_hiv = [result['hiv'] for result in list_results]
    list_comorb = [result['comorb'] for result in list_results]

    #####use mean to calculate residuals between simulated and historical targets
    dict_sim_mean = {} #add annual average as outcomes for residual comparison instead of cumulative ones
    dict_hist = {}
    dict_sim_cum_mean = {} #initialize dictionaries to save the averaged outcomes, no need to specify the keys if not appending list
    dict_hist_cum = {}
    arr_year = list_results[0]['hiv']['year'].values
    for key_j in ['new_art','new_art bf08','new_art 0811','new_art sin12','dead','S2Uoff','Uoff2S','S2Uon','Uon2S','Uon2Uoff','Uoff2Uon','A2D','S2D','Uon2D','Uoff2D']: 
        arr_sim = np.array([data_k[key_j][data_k['year']>arr_year[0]] for data_k in list_hiv])
        dict_sim_mean[key_j] = np.mean(arr_sim,axis=0)
        dict_hist[key_j] = data_hiv_hist[key_j][data_hiv_hist['year']>arr_year[0]].values
        cum_sim = np.array([data_k[key_j][data_k['year']>arr_year[0]].cumsum() for data_k in list_hiv])
        dict_sim_cum_mean[key_j] = np.mean(cum_sim,axis=0)
        assert (dict_sim_cum_mean[key_j].shape[0]==len(arr_year)-1), 'Wrong dimension for outcomes on average'
        dict_hist_cum[key_j] = data_hiv_hist[key_j][data_hiv_hist['year']>arr_year[0]].values.cumsum()
    for key_j in c.list_comorb:
        arr_sim = np.array([data_k[key_j][data_k['year']>arr_year[0]] for data_k in list_comorb])
        dict_sim_mean[key_j] = np.mean(arr_sim,axis=0)
        dict_hist[key_j] = data_comorb_hist[key_j][data_comorb_hist['year']>arr_year[0]].values
        cum_sim = np.array([data_k[key_j][data_k['year']>arr_year[0]].cumsum() for data_k in list_comorb])
        dict_sim_cum_mean[key_j] = np.mean(cum_sim,axis=0)
        assert (dict_sim_cum_mean[key_j].shape[0]==len(arr_year)-1), 'Wrong dimension for outcomes on average'
        dict_hist_cum[key_j] = data_comorb_hist[key_j][data_comorb_hist['year']>arr_year[0]].values.cumsum()

    ####focus mortality calibration up to 2015 since LTFU information was determined by 18months before the last contact date and overall mortality was underestimated in 2016
    if arg_cali in ['new_art','trans_art','dead']: 
        for key_j in ['dead','A2D','S2D','Uon2D','Uoff2D']: 
            dict_sim_mean[key_j] = dict_sim_mean[key_j][:-1]
            dict_hist[key_j] = dict_hist[key_j][:-1]
            dict_sim_cum_mean[key_j] = dict_sim_cum_mean[key_j][:-1]
            dict_hist_cum[key_j] = dict_hist_cum[key_j][:-1]

    ####focus ART initiation calibration up to 2015 due to how FARVDT was randomized
    if arg_cali in ['new_art','new_art bf08','new_art 0811','new_art sin12']:
        for key_j in ['new_art','new_art bf08','new_art 0811','new_art sin12']:
            dict_sim_mean[key_j] = dict_sim_mean[key_j][:-1]
            dict_hist[key_j] = dict_hist[key_j][:-1]
            dict_sim_cum_mean[key_j] = dict_sim_cum_mean[key_j][:-1]
            dict_hist_cum[key_j] = dict_hist_cum[key_j][:-1]
    
    ####create dictionary for sum of squared residual for each outcome
    dict_res = {k:(dict_sim_cum_mean[k][-1]-dict_hist_cum[k][-1])**2 for k in dict_sim_cum_mean.keys()}
    
    ####show adjusted coefficients and outcome residual
    if arg_cali in c.list_comorb:
        print ('Coefficients for one comorbidity: ', c.dic_prob_comorb_coeff)
        print ('Optimized func for comorbidities: ', dict_res[arg_cali])
    elif arg_cali in ['new_art bf08','new_art 0811','new_art sin12']:
        print ('Piece-wise coefficients for ART initiation: ', c.dic_coeff_adj_a2t_pw)
        print ('Coefficients for ART initiation: ', c.dic_coeff_adj_a2t)
        print ('Optimized func for ART initiation stratified by diagnosis year: ', dict_res[arg_cali])
        print ('Optimized func for overall ART initiation: ', dict_res['new_art'])
    elif arg_cali=='new_art':
        print ('Piece-wise coefficients for ART initiation: ', c.dic_coeff_adj_a2t_pw)
        print ('Coefficients for ART initiation: ', c.dic_coeff_adj_a2t)
        print ('Residual overall by compartments: ', [dict_res[k] for k in ['new_art bf08','new_art 0811','new_art sin12']]) #remove A2D which led to worsened estimates for a2t
        print ('Residual stratified by diagnosis year: ', [dict_res[k] for k in ['new_art bf08','new_art 0811','new_art sin12']])
    elif arg_cali=='trans_art':
        print ('Coefficients for transitions after ART initiaiton: ', c.dic_coeff_prob_art)
        print ('Optimized func for ART transitions: ', [dict_res[key_i] for key_i in ['S2Uoff','Uoff2S','S2Uon','Uon2S','Uon2Uoff','Uoff2Uon']]) #calibration on trans_art without fitting to dead
    elif arg_cali=='dead':
        print ('Coefficient for probability of all-cause mortality (art/noart): ', c.coeff_prob_dead_ltfu_art, c.coeff_prob_dead_ltfu_noart_pw,c.dic_prob_dead_art['pvl_dv'][0])
        print ('Optimized func for ART transitions and mortality: ', [dict_res[key_i] for key_i in ['A2D','S2D','Uon2D','Uoff2D']])

    if arg_cali=='all':
        return sum(list(dict_res.values())) #add list to make sure it always returns to a float number instead of an array
    elif arg_cali=='comorb':
        return sum([dict_res[k] for k in c.list_comorb])
    elif arg_cali in c.list_comorb:
        return dict_res[arg_cali]
    elif arg_cali in ['new_art bf08','new_art 0811','new_art sin12']:
        return dict_res[arg_cali]
    elif arg_cali=='new_art':
        return sum([dict_res[k] for k in ['new_art bf08','new_art 0811','new_art sin12']]) 
    elif arg_cali=='trans_art':
        return sum([dict_res[key_i] for key_i in ['S2Uoff','Uoff2S','S2Uon','Uon2S','Uon2Uoff','Uoff2Uon']]) 
    elif arg_cali=='dead':
        return sum([dict_res[key_i] for key_i in ['A2D','S2D','Uon2D','Uoff2D']]) 

def func_mc_filter_micro_all(arr_rand,dict_hist_target,dict_data,c):
    """return to residuals for comorbidities (11), new_art (5), trans_art (6) and dead (3), saved as dictionary
    arr_rand is random values between [0,1] generated by Latin-hypercube sampling"""

    ####separate dataframe of hiv and comorb for historical data
    data_hiv_hist = dict_hist_target['hiv']
    data_comorb_hist = dict_hist_target['comorb']

    ####adjust randomized coefficients after introducing piece-wise coefficients for new_art sin12 and coeff_prob_dead_ltfu_art 
    arr_rand_comorb = arr_rand[:11]
    arr_rand_a2t = arr_rand[11:11+6]
    arr_rand_art = arr_rand[11+6:11+6+6]
    arr_rand_dead = arr_rand[11+6+6:]
    assert len(arr_rand[11+6+6:])==4, 'Double check the number of randomized coefficients generated'
    dict_random = {} #initiazlie dictionary to save randomized coefficients for adjustment
    dict_random['comorb'] = {key_i: arr_rand_comorb[i]*(c.dic_mc_range['comorb'][key_i][1]-c.dic_mc_range['comorb'][key_i][0])+c.dic_mc_range['comorb'][key_i][0] for i,key_i in enumerate(c.dic_mc_range['comorb'].keys())}
    dict_random['new_art'] = {key_i: arr_rand_a2t[i*2:i*2+2]*(c.dic_mc_range['new_art'][key_i][1,:]-c.dic_mc_range['new_art'][key_i][0,:])+c.dic_mc_range['new_art'][key_i][0,:] for i,key_i in enumerate(['bf08','0811','sin12'])}
    dict_random['trans_art'] = arr_rand_art*(c.dic_mc_range['trans_art'][1,:]-c.dic_mc_range['trans_art'][0,:])+c.dic_mc_range['trans_art'][0,:]
    dict_random['dead'] = arr_rand_dead*(c.dic_mc_range['dead'][1,:]-c.dic_mc_range['dead'][0,:])+c.dic_mc_range['dead'][0,:]

    ####update adjusted coefficients based on random values
    for key_i in dict_random['comorb'].keys():
        c.dic_prob_comorb_coeff[key_i] = dict_random['comorb'][key_i] 
    c.dic_coeff_adj_a2t_pw['bf08'] = np.copy(dict_random['new_art']['bf08'])
    c.dic_coeff_adj_a2t_pw['0811'] = np.copy(dict_random['new_art']['0811']) 
    c.dic_coeff_adj_a2t_pw['sin12'] = np.copy(dict_random['new_art']['sin12']) 
    c.dic_coeff_prob_art = {k:dict_random['trans_art'][i] for i,k in enumerate(list(c.dic_coeff_prob_art))}
    c.coeff_prob_dead_ltfu_art = dict_random['dead'][0]
    c.coeff_prob_dead_ltfu_noart_pw = np.copy([dict_random['dead'][1],dict_random['dead'][2]]) 
    c.dic_prob_dead_art['pvl_dv'][0] = dict_random['dead'][-1]

    ####run microsimulation using multi-processing and estimate outcomes for residual estimation
    pool = Pool(processes=c.n_pool,maxtasksperchild=1)  
    results = [pool.apply_async(mp_microsim_outcomes_cali,args=(c,dict_data,c.arr_seed[i])) for i in range(c.num_sim)]
    list_results = [x.get() for x in results] 
    list_hiv = [result['hiv'] for result in list_results]
    list_comorb = [result['comorb'] for result in list_results]

    #####use mean to calculate residuals between simulated and historical targets
    dict_sim_mean = {} #add annual average as outcomes for residual comparison instead of cumulative ones
    dict_hist = {}
    dict_sim_cum_mean = {} #initialize dictionaries to save the averaged outcomes, no need to specify the keys if not appending list
    dict_hist_cum = {}
    arr_year = list_results[0]['hiv']['year'].values
    for key_j in ['new_art bf08','new_art 0811','new_art sin12','S2Uoff','Uoff2S','S2Uon','Uon2S','Uon2Uoff','Uoff2Uon','A2D','S2D','Uon2D','Uoff2D']: 
        arr_sim = np.array([data_k[key_j][data_k['year']>arr_year[0]] for data_k in list_hiv])
        dict_sim_mean[key_j] = np.mean(arr_sim,axis=0)
        dict_hist[key_j] = data_hiv_hist[key_j][data_hiv_hist['year']>arr_year[0]].values
        cum_sim = np.array([data_k[key_j][data_k['year']>arr_year[0]].cumsum() for data_k in list_hiv])
        dict_sim_cum_mean[key_j] = np.mean(cum_sim,axis=0)
        assert (dict_sim_cum_mean[key_j].shape[0]==len(arr_year)-1), 'Wrong dimension for outcomes on average'
        dict_hist_cum[key_j] = data_hiv_hist[key_j][data_hiv_hist['year']>arr_year[0]].values.cumsum()
    for key_j in c.list_comorb:
        arr_sim = np.array([data_k[key_j][data_k['year']>arr_year[0]] for data_k in list_comorb])
        dict_sim_mean[key_j] = np.mean(arr_sim,axis=0)
        dict_hist[key_j] = data_comorb_hist[key_j][data_comorb_hist['year']>arr_year[0]].values
        cum_sim = np.array([data_k[key_j][data_k['year']>arr_year[0]].cumsum() for data_k in list_comorb])
        dict_sim_cum_mean[key_j] = np.mean(cum_sim,axis=0)
        assert (dict_sim_cum_mean[key_j].shape[0]==len(arr_year)-1), 'Wrong dimension for outcomes on average'
        dict_hist_cum[key_j] = data_comorb_hist[key_j][data_comorb_hist['year']>arr_year[0]].values.cumsum()

    ####trans_art and mortality were limited to one year prior to the end due to underestimation of LTFU in the final year, which affecting the number of PLWH in each state
    for key_j in ['new_art bf08','new_art 0811','new_art sin12','A2D','S2D','Uon2D','Uoff2D','S2Uoff','Uoff2S','S2Uon','Uon2S','Uon2Uoff','Uoff2Uon']: 
        dict_sim_mean[key_j] = dict_sim_mean[key_j][:-1]
        dict_hist[key_j] = dict_hist[key_j][:-1]
        dict_sim_cum_mean[key_j] = dict_sim_cum_mean[key_j][:-1]
        dict_hist_cum[key_j] = dict_hist_cum[key_j][:-1]

    ####estimate residuals for each target
    dict_res = {k:(dict_sim_cum_mean[k][-1]-dict_hist_cum[k][-1])**2 for k in dict_sim_cum_mean.keys()} 

    #return dict_res
    return [dict_random, dict_res] #keep records of randomized parameters and corresponding residuals

def func_mc_filter_res(list_dict_res,dict_res_target):
    """compare residuals for each target and keep those smaller than the one without randomization
    return to a dictionary with indices filtered out for each target"""

    list_target = list(list_dict_res[0])
    dict_filter_index = {key_i:[] for key_i in list_target}
    dict_filter_index['new_art_all'] = [] #filter by the sum of compartment residuals
    dict_filter_index['trans_art_all'] = []
    dict_filter_index['comorb_all'] = []
    dict_filter_index['dead_all'] = []
    for i,dict_res in enumerate(list_dict_res):
        for key_i in list_target:
            if dict_res[key_i]<=dict_res_target[key_i]:
                dict_filter_index[key_i].append(i)

        ###keep the index with sum of residual smaller than the target for overall new_art/trans_art/comorb/dead
        list_key_a2t = ['new_art bf08','new_art 0811','new_art sin12']
        if sum([dict_res[key_i] for key_i in list_key_a2t])<=sum([dict_res_target[key_i] for key_i in list_key_a2t]):
            dict_filter_index['new_art_all'].append(i)
        list_key_art = ['S2Uoff','Uoff2S','S2Uon','Uon2S','Uon2Uoff','Uoff2Uon']
        if sum([dict_res[key_i] for key_i in list_key_art])<=sum([dict_res_target[key_i] for key_i in list_key_art]):
            dict_filter_index['trans_art_all'].append(i)
        list_key_comorb = ['cvd','htn','dm','oa','copd','ckd','cld','cancer','manx','sczo','prsn']
        if sum([dict_res[key_i] for key_i in list_key_comorb])<=sum([dict_res_target[key_i] for key_i in list_key_comorb]):
            dict_filter_index['comorb_all'].append(i)
        list_key_dead = ['A2D','S2D','Uon2D','Uoff2D']
        if sum([dict_res[key_i] for key_i in list_key_dead])<=sum([dict_res_target[key_i] for key_i in list_key_dead]):
            dict_filter_index['dead_all'].append(i)

    ###focus on new_art/trans_art/comorb/dead separately
    dict_filter_index['new_art'] = list(set(dict_filter_index['new_art bf08']).intersection(set(dict_filter_index['new_art 0811'])).intersection(set(dict_filter_index['new_art sin12'])))
    dict_filter_index['trans_art'] = dict_filter_index[list_key_art[0]]
    for key_i in list_key_art[1:]:
        dict_filter_index['trans_art'] = list(set(dict_filter_index['trans_art']).intersection(set(dict_filter_index[key_i])))
    dict_filter_index['comorb'] = dict_filter_index[list_key_comorb[0]]
    for key_i in list_key_comorb[1:]:
        dict_filter_index['comorb'] = list(set(dict_filter_index['comorb']).intersection(set(dict_filter_index[key_i])))
    dict_filter_index['dead'] = dict_filter_index[list_key_dead[0]]
    for key_i in list_key_dead[1:]:
        dict_filter_index['dead'] = list(set(dict_filter_index['dead']).intersection(set(dict_filter_index[key_i])))

    return dict_filter_index

def func_mc_param_range_update(list_coeff,dict_filter,c):
    """return to updated p.dic_mc_range"""

    #####initialize updated parameter range
    dict_mc_range_update = copy.deepcopy(c.dic_mc_range)

    #####Update parameters for new_art based on new_art_all (smaller total residual)
    print ('Number of sets with each residual of new_art below the target: ', len(dict_filter['new_art']))
    print ('Key and the number of filtered sets filtered by sum of residuals: ', 'new_art_all ', len(dict_filter['new_art_all']))
    if len(dict_filter['new_art'])>1: 
        print ('Residual used to update new_art: ', 'new_art')
        for key_res in ['bf08','0811','sin12']: 
            arr_coeff = np.array([list_coeff[i]['new_art'][key_res] for i in dict_filter['new_art']])
            dict_mc_range_update['new_art'][key_res] = np.array([[min(arr_coeff[:,0]),min(arr_coeff[:,1])],[max(arr_coeff[:,0]),max(arr_coeff[:,1])]])
    elif len(dict_filter['new_art_all'])>1:
        print ('Residual used to update new_art: ', 'new_art_all')
        for key_res in ['bf08','0811','sin12']: 
            arr_coeff = np.array([list_coeff[i]['new_art'][key_res] for i in dict_filter['new_art_all']])
            dict_mc_range_update['new_art'][key_res] = np.array([[min(arr_coeff[:,0]),min(arr_coeff[:,1])],[max(arr_coeff[:,0]),max(arr_coeff[:,1])]])
    else:
        print ('Not enough filtered samples from new_art_all')

    #####Update parameters for trans_art based on trans_art_all
    print ('Number of sets with each residual of trans_art below the target: ', len(dict_filter['trans_art']))
    print ('Key and the number of filtered sets filtered by sum of residuals: ', 'trans_art_all ', len(dict_filter['trans_art_all']))
    if len(dict_filter['trans_art'])>1:
        print ('Residual used to update trans_art: ', 'trans_art')
        for i in range(6):
            arr_coeff = np.array([list_coeff[j]['trans_art'][i] for j in dict_filter['trans_art']])
            dict_mc_range_update['trans_art'][:,i] = np.array([min(arr_coeff),max(arr_coeff)]) #LHS and RHS were shaped (2,)
    elif len(dict_filter['trans_art_all'])>1:
        print ('Residual used to update trans_art: ', 'trans_art_all')
        for i in range(6):
            arr_coeff = np.array([list_coeff[j]['trans_art'][i] for j in dict_filter['trans_art_all']])
            dict_mc_range_update['trans_art'][:,i] = np.array([min(arr_coeff),max(arr_coeff)]) #LHS and RHS were shaped (2,)
    else:
        print ('Not enough filtered samples from trans_art_all')

    #####Update parameters for dead based on dead_all
    print ('Key and the number of filtered sets filtered by each residual of dead: ', 'dead ', len(dict_filter['dead']))
    print ('Key and the number of filtered sets filtered by sum of residuals: ', 'dead_all ', len(dict_filter['dead_all']))
    if len(dict_filter['dead'])>1:
        print ('Residual used to update dead: ', 'dead')
        for i in range(3):
            arr_coeff = np.array([list_coeff[j]['dead'][i] for j in dict_filter['dead']])
            dict_mc_range_update['dead'][:,i] = np.array([min(arr_coeff),max(arr_coeff)])
    elif len(dict_filter['dead_all'])>1:
        print ('Residual used to update dead: ', 'dead_all')
        for i in range(3):
            arr_coeff = np.array([list_coeff[j]['dead'][i] for j in dict_filter['dead_all']])
            dict_mc_range_update['dead'][:,i] = np.array([min(arr_coeff),max(arr_coeff)])
    else:
        print ('Not enough filtered samples from dead_all')

    #####update comorbidity-related coefficient range based on dict_filter[comorb_i]
    for comorb_i in c.list_comorb:
        if len(dict_filter[comorb_i])>1:
            print ('Key and the number of filtered sets: ', comorb_i+' ', len(dict_filter[comorb_i]))
            arr_coeff = [list_coeff[i]['comorb'][comorb_i] for i in dict_filter[comorb_i]]
            dict_mc_range_update['comorb'][comorb_i] = [min(arr_coeff),max(arr_coeff)]
        else:
            print ('Not enough filtered samples for ', comorb_i)

    return dict_mc_range_update

def func_mc_param_range_update_res(list_coeff,list_res,c,p_filter):
    """update parameter range based on smallest residual sum, p_filter indicate the percentage to be selected"""

    ####estimate sum of residuals and sort the indices by ascending order
    arr_res_sum = np.array([sum(list(rand_res.values())) for rand_res in list_res])
    arr_sort_ind = np.argsort(arr_res_sum)
    len_filter = len(arr_sort_ind)*p_filter
    arr_ind_filter = np.copy(arr_sort_ind[:int(len_filter)])

    ####update parameter ranges
    dict_mc_range_update = copy.deepcopy(c.dic_mc_range)
    for key_res in ['bf08','0811','sin12']: 
        arr_coeff = np.array([list_coeff[i]['new_art'][key_res] for i in arr_ind_filter])
        dict_mc_range_update['new_art'][key_res] = np.array([[min(arr_coeff[:,0]),min(arr_coeff[:,1])],[max(arr_coeff[:,0]),max(arr_coeff[:,1])]])
    for i in range(6):
        arr_coeff = np.array([list_coeff[j]['trans_art'][i] for j in arr_ind_filter])
        dict_mc_range_update['trans_art'][:,i] = np.array([min(arr_coeff),max(arr_coeff)]) #LHS and RHS were shaped (2,)
    for i in range(3):
        arr_coeff = np.array([list_coeff[j]['dead'][i] for j in arr_ind_filter])
        dict_mc_range_update['dead'][:,i] = np.array([min(arr_coeff),max(arr_coeff)])
    for comorb_i in c.list_comorb:
        arr_coeff = [list_coeff[i]['comorb'][comorb_i] for i in arr_ind_filter]
        dict_mc_range_update['comorb'][comorb_i] = [min(arr_coeff),max(arr_coeff)]

    return dict_mc_range_update

def func_hist_prepare(c,dict_data):
    """return to dictionary with historical records and derived information for re-sampling"""

    #####rename dataframes/dictionaries for simplification purpose
    data_cohort = dict_data['cohort']
    data_ncd4 = dict_data['cd4']
    data_su = dict_data['su']
    data_hiv = dict_data['hiv']
    data_comorb = dict_data['comorb']

    #####derive the variables used for probability estimation
    data_char = data_cohort[['moh_id','sex_at_birth_dv','DOB','earliest_HIV','FARVDT','end_fu_dt']][data_cohort['moh_id'].isin(data_hiv['moh_id'])].reset_index(drop=True) 
    data_char['baseline_dt'] = data_char['earliest_HIV'].apply(lambda x: max(x,c.study_dt)) #start follow participants either from 2008-01-01 or earliest_HIV later than 2008-01-01
    data_char['sex_dv'] = data_char['sex_at_birth_dv'].apply(lambda x: 0 if x=='M' else 1) 
    data_char = pd.merge(data_char,data_su[['moh_id','IDU_all','set_baseline_dt','earliest_su_dt']],on='moh_id',how='left') #use how=left to make the information in line with the participants in the cohort, add set_baseline_dt to filter for those eligible for statistical analysis
    data_char['su_bsln_dv'] = data_char[['baseline_dt','IDU_all','earliest_su_dt']].apply(lambda x: 1 if (x['IDU_all']==1 or x['earliest_su_dt']<=x['baseline_dt']) else (0 if (x['IDU_all']==0 or (x['IDU_all']==9 and x['earliest_su_dt']>x['baseline_dt'] and x['earliest_su_dt']<c.dummy_end_dt)) else 9),axis=1) #see the definition of substance_use_baseline 
    data_char = pd.merge(data_char,data_ncd4[['moh_id','ncd4_bsln']],how='left',on='moh_id')
    data_char['ncd4_bsln_dv'] = data_char['ncd4_bsln'].apply(lambda x: 0 if x>=200 else (1 if x<200 else 9)) #assign missing values as 9, same as su_bsln_dv
    data_char['year_diag_dv'] = data_char['earliest_HIV'].dt.year-c.min_year_diag #derived variable of year of diagnosis, 1979 is the earliest diagnosis 
    data_char['age_bsln_scale10_dv'] = (data_char['baseline_dt']-data_char['DOB']).dt.total_seconds()/(60*60*24*c.def_year)/10 #add baseline age for rate estimates
    data_char['year_diag_dv_a2t'] = data_char['earliest_HIV'].dt.year-c.min_year_diag_a2t #different from the one used for comorbidity analysis
    data_char['a2t_cat'] = data_char['earliest_HIV'].apply(lambda x: 'bf08' if x.year<2008 else ('0811' if x.year>=2008 and x.year<2012 else 'sin12')) #assign model category for probability from A to T to each individual based on year at diagnosis
    data_char['art_baseline_dt'] = data_char['FARVDT'].apply(lambda x: max(x,c.study_dt))
    data_char['age_art_bsln_scale10_dv'] = (data_char['art_baseline_dt']-data_char['DOB']).dt.total_seconds()/(60*60*24*c.def_year)/10 #add baseline age for rate estimates
    data_char = data_char.sort_values(by='moh_id').reset_index(drop=True) 

    #####rename column for baseline_dt in data_comorb, set dimension/moh_id in data_comorb_update the same as data_char
    data_comorb_update = data_comorb[data_comorb['moh_id'].isin(data_hiv['moh_id'])].sort_values(by='moh_id').reset_index(drop=True).rename(columns={'micro_baseline_dt':'baseline_dt'})

    #####recreate dictionary with records ready for initialization of microsimulation
    dict_update_data = {
        'char':data_char,
        'hiv':dict_data['hiv'],
        'comorb':data_comorb_update,
        'pvl':dict_data['pvl'],
        'reg_status':dict_data['reg_status'],
        'reg_prop':dict_data['reg_prop']
    }

    return dict_update_data

def func_init_pop(c,dict_data):
    """create initial population of plwh starting from 2008-01-01 with necessary characteristics for microsimulation
    dict_data is the dictionary with updated historical data with derived information in dict_data['char']
    c.rand_pop indicates how the population was created: 
        rand_pop=hist08, only considering baseline=set_baseline_dt=2008-01-01 with known ncd4/su/comorb at baseline
        rand_pop=hist, considering all PLWH in the cohort starting from 2008-01-01, with randomized characteristics if unknown (sampled from plwh on 2008-01-01 or new diagnosis)
        rand_pop=random, considering randomized characteristics for all PLWH starting from 2008-01-01"""

    #####rename dataframes/dictionaries for simplification purpose
    data_char = dict_data['char']
    data_comorb = dict_data['comorb']
    data_hiv = dict_data['hiv']
    data_status_pvl = dict_data['pvl']
    dict_status_reg = dict_data['reg_status'] #need type of regimen while on ART on 2008-01-01 to estimate further proportion time of specific regimen while on ART
    dict_prop_reg = dict_data['reg_prop']

    #####initialize dataframes/dictionaries to hold information for probability estimation
    data_comorb_sim = data_comorb.copy()
    for comorb_i in c.list_comorb:
        data_comorb_sim.loc[data_comorb_sim['earliest_'+comorb_i+'_dt']>data_comorb_sim['baseline_dt'],'earliest_'+comorb_i+'_dt'] = c.dummy_end_dt
    data_hiv_sim = data_hiv[['moh_id',str(c.t0)[:10]]].sort_values(by='moh_id').reset_index(drop=True) #only need status on t0=2008-01-01
    data_status_pvl_sim = data_status_pvl[['moh_id',str(c.t0)[:10]]].sort_values(by='moh_id').reset_index(drop=True)
    dict_status_reg_sim = {}
    dict_prop_reg_sim = {}
    for reg_i in ['tdf','nnrti','pi','insti']:
        data_status_reg_i = dict_status_reg[reg_i].copy()
        dict_status_reg_sim[reg_i] = data_status_reg_i[['moh_id',str(c.t0)[:10]]].sort_values(by='moh_id').reset_index(drop=True)
        data_prop_reg_i = dict_prop_reg[reg_i].copy()
        dict_prop_reg_sim[reg_i] = data_prop_reg_i[['moh_id',str(c.t0)[:10]]].sort_values(by='moh_id').reset_index(drop=True)

    #####create dictioanry to save dataframes which can be updated at each time step
    if c.rand_pop=='hist08':
        arr_id_init = data_char['moh_id'][(data_char['ncd4_bsln_dv']!=9)&(data_char['su_bsln_dv']!=9)&(data_char['baseline_dt']==data_char['set_baseline_dt'])&(data_char['earliest_HIV']<c.t0)].values #add condition on earliest_HIV to exclude those diagnosed on c.t0 as t0 was defined at 00:00:00
        dict_data_sim = {
            'char':data_char[data_char['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True),
            'hiv':data_hiv_sim[data_hiv_sim['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True), 
            'comorb':data_comorb_sim[data_comorb_sim['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True),
            'pvl':data_status_pvl_sim[data_status_pvl_sim['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True), 
            'reg_status':{k:dict_status_reg_sim[k][dict_status_reg_sim[k]['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True) for k in dict_status_reg_sim.keys()},
            'reg_prop':{k:dict_prop_reg_sim[k][dict_prop_reg_sim[k]['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True) for k in dict_prop_reg_sim.keys()} 
        }
    elif c.rand_pop=='hist':
        arr_id_init = data_char['moh_id'][(data_char['baseline_dt']==c.t0)&(data_char['earliest_HIV']<c.t0)].values 
        dict_data_sim = {
            ####sort_values makes sure the order of each dataframe is the same
            'char':data_char[data_char['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True),
            'hiv':data_hiv_sim[data_hiv_sim['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True), 
            'comorb':data_comorb_sim[data_comorb_sim['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True),
            'pvl':data_status_pvl_sim[data_status_pvl_sim['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True), 
            'reg_status':{k:dict_status_reg_sim[k][dict_status_reg_sim[k]['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True) for k in dict_status_reg_sim.keys()}, 
            'reg_prop':{k:dict_prop_reg_sim[k][dict_prop_reg_sim[k]['moh_id'].isin(arr_id_init)].sort_values(by='moh_id').reset_index(drop=True) for k in dict_prop_reg_sim.keys()} 
        }
        p_ncd4_200 = dict_data_sim['char'][dict_data_sim['char']['ncd4_bsln_dv']==1].shape[0]/dict_data_sim['char'][dict_data_sim['char']['ncd4_bsln_dv']!=9].shape[0]
        dict_data_sim['char']['ncd4_bsln_dv'] = dict_data_sim['char']['ncd4_bsln_dv'].apply(lambda x: x if x!=9 else random.choices([0,1],weights=[1-p_ncd4_200,p_ncd4_200],k=1)[0])
        p_su_bsln = dict_data_sim['char'][dict_data_sim['char']['su_bsln_dv']==1].shape[0]/dict_data_sim['char'][dict_data_sim['char']['su_bsln_dv']!=9].shape[0]
        dict_data_sim['char']['su_bsln_dv'] = dict_data_sim['char']['su_bsln_dv'].apply(lambda x: x if x!=9 else random.choices([0,1],weights=[1-p_su_bsln,p_su_bsln],k=1)[0])
        data_comorb_bsln = data_comorb[data_comorb['moh_id'].isin(data_hiv['moh_id'])].sort_values(by='moh_id').reset_index(drop=True) 
        arr_id_art = dict_data_sim['hiv']['moh_id'][dict_data_sim['hiv'][str(c.t0)[:10]].isin(['S','U_on'])].values #possible missing regimen information even though on ART (S/U_on)
        arr_id_reg_unknown = dict_data_sim['reg_status']['tdf']['moh_id'][(dict_data_sim['reg_status']['tdf']['moh_id'].isin(arr_id_art))&(pd.isnull(dict_data_sim['reg_status']['tdf'][str(c.t0)[:10]]))].values
        arr_id_reg_known = set(arr_id_art)-set(arr_id_reg_unknown)
        for reg_i in dict_data_sim['reg_status'].keys():
            dict_data_sim['reg_status'][reg_i].loc[dict_data_sim['reg_status'][reg_i]['moh_id'].isin(arr_id_reg_unknown),str(c.t0)[:10]] = np.random.choice(dict_data_sim['reg_status'][reg_i][str(c.t0)[:10]][dict_data_sim['reg_status'][reg_i]['moh_id'].isin(arr_id_reg_known)],size=len(arr_id_reg_unknown),replace=True) #replace nan by sampled known values
            arr_id_sample_prop = dict_data_sim['reg_status'][reg_i]['moh_id'][(dict_data_sim['reg_status'][reg_i]['moh_id'].isin(arr_id_reg_unknown))&(dict_data_sim['reg_status'][reg_i][str(c.t0)[:10]]==1)].values #only need to resample proportion time on regimen if reg_status=1
            arr_id_sampled_prop = dict_data_sim['reg_status'][reg_i]['moh_id'][(dict_data_sim['reg_status'][reg_i]['moh_id'].isin(arr_id_reg_known))&(dict_data_sim['reg_status'][reg_i][str(c.t0)[:10]]==1)].values #sampled from those with unknow status and status=1
            dict_data_sim['reg_prop'][reg_i].loc[dict_data_sim['reg_prop'][reg_i]['moh_id'].isin(arr_id_sample_prop),str(c.t0)[:10]] = np.random.choice(dict_data_sim['reg_prop'][reg_i][str(c.t0)[:10]][dict_data_sim['reg_prop'][reg_i]['moh_id'].isin(arr_id_sampled_prop)],size=len(arr_id_sample_prop),replace=True)
        arr_comorb_unknown = dict_data_sim['comorb']['moh_id'][(dict_data_sim['comorb']['micro_baseline_comorb']=='unknown')].values
        indices_comorb_unknown_refill = np.random.choice(data_comorb_bsln.index,size=len(arr_comorb_unknown),replace=True) #resample index instead of moh_id to use loc to repleat rows
        for comorb_i in c.list_comorb:
            data_comorbi_bsln = data_comorb_bsln.loc[indices_comorb_unknown_refill]
            data_comorbi_bsln['adj_earliest_'+comorb_i+'_dt'] = data_comorbi_bsln[['earliest_'+comorb_i+'_dt','set_baseline_dt']].apply(lambda x: c.t0 if x['earliest_'+comorb_i+'_dt']<=x['set_baseline_dt'] else c.dummy_end_dt,axis=1)
            dict_data_sim['comorb'].loc[dict_data_sim['comorb']['moh_id'].isin(arr_comorb_unknown),'earliest_'+comorb_i+'_dt'] = data_comorbi_bsln['adj_earliest_'+comorb_i+'_dt'].values
    elif c.rand_pop=='random':
        arr_index_init = data_char['moh_id'][(data_char['baseline_dt']==c.t0)&(data_char['earliest_HIV']<c.t0)].index 
        arr_index_init_sample = np.random.choice(arr_index_init,size=len(arr_index_init),replace=True)
        dict_data_sim = {
            'char':data_char.loc[arr_index_init_sample].reset_index(drop=True), 
            'hiv':data_hiv_sim.loc[arr_index_init_sample].reset_index(drop=True),
            'comorb':data_comorb_sim.loc[arr_index_init_sample].reset_index(drop=True),
            'pvl':data_status_pvl_sim.loc[arr_index_init_sample].reset_index(drop=True),
            'reg_status':{k:dict_status_reg_sim[k].loc[arr_index_init_sample].reset_index(drop=True) for k in dict_status_reg_sim.keys()},
            'reg_prop':{k:dict_prop_reg_sim[k].loc[arr_index_init_sample].reset_index(drop=True) for k in dict_prop_reg_sim.keys()}
        }
        p_ncd4_200 = dict_data_sim['char'][dict_data_sim['char']['ncd4_bsln_dv']==1].shape[0]/dict_data_sim['char'][dict_data_sim['char']['ncd4_bsln_dv']!=9].shape[0]
        dict_data_sim['char']['ncd4_bsln_dv'] = dict_data_sim['char']['ncd4_bsln_dv'].apply(lambda x: x if x!=9 else random.choices([0,1],weights=[1-p_ncd4_200,p_ncd4_200],k=1)[0])
        p_su_bsln = dict_data_sim['char'][dict_data_sim['char']['su_bsln_dv']==1].shape[0]/dict_data_sim['char'][dict_data_sim['char']['su_bsln_dv']!=9].shape[0]
        dict_data_sim['char']['su_bsln_dv'] = dict_data_sim['char']['su_bsln_dv'].apply(lambda x: x if x!=9 else random.choices([0,1],weights=[1-p_su_bsln,p_su_bsln],k=1)[0])
        arr_index_art = dict_data_sim['hiv'][dict_data_sim['hiv'][str(c.t0)[:10]].isin(['S','U_on'])].index 
        arr_index_reg_unknown = dict_data_sim['reg_status']['tdf'][(dict_data_sim['reg_status']['tdf'].index.isin(arr_index_art))&(pd.isnull(dict_data_sim['reg_status']['tdf'][str(c.t0)[:10]]))].index
        arr_index_reg_known = set(arr_index_art)-set(arr_index_reg_unknown)
        for reg_i in dict_data_sim['reg_status'].keys():
            dict_data_sim['reg_status'][reg_i].loc[dict_data_sim['reg_status'][reg_i].index.isin(arr_index_reg_unknown),str(c.t0)[:10]] = np.random.choice(dict_data_sim['reg_status'][reg_i][str(c.t0)[:10]][dict_data_sim['reg_status'][reg_i].index.isin(arr_index_reg_known)],size=len(arr_index_reg_unknown),replace=True) 
            arr_index_sample_prop = dict_data_sim['reg_status'][reg_i][(dict_data_sim['reg_status'][reg_i].index.isin(arr_index_reg_unknown))&(dict_data_sim['reg_status'][reg_i][str(c.t0)[:10]]==1)].index 
            arr_index_sampled_prop = dict_data_sim['reg_status'][reg_i][(dict_data_sim['reg_status'][reg_i].index.isin(arr_index_reg_known))&(dict_data_sim['reg_status'][reg_i][str(c.t0)[:10]]==1)].index 
            dict_data_sim['reg_prop'][reg_i].loc[dict_data_sim['reg_prop'][reg_i].index.isin(arr_index_sample_prop),str(c.t0)[:10]] = np.random.choice(dict_data_sim['reg_prop'][reg_i][str(c.t0)[:10]][dict_data_sim['reg_prop'][reg_i].index.isin(arr_index_sampled_prop)],size=len(arr_index_sample_prop),replace=True)
        data_comorb_bsln = data_comorb[data_comorb['moh_id'].isin(data_hiv['moh_id'])].sort_values(by='moh_id').reset_index(drop=True) 
        arr_comorb_unknown = dict_data_sim['comorb']['moh_id'][(dict_data_sim['comorb']['micro_baseline_comorb']=='unknown')].values
        indices_comorb_unknown_refill = np.random.choice(data_comorb_bsln.index,size=len(arr_comorb_unknown),replace=True) 
        for comorb_i in c.list_comorb:
            data_comorbi_bsln = data_comorb_bsln.loc[indices_comorb_unknown_refill]
            data_comorbi_bsln['adj_earliest_'+comorb_i+'_dt'] = data_comorbi_bsln[['earliest_'+comorb_i+'_dt','set_baseline_dt']].apply(lambda x: c.t0 if x['earliest_'+comorb_i+'_dt']<=x['set_baseline_dt'] else c.dummy_end_dt,axis=1)
            dict_data_sim['comorb'].loc[dict_data_sim['comorb']['moh_id'].isin(arr_comorb_unknown),'earliest_'+comorb_i+'_dt'] = data_comorbi_bsln['adj_earliest_'+comorb_i+'_dt'].values       
    else:
        sys.exit("Wrong indicator to create cohort for microsimulation") #stop if passing wrong rand_pop

    #####randomly sample health behavior parameters (pa-physical activity, smk-smoke and alc-alcohol) if needed
    if c.rand_hb==1:
        dict_data_sim['char']['pa'] = pd.Series(random.choices(c.dic_dist_pa['val'],weights=c.dic_dist_pa['prop'],k=dict_data_sim['char'].shape[0]))
        arr_smk_alc = random.choices(c.dic_dist_smk_alc['val'],weights=c.dic_dist_smk_alc['prop'],k=dict_data_sim['char'].shape[0])
        list_smk,list_alc = zip(*list(arr_smk_alc))
        dict_data_sim['char']['smk'] = pd.Series(list_smk)
        dict_data_sim['char']['alc'] = pd.Series(list_alc)

    return dict_data_sim

def func_init_new(c,dict_data,t_start,t_stop):
    """initialize characteristics for new diagnosis with or without randomly assigned health behaviours
    dict_data is the dictionary with updated historical data with derived information in dict_data['char'];
    t_start and t_stop indicate the period of new diagnoses
    c.rand_pop indicates how the population was created:
        rand_pop=hist08, return to empty dictionary
        rand_pop=hist, use historical records and randomly sampled ones among new diagnosis since 2008-01-01 if unknown
        rand_pop=random, use randomized charactersitics for all new diagnosis, sampled among new diagnosis since 2008-01-01"""

    #####rename dataframes/dictionaries for simplification purpose
    data_char = dict_data['char']
    data_comorb = dict_data['comorb']

    #####create comorbidity sample pool using the comorbidity status at set_baseline_dt to initialize comorbidity status for new diagnosis
    data_comorb_sample = data_comorb.copy()
    for comorb_i in c.list_comorb:
        data_comorb_sample.loc[data_comorb_sample['earliest_'+comorb_i+'_dt']>data_comorb_sample['set_baseline_dt'],'earliest_'+comorb_i+'_dt'] = c.dummy_end_dt #only keep prevalent case at set_baseline_dt for re-sampling purpose

    #####create characteristic sample pool using all new diagnosis since 2008 for missing CD4 and/or substance use 
    arr_id_new_all = data_char['moh_id'][(data_char['earliest_HIV']>=c.t0)].values #new diagnosis since 2008-01-01 as the pool for characteristic sampling
    data_char_sample = data_char[data_char['moh_id'].isin(arr_id_new_all)].sort_values(by='moh_id').reset_index(drop=True)
    p_ncd4_200 = data_char_sample[data_char_sample['ncd4_bsln_dv']==1].shape[0]/data_char_sample[data_char_sample['ncd4_bsln_dv']!=9].shape[0] #refill missing cd4 based on proportion of cd4<200 among all new diagnosis since 2008
    p_su_bsln = data_char_sample[data_char_sample['su_bsln_dv']==1].shape[0]/data_char_sample[data_char_sample['su_bsln_dv']!=9].shape[0] #refill missing su based on proportion of su=1 among all new diagnosis

    #####create dictioanry to save dataframes which can be updated at each time step
    if c.rand_pop=='hist08':
        dict_data_sim = {k: pd.DataFrame([]) for k in dict_data.keys()} #no new diagnosis if rand_pop=hist08
        dict_data_sim['reg_status'] = {k:pd.DataFrame([]) for k in ['tdf','nnrti','pi','insti']} #use keys in dict_status_reg_sim to exclude 'aba' in historical data
        dict_data_sim['reg_prop'] = {k:pd.DataFrame([]) for k in ['tdf','nnrti','pi','insti']}
    elif c.rand_pop=='hist':
        arr_id_new = data_char['moh_id'][(data_char['earliest_HIV']>=t_start)&(data_char['earliest_HIV']<t_stop)].values
        dict_data_sim = {
            'char':data_char[data_char['moh_id'].isin(arr_id_new)].sort_values(by='moh_id').reset_index(drop=True),
            'hiv':pd.DataFrame({'moh_id':arr_id_new,str(t_start)[:10]:['U']*len(arr_id_new),str(t_stop)[:10]:['A']*len(arr_id_new)}).sort_values(by='moh_id').reset_index(drop=True), #add unaware status at t_start for new diagnosis for the derivation of pvl at the step after t_stop
            'pvl':pd.DataFrame({'moh_id':arr_id_new,str(t_stop)[:10]:np.ones(len(arr_id_new))}).sort_values(by='moh_id').reset_index(drop=True), #assume unsuppression at t_stop for new diagnosis
            'reg_status':{k:pd.DataFrame({'moh_id':arr_id_new,str(t_stop)[:10]:np.nan}).sort_values(by='moh_id').reset_index(drop=True) for k in ['tdf','nnrti','pi','insti']}, 
            'reg_prop':{k:pd.DataFrame({'moh_id':arr_id_new,str(t_stop)[:10]:np.zeros(len(arr_id_new))}).sort_values(by='moh_id').reset_index(drop=True) for k in['tdf','nnrti','pi','insti']} 
        }
        dict_data_sim['comorb'] = data_comorb[['moh_id','earliest_HIV','baseline_dt']][data_comorb['moh_id'].isin(arr_id_new)].sort_values(by='moh_id').reset_index(drop=True) #keep earliest_HIV and baseline_dt information for new diagnosis to identify comorbidity incidence based on data_comorb_sim, keep sort_values to be consistent with the other dataframes
        indices_comorb_unknown_refill = np.random.choice(data_comorb_sample.index,size=len(arr_id_new),replace=True) #comorb is unknown for all new diagnosis since set_baseline_dt should be at least one year after diagnosis
        for comorb_i in c.list_comorb:
            data_comorbi_sample = data_comorb_sample.loc[indices_comorb_unknown_refill]
            data_comorbi_sample['adj_earliest_'+comorb_i+'_dt'] = data_comorbi_sample[['earliest_'+comorb_i+'_dt','set_baseline_dt']].apply(lambda x: t_start if x['earliest_'+comorb_i+'_dt']<=x['set_baseline_dt'] else c.dummy_end_dt,axis=1) #use t_start as the dummy diagnosis date if it's prevalent at t_stop
            dict_data_sim['comorb']['earliest_'+comorb_i+'_dt'] = data_comorbi_sample['adj_earliest_'+comorb_i+'_dt'].values
        dict_data_sim['char']['ncd4_bsln_dv'] = dict_data_sim['char']['ncd4_bsln_dv'].apply(lambda x: x if x!=9 else random.choices([0,1],weights=[1-p_ncd4_200,p_ncd4_200],k=1)[0]) 
    elif c.rand_pop=='random':
        n_sample = data_char['moh_id'][(data_char['earliest_HIV']>=t_start)&(data_char['earliest_HIV']<t_stop)].shape[0]
        arr_index_new_sample = np.random.choice(data_char_sample.index,size=n_sample,replace=True)
        arr_id_new_sample = data_char_sample['moh_id'].loc[arr_index_new_sample].values
        dict_data_sim = {
            'char':data_char_sample.loc[arr_index_new_sample].reset_index(drop=True),
            'hiv':pd.DataFrame({'moh_id':arr_id_new_sample,str(t_start)[:10]:['U']*len(arr_id_new_sample),str(t_stop)[:10]:['A']*len(arr_id_new_sample)}).reset_index(drop=True), 
            'pvl':pd.DataFrame({'moh_id':arr_id_new_sample,str(t_stop)[:10]:np.ones(len(arr_id_new_sample))}).reset_index(drop=True),
            'reg_status':{k:pd.DataFrame({'moh_id':arr_id_new_sample,str(t_stop)[:10]:np.nan}).reset_index(drop=True) for k in ['tdf','nnrti','pi','insti']},
            'reg_prop':{k:pd.DataFrame({'moh_id':arr_id_new_sample,str(t_stop)[:10]:np.zeros(len(arr_id_new_sample))}).reset_index(drop=True) for k in['tdf','nnrti','pi','insti']} 
        }
        arr_t_sim = np.random.uniform(0,1,len(arr_index_new_sample))#assume earliest_HIV between t_start and t_stop 
        dict_data_sim['char']['earliest_HIV'] = pd.to_datetime(pd.Series(t_start+pd.to_timedelta(c.def_year*c.dt*arr_t_sim,unit='d')).dt.date)
        dict_data_sim['char']['baseline_dt'] = dict_data_sim['char']['earliest_HIV'].apply(lambda x: max(x,c.t0))
        dict_data_sim['char']['year_diag_dv'] = dict_data_sim['char']['earliest_HIV'].dt.year-c.min_year_diag #derived variable of year of diagnosis
        dict_data_sim['char']['age_bsln_scale10_dv'] = (dict_data_sim['char']['baseline_dt']-dict_data_sim['char']['DOB']).dt.total_seconds()/(60*60*24*c.def_year)/10 #add baseline age
        dict_data_sim['char']['year_diag_dv_a2t'] = dict_data_sim['char']['earliest_HIV'].dt.year-c.min_year_diag_a2t #different from the one used for comorbidity analysis
        dict_data_sim['char']['a2t_cat'] = dict_data_sim['char']['earliest_HIV'].apply(lambda x: 'bf08' if x.year<2008 else ('0811' if x.year>=2008 and x.year<2012 else 'sin12'))
        dict_data_sim['comorb'] = dict_data_sim['char'][['moh_id','earliest_HIV','baseline_dt']].copy() 
        indices_comorb_unknown_refill = np.random.choice(data_comorb_sample.index,size=len(arr_id_new_sample),replace=True) #comorb is unknown for all new diagnosis since set_baseline_dt should be at least one year after diagnosis
        for comorb_i in c.list_comorb:
            data_comorbi_sample = data_comorb_sample.loc[indices_comorb_unknown_refill]
            data_comorbi_sample['adj_earliest_'+comorb_i+'_dt'] = data_comorbi_sample[['earliest_'+comorb_i+'_dt','set_baseline_dt']].apply(lambda x: t_start if x['earliest_'+comorb_i+'_dt']<=x['set_baseline_dt'] else c.dummy_end_dt,axis=1) 
            dict_data_sim['comorb']['earliest_'+comorb_i+'_dt'] = data_comorbi_sample['adj_earliest_'+comorb_i+'_dt'].values
        dict_data_sim['char']['ncd4_bsln_dv'] = dict_data_sim['char']['ncd4_bsln_dv'].apply(lambda x: x if x!=9 else random.choices([0,1],weights=[1-p_ncd4_200,p_ncd4_200],k=1)[0]) 
        dict_data_sim['char']['su_bsln_dv'] = dict_data_sim['char']['su_bsln_dv'].apply(lambda x: x if x!=9 else random.choices([0,1],weights=[1-p_su_bsln,p_su_bsln],k=1)[0])

    if c.rand_hb==1:
        dict_data_sim['char']['pa'] = pd.Series(random.choices(c.dic_dist_pa['val'],weights=c.dic_dist_pa['prop'],k=dict_data_sim['char'].shape[0]))
        arr_smk_alc = random.choices(c.dic_dist_smk_alc['val'],weights=c.dic_dist_smk_alc['prop'],k=dict_data_sim['char'].shape[0])
        list_smk,list_alc = zip(*list(arr_smk_alc))
        dict_data_sim['char']['smk'] = pd.Series(list_smk)
        dict_data_sim['char']['alc'] = pd.Series(list_alc)
        
    return dict_data_sim

def func_init_new_rand(c,dict_data,t_start,t_stop,n_newdiag):
    """Given the number of new diagnosis derived from undiagnosed PLWH and new infections, generate characteristic for new diagnoses"""

    #####rename dataframes/dictionaries for simplification purpose
    data_char = dict_data['char']
    data_comorb = dict_data['comorb']

    #####create comorbidity sample pool using the comorbidity status at set_baseline_dt to initialize comorbidity status for new diagnosis
    data_comorb_sample = data_comorb.copy()
    for comorb_i in c.list_comorb:
        data_comorb_sample.loc[data_comorb_sample['earliest_'+comorb_i+'_dt']>data_comorb_sample['set_baseline_dt'],'earliest_'+comorb_i+'_dt'] = c.dummy_end_dt #only keep prevalent case at set_baseline_dt for re-sampling purpose

    #####create characteristic sample pool using all new diagnosis since 2008 for missing CD4 and/or substance use 
    arr_id_new_all = data_char['moh_id'][(data_char['earliest_HIV']>=c.t0)].values #new diagnosis since 2008-01-01 as the pool for characteristic sampling
    data_char_sample = data_char[data_char['moh_id'].isin(arr_id_new_all)].sort_values(by='moh_id').reset_index(drop=True)
    data_char_sample['age_diag'] = (data_char_sample['earliest_HIV']-data_char_sample['DOB']).dt.total_seconds()/(60*60*24)/c.def_year #use age_diag instead of DOB to derive DOB based on earliest_HIV and age_diag especially for new diag after 2016
    p_ncd4_200 = data_char_sample[data_char_sample['ncd4_bsln_dv']==1].shape[0]/data_char_sample[data_char_sample['ncd4_bsln_dv']!=9].shape[0] #refill missing cd4 based on proportion of cd4<200 among all new diagnosis since 2008
    p_su_bsln = data_char_sample[data_char_sample['su_bsln_dv']==1].shape[0]/data_char_sample[data_char_sample['su_bsln_dv']!=9].shape[0] #refill missing su based on proportion of su=1 among all new diagnosis

    #####given the number of new diagnosis from undiagnosed PLWH and new infections, characteristics can only be generated randomly
    n_sample = n_newdiag
    arr_index_new_sample = np.random.choice(data_char_sample.index,size=n_sample,replace=True)
    arr_id_new_sample = data_char_sample['moh_id'].loc[arr_index_new_sample].values
    dict_data_sim = {
        'char':data_char_sample.loc[arr_index_new_sample].reset_index(drop=True), #no need to sort_values as the indices were chosen randomly and can be repetitive
        'hiv':pd.DataFrame({'moh_id':arr_id_new_sample,str(t_start)[:10]:['U']*len(arr_id_new_sample),str(t_stop)[:10]:['A']*len(arr_id_new_sample)}).reset_index(drop=True), #add unaware status at t_start for new diagnosis for the derivation of pvl at the step after t_stop
        'pvl':pd.DataFrame({'moh_id':arr_id_new_sample,str(t_stop)[:10]:np.ones(len(arr_id_new_sample))}).reset_index(drop=True),
        'reg_status':{k:pd.DataFrame({'moh_id':arr_id_new_sample,str(t_stop)[:10]:np.nan}).reset_index(drop=True) for k in ['tdf','nnrti','pi','insti']},
        'reg_prop':{k:pd.DataFrame({'moh_id':arr_id_new_sample,str(t_stop)[:10]:np.zeros(len(arr_id_new_sample))}).reset_index(drop=True) for k in['tdf','nnrti','pi','insti']} 
    }
    arr_t_sim = np.random.uniform(0,1,len(arr_index_new_sample))#assume earliest_HIV between t_start and t_stop 
    dict_data_sim['char']['earliest_HIV'] = pd.to_datetime(pd.Series(t_start+pd.to_timedelta(c.def_year*c.dt*arr_t_sim,unit='d')).dt.date)
    dict_data_sim['char']['baseline_dt'] = dict_data_sim['char']['earliest_HIV'].apply(lambda x: max(x,c.t0))
    dict_data_sim['char']['year_diag_dv'] = dict_data_sim['char']['earliest_HIV'].dt.year-c.min_year_diag #derived variable of year of diagnosis, 1979 is the earliest diagnosis 
    dict_data_sim['char']['age_bsln_scale10_dv'] = (dict_data_sim['char']['baseline_dt']-dict_data_sim['char']['DOB']).dt.total_seconds()/(60*60*24*c.def_year)/10 #add baseline age for rate estimates
    dict_data_sim['char']['year_diag_dv_a2t'] = dict_data_sim['char']['earliest_HIV'].dt.year-c.min_year_diag_a2t #different from the one used for comorbidity analysis
    dict_data_sim['char']['a2t_cat'] = dict_data_sim['char']['earliest_HIV'].apply(lambda x: 'bf08' if x.year<2008 else ('0811' if x.year>=2008 and x.year<2012 else 'sin12'))
    dict_data_sim['comorb'] = dict_data_sim['char'][['moh_id','earliest_HIV','baseline_dt']].copy() #keep earliest_HIV and baseline_dt information consistent with data_char for new diagnosis to identify comorbidity incidence based on data_comorb_sim
    dict_data_sim['char']['DOB'] = dict_data_sim['char']['earliest_HIV']-pd.to_timedelta(dict_data_sim['char']['age_diag']*c.def_year,unit='d')
    indices_comorb_unknown_refill = np.random.choice(data_comorb_sample.index,size=len(arr_id_new_sample),replace=True) #comorb is unknown for all new diagnosis since set_baseline_dt should be at least one year after diagnosis
    for comorb_i in c.list_comorb:
        data_comorbi_sample = data_comorb_sample.loc[indices_comorb_unknown_refill]
        data_comorbi_sample['adj_earliest_'+comorb_i+'_dt'] = data_comorbi_sample[['earliest_'+comorb_i+'_dt','set_baseline_dt']].apply(lambda x: t_start if x['earliest_'+comorb_i+'_dt']<=x['set_baseline_dt'] else c.dummy_end_dt,axis=1) #use t_start as the dummy diagnosis date if it's prevalent at t_stop
        dict_data_sim['comorb']['earliest_'+comorb_i+'_dt'] = data_comorbi_sample['adj_earliest_'+comorb_i+'_dt'].values
    dict_data_sim['char']['ncd4_bsln_dv'] = dict_data_sim['char']['ncd4_bsln_dv'].apply(lambda x: x if x!=9 else random.choices([0,1],weights=[1-p_ncd4_200,p_ncd4_200],k=1)[0]) 
    dict_data_sim['char']['su_bsln_dv'] = dict_data_sim['char']['su_bsln_dv'].apply(lambda x: x if x!=9 else random.choices([0,1],weights=[1-p_su_bsln,p_su_bsln],k=1)[0])

    #####randomly sample health behavior parameters (pa-physical activity, smk-smoke and alc-alcohol) if needed
    if c.rand_hb==1:
        dict_data_sim['char']['pa'] = pd.Series(random.choices(c.dic_dist_pa['val'],weights=c.dic_dist_pa['prop'],k=dict_data_sim['char'].shape[0]))
        arr_smk_alc = random.choices(c.dic_dist_smk_alc['val'],weights=c.dic_dist_smk_alc['prop'],k=dict_data_sim['char'].shape[0])
        list_smk,list_alc = zip(*list(arr_smk_alc))
        dict_data_sim['char']['smk'] = pd.Series(list_smk)
        dict_data_sim['char']['alc'] = pd.Series(list_alc)
        
    return dict_data_sim

def func_params_update(c):
    """update calibrated and derived parameters here to avoid re-paste multiple times once the estimates change"""

    #####introduce random terms for final probability estimation of a2t and transitions after ART initiation
    c.rand_blup_a2t = 1. #0. # #used to determine whether to use randomized blup terms in probability estimation
    if c.rand_blup_a2t==1:
        print ('Consider random term for each individual for the probability a2t')
    elif c.rand_blup_a2t==0:
        print ('Consider no random terms for indiviudals for the probability a2t')
    else:
        print ('Wrong parameter value for random term in probability estimation')
    c.rand_blup_art = 1 #0/1 to indicate whether to use randomized terms for each individual
    if c.rand_blup_art==1:
        print ('Consider random term for each individual for probabilities after ART initiation')
    elif c.rand_blup_art==0:
        print ('Consider no random terms for indiviudals for probabilities after ART initiation')
    else:
        print ('Wrong parameter value for random term in probability estimation')

    #####define status set for probability update (arr_state_alive) and factors selected for probability estimation ([never] on ART)
    c.arr_state_alive = ['U','A','S','U_on','U_off'] 
    c.arr_state_art = ['S','U_on','U_off']

    #####load derived regimen distribution among PLWH/new ART initiations over time and probability of regimen switch while staying on ART
    excel_data = pd.ExcelFile('data/prob_reg_switch.xlsx')
    excel_sheet = excel_data.sheet_names
    data_init_reg = excel_data.parse(excel_sheet[0]) #used for regimen at ART initiation and those back on ART
    data_plwh_reg = excel_data.parse(excel_sheet[2]) #used for regimen of PLWH at the beginning of the simulation (2008-01-01)
    data_plwh_super = excel_data.parse(excel_sheet[3]) #used for regimen assigned based on super-HAART if switched/initiated
    data_reg_trans = excel_data.parse(excel_sheet[4]) #used for regimen switch while on ART
    c.reg_dist = {'init':data_init_reg.copy(), 'plwh':data_plwh_reg.copy(),'super':data_plwh_super.copy()}
    c.year_split_reg = 2014 #use piece-wise average for regimen switch while on ART for simplification
    year_split_tdf = c.year_split_reg
    p_tdf2other_bf = data_reg_trans['p tdf2other'][data_reg_trans['date2']<str(year_split_tdf)+'-01-01'].mean()
    p_tdf2other_sin = data_reg_trans['p tdf2other'][data_reg_trans['date2']>=str(year_split_tdf)+'-01-01'].mean()
    dic_reg_trans_p = {'tdf':[year_split_tdf,[1-p_tdf2other_bf,p_tdf2other_bf],[1-p_tdf2other_sin,p_tdf2other_sin]]}
    p_other2tdf_bf = data_reg_trans['p other2tdf'][data_reg_trans['date2']<str(year_split_tdf)+'-01-01'].mean()
    p_other2tdf_sin = data_reg_trans['p other2tdf'][data_reg_trans['date2']>=str(year_split_tdf)+'-01-01'].mean()
    dic_reg_trans_other = {'other':[year_split_tdf,[1-p_other2tdf_bf,p_other2tdf_bf],[1-p_other2tdf_sin,p_other2tdf_sin]]}
    dic_reg_trans_p.update(dic_reg_trans_other)
    year_split_nnrti = c.year_split_reg
    arr_p_nnrti2rest_bf = data_reg_trans[['p nnrti2pi','p nnrti2insti','p nnrti2super']][data_reg_trans['date2']<str(year_split_nnrti)+'-01-01'].mean().values
    arr_p_nnrti2rest_sin = data_reg_trans[['p nnrti2pi','p nnrti2insti','p nnrti2super']][data_reg_trans['date2']>=str(year_split_nnrti)+'-01-01'].mean().values
    dic_reg_trans_nnrti = {'nnrti':[year_split_nnrti,[1-sum(arr_p_nnrti2rest_bf)]+list(arr_p_nnrti2rest_bf),[1-sum(arr_p_nnrti2rest_sin)]+list(arr_p_nnrti2rest_sin)]}
    dic_reg_trans_p.update(dic_reg_trans_nnrti)
    year_split_pi = c.year_split_reg
    arr_p_pi2rest_bf = data_reg_trans[['p pi2nnrti','p pi2insti','p pi2super']][data_reg_trans['date2']<str(year_split_pi)+'-01-01'].mean().values
    arr_p_pi2rest_sin = data_reg_trans[['p pi2nnrti','p pi2insti','p pi2super']][data_reg_trans['date2']>=str(year_split_pi)+'-01-01'].mean().values
    dic_reg_trans_pi = {'pi':[year_split_pi,[1-sum(arr_p_pi2rest_bf)]+list(arr_p_pi2rest_bf),[1-sum(arr_p_pi2rest_sin)]+list(arr_p_pi2rest_sin)]}
    dic_reg_trans_p.update(dic_reg_trans_pi)
    year_split_insti = c.year_split_reg
    arr_p_insti2rest_bf = data_reg_trans[['p insti2nnrti','p insti2pi','p insti2super']][data_reg_trans['date2']<str(year_split_insti)+'-01-01'].mean().values
    arr_p_insti2rest_sin = data_reg_trans[['p insti2nnrti','p insti2pi','p insti2super']][data_reg_trans['date2']>=str(year_split_insti)+'-01-01'].mean().values
    dic_reg_trans_insti = {'insti':[year_split_insti,[1-sum(arr_p_insti2rest_bf)]+list(arr_p_insti2rest_bf),[1-sum(arr_p_insti2rest_sin)]+list(arr_p_insti2rest_sin)]}
    dic_reg_trans_p.update(dic_reg_trans_insti)
    year_split_super = c.year_split_reg
    arr_p_super2rest_bf = data_reg_trans[['p super2nnrti','p super2pi','p super2insti']][data_reg_trans['date2']<str(year_split_super)+'-01-01'].mean().values
    arr_p_super2rest_sin = data_reg_trans[['p super2nnrti','p super2pi','p super2insti']][data_reg_trans['date2']>=str(year_split_super)+'-01-01'].mean().values
    dic_reg_trans_super = {'super':[year_split_super,[1-sum(arr_p_super2rest_bf)]+list(arr_p_super2rest_bf),[1-sum(arr_p_super2rest_sin)]+list(arr_p_super2rest_sin)]}
    dic_reg_trans_p.update(dic_reg_trans_super)
    c.dic_reg_trans_p = copy.deepcopy(dic_reg_trans_p) #keys represent the initial state of the transition 
    c.list_bb = ['tdf','other'] #represents the order of the probabilities from certain state for microsimulation
    c.list_nbb = ['nnrti','pi','insti','super']
    c.super_state = ['nnrti_pi','nnrti_insti','pi_insti','nnrti_pi_insti'] #used to sample combination of superHAART regimen

    #####introduce probability to different ART states (S/U_on/U_off) if ART initiation happened, details can be found in "prob ART init to diff. status"
    c.prob_A2S = 0.501024823502619 #use A2S_all for overall impact
    c.prob_A2Uon = 0.4618537918469597

    #####introduce simple estimation for proportion of time on specific regimen over one-year period
    c.dic_reg_prop = {
            '000':0., '001':0.25, '010':0.5, '011':0.75, '100':0.25, '101':0.5, '110':0.75, 
            '111':1., '400':0., '401':0.25, '410':0.5, '411':0.75,
            '440':0., '441':0.25, '444':0.,
        }

    #####introduce age category for outcome derivation
    c.dic_age_cat = {'<30':[0,30],'30-40':[30,40],'40-50':[40,50],'50-60':[50,60],'60-70':[60,70],'>=70':[70,200]}

    return c

class Params():
    """predetermined parameters that can be modified in the main code"""
    
    def __init__(self):
        
        self.def_year = 365.25
        self.dummy_end_dt = pd.Timestamp('2100-12-31') #a date far after the end_fu_dt 
        self.study_dt = pd.Timestamp('2008-01-01') #set the start date of the study, different from t0 which can be from anytime after study_dt with known information

        #####parameters for discrete system
        self.pi = 0.0252464 #inflow rate of general population in BC age>=15
        self.mu_s = 0.00846171 #mortality rate of general population in BC, age>=15, details in 'pop_bc_growth_death'
        self.N2017 = 4299128.476601299 #size of general population in BC by the end of 2017

        #####parameters related to characteristics of the population
        self.list_comorb = ['cvd','htn','dm','oa','copd','ckd','cld','cancer','manx','sczo','prsn']

        #####introduce coefficients for probability from A to T 
        self.dic_prob_a2t = {
            'beta0': 0.1556,
            'sex_cat': -0.239, #variable name is the same as those in the dataframe with characteristics
            'age_diag': 0.2331,
            'ncd4_diag_sim': -0.2995,
            'ncomorb_diag': -0.2354,
            '2012_diag': 0.4343, #be aware that year of diagnosis is a continuous variable starting 2008, which cannot be implemented as there were PLWH diagnosed much earlier than 2008
            'only_mhd_diag': -0.1858,
            'only_oud_diag': -1.1013,
            'mhd_oud_diag': -0.8834
        }

        #####introduce parameters for the impact of health-related factors (physical activity, alcohol and smoking) on incidence of comorbidities and all-cause mortality rates
        self.dic_coeff_pa = { #represents the additional risk among those less active/sedentary to develop comorbidities
            'cvd':[1.48,1.07,2.06],
            'htn':[1.28,1.1,1.49],
            'dm':[1.49,1.14,1.94],
            'oa':[1.,1.,1.], #the additional risk for developing OA is on hold since lack of literature
            'copd':[1.35,1.22,1.52],
            'ckd':[1.34,1.1,1.63],
            'cld':[1.33,1.19,1.49],
            'cancer':[1.18,1.04,1.32],
            'manx':[1.27,1.01,1.59],
            'sczo':[1.,1.,1.],
            'prsn':[1.,1.,1.],
            'mortality':[1.41,1.39,1.45]
        }
        self.dic_coeff_alc = { #represents the additional risk among those heavy drinker/binge drinker to develop comorbidities
            'cvd':[1.62,1.32,1.98],
            'htn':[1.4,1.03,1.87],
            'dm':[1.,1.,1.],
            'oa':[1.43,1.01,2.02],
            'copd':[1.,1.,1.],
            'ckd':[1.,1.,1.],
            'cld':[3.8,2.4,5.8],
            'cancer':[1.22,1.08,1.38],
            'manx':[1.68,1.17,2.42],
            'sczo':[1.,1.,1.],
            'prsn':[1.,1.,1.],
            'mortality':[1.9,1.3,2.7]
        }
        self.dic_coeff_smk = { #row1 represents current smoker and row2 represents former/experimental smoker
            'cvd':np.array([[2.19,1.80,2.67],[1.39,1.10,1.76]]),
            'htn':np.array([[1.,1.,1.],[1.,1.,1.]]),
            'dm':np.array([[1.37,1.33,1.42],[1.14,1.1,1.18]]),
            'oa':np.array([[1.,1.,1.],[1.,1.,1.]]),
            'copd':np.array([[2.41,1.7,3.42],[2.41,1.7,3.42]]),
            'ckd':np.array([[1.34,1.23,1.47],[1.15,1.08,1.23]]),
            'cld':np.array([[1.,1.,1.],[1.,1.,1.]]),
            'cancer':np.array([[1.45,1.17,1.79],[1.06,0.82,1.36]]),
            'manx':np.array([[1.99,1.71,2.32],[1.99,1.71,2.32]]),
            'sczo':np.array([[2.27,1.67,3.08],[2.27,1.67,3.08]]),
            'prsn':np.array([[1.,1.,1.],[1.,1.,1.]]),
            'mortality':np.array([[1.28,1.13,1.46],[1.,1.,1.]])
        }

        #####introduce coefficients for each comorbidity/mortality to adjust for the impact of health behaviours
        self.dic_prob_comorb_coeff = {key_i:1. for key_i in ['cvd','htn','dm','oa','copd','ckd','cld','cancer','manx','sczo','prsn','dead']} #key names are consistent with keys for each comorbidity simulation so that each coefficient can be calibrated for each comorbidity separately
        self.dt_split_prob_dead_noart = pd.Timestamp('2012-12-31') #use 2012-12-31 which led to reasonable annual and cum results
        self.coeff_prob_dead_ltfu_noart_pw = np.array([1.,1.]) #introduce piece wise coefficients to adjust probability of dead among those without ART initiation to account the elevated overdose risk since 2014
        self.coeff_prob_dead_ltfu_art = self.coeff_prob_dead_ltfu_noart = 1. #separate coeff_prob_dead_ltfu to improve fitting for deaths
        self.coeff_prob_dead_ltfu = 1. #introduce dic_prob_dead_ltfu to adjust mortality rate to account for ltfu as dead

        #####introduce parameters representing proportion of physical activity, alcohol, smoking status, derived from CCHS data
        self.dic_dist_pa = {
            'val':[0., 1.], #0 represents physical active and 1 represents not active
            'prop': [0.757565, 0.242435]
        }
        self.dic_dist_smk_alc = {
            'val':[(0.,0.),(0.,1.),(1.,0.),(1.,1.),(2.,0.),(2.,1.)], #smoke:0 for never, 1 for current, 2 for former and alcohol: 0 for No and 1 for Yes
            'prop':[0.3761326,0.0307832,0.0979346,0.0532086,0.3700383,0.0719027]
        }

        #####add dictionary include the name of probability function for each comorbidity
        self.dic_prob_comorb = {
            'cvd':func_prob_update_cvd,
            'htn':func_prob_update_htn,
            'dm':func_prob_update_dm,
            'oa':func_prob_update_oa,
            'copd':func_prob_update_copd,
            'ckd':func_prob_update_ckd,
            'cld':func_prob_update_cld,
            'cancer':func_prob_update_cancer,
            'manx':func_prob_update_manx,
            'sczo':func_prob_update_sczo,
            'prsn':func_prob_update_prsn,
            'dead':func_prob_update_mortality #include function to update mortality probability
        }

        #####introduce coefficients for probability of COPD, separated by ART initiation status
        self.min_year_diag = 1979 
        self.dic_prob_copd_art = {
            'beta0':[-14.682,-15.5233,-13.8406],
            'sex_dv':[0.5286,0.2833,0.774], 
            'su_bsln_dv':[1.4164,1.1639,1.6689], 
            'age_bsln_scale10_dv':[0.6806,0.5667,0.7945], 
            'manx_dv':[0.3902,0.1714,0.6089], 
            'prsn_dv':[0.337,-0.03084,0.7048], 
            'ncd4_bsln_dv':[0.2775,0.05787,0.4971], 
            'year_diag_dv':[-0.02743,-0.04715,-0.00771], 
            'counter':[0.0384,-0.00302,0.07983],
        }
        self.dic_prob_copd_noart = {
            'beta0':[-16.0766,-18.6271,-13.5262], 
            'sex_dv':[0.608,0.003522,1.2124], 
            'su_bsln_dv':[0.9698,0.124,1.8156], 
            'age_bsln_scale10_dv':[1.122,0.7845,1.4595], 
            'manx_dv':[0.9395,0.1896,1.6894], 
            'year_diag_dv':[-0.04952,-0.1107,0.0117], 
            'counter':[0.1,-0.01325,0.2133],
        }

        #####introduce coefficients for probability of comorbidities other than COPD, separated by ART initiation status
        self.dic_prob_htn_art = {
            'beta0':[-11.0374,-11.7878,-10.287], 
            'su_bsln_dv':[-0.2818,-0.4646,-0.09903], 
            'age_bsln_scale10_dv':[0.499,0.4042,0.5939], 
            'year_diag_dv':[-0.05742,-0.07251,-0.04234], 
            'counter':[-0.05306,-0.09022,-0.01591],
            'dm_dv':[0.6454,0.404,0.8867], 
            'ckd_dv':[0.6266,0.3745,0.8787], 
            'sczo_dv':[0.6054,0.2482,0.9627], 
            'prsn_dv':[-0.356,-0.8128,0.1008],
        }
        self.dic_prob_htn_noart = {
            'beta0':[-12.9107,-14.4641,-11.3574], 
            'age_bsln_scale10_dv':[0.5393,0.2048,0.8739], 
            'dm_dv':[1.5294,0.7287,2.3301], 
            'ckd_dv':[0.8658,-0.02477,1.7564], 
        }
        self.dic_prob_dm_art = { 
            'beta0':[-10.9992,-11.7747,-10.2237], 
            'age_bsln_scale10_dv':[0.2543,0.1412,0.3674], 
            'htn_dv':[0.9095,0.6663,1.1527], 
            'manx_dv':[0.186,-0.02457,0.3965], 
            'pvl_dv':[-0.3175,-0.6444,0.009428], 
            'year_diag_dv':[-0.03578,-0.05459,-0.01698], 
            'counter':[-0.05651,-0.09987,-0.01315],
        }
        self.dic_prob_dm_noart = {
            'beta0':[-14.1014,-16.9256,-11.2772], 
            'sex_dv':[1.,-0.0829,2.0829],
            'age_bsln_scale10_dv':[0.3994,-0.1576,0.9564],
            'manx_dv':[0.8612,-0.4291,2.1514],
        }
        self.dic_prob_oa_art = { 
            'beta0':[-14.7318,-15.5403,-13.9232],
            'sex_dv':[0.4963,0.1721,0.8205], 
            'su_bsln_dv':[0.4285,0.1405,0.7164], 
            'ncd4_bsln_dv':[0.3568,0.0733,0.6403],
            'age_bsln_scale10_dv':[0.6486,0.516,0.7813],
            'manx_dv':[0.6604,0.3866,0.9341], 
            'counter':[-0.05116,-0.105,0.002642],
        }
        self.dic_prob_oa_noart = {
            'beta0':[-14.1113,-15.904,-12.3186], 
            'age_bsln_scale10_dv':[0.6574,0.3164,0.9984], 
            'manx_dv':[1.1009,0.2604,1.9413], 
        }
        self.dic_prob_cvd_art = { 
            'beta0':[-14.0421,-14.7305,-13.3536], 
            'sex_dv':[-0.2877,-0.6102,0.03484], 
            'age_bsln_scale10_dv':[0.4499,0.3353,0.5644], 
            'su_bsln_dv':[0.3393,0.1018,0.5768], 
            'copd_dv':[0.7596,0.4712,1.0479], 
            'htn_dv':[1.3496,1.119,1.5802], 
            'dm_dv':[0.267,-0.01264,0.5467], 
            'ckd_dv':[0.3864,0.1116,0.6613], 
            'manx_dv':[0.3864,0.1676,0.6052], 
            'pvl_dv':[0.317,0.03485,0.5991], 
            'ncd4_bsln_dv':[0.3354,0.09989,0.5709], 
            'pi_dv':[0.1828,-0.05063,0.4163], 
            'counter':[0.0398,-0.00239,0.08198], 
        }
        self.dic_prob_cvd_noart = { 
            'beta0':[-16.6568,-19.742,-13.5717], 
            'sex_dv':[1.0184,0.09283,1.9441],
            'su_bsln_dv':[1.5295,0.01269,3.0464],
            'age_bsln_scale10_dv':[0.8163,0.3212,1.3114], 
            'htn_dv':[0.8782,-0.1937,1.9501], 
            'dm_dv':[1.6018,0.5659,2.6377],
        }
        self.dic_prob_ckd_art = { 
            'beta0':[-13.1993,-13.9125,-12.4861], 
            'sex_dv':[0.2015,-0.06476,0.4677],
            'su_bsln_dv':[0.6325,0.3912,0.8737], 
            'age_bsln_scale10_dv':[0.372,0.254,0.4901], 
            'htn_dv':[0.7619,0.5034,1.0203],
            'dm_dv':[0.6564,0.3746,0.9381], 
            'manx_dv':[0.3051,0.08525,0.5249], 
            'sczo_dv':[0.3238,-0.07093,0.7186], 
            'pvl_dv':[0.7909,0.5495,1.0323],
            'ncd4_bsln_dv':[0.7258,0.5013,0.9502], 
            'tdf_dv':[-0.7788,-1.0268,-0.5307], 
            'pi_dv':[0.1884,-0.05466,0.4315], 
            'insti_dv':[0.3282,0.03743,0.6189], 
            'counter':[-0.02864,-0.07203,0.01475],
        }
        self.dic_prob_ckd_noart = { 
            'beta0':[-15.7112,-19.27,-12.1562], 
            'sex_dv':[1.1679,0.2716,2.0643], 
            'age_bsln_scale10_dv':[0.727,0.2668,1.1872], 
            'htn_dv':[1.368,0.1884,2.5476], 
            'manx_dv':[0.8859,-0.1673,1.939],
        }
        self.dic_prob_cld_art = { 
            'beta0':[-11.9597,-12.6858,-11.2337], 
            'sex_dv':[0.3542,0.1654,0.543], 
            'su_bsln_dv':[1.6307,1.4182,1.8432], 
            'age_bsln_scale10_dv':[0.2443,0.1517,0.337], 
            'manx_dv':[0.2671,0.09925,0.4349], 
            'sczo_dv':[0.4294,0.1596,0.6992], 
            'pvl_dv':[0.7201,0.5328,0.9075], 
            'ncd4_bsln_dv':[0.4651,0.2976,0.6326], 
            'year_diag_dv':[-0.01557,-0.03131,0.000177], 
            'counter':[-0.02955,-0.06321,0.004112], 
            'nnrti_dv':[-0.2868,-0.5415,-0.03212], 
            'pi_dv':[-0.1788,-0.4116,0.05397], 
        }
        self.dic_prob_cld_noart = { 
            'beta0':[-11.3054,-12.0855,-10.5253], 
            'su_bsln_dv':[1.39,0.6482,2.1317], 
            'manx_dv':[1.0019,0.4305,1.5734], 
        }
        self.dic_prob_cancer_art = { 
            'beta0':[-13.6325,-14.5438,-12.7213], 
            'sex_dv':[-0.3249,-0.6201,-0.02961], 
            'age_bsln_scale10_dv':[0.5774,0.473,0.6818], 
            'copd_dv':[0.4413,0.1529,0.7296], 
            'cld_dv':[0.6282,0.3972,0.8592], 
            'dm_dv':[0.3603,0.1021,0.6185], 
            'ckd_dv':[0.5837,0.3288,0.8387], 
            'manx_dv':[0.1639,-0.03297,0.3607],
            'pvl_dv':[0.698,0.4707,0.9253], 
            'year_diag_dv':[-0.01649,-0.03425,0.001266],
            'counter':[0.06146,0.02113,0.1018],
        }
        self.dic_prob_cancer_noart = { 
            'beta0':[-15.0856,-17.1326,-13.0386], 
            'sex_dv':[-0.8201,-1.7973,0.1571], 
            'age_bsln_scale10_dv':[0.8466,0.4769,1.2162], 
            'cld_dv':[0.8376,0.08995,1.5853],
            'manx_dv':[0.822,-0.04068,1.6848], 
        }
        self.dic_prob_manx_art = { 
            'beta0':[-9.0014,-9.3786,-8.6242], 
            'sex_dv':[0.2879,0.1072,0.4687], 
            'su_bsln_dv':[0.702,0.5554,0.8487], 
            'age_bsln_scale10_dv':[-0.09675,-0.1719,-0.02161], 
            'counter':[-0.07105,-0.1008,-0.04133], 
            'prsn_dv':[0.7976,0.4173,1.1778], 
            'sczo_dv':[1.0958,0.7957,1.396], 
        }
        self.dic_prob_manx_noart = { 
            'beta0':[-8.0022,-9.0697,-6.9347], 
            'sex_dv':[0.4227,-0.02452,0.87], 
            'su_bsln_dv':[0.7807,0.2942,1.2673], 
            'age_bsln_scale10_dv':[-0.2371,-0.4804,0.006282],
            'counter':[-0.1579,-0.2623,-0.05342],
            'prsn_dv':[0.7497,-0.1567,1.656], 
        }
        self.dic_prob_prsn_art = { 
            'beta0':[-12.3436,-13.3361,-11.3511], 
            'su_bsln_dv':[1.0991,0.7144,1.4838], 
            'age_bsln_scale10_dv':[-0.269,-0.4429,-0.09506], 
            'counter':[-0.07555,-0.1386,-0.01253], 
            'manx_dv':[2.3393,1.7832,2.8954], 
            'sczo_dv':[1.2226,0.8645,1.5808], 
        }
        self.dic_prob_prsn_noart = {
            'beta0':[-10.6453,-11.4959,-9.7947], 
            'sex_dv':[-0.6208,-1.3082,0.06668],
            'manx_dv':[1.5199,0.6901,2.3498], 
            'sczo_dv':[0.9402,0.2413,1.6392], 
            'counter':[-0.1434,-0.2747,-0.01206],
        }
        self.dic_prob_sczo_art = { 
            'beta0':[-12.3298,-13.4959,-11.1638], 
            'sex_dv':[-0.3634,-0.7209,-0.00598], 
            'su_bsln_dv':[0.973,0.6384,1.3075], 
            'age_bsln_scale10_dv':[-0.3067,-0.4774,-0.136], 
            'year_diag_dv':[0.02852,0.002115,0.05493], 
            'manx_dv':[1.855,1.4309,2.2792], 
            'prsn_dv':[0.9903,0.6342,1.3464], 
        }
        self.dic_prob_sczo_noart = { 
            'beta0':[-11.5932,-14.0605,-9.126], 
            'age_bsln_scale10_dv':[-0.2745,-0.6473,0.09828],
            'counter':[-0.1464,-0.3032,0.01032],
            'manx_dv':[3.3193,1.325,5.3136], 
        }
        
        #####introduce coefficients for probability of mortality, separated by ART initiation status
        self.dic_prob_dead_art = {
            'beta0':[-13.2381,-13.7311,-12.7451], 
            'su_bsln_dv':[0.5332,0.3197,0.7467], 
            'age_bsln_scale10_dv':[0.3735,0.2826,0.4643], 
            'copd_dv':[0.6036,0.377,0.8303], 
            'dm_dv':[0.1907,-0.0451,0.4266], 
            'cvd_dv':[0.7516,0.529,0.9742], 
            'ckd_dv':[0.7483,0.5518,0.9449], 
            'cld_dv':[0.8569,0.6569,1.0568], 
            'cancer_dv':[1.1348,0.9256,1.344], 
            'pvl_dv':[0.7405,0.5543,0.9266], 
            'ncd4_bsln_dv':[0.351,0.174,0.528], 
        }
        self.dic_prob_dead_noart = { 
            'beta0':[-10.697,-11.657,-9.7371], 
            'su_bsln_dv':[0.5468,0.0027,1.0908],
            'age_bsln_scale10_dv':[0.1976,-0.011,0.4061], 
            'cvd_dv':[1.1667,0.6321,1.7013],
            'cld_dv':[0.7716,0.3159,1.2274], 
            'cancer_dv':[1.4781,0.8408,2.1153],
            'sczo_dv':[-0.5289,-1.2398,0.182],
            'counter':[-0.099,-0.1822,-0.0159],
        }

        #####introduce coefficients for probability of Dx2Tx, separated by diagnosis year
        self.min_year_diag_a2t = 1980 #different from comorbidity analysis
        self.dic_prob_a2t_bf08 = {
            'beta0':[-3.7534,-4.6908,-2.8159],
            'sex_dv':[-0.2439,-0.466,-0.02173],
            'ncd4_bsln_dv':[1.8301,1.503,2.1572],
            'age_bsln_scale10_dv':[0.2043,0.1028,0.3057],
            'year_diag_dv_a2t':[0.04718,0.01959,0.07478], #change key name so that it's different from the one used for comorbidity rate
            'counter':[0.05632,0.005707,0.1069],
        }
        self.dic_prob_a2t_0811 = {
            'beta0':[-16.9277,-21.7653,-12.0901],
            'ncd4_bsln_dv':[3.0735,2.5103,3.6367],
            'su_bsln_dv':[-0.5661,-0.9103,-0.2219],
            'age_bsln_scale10_dv':[0.3828,0.2267,0.5388],
            'year_diag_dv_a2t':[0.4989,0.3416,0.6563], #change key name so that it's different from the one used for comorbidity rate
            'counter':[0.1623,0.07528,0.2493],
        }
        self.dic_prob_a2t_sin12 = {
            'beta0':[-14.4234,-23.8426,-5.0042],
            'sex_dv':[-1.0804,-1.965,-0.1957],
            'ncd4_bsln_dv':[2.1155,1.1924,3.0386],
            'su_bsln_dv':[-2.0601,-2.9387,-1.1815],
            'year_diag_dv_a2t':[0.5448,0.2581,0.8316], #change key name so that it's different from the one used for comorbidity rate
            'counter':[0.246,0.0121,0.4798],
        }
        self.dic_random_a2t = {
            'bf08':[0.5917,0.158], #estimated mean and standard error for the random term (variance of a normal distribution with mean=0) for each individual in the model
            '0811':[1.7134,0.205],
            'sin12':[3.35,0.4129],
        }
        self.dic_coeff_adj_a2t = {'bf08':1., '0811':1., 'sin12':1.} #add to adjust probabilities estimated for each period, default=1, changed in calibration process
        self.dt_split_prob_a2t_sin12 = pd.Timestamp('2013-12-31') #separate split time for prob a2t by HIV diagnosis year
        self.dt_split_prob_a2t_bf08 = self.dt_split_prob_a2t_0811 = pd.Timestamp('2012-01-01') #probability of ART initiation split in 2012 for bf08 and 0811
        self.dic_coeff_adj_a2t_pw = {'bf08':np.ones(2), '0811':np.ones(2), 'sin12':np.ones(2)} #introduce piece-wise probability adjustment before and since 2012

        #####introduce coefficients for probabilities to Unsuppressed(U_off/U_on) from Suppressed(S)/U_off/U_on
        self.dic_prob_u_off = {
            'beta0':[-3.9046,-4.3027,-3.5064],
            'sex_dv':[0.5348,0.3716,0.698],
            'ncd4_bsln_dv':[0.4314,0.2829,0.5798],
            'su_bsln_dv':[0.8061,0.6486,0.9637],
            'age_art_bsln_scale10_dv':[-0.3329,-0.4139,-0.2519], 
            'art_counter':[-0.05679,-0.07077,--0.04281], #art_counter is different from counter for baseline_dt
            'u_off_dv':[4.604,4.4269,4.7811], #u_off_dv indicating whether previous step is at state U_off(1) or not(0)
            'u_on_dv':[2.6261,2.4587,2.7935],
            'manx_dv':[0.0922,-0.05188,0.2363],
            'prsn_dv':[-0.2136,-0.5005,0.07323],
        }
        self.dic_prob_u_on = {
            'beta0':[-4.0709,-4.3778,-3.764],
            'sex_dv':[0.3246,0.1879,0.4612],
            'ncd4_bsln_dv':[0.6242,0.5081,0.7403],
            'su_bsln_dv':[0.9543,0.8354,1.0733],
            'age_art_bsln_scale10_dv':[-0.1361,-0.1951,-0.07717],
            'art_counter':[-0.02164,-0.03048,-0.0128], 
            'u_off_dv':[3.1126,2.9509,3.2743], 
            'u_on_dv':[2.2239,2.1083,2.3395],
            'manx_dv':[0.1825,0.07313,0.2919],
            'prsn_dv':[0.2014,-0.00547,0.4083],
        }
        self.dic_random_t = {
            'u_off':[0.809,0.1082], #estimated mean and standard error for the random term (variance of a normal distribution with mean=0) for each individual in the model
            'u_on':[1.2323,0.09071],
        }
        self.dic_coeff_prob_art = {'s_off':1., 's_on':1., 'off_s':1., 'off_on':1., 'on_on':1., 'on_off':1.} #adjust derived probabilities of ART transitions, use on_on instead of on_s with a smaller probability

        #####introduce coefficient to adjust probability of comorbidity incidence, default=1
        self.prob_comorb_coeff = 1.

        #####introduce parameters such as mortality rate for undiagnosed PLWH to estimate prevalence over time
        self.d_U = 1/12.1 #same for mono-/co-infected HIV, (PMID:28301424) median survival time from seroconversion at age 30 (29 as median age at seroconversion) to death without ART among male from Europe/North America
        self.sig_iur_params = np.array([5.38300232e-02, 3.19324163e-02, 5.07343216e+00, 2.01477145e+03]) #in the order of y0,y1,tm,t0
        self.sig_iur_params_lb = np.array([0.05263581, 0.02736219]) #introduce uncertainty in IUR estimation based on the credible interval from BC-CfE modeling
        self.sig_iur_params_ub = np.array([0.05502749, 0.03724505])
        self.sig_rdiag_params = np.array([1.05539568e-01, 2.00892319e-01, 2.88342681e+00, 2.01116228e+03]) #fit to diagnosis from BCCDC
        self.coeff_rdiag = 1.02912445 #probability of diagnosis adjusted based on cum. new diagnoses in 2016
        self.sig_migr_params = np.array([2.84370282e+02, 2.46489938e+02, 1.43629892e+00, 2.01282960e+03])
        self.init_prev = 9381.033992473947 #based on PHAC's estimates 

        #####introduce parameters to adjust comorbidity incidence probability for sensitivity analysis
        self.dic_prob_comorb_sens_coeff_range = {'plus':1.25,'minus':0.75} #scaling for each comorbidity probability
        self.dic_prob_comorb_sens_coeff = {
            'cvd':1., #will be replaced by dic_prob_comorb_sens_coeff_range if specifying plus or minus scenario
            'htn':1.,
            'dm':1.,
            'oa':1.,
            'copd':1.,
            'ckd':1.,
            'cld':1.,
            'cancer':1.,
            'manx':1.,
            'sczo':1.,
            'prsn':1.,
        }

if __name__ == "__main__":

    #####Files derived using data at individual level for the microsimulation model
    runs = ['record_cd4_nadir'] #derive CD4 nadir at baseline: cd4_nadir_baseline.pkl
    runs = ['record_derived_comorbidites'] #derive comorbidity status: comorbidities.pkl
    runs = ['record_oud_all'] #derive OUD diagnosis: oud_record.pkl
    runs = ['record_pwid_all'] #derive PWID diagnosis: pwid_record.pkl
    runs = ['record_derived_substance_use'] #derive substance use information: substance_use.pkl
    runs = ['record_spvl_rebound'] #derive information about spvl/rebound: spvl_rebound_dt.pkl
    runs = ['record_rebound_interruption_failure'] #derive ART interruption/failure information for STOP: spvl_rebound_status.pkl
    runs = ['record_spvl_rebound_interruption/failure_dtp'] #derive spvl, ART interruption/failure information for DTP: dtp_spvl_rebound_status.pkl
    runs = ['record_hiv_status_every6m'] #derive HIV status for STOP: hiv_status_pre.pkl
    runs = ['record_hiv_status_every6m_dtp'] #derive HIV status for DTP: dtp_hiv_status_pre.pkl
    runs = ['record_hiv_status_every6m_adj'] #adjust HIV status for STOP and DTP: hiv_status.pkl and dtp_hiv_status.pkl
    runs = ['record_spec_art_regimen'] #derive art_class.pkl
    runs = ['record_art_regimen_by_timestep'] #derive information of ART regimen: art_reg_status.pkl
    runs = ['record_status_pvl_reg_1yr'] #derive ART regimen related to pvl status at each 1-year interval: pvl_status.pkl and prop_time_art_reg.pkl
    runs = ['record_art_regimen_dist'] #derive regimen distribution information for parameter update: prob_reg_switch.xlsx
    runs = ['record_prob_ART_init'] #derive probability of suprression/failure after ART initiation as prob_A2S and prob_A2Uon in func_params_update
    
    #####Model calibration and validation
    runs = ['cali_new_diag'] #model new diagnosed PLWH in BC (local diagnosis and migrants), Figures S1-S3
    runs = ['microsim_full_hist_cali'] #model calibration, Figures S4-S7 for the Supplementary Information
    runs = ['microsim_full_rand_vali'] #model validation, Figures S8-S10 for the SI
    
    #####Main model
    runs = ['microsim_rand_sim_results'] #main model

    #####Results and sensitivity analysis
    runs = ['figure_dplwh_age_comorb'] #create Figures 2-5 for the manuscript
    runs = ['figure_ranking_comorbidities'] #create Figure 6 for the manuscript and Figures S12-S13 for the SI
    runs = ['figure_ranking_comb_comorbidities'] #create Figures 7-8 and Figure S14
    runs = ['figure_sens_comorb_prob'] #create Figure S15 for sensitivity analysis

    #####define color palettes
    dic_color_palette = {'blue':['#%02x%02x%02x'%(33,89,103),'#%02x%02x%02x'%(75,172,198),'#%02x%02x%02x'%(183,222,232)],
            'green':['#%02x%02x%02x'%(79,98,40),'#%02x%02x%02x'%(155,187,89),'#%02x%02x%02x'%(216,228,188)],
            'red':['#%02x%02x%02x'%(128,0,0),'#%02x%02x%02x'%(192,0,0),'#%02x%02x%02x'%(230,184,183)],
            'purple':['#%02x%02x%02x'%(131,0,131),'#%02x%02x%02x'%(148,116,180),'#%02x%02x%02x'%(204,191,222)],
            'pink':['#%02x%02x%02x'%(193,116,142),'#%02x%02x%02x'%(217,171,187),'#%02x%02x%02x'%(236,213,221)]}

    list_color_add = ['#%02x%02x%02x'%(171,87,121),'#%02x%02x%02x'%(240,218,229),'#%02x%02x%02x'%(250,243,247),'#%02x%02x%02x'%(192,174,183)] #additional color choices

    if 'record_oud_all' in runs:
        """find all the records in MSP, DAD and PharmaNet relevant to OUD definition, to apply look-back window to re-define earliest OUD diagnoses date"""

        #####create dataframe to save OUD information with moh_id, source(MSP,DAD or PNET), dt(diagnosis_date) and type(OUD)
        dat_oud_record = pd.DataFrame([],columns=['moh_id','dt','source','type'])

        ######load data of MSP records from msp_app 
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        dat_msp_app = pd.read_sas('data/msp_app.sas7bdat',encoding='latin-1') #file for MSP Alternate Payment Program
        file_data = open('data/moh/msp_ffs.pkl','rb')
        dat_msp_ffs = pickle.load(file_data) 

        #####merge ffs and app before applying OUD algorithm
        dat_msp = pd.concat([dat_msp_app[list(dat_msp_ffs)], dat_msp_ffs],ignore_index=True)

        #####identify patient with records relevant to OUD diagnosis, output as oud_record_msp.pkl
        arr_fitm = [39, 15039] #OAT billing code, certain for OUD, should be considered separately from DIAGCD
        arr_diagcd = ['3040','3047','3055','9650','E8500','E8501','E8502','F11'] #X42,X62 etc. should only be considered in DAD data
        dat_msp_study = dat_msp[(dat_msp.FITM.isin(arr_fitm))|(dat_msp.DIAGCD.str.startswith(tuple(arr_diagcd)))].reset_index(drop=True) #consider all the records with DIAGCD starting with the specific codes instead of equal
        dat_msp_study_rmdup1 = dat_msp_study.groupby(['moh_id','SERVDT','CLMSPEC','DIAGCD','FITM']).size().reset_index().rename(columns={0:'count'}) 
        dat_msp_study_rmdup2 = dat_msp_study_rmdup1.groupby(['moh_id','SERVDT']).size().reset_index().rename(columns={0:'count'}) #unique combination of moh_id and SERVDT
        arr_id_msp = [id_i for id_i in dat_msp_study_rmdup2.moh_id.unique() if dat_msp_study_rmdup2[dat_msp_study_rmdup2.moh_id==id_i].shape[0]>=3] 
        dat_oud_msp = pd.DataFrame([],columns=['moh_id','dt'])
        for id_i in arr_id_msp:
            dat_msp_i = dat_msp_study[['moh_id','SERVDT']][dat_msp_study['moh_id']==id_i].sort_values(by='SERVDT').reset_index(drop=True) #need to sort by service date again since the order may be changed due to searching order of diag_cd
            dat_msp_i = dat_msp_i.head(-2).rename(columns={'SERVDT':'dt'}) #exclude the last two records since those alone cannot determine OUD diagnosis
            dat_oud_msp = pd.concat([dat_oud_msp,dat_msp_i],ignore_index=True)
        dat_oud_msp['source'] = 'MSP'

        #####load data of DAD records and record relevant ICD-9/10 codes for OUD diagnosis
        dat_dad = pd.read_sas('data/dad.sas7bdat',encoding='latin-1') #file for hospitalization 
        arr_dad_col9 = ['DIAG'+str(i) for i in range(1,17)]
        arr_dad_col10 = ['DIAGX'+str(i) for i in range(1,26)]
        arr_diag9 = ['3040','3047','3055','9650','E8500','E8501','E8502']
        arr_diag10_1 = ['F11','X42','X62','Y12'] #F11 should be used solely, and the rest should be combined with arr_diag10_2
        arr_diag10_2 = ['T400','T401','T402','T403','T404','T406']

        #####identify patient with OUD-related hospitalization, save all the relevant records for each individual
        dat_dad_oud = pd.DataFrame([],columns=['moh_id','addate'])
        for id_i in dat_dad.moh_id.unique():
            
            dat_dad_i = dat_dad[dat_dad.moh_id==id_i].sort_values(by='addate').reset_index(drop=True) #start check from their earliest record
            dat_dad_i[arr_dad_col9] = dat_dad_i[arr_dad_col9].fillna('') #replace nan by '' so the string check works for the whole list
            dat_dad_i[arr_dad_col10] = dat_dad_i[arr_dad_col10].fillna('') 
            for index_j in dat_dad_i.index:
                
                ###any ICD-9 code in arr_diag9 shown up in the diagnosis
                if any([item.startswith(tuple(arr_diag9)) for item in dat_dad_i[arr_dad_col9][dat_dad_i.index==index_j].values[0]]): 
                    dic_oud = {'moh_id':id_i,'addate':dat_dad_i.addate[dat_dad_i.index==index_j].values[0]}
                    dat_dad_oud = pd.concat([dat_dad_oud,pd.DataFrame(dic_oud,index=[0])],ignore_index=True) 

                ###F11 shown up in any arr_dad_col10
                elif any([item.startswith('F11') for item in dat_dad_i[arr_dad_col10][dat_dad_i.index==index_j].values[0]]):
                    dic_oud = {'moh_id':id_i,'addate':dat_dad_i.addate[dat_dad_i.index==index_j].values[0]}
                    dat_dad_oud = pd.concat([dat_dad_oud,pd.DataFrame(dic_oud,index=[0])],ignore_index=True)

                ###code in arr_diag10_1[1:] and arr_diag10_2 showed up in the same hospitalization record
                elif any([item.startswith(tuple(arr_diag10_1[1:])) for item in dat_dad_i[arr_dad_col10][dat_dad_i.index==index_j].values[0]]) and any([item.startswith(tuple(arr_diag10_2)) for item in dat_dad_i[arr_dad_col10][dat_dad_i.index==index_j].values[0]]): 
                    dic_oud = {'moh_id':id_i,'addate':dat_dad_i.addate[dat_dad_i.index==index_j].values[0]}
                    dat_dad_oud = pd.concat([dat_dad_oud,pd.DataFrame(dic_oud,index=[0])],ignore_index=True)

        dat_dad_oud = dat_dad_oud.rename(columns={'addate':'dt'})
        dat_dad_oud['source'] = 'DAD'

        #####load PharmaNet data 
        file_data = open('data/pnet.pkl','rb')
        dat_pnet = pickle.load(file_data) 
        dat_pnet = dat_pnet[pd.isnull(dat_pnet['date_of_service'])==False].reset_index(drop=True) 
        
        #####identify patients with OAT dispense record from PharmaNet, save all relevant records to dat_oud
        arr_din_pin = [999792, 999793, 66999990, 66999991, 66999992,66999993, 
                        66999997, 66999998, 66999999, 67000000, 67000001,67000002,
                        67000003, 67000004, 67000005, 67000006, 67000007, 67000008,
                        2295695, 2295709, 2408090, 2408104, 2424851, 2424878, 2453908, 2453916, 2468085, 2468093,
                        2242962, 2242963, 2242964,66999994,66999995, 66999996,
                        22123349, 22123346, 22123347, 22123348] 
        dat_pnet_study = dat_pnet[['moh_id','date_of_service']][dat_pnet.DIN_PIN.isin(arr_din_pin)].reset_index(drop=True).rename(columns={'date_of_service':'dt'}) 
        dat_pnet_study['source'] = 'PNET'
        dat_pnet_study = dat_pnet_study[['moh_id','dt','source']].copy() #only keep date and source=PNET
        
        #####combine records from MSP, DAD and PNET together
        dat_oud_record = pd.concat([dat_oud_msp,dat_dad_oud,dat_pnet_study],ignore_index=True) 
        dat_oud_record['type'] = 'OUD'
        dat_oud_record = dat_oud_record.sort_values(['moh_id','dt']).reset_index(drop=True) 
        dat_oud_record['dt'] = pd.to_datetime(dat_oud_record['dt']) #change the format of dates to datetime

        file_data = open('data/oud_record.pkl','wb') 
        pickle.dump(dat_oud_record, file_data)
        file_data.close()

    if 'record_pwid_all' in runs:
        """save all PWID-related diagnosis dates in one file, used to determine substance use behaviour based on look-back window"""

        p = Params()

        ######load data of MSP records from msp
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        dat_msp_app = pd.read_sas('data/msp_app.sas7bdat',encoding='latin-1') #file for MSP Alternate Payment Program
        file_data = open('data/msp_ffs.pkl','rb')
        dat_msp_ffs = pickle.load(file_data) 
        
        #####merge ffs and app before applying OUD algorithm
        dat_msp = dat_msp_app[list(dat_msp_ffs)].append(dat_msp_ffs,ignore_index=True) 

        #####identify MSP records related to drug-related diagnosis
        arr_diagcd = ['292','970','3040','3041','3042','3044','3045','3046','3047','3048','3049','3054','3055','3056','3057','3059',
                '6483','7960','9621','9650','9658','9663','9664','9670','9684','9685','9694','9696','9697','9698','9699','E8500','V6542'] #MSP ICD-9 diagnostic codes starting with those numbers
        dat_msp_study = pd.DataFrame([])
        for code_i in arr_diagcd:
            dat_msp_i = dat_msp[dat_msp.DIAGCD.str.startswith(code_i)].reset_index(drop=True) 
            if dat_msp_i.empty:
                print (code_i) #show which code number does not exist in our MSP records
            dat_msp_study = pd.concat([dat_msp_study,dat_msp_i],ignore_index=True) #service date may not be in ascending order since searching in order of diagcd

        #####drop the records with the same CLMSPEC+DIAGCD+SERVDT+FITM for each individual and drop the records with the same moh_id+SERVDT
        dat_msp_study_rmdup1 = dat_msp_study.groupby(['moh_id','SERVDT','CLMSPEC','DIAGCD','FITM']).size().reset_index().rename(columns={0:'count'}) 
        dat_msp_study_rmdup2 = dat_msp_study_rmdup1.groupby(['moh_id','SERVDT']).size().reset_index().rename(columns={0:'count'}) #unique combination of moh_id and SERVDT
        list_id_msp = [id_i for id_i in dat_msp_study_rmdup2.moh_id.unique() if dat_msp_study_rmdup2[dat_msp_study_rmdup2.moh_id==id_i].shape[0]>=2] 

        #####save all the PWID-related MSP records
        dat_pwid_msp = pd.DataFrame([],columns=['moh_id','dt'])
        for id_i in list_id_msp:
            dat_msp_i = dat_msp_study[['moh_id','SERVDT']][dat_msp_study['moh_id']==id_i].sort_values(by='SERVDT').reset_index(drop=True) 
            dat_msp_i = dat_msp_i.head(-1).rename(columns={'SERVDT':'dt'}) #exclude the final record since it alone cannot determine PWID diagnosis
            dat_pwid_msp = dat_pwid_msp.append(dat_msp_i,ignore_index=True)
        dat_pwid_msp['source'] = 'MSP'
        dat_pwid_msp['DOB'] = pd.Series(dat_cohort['DOB'][dat_cohort['moh_id']==id_i].values[0] for id_i in dat_pwid_msp['moh_id'])
        dat_pwid_msp['age_diag'] = pd.Series((dat_pwid_msp['dt']-dat_pwid_msp['DOB']).dt.days/p.def_year)
        dat_pwid_msp = dat_pwid_msp[(dat_pwid_msp['age_diag']>=11)&(dat_pwid_msp['age_diag']<=65)].reset_index(drop=True) 

        #####load hospitalization records and all pwid-related codes
        dat_dad = pd.read_sas('data/dad.sas7bdat',encoding='latin-1') #file for hospitalization discharge
        arr_dad_col9 = ['DIAG'+str(i) for i in range(1,17)]
        arr_dad_col10 = ['DIAGX'+str(i) for i in range(1,26)]
        arr_diag9_1 = ['292','970','3040','3041','3042','3044','3045','3046','3047','3048','3049','3054','3055','3056','3057','3059',
                '6483','7960','9621','9650','9658','9663','9664','9670','9684','9685','9694','9696','9697','9698','9699','E8500'] #same as codes in MSP data
        arr_diag9_2 = ['V6542'] #search for ICD-9 codes exactly the same as in arr_diag9_2
        arr_diag10_1 = ['F11','F13','F14','F15','F19'] #search for codes begin with those
        arr_diag10_2 = ['R781','R782','T387','T400','T401','T402','T403','T404','T405','T406','T408','T409','T412',
                'T423','T424','T425','T426','T427','T428','T436','T437','T438','T439','T507'] #search for codes exactly as those
        arr_diag10_nacrs = ['751','753'] #search for codes beginning with those NACRS ICD 10 complaint codes

        #####identify patient with drug-related hospitalization/emergency department visit, save id and all related addate to dat_pwid
        dat_pwid_hosp = pd.DataFrame([],columns=['moh_id','DAD_addate'])
        for id_i in dat_dad.moh_id.unique():
            
            dat_dad_i = dat_dad[dat_dad.moh_id==id_i].sort_values(by='addate').reset_index(drop=True) #start check from their earliest record, for evaluate diagnosis age later
            
            for index_j in dat_dad_i.index: 
                
                ###any ICD-9 code beginning with codes in arr_diag9_1 
                k = 0
                while k<len(arr_diag9_1) and any(np.char.startswith(dat_dad_i[arr_dad_col9][dat_dad_i.index==index_j].values[0].astype(str),arr_diag9_1[k]))==False:
                    k += 1
                if k<len(arr_diag9_1): #confirm the while loop stopped when the condition is satified
                    dic_pwid_hosp = {'moh_id':id_i,'DAD_addate':dat_dad_i.addate[dat_dad_i.index==index_j].values[0]}
                    dat_pwid_hosp = dat_pwid_hosp.append(dic_pwid_hosp,ignore_index=True)
                
                ###check if exact value from arr_diag9_2 can be found
                if any([item in dat_dad_i[arr_dad_col9][dat_dad_i.index==index_j].values for item in arr_diag9_2]):
                    dic_pwid_hosp = {'moh_id':id_i,'DAD_addate':dat_dad_i.addate[dat_dad_i.index==index_j].values[0]}
                    dat_pwid_hosp = dat_pwid_hosp.append(dic_pwid_hosp,ignore_index=True)

                ###any ICD-10 code begins with codes in arr_diag10_1 
                k = 0
                while k<len(arr_diag10_1) and any(np.char.startswith(dat_dad_i[arr_dad_col10][dat_dad_i.index==index_j].values[0].astype(str),arr_diag10_1[k]))==False:
                    k += 1
                if k<len(arr_diag10_1): #confirm the while loop stopped when the condition is satified
                    dic_pwid_hosp = {'moh_id':id_i,'DAD_addate':dat_dad_i.addate[dat_dad_i.index==index_j].values[0]}
                    dat_pwid_hosp = dat_pwid_hosp.append(dic_pwid_hosp,ignore_index=True)

                ###any exact ICD-10 code in arr_diag10_2 
                k = 0
                while k<len(arr_diag10_2) and (arr_diag10_2[k] in dat_dad_i[arr_dad_col10][dat_dad_i.index==index_j].values[0])==False: 
                    k += 1
                if k<len(arr_diag10_2): #confirm the while loop stopped when the condition is satified
                    dic_pwid_hosp = {'moh_id':id_i,'DAD_addate':dat_dad_i.addate[dat_dad_i.index==index_j].values[0]}
                    dat_pwid_hosp = dat_pwid_hosp.append(dic_pwid_hosp,ignore_index=True)
                    

                ###any ICD-10 code begins with codes in arr_diag10_nacrs 
                k = 0
                while k<len(arr_diag10_nacrs) and any(np.char.startswith(dat_dad_i[arr_dad_col10][dat_dad_i.index==index_j].values[0].astype(str),arr_diag10_nacrs[k]))==False:
                    k += 1
                if k<len(arr_diag10_nacrs): #confirm the while loop stopped when the condition is satified
                    dic_pwid_hosp = {'moh_id':id_i,'DAD_addate':dat_dad_i.addate[dat_dad_i.index==index_j].values[0]}
                    dat_pwid_hosp = dat_pwid_hosp.append(dic_pwid_hosp,ignore_index=True)
                    
        dat_pwid_hosp = dat_pwid_hosp.rename(columns={'DAD_addate':'dt'})
        dat_pwid_hosp['source'] = 'DAD'
        dat_pwid_hosp['DOB'] = pd.Series(dat_cohort['DOB'][dat_cohort['moh_id']==id_i].values[0] for id_i in dat_pwid_hosp['moh_id'])
        dat_pwid_hosp['age_diag'] = pd.Series((dat_pwid_hosp['dt']-dat_pwid_hosp['DOB']).dt.days/p.def_year)
        dat_pwid_hosp_age = dat_pwid_hosp[(dat_pwid_hosp['age_diag']>=11)&(dat_pwid_hosp['age_diag']<=65)].reset_index(drop=True) 

        #####combine records from MSP and DAD
        dat_pwid_record = dat_pwid_msp.append(dat_pwid_hosp_age,ignore_index=True) 
        dat_pwid_record['type'] = 'PWID'
        dat_pwid_record = dat_pwid_record.sort_values(['moh_id','dt']).reset_index(drop=True) 

        file_data = open('data/pwid_record.pkl','wb') 
        pickle.dump(dat_pwid_record, file_data)
        file_data.close()

    if 'record_derived_comorbidites' in runs:
        """Apply 5-years lookback window for prevalent/incident comorbidity cases using 
            set_baseline_dt=max(2008-01-01,earliest_HIV+1yr,FARVDT+1yr,earliest_record+5yr)"""

        p = Params()

        ######load data of STOP7 and appliy eligibility criteria
        dat_cohort = pd.read_sas('data/demo/stop.sas7bdat',encoding='latin-1')
        dat_record = pd.read_sas('data/earliest_record.sas7bdat',encoding='latin-1') 
        dat_cohort = pd.merge(dat_cohort,dat_record[dat_record.index<dat_record.index.max()],how='left',on='moh_id')
        dat_check = dat_cohort[(dat_cohort.end_fu_dt.dt.year>=2008)].reset_index(drop=True) 
        dat_check = dat_check[(pd.isnull(dat_check.sex_at_birth_dv)==False)&(pd.isnull(dat_check.DOB)==False)].reset_index(drop=True) 
        dat_check = dat_check[(dat_check.earliest_HIV-dat_check.DOB).dt.days/p.def_year>=18].reset_index(drop=True) 
        dat_check['FARVDT'] = dat_check['FARVDT'].apply(lambda x: x if x<=pd.Timestamp('2017-03-31') else np.nan) #remove FARVDT if it happened after the cut-off date of STOP7

        #####load data of comorbidities 
        list_comorb_names = ['cvd','htn','dm','oa','copd','ckd','cld_cmb','manx','sczo_moh','prsn'] 
        dic_comorb = {}
        for name_i in list_comorb_names:
            dic_comorb[name_i] = pd.read_sas('data/'+name_i+'.sas7bdat',encoding='latin-1')
            col_date = [x for x in list(dic_comorb[name_i]) if 'dt' in x][0]
            dic_comorb[name_i] = dic_comorb[name_i].rename(columns={col_date:'dt'})
        dic_comorb['sczo'] = dic_comorb['sczo_moh'].copy()
        dic_comorb['cld'] = dic_comorb['cld_cmb'].copy()

        #####load cancer file after removing AIDS-related cancer
        dic_comorb['cancer'] = pd.read_sas('data/cancer_noaids.sas7bdat',encoding='latin-1')
        dic_comorb['cancer'] = dic_comorb['cancer'].rename(columns={'Cancer_dt':'dt'})

        #####use anchor date as the latest date among ealriest_HIV+1yr, FARVDT+1yr and earliest_record_dt+5yr 
        window_lookback = p.def_year*5 #define length of lookback window
        dat_check['1yr_after_eHIV'] = pd.Series(dat_check['earliest_HIV']+pd.to_timedelta(p.def_year,unit='d'))
        dat_check['1yr_after_FARVDT'] = pd.Series(dat_check['FARVDT']+pd.to_timedelta(p.def_year,unit='d'))
        dat_check['2yr_after_eRecord'] = pd.Series(dat_check['earliest_record_dt']+pd.to_timedelta(window_lookback,unit='d'))        
        dat_check['study_begin_dt'] = pd.Series([pd.Timestamp('2008-01-01')]*dat_check.shape[0])
        dat_check['set_baseline_dt'] = dat_check[['1yr_after_eHIV','1yr_after_FARVDT','2yr_after_eRecord','study_begin_dt']].apply(max,axis=1)
        dat_final = dat_check.copy() 
        dat_final['lookback_dt'] = pd.Series(dat_final['set_baseline_dt']-pd.to_timedelta(window_lookback,unit='d'))

        #####create dataframe to save earliest_comorb_dt for prevalent case if comorb_dt in lookback_dt-baseline_dt and for incident case if comorb_dt only after baseline_dt
        dat_final_comorb = dat_final[['moh_id','earliest_HIV','lookback_dt','set_baseline_dt']].copy()
        list_comorb = ['cvd','htn','dm','oa','copd','ckd','cld','cancer','manx','sczo','prsn']
        for comorb_i in list_comorb:
            dat_comorb_i = dic_comorb[comorb_i]
            dic_record_i = {'moh_id':[], 'earliest_'+comorb_i+'_dt':[]}
            for id_j in dat_comorb_i['moh_id'][dat_comorb_i['moh_id'].isin(dat_final_comorb['moh_id'])].unique():
                start_dt = dat_final_comorb['lookback_dt'][dat_final_comorb['moh_id']==id_j].values[0]
                stop_dt = dat_final_comorb['set_baseline_dt'][dat_final_comorb['moh_id']==id_j].values[0]
                if not dat_comorb_i[(dat_comorb_i['moh_id']==id_j)&(dat_comorb_i['dt']>=start_dt)&(dat_comorb_i['dt']<stop_dt)].empty:
                    dic_record_i['moh_id'].append(id_j)
                    dic_record_i['earliest_'+comorb_i+'_dt'].append(dat_comorb_i['dt'][(dat_comorb_i['moh_id']==id_j)&(dat_comorb_i['dt']>=start_dt)&(dat_comorb_i['dt']<stop_dt)].min()) #keep the earliest record found in the look-back window 
                elif not dat_comorb_i[(dat_comorb_i['moh_id']==id_j)&(dat_comorb_i['dt']>=stop_dt)].empty:
                    dic_record_i['moh_id'].append(id_j)
                    dic_record_i['earliest_'+comorb_i+'_dt'].append(dat_comorb_i['dt'][(dat_comorb_i['moh_id']==id_j)&(dat_comorb_i['dt']>=stop_dt)].min())
                else:
                    dic_record_i['moh_id'].append(id_j)
                    dic_record_i['earliest_'+comorb_i+'_dt'].append(np.nan)
            dat_record_i = pd.DataFrame.from_dict(dic_record_i)
            dat_final_comorb = pd.merge(dat_final_comorb,dat_record_i,how='left',on='moh_id')

        #####save the derived earliest_comorb_dt based on washout period to .pkl file
        file_data = open('data/comorbidities.pkl','wb') 
        pickle.dump(dat_final_comorb, file_data)

    if 'record_derived_substance_use' in runs:
        """derive prevalent and incident cases of substance use (IDU/OUD/PWID) using the same 5-year washout period, consistent with the case-finding algorithm for comorbidities"""

        p = Params()

        ######load data of STOP
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        dat_record = pd.read_sas('data/earliest_health_record.sas7bdat',encoding='latin-1') #dates of earliest health records
        dat_cohort = pd.merge(dat_cohort,dat_record[dat_record.index<dat_record.index.max()],how='left',on='moh_id')
        dat_check = dat_cohort[(dat_cohort.end_fu_dt.dt.year>=2008)].reset_index(drop=True) 
        dat_check = dat_check[(pd.isnull(dat_check.sex_at_birth_dv)==False)&(pd.isnull(dat_check.DOB)==False)].reset_index(drop=True) 
        dat_check = dat_check[(dat_check.earliest_HIV-dat_check.DOB).dt.days/p.def_year>=18].reset_index(drop=True) #at least 18 years old at diagnosis
        dat_check['FARVDT'] = dat_check['FARVDT'].apply(lambda x: x if x<=pd.Timestamp('2017-03-31') else np.nan) #remove FARVDT if it happened after the cut-off date of STOP
        
        #####load OUD and PWID diagnosis dates based on case-finding algorithm from PMID: 31899565 and PMID: 34097994
        file_data = open('data/oud_record.pkl','rb') #dates for OUD-related health visits
        dat_oud = pickle.load(file_data)
        file_data = open('data/pwid_record.pkl','rb') #dates for PWID-related health visits
        dat_pwid = pickle.load(file_data)

        #####ignore oud and pwid records after 2016-12-31 for those eligible for statisitcal analysis
        dat_oud = dat_oud[dat_oud['dt']<=cutoff_dt].reset_index(drop=True)
        dat_pwid = dat_pwid[dat_pwid['dt']<=cutoff_dt].reset_index(drop=True)

        #####define anchor date as the latest date among ealriest_HIV+1yr, FARVDT+1yr, earliest_record_dt+2yr and study_begin_dt (2008-01-01) 
        window_lookback = p.def_year*5 #define length of lookback window
        dat_check['1yr_after_eHIV'] = pd.Series(dat_check['earliest_HIV']+pd.to_timedelta(p.def_year,unit='d'))
        dat_check['1yr_after_FARVDT'] = pd.Series(dat_check['FARVDT']+pd.to_timedelta(p.def_year,unit='d'))
        dat_check['2yr_after_eRecord'] = pd.Series(dat_check['earliest_record_dt']+pd.to_timedelta(window_lookback,unit='d'))
        dat_check['study_begin_dt'] = pd.Series([pd.Timestamp('2008-01-01')]*dat_check.shape[0])
        dat_check['set_baseline_dt'] = dat_check[['1yr_after_eHIV','1yr_after_FARVDT','2yr_after_eRecord','study_begin_dt']].apply(max,axis=1)
        dat_final = dat_check.copy() #derive earliest_comorb_dt for all participants in dat_cohort without limitation on follow-up time
        dat_final['lookback_dt'] = pd.Series(dat_final['set_baseline_dt']-pd.to_timedelta(window_lookback,unit='d'))

        #####create dataframe to save earliest_su_dt for prevalent case if IDU=1 or dat_oud[dt]/dat_pwid[dt] in lookback_dt-baseline_dt and for incident case if dt only after baseline_dt
        dat_final_su = dat_final[['moh_id','IDU_all','earliest_HIV','FARVDT','earliest_record_dt','lookback_dt','set_baseline_dt']].copy() #keep all the records relevant to lookback window and substance use status
        dic_record_i = {'moh_id':[], 'earliest_su_dt':[]}
        for id_j in dat_final_su['moh_id'].values:
            if dat_final_su['IDU_all'][dat_final_su['moh_id']==id_j].values[0]==1:
                dic_record_i['moh_id'].append(id_j)
                dic_record_i['earliest_su_dt'].append(dat_final_su['earliest_HIV'][dat_final_su['moh_id']==id_j].values[0]-pd.to_timedelta(1.,unit='d')) #mark earliest_su_dt<earliest_HIV for IDU=1 to make sure it will be identified as prevalent case
            else:
                start_dt = dat_final_su['lookback_dt'][dat_final_su['moh_id']==id_j].values[0]
                stop_dt = dat_final_su['set_baseline_dt'][dat_final_su['moh_id']==id_j].values[0]
                if (not dat_oud[(dat_oud['moh_id']==id_j)&(dat_oud['dt']>=start_dt)&(dat_oud['dt']<stop_dt)].empty) or (not dat_pwid[(dat_pwid['moh_id']==id_j)&(dat_pwid['dt']>=start_dt)&(dat_pwid['dt']<stop_dt)].empty):
                    dic_record_i['moh_id'].append(id_j)
                    arr_su_dt = np.concatenate((dat_oud['dt'][(dat_oud['moh_id']==id_j)&(dat_oud['dt']>=start_dt)&(dat_oud['dt']<stop_dt)].values,dat_pwid['dt'][(dat_pwid['moh_id']==id_j)&(dat_pwid['dt']>=start_dt)&(dat_pwid['dt']<stop_dt)].values))
                    dic_record_i['earliest_su_dt'].append(arr_su_dt.min()) #keep the earliest OUD/PWID record captured in the lookback window
                elif (not dat_oud[(dat_oud['moh_id']==id_j)&(dat_oud['dt']>=stop_dt)].empty) or (not dat_pwid[(dat_pwid['moh_id']==id_j)&(dat_pwid['dt']>=stop_dt)].empty):
                    arr_dates = np.concatenate((dat_oud['dt'][(dat_oud['moh_id']==id_j)&(dat_oud['dt']>=stop_dt)].values,dat_pwid['dt'][(dat_pwid['moh_id']==id_j)&(dat_pwid['dt']>=stop_dt)].values))
                    dic_record_i['moh_id'].append(id_j)
                    dic_record_i['earliest_su_dt'].append(arr_dates.min())
                else:
                    dic_record_i['moh_id'].append(id_j)
                    dic_record_i['earliest_su_dt'].append(np.nan)
        dat_record_i = pd.DataFrame.from_dict(dic_record_i)
        dat_final_su = pd.merge(dat_final_su,dat_record_i,how='left',on='moh_id')
        
        #####save records as .pkl file
        file_data = open('data/substance_use.pkl','wb') 
        pickle.dump(dat_final_su, file_data)

    if 'record_cd4_nadir' in runs:
        """create CD4 nadir information at baseline (i.e., max(earliest_HIV,2008-01-01)) for each STOP participants
        definition of CD4 nadir at diagnosis: lowest CD4 counts within +/- 6 months of earliest_HIV"""

        def_year = 365.25 #year definition for time window

        ######load data of STOP and keep the necessary information in the new dataframe
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        dat_cd4 = pd.read_sas('data/stop_cd4.sas7bdat',encoding='latin-1')
        dat_cohort['baseline_dt'] = dat_cohort['earliest_HIV'].apply(lambda x: max(x,pd.Timestamp('2008-01-01'))) #start follow participants either from 2008-01-01 or earliest_HIV later than 2008-01-01
        dat_ncd4 = dat_cohort[['moh_id','baseline_dt']].copy() 

        #####define time window to identify lowest CD4 near the baseline
        t_window = def_year #define cd4 nadir within 1 year around baseline (+1/-1 year around baseline)

        #####for each participant, find lowest cd4 count within the time window if available
        arr_ncd4 = []
        arr_ncd4_dt = []
        for id_i in dat_ncd4.moh_id:
            t1 = dat_cohort['baseline_dt'][dat_cohort.moh_id==id_i].values[0]-pd.to_timedelta(t_window,unit='d') 
            t2 = dat_cohort['baseline_dt'][dat_cohort.moh_id==id_i].values[0]+pd.to_timedelta(t_window,unit='d')
            arr_cd4 = dat_cd4['RESULT'][(dat_cd4.moh_id==id_i)&(dat_cd4.TESTDATE>=t1)&(dat_cd4.TESTDATE<=t2)].values
            if dat_cd4[dat_cd4.moh_id==id_i].shape[0]>0 and len(arr_cd4)==0:
                print ('No CD4 available: ', id_i)
            arr_ncd4.append(min(arr_cd4) if len(arr_cd4)>0 else np.nan)
            arr_ncd4_dt.append(dat_cd4['TESTDATE'][(dat_cd4.moh_id==id_i)&(dat_cd4.TESTDATE>=t1)&(dat_cd4.TESTDATE<=t2)&(dat_cd4.RESULT==min(arr_cd4))].values[0] if len(arr_cd4)>0 else np.nan)
        dat_ncd4['ncd4_bsln_dt'] = pd.Series(arr_ncd4_dt) 
        dat_ncd4['ncd4_bsln'] = pd.Series(arr_ncd4)

        #####save the record of cd4 nadir at baseline to .pkl file
        file_data = open('data/cd4_nadir_baseline.pkl','wb')  
        pickle.dump(dat_ncd4, file_data)
        file_data.close()

    if 'record_spvl_rebound' in runs:
        """determine the date of viral suppression and viral rebound longitudinally at individual level for individuals in STOP
        define viral suppression as two consecutive pvl<500 if COLDATE<1997-04-18 and pvl<400 if 1997-04-18<=COLDATE<1999-04-01"""

        p = Params()

        #####introduce different viral load threshold to define viral suppression based on test date
        p.dic_spvl_thrh = {'1997-04-18':500,'1999-04-01':400} #pvl<500 if COLDATE<1997-04-18, pvl<200 if COLDATE>=1999-04-01

        #####load data to get characteristics needed to estimate transition probability
        dat_cohort = pd.read_sas('data/stop7.sas7bdat',encoding='latin-1') 
        dat_pvl = pd.read_sas('data/virload.sas7bdat',encoding='latin-1') #viral load test results for STOP

        #####define the earliest year of FARVDT to begin with, and initialize the dataframe for viral suppression and rebound date
        dat_record = dat_cohort[['moh_id','FARVDT','end_fu_dt']].copy() #no eligibility criteria based on ART initiation date
        
        #####define viral suppression and rebound alternertively
        dat_spvl_dt = func_spvl_dt(p,dat_record[pd.isnull(dat_record.FARVDT)==False].reset_index(drop=True),dat_pvl,'FARVDT','end_fu_dt',1) 
        dat_record = pd.merge(dat_record,dat_spvl_dt,how='left',on='moh_id') 
        dat_rebound_dt = func_rebound_dt(p,dat_record[pd.isnull(dat_record['1_spvl_dt'])==False].reset_index(drop=True),dat_pvl,'1_spvl_dt','end_fu_dt',1)
        dat_record = pd.merge(dat_record,dat_rebound_dt,how='left',on='moh_id')
        i=1
        while dat_record[(pd.isnull(dat_record[str(i)+'_rebound_dt'])==False)].empty==False:
            dat_spvl_dt = func_spvl_dt(p,dat_record[pd.isnull(dat_record[str(i)+'_rebound_dt'])==False].reset_index(drop=True),dat_pvl,str(i)+'_rebound_dt','end_fu_dt',i+1)
            dat_record = pd.merge(dat_record,dat_spvl_dt,how='left',on='moh_id')
            i+=1
            if dat_record[pd.isnull(dat_record[str(i)+'_spvl_dt'])==False].empty:
                break;
            else:
                dat_rebound_dt = func_rebound_dt(p,dat_record[pd.isnull(dat_record[str(i)+'_spvl_dt'])==False].reset_index(drop=True),dat_pvl,str(i)+'_spvl_dt','end_fu_dt',i)
                dat_record = pd.merge(dat_record,dat_rebound_dt,how='left',on='moh_id')

        #####output the spvl and rebound dates as .pkl file
        file_data = open('data/spvl_rebound_dt.pkl','wb') 
        pickle.dump(dat_record, file_data)

    if 'record_rebound_interruption_failure' in runs:
        """classify each date of rebound as ART interruption ([PMID: 29048508]: >=90days gap between previous stop date and current start date) 
        or ART failure (other sceanrios)"""

        p = Params()

        #####load data of characteristics and ART dispension information
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        dat_art = pd.read_sas('data/art_records.sas7bdat',encoding='latin-1') 

        #####load data of spvl/rebound information
        file_data = open('data/spvl_rebound_dt.pkl','rb') #load derived spvl/rebound information
        dat_spvl_rebound = pickle.load(file_data)
        assert (dat_spvl_rebound[pd.isnull(dat_spvl_rebound['10_rebound_dt'])==False].shape[0]==0), 'Exist individuals with non-NaN 10_rebound_dt'
        list_spvl_rebound = [[str(i)+'_spvl_dt',str(i)+'_rebound_dt'] for i in range(1,11)]
        list_columns = [item for sublist in list_spvl_rebound for item in sublist]
        dat_study = dat_spvl_rebound[['moh_id','FARVDT','end_fu_dt']+list_columns].copy()

        #####among those with n_rebound_dt available, define ART interruption/failure based on ART dispension information
        for n_rebound in range(1,10):
            dic_rebound_art = {'I':[], 'F':[], 'U':[]} #I indicates ART interruption as n_rebound_dt in the period where the gap between stopdate and next startdate(no startdate) >=90 days, F indicates ART failure for the other scenarios, U indicates no ART dispension information available during the period spvl-rebound-spvl
            list_col_rebound = [str(i)+'_rebound_dt' for i in range(1,11)]
            for id_i in dat_study['moh_id'][pd.isnull(dat_study[str(n_rebound)+'_rebound_dt'])==False].values:
                t_rebound = pd.Timestamp(dat_study[str(n_rebound)+'_rebound_dt'][dat_study.moh_id==id_i].values[0])
                dat_art_i = dat_art[dat_art.moh_id==id_i].reset_index(drop=True)
                if dat_art_i.empty:
                    dic_rebound_art['U'].append(id_i) #no ART dispension information at all
                else:
                    assert (dat_art_i['STARTDATE'].min()<=t_rebound),'No ART dispension information for 1_rebound'
                    for index_j in dat_art_i.index:
                        if index_j<dat_art_i.index.max():
                            t_start = pd.Timestamp(dat_art_i['STARTDATE'][dat_art_i.index==index_j].values[0])
                            t_stop = pd.Timestamp(dat_art_i['STOPDATE'][dat_art_i.index==index_j].values[0])
                            t_start_new = pd.Timestamp(dat_art_i['STARTDATE'][dat_art_i.index==index_j+1].values[0])
                            if t_start<t_rebound<=t_stop or (t_stop<t_rebound<=t_start_new and (t_start_new-t_stop).days<90):
                                dic_rebound_art['F'].append(id_i)
                                break;
                            elif t_stop<t_rebound<=t_start_new and (t_start_new-t_stop).days>=90:
                                dic_rebound_art['I'].append(id_i)
                                break;
                        else:
                            #####need specify final t_start and t_stop separately as t_start_new does not exist
                            t_start = pd.Timestamp(dat_art_i['STARTDATE'][dat_art_i.index==index_j].values[0])
                            t_stop = pd.Timestamp(dat_art_i['STOPDATE'][dat_art_i.index==index_j].values[0])
                            if t_rebound>t_stop:
                                dic_rebound_art['I'].append(id_i) #define interruption when rebound_dt>final stopdate and no new startdate
                            elif t_rebound<t_start:
                                dic_rebound_art['I'].append(id_i)
                            else:
                                dic_rebound_art['F'].append(id_i)
            dat_study[str(n_rebound)+'_rebound_status'] = dat_study.apply(func_apply_key_by_id,args=(dic_rebound_art,),axis=1)

        #####save ART interruption/failure status as .pkl file
        file_data = open('data/spvl_rebound_status.pkl','wb') 
        pickle.dump(dat_study, file_data)

    if 'record_spvl_rebound_interruption/failure_dtp' in runs:
        """Determine the date of viral suppression and viral rebound longitudinally at individual level for individuals in STOP;
        Define viral suppression as two consecutive pvl<500 if COLDATE<1997-04-18 and pvl<400 if 1997-04-18<=COLDATE<1999-04-01 similar setting for viral rebound;
        Classify each date of rebound as ART interruption ([PMID: 29048508]: >=90days gap between previous stop date and current start date) or ART failure (other sceanrios)
        Note that rebound status is determined at the first rebound dt"""

        p = Params()

        #####introduce different viral load threshold to define viral suppression based on test date
        p.dic_spvl_thrh = {'1997-04-18':500,'1999-04-01':400} #pvl<500 if COLDATE<1997-04-18, pvl<200 if COLDATE>=1999-04-01

        #####load data to get characteristics 
        dat_cohort = pd.read_sas('data/dtp.sas7bdat',encoding='latin-1') 
        dat_cohort = dat_cohort.rename(columns={'PSEUDO':'moh_id'})
        cutoff_dt = pd.Timestamp('2019-12-31')
        dat_cohort['end_fu_dt'] = dat_cohort.apply(lambda x: x['LASTCTDT'] if x['LASTCTDT']==x['DTHDT'] or (cutoff_dt-x['LASTCTDT']).days>365.25/2*3 else cutoff_dt, axis=1) #apply last_contact_dt similar to STOP, to leave enough time to make sure the person is truely 'lost to follow up', LASTCTDT should be conservative as not many contact resources are linked to DTP (unlike STOP)
        dat_art = pd.read_sas('data/dtp_art_records.sas7bdat',encoding='latin-1') 
        dat_art = dat_art.rename(columns={'PSEUDO':'moh_id'})
        dat_pvl = pd.read_sas('data/dtp_virload.sas7bdat',encoding='latin-1') #viral load information for DTP
        dat_pvl = dat_pvl.rename(columns={'PSEUDO':'moh_id'})

        #####define eligible population to determine viral status 
        dat_record = dat_cohort[['moh_id','FARVDT','end_fu_dt']].copy() #no eligibility criteria based on ART initiation date

        #####define viral suppression and rebound alternertively
        dat_spvl_dt = func_spvl_dt(p,dat_record,dat_pvl,'FARVDT','end_fu_dt',1) #derive first viral suppression date, all participants from DTP with FARVDT
        dat_record = pd.merge(dat_record,dat_spvl_dt,how='left',on='moh_id')
        dat_rebound_dt = func_rebound_dt(p,dat_record[pd.isnull(dat_record['1_spvl_dt'])==False].reset_index(drop=True),dat_pvl,'1_spvl_dt','end_fu_dt',1)
        dat_record = pd.merge(dat_record,dat_rebound_dt,how='left',on='moh_id')
        i=1
        while dat_record[(pd.isnull(dat_record[str(i)+'_rebound_dt'])==False)].empty==False:
            dat_spvl_dt = func_spvl_dt(p,dat_record[pd.isnull(dat_record[str(i)+'_rebound_dt'])==False].reset_index(drop=True),dat_pvl,str(i)+'_rebound_dt','end_fu_dt',i+1)
            dat_record = pd.merge(dat_record,dat_spvl_dt,how='left',on='moh_id')
            i+=1
            if dat_record[pd.isnull(dat_record[str(i)+'_spvl_dt'])==False].empty:
                break;
            else:
                dat_rebound_dt = func_rebound_dt(p,dat_record[pd.isnull(dat_record[str(i)+'_spvl_dt'])==False].reset_index(drop=True),dat_pvl,str(i)+'_spvl_dt','end_fu_dt',i)
                dat_record = pd.merge(dat_record,dat_rebound_dt,how='left',on='moh_id')

        #####load data of spvl/rebound information
        dat_spvl_rebound = dat_record.copy()
        assert (dat_spvl_rebound[pd.isnull(dat_spvl_rebound['11_rebound_dt'])==False].shape[0]==0), 'Exist individuals with non-NaN 11_rebound_dt'
        list_spvl_rebound = [[str(i)+'_spvl_dt',str(i)+'_rebound_dt'] for i in range(1,12)]
        list_columns = [item for sublist in list_spvl_rebound for item in sublist]
        dat_study = dat_spvl_rebound[['moh_id','FARVDT','end_fu_dt']+list_columns].copy()

        #####among those with n_rebound_dt available, define ART interruption/failure based on ART dispension information
        for n_rebound in range(1,11): #no non-NaN 11_rebound_dt 
            dic_rebound_art = {'I':[], 'F':[], 'U':[]} #I indicates ART interruption as n_rebound_dt in the period where the gap between stopdate and next startdate(no startdate) >=90 days, F indicates ART failure for the other scenarios, U indicates no ART dispension information available during the period spvl-rebound-spvl
            list_col_rebound = [str(i)+'_rebound_dt' for i in range(1,11)]
            for id_i in dat_study['moh_id'][pd.isnull(dat_study[str(n_rebound)+'_rebound_dt'])==False].values:
                t_rebound = pd.Timestamp(dat_study[str(n_rebound)+'_rebound_dt'][dat_study.moh_id==id_i].values[0])
                dat_art_i = dat_art[dat_art.moh_id==id_i].sort_values(by='STARTDATE').reset_index(drop=True) #unlike STOP, need to sort dispension dates in ascending order
                if dat_art_i.empty:
                    dic_rebound_art['U'].append(id_i) #no ART dispension information at all
                else:
                    assert (dat_art_i['STARTDATE'].min()<=t_rebound),'No ART dispension information for 1_rebound'
                    for index_j in dat_art_i.index:
                        if index_j<dat_art_i.index.max():
                            t_start = pd.Timestamp(dat_art_i['STARTDATE'][dat_art_i.index==index_j].values[0])
                            t_stop = pd.Timestamp(dat_art_i['STOPDATE'][dat_art_i.index==index_j].values[0])
                            t_start_new = pd.Timestamp(dat_art_i['STARTDATE'][dat_art_i.index==index_j+1].values[0])
                            if t_start<t_rebound<=t_stop or (t_stop<t_rebound<=t_start_new and (t_start_new-t_stop).days<90):
                                dic_rebound_art['F'].append(id_i)
                                break;
                            elif t_stop<t_rebound<=t_start_new and (t_start_new-t_stop).days>=90:
                                dic_rebound_art['I'].append(id_i)
                                break;
                        else:
                            #####need specify final t_start and t_stop separately as t_start_new does not exist
                            t_start = pd.Timestamp(dat_art_i['STARTDATE'][dat_art_i.index==index_j].values[0])
                            t_stop = pd.Timestamp(dat_art_i['STOPDATE'][dat_art_i.index==index_j].values[0])
                            if t_rebound>t_stop:
                                dic_rebound_art['I'].append(id_i) #define interruption when rebound_dt>final stopdate and no new startdate
                            elif t_rebound<t_start:
                                dic_rebound_art['I'].append(id_i)
                            else:
                                dic_rebound_art['F'].append(id_i)
            dat_study[str(n_rebound)+'_rebound_status'] = dat_study.apply(func_apply_key_by_id,args=(dic_rebound_art,),axis=1)

        #####save rebound status as .pkl file
        file_data = open('data/dtp_spvl_rebound_status.pkl','wb') 
        pickle.dump(dat_study, file_data)

    if 'record_spec_art_regimen' in runs:
        """generate records of specific ART regimen status for each ART disribution record, in order to classify ART for PLWH alive by 2008-01-01 or new diagnosis
        (probability of ART initiation on specific regimen, time on each regimen, possible regimen change due to ART interruption/failure)"""

        p = Params()

        #####load data of characteristics and ART dispension information
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        dat_art = pd.read_sas('data/art_records.sas7bdat',encoding='latin-1') 
        dat_art_study = dat_art[['moh_id','STARTDATE','STOPDATE','REGIMEN','NON_BACKBONE_ARV','BACKBONE_ARV','NON_BACKBONE_CLASSES']].copy() #only keep those information necessary for our analysis

        #####create new column for Abacavir (ABA|KIV|TRZ|TRQ)
        dat_art_study['reg_aba'] = dat_art_study['REGIMEN'].str.contains('ABA|KIV|TRZ|TRQ').astype(int) #convert True/False to 1/0

        #####create new variable var_bb_aba for Abacavir (aba), other types of backbone ARV (other) and no backbone ARV dispension in the specific record (nan), nan is necessary if percent time on other backbone ARV needed to be calculated
        dat_art_study['var_bb_aba'] = dat_art_study[['BACKBONE_ARV','reg_aba']].apply(lambda x: 'aba' if x['reg_aba']==1 else ('other' if pd.isnull(x['BACKBONE_ARV'])==False else np.nan),axis=1)

        #####create new column for third class drug NNRTI (ATP|DMP|NEV|RPV|ODF|CPL|DEL|DMP|ETV|MKC|NEV|RPV|ODF|TMC)
        dat_art_study['class_nnrti'] = dat_art_study['REGIMEN'].str.contains('ATP|DMP|NEV|RPV|ODF|CPL|DEL|DMP|ETV|MKC|NEV|RPV|ODF|TMC').astype(int) 

        #####create new column for thrid class drug PI (EAP|ABT|AMV|ATA|BRV|DRV|IND|NEL|PZC|RIT|SAQ|TIP)
        dat_art_study['class_pi'] = dat_art_study['REGIMEN'].str.contains('EAP|ABT|AMV|ATA|BRV|DRV|IND|NEL|PZC|RIT|SAQ|TIP').astype(int) 

        #####create new column for third class drug INSTI (BCG|DTG|EGV|MKP|MKS|TRQ|STR|GNV)
        dat_art_study['class_insti'] = dat_art_study['REGIMEN'].str.contains('BCG|DTG|EGV|MKP|MKS|TRQ|STR|GNV').astype(int) 

        #####create new column var_class3 for NNRTI only, PI only, INSTI only, and other (CAN, super HAART included), NAN if NON_BACKBONE_ARV is empty, no value as not on ART since all the records indicating on ART
        list_nbb_all = [x.split('/') for x in dat_art_study['NON_BACKBONE_ARV'][pd.isnull(dat_art_study['NON_BACKBONE_ARV'])==False].unique()]
        list_nbb_all_flat = [item for sublist in list_nbb_all for item in sublist]
        set_nbb_all_flat = set(list_nbb_all_flat)
        set_nbb_all_flat.remove('') #remove the empty string due to string split operation
        list_nbb_nnr_pi_iin = 'ATP|DMP|NEV|RPV|ODF|CPL|DEL|DMP|ETV|MKC|NEV|RPV|ODF|TMC|EAP|ABT|AMV|ATA|BRV|DRV|IND|NEL|PZC|RIT|SAQ|TIP|BCG|DTG|EGV|MKP|MKS|TRQ|STR|GNV'.split('|') #short names for non-backbone ARVs for NNRTI, PI and INSTI
        set_nbb_other = set_nbb_all_flat - set(list_nbb_nnr_pi_iin) #to confirm that short names for the regimens other than NNRTI, PI and INSTI are COB, LOP, MVC and T20
        dat_art_study['class_other'] = dat_art_study['REGIMEN'].str.contains('MVC|T20').astype(int) #define other for non-backbone ARVs other than NNRTI, PI and INSTI (MVC and T20)
        dat_art_study['var_class3'] = dat_art_study[['class_nnrti','class_pi','class_insti','class_other']].apply(lambda x: 'other' if (x[['class_nnrti','class_pi','class_insti']].sum()>1 or x['class_other']==1) else (np.nan if (x.sum()<1) else ('nnrti' if x['class_nnrti']==1 else ('pi' if x['class_pi']==1 else 'insti'))), axis=1)

        #####create new column for TDF (ATP|CPL|TRU|TDF|STR)
        dat_art_study['reg_tdf'] = dat_art_study['REGIMEN'].str.contains('ATP|CPL|TRU|TDF|STR').astype(int) #convert True/False to 1/0

        #####create new column for TAF (DSC|ODF|TAF|GNV|BVY)
        dat_art_study['reg_taf'] = dat_art_study['REGIMEN'].str.contains('DSC|ODF|TAF|GNV|BVY').astype(int) #convert True/False to 1/0

        #####create new column var_bb_tdf for TDF, TAF , neither (including the scenario with other backbone ARV and the scenario without any backbone ARV dispension)
        dat_art_study[(dat_art_study['reg_tdf']==1)&(dat_art_study['reg_taf']==1)].shape #confirm that no records with reg_tdf=1 and reg_taf=1 together
        dat_art_study['var_bb_tdf'] = dat_art_study[['BACKBONE_ARV','reg_tdf','reg_taf']].apply(lambda x: np.nan if pd.isnull(x['BACKBONE_ARV']) else ('other' if x[['reg_tdf','reg_taf']].sum()==0 else ('tdf' if x['reg_tdf']==1 else 'taf')), axis=1) #similar to var_bb_aba, specify those without backbone ART in the record as nan

        #####save derived information of ART classification as .pkl file
        file_data = open('data/art_class.pkl','wb') 
        pickle.dump(dat_art_study, file_data)

    if 'record_art_regimen_dist' in runs:
        """estimate transition probability among those staying on ART after one time-step"""

        p = Params()

        ######load data of STOP7 and derived information
        dat_stop = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        dat_stop_hiv = pd.read_pickle('data/hiv_status.pkl') 
        list_dates = list(dat_stop_hiv)[4:][::2] #consider days after every one year of follow-up
        file_data = open('data/spvl_rebound_status.pkl','rb') #load derived spvl/rebound(status included) information
        dat_stop_spvl_rebound = pickle.load(file_data)
        list_spvl_rebound = [[str(i)+'_spvl_dt',str(i)+'_rebound_dt'] for i in range(1,11)]
        list_columns = [item for sublist in list_spvl_rebound for item in sublist] #columns with spvl/rebound date
        dat_stop_spvl_rebound[list_columns] = dat_stop_spvl_rebound[list_columns].fillna(pd.Timestamp('2050-01-01')) #replace nan by dummy dates for comparison purpose
        list_stop_status_columns = ['FARVDT']+list_columns #consider period from FARVDT to last spvl/rebound
        file_data = open('data/art_class.pkl','rb') #load derived information of ART classification 
        dat_stop_reg_art = pickle.load(file_data)
        dat_stop_reg_art = dat_stop_reg_art.sort_values(by=['moh_id','STARTDATE']).reset_index(drop=True) #make sure ART information is sorted by moh_id and STARTDATE

        #####focus on PLWH eligible for microsimulation since 2008
        dat_check = dat_stop[(dat_stop.end_fu_dt.dt.year>=2008)].reset_index(drop=True) 
        dat_check = dat_check[(pd.isnull(dat_check.sex_at_birth_dv)==False)&(pd.isnull(dat_check.DOB)==False)].reset_index(drop=True) 
        dat_check_reg = dat_stop_reg_art[dat_stop_reg_art['moh_id'].isin(dat_check['moh_id'])].reset_index(drop=True) 
        dat_check_reg['class_bb'] = dat_check_reg[['var_bb_aba','var_bb_tdf']].apply(lambda x: 'tdf' if x['var_bb_tdf']=='tdf' else (np.nan if pd.isnull(x['var_bb_tdf']) else 'other'), axis=1) #only keep tdf for backbone classes based on covariates left in the comorbidity models
        dat_check_reg['class_nbb'] = dat_check_reg[['class_nnrti','class_pi','class_insti','class_other']].apply(lambda x:'nnrti' if x['class_nnrti']==1 and sum(x[['class_pi','class_insti']])==0 else ('pi' if x['class_pi']==1 and sum(x[['class_nnrti','class_insti']])==0 else ('insti' if x['class_insti']==1 and x[['class_nnrti','class_pi']].sum()==0 else ('super' if sum([x['class_nnrti'],x['class_pi'],x['class_insti']])>=2 else (np.nan if x[['class_nnrti','class_pi','class_insti','class_other']].sum()==0 else 'other')))),axis=1) 
        dat_check_reg['class_nbb_super'] = dat_check_reg[['class_nbb','class_nnrti','class_pi','class_insti']].apply(lambda x: np.nan if x['class_nbb']!='super' else ('nnrti_pi_insti' if sum(x[['class_nnrti','class_pi','class_insti']])==3 else ('nnrti_pi' if x['class_nnrti']+x['class_pi']==2 else ('nnrti_insti' if x['class_nnrti']+x['class_insti']==2 else 'pi_insti'))),axis=1)
        dat_class_count = dat_check_reg.groupby(['class_bb','class_nbb']).size().reset_index().rename(columns={0:'count'}) #count combination of class_bb and class_nbb as a new dataframe
        dat_class_count['percent'] = dat_class_count['count']/dat_class_count['count'].sum() #show percentage of each combination
        dat_check_reg = dat_check_reg[dat_check_reg['STOPDATE']>pd.Timestamp('2008-01-01')].reset_index(drop=True) 
        dat_reg_super = pd.DataFrame([],columns=['moh_id','days']) #estimate days on ART initiated with super HAART before switing to other regimens
        for id_i in arr_id_check:
            farv_dt = dat_stop['FARVDT'][dat_stop['moh_id']==id_i].values[0]
            dat_check_reg_i = dat_check_reg[dat_check_reg['moh_id']==id_i].reset_index(drop=True)
            if (not dat_check_reg_i[(dat_check_reg_i['STARTDATE']==farv_dt)].empty) and (dat_check_reg_i['class_nbb'][dat_check_reg_i.index==0].values[0]=='super'):
                sum_days = ((dat_check_reg_i['STOPDATE']-dat_check_reg_i['STARTDATE']).dt.total_seconds()/(60*60*24)).values[0]+1 #first row,+1 to include both ends of the period so that one-day on ART if STARTDATE=STOPDATE
                j = 1
                while j<dat_check_reg_i.shape[0] and dat_check_reg_i['class_nbb'][dat_check_reg_i.index==j].values[0]=='super':
                    sum_days += ((dat_check_reg_i['STOPDATE'][dat_check_reg_i.index==j]-dat_check_reg_i['STARTDATE'][dat_check_reg_i.index==j]).dt.total_seconds()/(60*60*24)).values[0]+1
                    j+=1
                dat_reg_super = pd.concat([dat_reg_super,pd.DataFrame({'moh_id':id_i,'days':sum_days},index=[0])],ignore_index=True)
        dat_reg_super['days'] =pd.to_numeric( dat_reg_super['days'])
        
        #####estimate count and percentage of each regimen/regimen combination for ART initiation in each year
        dat_init_reg = pd.DataFrame([],columns=['date','class_bb','class_nbb','count','percent'])
        dat_init_super = pd.DataFrame([]) #consider distribution of supper HAART at ART initiation since 2008-01-01
        for i,date_i in enumerate(list_dates[1:]):
            arr_id_check = dat_check['moh_id'][(dat_check['FARVDT']>=pd.Timestamp(list_dates[i]))&(dat_check['FARVDT']<pd.Timestamp(date_i))].values
            dat_check_reg_i = dat_check_reg[dat_check_reg['moh_id'].isin(arr_id_check)].reset_index(drop=True)
            dat_check_reg_i_init = dat_check_reg_i.groupby('moh_id',as_index=False).nth(0,dropna=False).reset_index() 
            dat_check_reg_i_init = dat_check_reg_i_init[(pd.isnull(dat_check_reg_i_init['class_bb'])==False)&(pd.isnull(dat_check_reg_i_init['class_nbb'])==False)&(dat_check_reg_i_init['class_nbb']!='other')].reset_index(drop=True)
            dat_class_count_i = dat_check_reg_i_init.groupby(['class_bb','class_nbb']).size().reset_index().rename(columns={0:'count'})
            dat_class_count_i['percent'] = dat_class_count_i['count']/dat_class_count_i['count'].sum()
            dat_class_count_i['date'] = date_i
            dat_init_reg = pd.concat([dat_init_reg,dat_class_count_i],ignore_index=True)

            ####estimate distribution of different combinations of super HAART at ART initiation since 2008-01-01
            dat_super_count_i = dat_check_reg_i_init.groupby(['class_nbb_super']).size().reset_index().rename(columns={0:'count'})
            dat_super_count_i['percent'] = dat_super_count_i['count']/dat_super_count_i['count'].sum()
            dat_super_count_i['date'] = date_i
            dat_init_super = pd.concat([dat_init_super,dat_super_count_i],ignore_index=True)

        #####estimate regimen distribution among PLWH on ART (S+U_on) at the end of time steps
        dat_cohort_hiv = dat_stop_hiv.copy() 
        dat_plwh_reg = pd.DataFrame([],columns=['date','class_bb','class_nbb','count','percent'])
        dat_plwh_super = pd.DataFrame([])
        for date_k in list_dates:
            arr_id_check = dat_cohort_hiv['moh_id'][dat_cohort_hiv[date_k].isin(['S','U_on'])].values
            arr_state_check = dat_cohort_hiv[date_k][dat_cohort_hiv['moh_id'].isin(arr_id_check)].values
            arr_startdt_check = [dat_check_reg['STARTDATE'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']<pd.Timestamp(date_k))].max() if not dat_check_reg[(dat_check_reg['moh_id']==id_i)].empty else np.nan for id_i in arr_id_check]
            arr_bb_check = [dat_check_reg['class_bb'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']==startdt_j)].values[0] if not pd.isnull(startdt_j) else np.nan for id_i,startdt_j in zip(arr_id_check,arr_startdt_check)]
            arr_nbb_check = [dat_check_reg['class_nbb'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']==startdt_j)].values[0] if not pd.isnull(startdt_j) else np.nan for id_i,startdt_j in zip(arr_id_check,arr_startdt_check)]
            arr_super_check = [dat_check_reg['class_nbb_super'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']==startdt_j)].values[0] if not pd.isnull(startdt_j) else np.nan for id_i,startdt_j in zip(arr_id_check,arr_startdt_check)]
            dat_reg_check = pd.DataFrame({'moh_id':arr_id_check,'state':arr_state_check,'startdt':arr_startdt_check,'class_bb':arr_bb_check,'class_nbb':arr_nbb_check,'class_nbb_super':arr_super_check})
            dat_reg_check = dat_reg_check[(pd.isnull(dat_reg_check['class_bb'])==False)&(pd.isnull(dat_reg_check['class_nbb'])==False)].reset_index(drop=True)
            dat_count_i = dat_reg_check.groupby(['class_bb','class_nbb']).size().reset_index().rename(columns={0:'count'})
            dat_count_i['percent'] = dat_count_i['count']/dat_count_i['count'].sum()
            dat_count_i['date'] = date_k
            dat_plwh_reg = pd.concat([dat_plwh_reg,dat_count_i],ignore_index=True)
            dat_super_count_i = dat_reg_check.groupby(['class_nbb_super']).size().reset_index().rename(columns={0:'count'})
            dat_super_count_i['percent'] = dat_super_count_i['count']/dat_super_count_i['count'].sum()
            dat_super_count_i['date'] = date_k
            dat_plwh_super = pd.concat([dat_plwh_super,dat_super_count_i],ignore_index=True)

        ####estimate probability of ART regimen switch if on ART at both time steps
        list_bb = ['tdf','other'] 
        list_nbb = ['nnrti','pi','insti','super']
        list_dates_6m = list(dat_stop_hiv)[4:]
        dat_reg_trans = pd.DataFrame([],columns=['date1','date2','n stay on','n tdf','n other','n nnrti','n pi','n insti','n super'])
        for k,date_k in enumerate(list_dates_6m[1:]):
            
            ####create dataframe containing regimen information at list_dates_6m[k] and date_k
            arr_id_check = dat_cohort_hiv['moh_id'][(dat_cohort_hiv[list_dates_6m[k]].isin(['S','U_on']))&(dat_cohort_hiv[date_k].isin(['S','U_on']))].values
            arr_startdt1_check = [dat_check_reg['STARTDATE'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']<pd.Timestamp(list_dates_6m[k]))].max() if not dat_check_reg[(dat_check_reg['moh_id']==id_i)].empty else np.nan for id_i in arr_id_check]
            arr_startdt2_check = [dat_check_reg['STARTDATE'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']<pd.Timestamp(date_k))].max() if not dat_check_reg[(dat_check_reg['moh_id']==id_i)].empty else np.nan for id_i in arr_id_check]
            arr_bb1_check = [dat_check_reg['class_bb'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']==startdt_j)].values[0] if not pd.isnull(startdt_j) else np.nan for id_i,startdt_j in zip(arr_id_check,arr_startdt1_check)]
            arr_nbb1_check = [dat_check_reg['class_nbb'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']==startdt_j)].values[0] if not pd.isnull(startdt_j) else np.nan for id_i,startdt_j in zip(arr_id_check,arr_startdt1_check)]
            arr_super1_check = [dat_check_reg['class_nbb_super'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']==startdt_j)].values[0] if not pd.isnull(startdt_j) else np.nan for id_i,startdt_j in zip(arr_id_check,arr_startdt1_check)]
            arr_bb2_check = [dat_check_reg['class_bb'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']==startdt_j)].values[0] if not pd.isnull(startdt_j) else np.nan for id_i,startdt_j in zip(arr_id_check,arr_startdt2_check)]
            arr_nbb2_check = [dat_check_reg['class_nbb'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']==startdt_j)].values[0] if not pd.isnull(startdt_j) else np.nan for id_i,startdt_j in zip(arr_id_check,arr_startdt2_check)]
            arr_super2_check = [dat_check_reg['class_nbb_super'][(dat_check_reg['moh_id']==id_i)&(dat_check_reg['STARTDATE']==startdt_j)].values[0] if not pd.isnull(startdt_j) else np.nan for id_i,startdt_j in zip(arr_id_check,arr_startdt2_check)]
            dat_reg_check = pd.DataFrame({
                'moh_id':arr_id_check,'startdt1':arr_startdt1_check,'startdt2':arr_startdt2_check,
                'class_bb1':arr_bb1_check,'class_bb2':arr_bb2_check,'class_nbb1':arr_nbb1_check,'class_nbb2':arr_nbb2_check,
                'class_nbb_super1':arr_super1_check,'class_nbb_super2':arr_super2_check})
            dat_reg_check = dat_reg_check[(pd.isnull(dat_reg_check['class_bb1'])==False)&(pd.isnull(dat_reg_check['class_nbb1'])==False)&(pd.isnull(dat_reg_check['class_bb2'])==False)&(pd.isnull(dat_reg_check['class_nbb2'])==False)].reset_index(drop=True)

            ####estimate probability of regimen switch for bb and nbb classes
            n_tdf1 = dat_reg_check[dat_reg_check['class_bb1']=='tdf'].shape[0]
            n_tdf2other = dat_reg_check[(dat_reg_check['class_bb1']=='tdf')&(dat_reg_check['class_bb2']=='other')].shape[0]
            n_other1 = dat_reg_check[dat_reg_check['class_bb1']=='other'].shape[0]
            n_other2tdf = dat_reg_check[(dat_reg_check['class_bb1']=='other')&(dat_reg_check['class_bb2']=='tdf')].shape[0]
            dic_reg_trans = {
                'date1':list_dates_6m[k], 'date2':date_k, 'n stay on':dat_reg_check.shape[0], 
                'n tdf':n_tdf1, 'n tdf2other':n_tdf2other, 'p tdf2other':n_tdf2other/n_tdf1,
                'n other':n_tdf1, 'n other2tdf':n_other2tdf, 'p other2tdf':n_other2tdf/n_other1,
            }
            for nbb_i in list_nbb:
                dic_reg_trans['n '+nbb_i] = dat_reg_check[dat_reg_check['class_nbb1']==nbb_i].shape[0]
                list_nbb_rest = list_nbb[:]
                list_nbb_rest.remove(nbb_i)
                for nbb_j in list_nbb_rest:
                    dic_reg_trans['n '+nbb_i+'2'+nbb_j] = dat_reg_check[(dat_reg_check['class_nbb1']==nbb_i)&(dat_reg_check['class_nbb2']==nbb_j)].shape[0]
                    dic_reg_trans['p '+nbb_i+'2'+nbb_j] = dic_reg_trans['n '+nbb_i+'2'+nbb_j]/dic_reg_trans['n '+nbb_i]
            dat_reg_trans = pd.concat([dat_reg_trans,pd.DataFrame(dic_reg_trans,index=[0])],ignore_index=True)

        #####output estimated transition probability among those staying on ART after one time-step as .xlsx file
        writer = pd.ExcelWriter('data/prob_reg_swtich.xlsx',engine='openpyxl') 
        dat_init_reg.to_excel(writer,'reg_dist_init',index=False)
        dat_init_super.to_excel(writer,'super_dist_init',index=False) 
        dat_plwh_reg.to_excel(writer,'reg_dist_plwh',index=False)
        dat_plwh_super.to_excel(writer,'super_dist_plwh',index=False)
        dat_reg_trans.to_excel(writer,'prob_reg_switch',index=False)
        writer.save()

    if 'record_art_regimen_by_timestep' in runs:
        """specify status of abacavir (aba), tenofovir (tdf) and third-class regimen (nnrti, pi, insti) by the end of each time step, depending on ART dispensation history
        note that status AF can be extended instead of moving to AI if ART was dispensed after AF but not achieved suppression
        specify the regimen as NaN if no ART dispensation history found by the end of the time step, or without ART initiation, or dead"""

        p = Params()

        #####load information of ART initiation, HIV status, and ART classification information 
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        file_data = open('data/hiv_status.pkl','rb') 
        dat_hiv = pickle.load(file_data)
        file_data = open('data/art_class.pkl','rb') 
        dat_art_reg = pickle.load(file_data)
        dat_art_reg = dat_art_reg.rename(columns={'class_nnrti':'reg_nnrti','class_pi':'reg_pi','class_insti':'reg_insti'}) #rename columns for third class regimen classification to be consistent with aba and tdf

        #####create dictionary to save dataframes with the status of aba, tdf, third class regimens by the end of each time step
        dic_dat_reg = {
            'aba':dat_hiv[['moh_id','earliest_HIV','FARVDT']].copy(),
            'tdf':dat_hiv[['moh_id','earliest_HIV','FARVDT']].copy(),
            'nnrti':dat_hiv[['moh_id','earliest_HIV','FARVDT']].copy(),
            'pi':dat_hiv[['moh_id','earliest_HIV','FARVDT']].copy(),
            'insti':dat_hiv[['moh_id','earliest_HIV','FARVDT']].copy()} 

        #####at each time step per 6 months, determine status of specified regimen by hiv status and reg_aba/tdf/nnrti/pi/insti closest to and ahead of the time step
        p.t0 = pd.Timestamp('2008-01-01') 
        p.t_end = pd.Timestamp('2017-01-01') 
        p.dt = 0.5
        p.n_step = int((p.t_end.year-p.t0.year)/p.dt)
        for i in range(0,p.n_step+1):
            t_stop = p.t0+pd.to_timedelta(p.def_year/2*i,unit='d') 
            for key_j in dic_dat_reg.keys():
                data_j = dic_dat_reg[key_j]
                data_j[str(t_stop)[:10]] = pd.Series(np.zeros(data_j.shape[0])) #initialize regimen information as zeros for t_stop
            list_id = dat_hiv['moh_id'].values #include all moh_id for each time step to specify values as NaN if no ART dispensation history was found for the period
            for id_i in list_id:
                dat_art_i = dat_art_reg[dat_art_reg['moh_id']==id_i].reset_index(drop=True) #possilbe no art historiy can be found
                if not dat_art_i.empty:
                    if dat_art_i[(dat_art_i['STARTDATE']<=t_stop)].shape[0]==0: #All ART start dates after t_stop, indicating no ART at t_stop
                        for key_j in dic_dat_reg.keys():
                            data_j = dic_dat_reg[key_j]
                            data_j.loc[data_j['moh_id']==id_i,str(t_stop)[:10]] = np.nan 
                    else: #note that regimen information possibly available if just becoming dead by t_stop, but the status became NaN moving forward 
                        index_dt = dat_art_i.index[(dat_art_i['STARTDATE']<=t_stop)].max() #find the record closest to and ahead of t_stop
                        art_start_dt = pd.Timestamp(dat_art_i['STARTDATE'][dat_art_i.index==index_dt].values[0])
                        art_stop_dt = pd.Timestamp(dat_art_i['STOPDATE'][dat_art_i.index==index_dt].values[0])
                        if art_start_dt<=t_stop<=art_stop_dt or (t_stop-art_stop_dt).total_seconds()/(60*60*24)<90:
                            for key_j in dic_dat_reg.keys():
                                data_j = dic_dat_reg[key_j]
                                data_j.loc[data_j['moh_id']==id_i,str(t_stop)[:10]] = dat_art_i['reg_'+key_j][dat_art_i.index==index_dt].values[0]
                else:
                    for key_j in dic_dat_reg.keys():
                        data_j = dic_dat_reg[key_j]
                        data_j.loc[data_j['moh_id']==id_i,str(t_stop)[:10]] = np.nan

        #####save derived information of ART classification as .pkl file
        file_data = open('data/art_reg_status.pkl','wb') 
        pickle.dump(dic_dat_reg, file_data)

    if 'record_status_pvl_reg_1yr' in runs:
        """derive pvl and regimen status over each one year step of microsimulation, used for probability of comorbidity incidence;
        derive proportion of time on specific regimen over one year time step (one year or ended by the end of follow-up), count third classes in super-HAART separately"""

        p = Params()

        #####load data to get characteristics needed to estimate transition probability
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        file_data = open('data/hiv_status_pre.pkl','rb') #HIV status without adjustment
        dat_hiv = pickle.load(file_data)
        file_data = open('data/art_reg_status.pkl','rb') #load regimen history for STOP participants
        dic_reg = pickle.load(file_data) #format as dictionary with keys=aba,tdf,nnrti,pi and insti
        file_data = open('data/art_class.pkl','rb') #load information of regimen classification
        dat_art_reg = pickle.load(file_data)

        #####introduce dictionary to specify proportion of regimen time over total time on treatment during the 1-year period
        dic_reg_val = {
            '000':0., '001':0.25, '010':0.5, '011':0.75, '100':0.25, '101':0.5, '110':0.75, 
            '111':1., '400':0., '401':0.25/0.75, '410':0.5/0.75, '411':0.75/0.75,
            '440':0., '441':0.25/0.25, '444':0.
        }

        #####create necessary information at the beginning for future step of microsimulation
        p.t0 = pd.Timestamp('2008-01-01') 
        p.t_end = pd.Timestamp('2017-01-01') 
        p.dt = 1. 
        p.n_step = int((p.t_end.year-p.t0.year)/p.dt)

        #####for each 1-year interval, evaluate status of viral suppression and ART history, saved as dataframe at the beginning of the interval
        dat_status_pvl = dat_hiv[['moh_id','earliest_HIV','FARVDT','end_fu_dt']].copy()
        dic_status_reg = {k:dic_reg[k][['moh_id','earliest_HIV','FARVDT']].copy() for k in dic_reg.keys()}
        dat_hiv = dat_hiv.fillna('_') #replace np.nan by string for consistency
        dic_reg = {k:dic_reg[k].fillna(4) for k in dic_reg.keys()} #replace before FARVDT by 4
        list_dates = list(dat_hiv)[4:]
        for i in range(1,p.n_step+1):
        
            ####define the beginning and the end of the time step
            t_start = p.t0+pd.to_timedelta(p.def_year*p.dt*(i-1),unit='d')
            t_stop = t_start+pd.to_timedelta(p.def_year*p.dt,unit='d')
            list_cols = [t for t in list_dates if t>=str(t_start)[:10] and t<=str(t_stop)[:10]] 

            ####only three "S" for the three step will be considered as viral suppression in the interval
            dat_hiv['comb'] = dat_hiv[list_cols].apply(lambda x: "".join(list(x)),axis=1) 
            dat_hiv['status'] = dat_hiv['comb'].apply(lambda x: 0 if x=='SSS' else 1) #0 if viral suppression
            dat_status_pvl = pd.merge(dat_status_pvl,dat_hiv[['moh_id','status']].copy(),how='left',on='moh_id').rename(columns={'status':str(t_start)[:10]})

            ####apply proportion of ART regimen over total time on ART during 1-year period
            for k in dic_reg.keys():
                dic_reg[k]['comb'] = dic_reg[k][list_cols].apply(lambda x: "".join([str(int(ele)) for ele in x]),axis=1)
                dic_reg[k]['status'] = dic_reg[k]['comb'].apply(lambda x: dic_reg_val[x]) 
                dic_status_reg[k] = pd.merge(dic_status_reg[k],dic_reg[k][['moh_id','status']].copy(),how='left',on='moh_id').rename(columns={'status':str(t_start)[:10]})
                dic_status_reg[k].loc[dic_status_reg[k]['FARVDT']==4,'FARVDT'] = np.nan #replace missing FARVDT back to NAN

        #####save derived dataframe as .pkl file
        file_data = open('data/pvl_status.pkl','wb') 
        pickle.dump(dat_status_pvl, file_data)
        file_data.close()

        #####for each 1-year interval, estimate the proportion of time on specific regimen over 1 year/until end of follow-up
        dic_art_reg_name = {'aba':'reg_aba','tdf':'reg_tdf','nnrti':'class_nnrti','pi':'class_pi','insti':'class_insti'} #specify which column to use for specific regimen in dat_art_reg
        dic_status_reg = {k:dat_hiv[['moh_id','earliest_HIV','FARVDT','end_fu_dt']].copy() for k in dic_reg.keys()}
        for i in range(1,p.n_step+1):
            
            ####define the beginning and the end of the time step
            t_start = p.t0+pd.to_timedelta(p.def_year*p.dt*(i-1),unit='d')
            t_stop = t_start+pd.to_timedelta(p.def_year*p.dt,unit='d') #consistent with the list_dates from dat_hiv

            ####derive proportion of time on specific regimen during follow-up peirod
            dic_time_reg = {k:[] for k in dic_reg.keys()}
            for id_i in dat_hiv['moh_id']:
                dat_art_i = dat_art_reg[dat_art_reg['moh_id']==id_i].reset_index(drop=True)
                if not dat_art_i.empty:
                    if dat_art_i[dat_art_i['STARTDATE']<=t_stop].shape[0]==0 or dat_art_i[dat_art_i['STOPDATE']>=t_start].shape[0]==0:
                        for k,col_k in dic_art_reg_name.items():
                            dic_time_reg[k].append(0.) #0% on specific regimen if no ART history captured between t_start and t_stop
                    else: #exist STARTDATE before t_stop and STOPDATE after t_start
                        t1 = dat_art_i['STOPDATE'][dat_art_i['STOPDATE']>=t_start].min() 
                        index1 = dat_art_i[dat_art_i['STOPDATE']==t1].index.values[0]
                        t2 = dat_art_i['STARTDATE'][dat_art_i['STARTDATE']<=t_stop].max()
                        index2 = dat_art_i[dat_art_i['STARTDATE']==t2].index.values[0]
                        if dat_art_i['STARTDATE'][dat_art_i.index==index1].values[0]<t_start: #Time on ART was truncated by t_start/stop
                            dat_art_i.loc[dat_art_i.index==index1,'STARTDATE'] = t_start
                        if dat_art_i['STOPDATE'][dat_art_i.index==index2].values[0]>t_stop:
                            dat_art_i.loc[dat_art_i.index==index2,'STOPDATE'] = t_stop
                        time_art = (dat_art_i['STOPDATE'][(dat_art_i.index>=index1)&(dat_art_i.index<=index2)]-dat_art_i['STARTDATE'][(dat_art_i.index>=index1)&(dat_art_i.index<=index2)]).dt.total_seconds()/(60*60*24)+1 #include the end date of each period
                        time_all = (min(t_stop,dat_hiv['end_fu_dt'][dat_hiv['moh_id']==id_i].values[0])-t_start).total_seconds()/(60*60*24)+1 #time from t_start to min(t_stop,end_fu_dt)
                        for k,col_k in dic_art_reg_name.items():
                            time_reg = time_art*dat_art_i[col_k][(dat_art_i.index>=index1)&(dat_art_i.index<=index2)] #focus on time on specific regimen
                            dic_time_reg[k].append(time_reg.sum()/time_all)                           
                else:
                    for k in dic_art_reg_name.keys():
                        dic_time_reg[k].append(0.)

            ####for each regimen, attach the list of time for t_start
            for k in dic_art_reg_name.keys():
                assert (len(dic_time_reg[k])==dic_status_reg[k].shape[0]), 'Inconsistent number of patients between derived proportion time and historical records'
                dic_status_reg[k][str(t_start)[:10]] = pd.Series(dic_time_reg[k])
        
        file_data = open('data/prop_time_art_reg.pkl','wb') 
        pickle.dump(dic_status_reg, file_data)
        file_data.close()

    if 'cali_new_diag' in runs:
        """calibration process for HIV incidence and new diagnosis simulation, using incidence-prevalence ratio (IPR) and diagnosis probability over 6 months instead of diagnosis rate"""

        #####load updated PHAC estimates for HIV incidence and prevalence and historical pvl from R0 project
        excel_record = pd.ExcelFile('data/phac_inc_prev.xlsx') #BC incidence, prevalence and IPR information
        excel_sheet = excel_record.sheet_names
        dat_inc = excel_record.parse(excel_sheet[0]) #annual HIV new infections with uncertainty range
        dat_prev = excel_record.parse(excel_sheet[1]) #HIV prevalence in each year with uncertianty range
        dat_pvl = excel_record.parse(excel_sheet[2]) #number of PLWH with suppressed viral load in each year
        dat_all = dat_inc.merge(dat_prev,how='left',on='year').merge(dat_pvl[['year','200']].copy().rename(columns={'200':'pvl_R'}),how='left',on='year')

        #####load estimated incidence-unsuppressed-ratio 
        excel_record = pd.ExcelFile('data/table_ipr.xlsx') #incidence-prevlanece/unsuppressed ratio until 2022 with uncertainty range
        excel_sheet = excel_record.sheet_names
        dat_iur = excel_record.parse(excel_sheet[-1]) #only need incidence-unsuppressed(200) ratio
        dat_all = dat_all.merge(dat_iur[['year','median']].copy(),how='left',on='year').rename(columns={'median':'iur200'})
        dat_iur_fit = dat_iur[dat_iur['year']>=2007].copy() #year 2007 refers to t0=0
        init_params = [0.055,0.02,4.,2014.8]
        match_N = minimize(func_target_sigmoid_iur,init_params,args=(dat_iur_fit),method='nelder-mead',options={'xatol': 1e-5, 'disp': True})
        opt_params = match_N.x
        opt_params = np.array([5.38300232e-02, 3.19324163e-02, 5.07343216e+00, 2.01477145e+03]) #save records after optimizing y0,y1,tm,t0

        #####estimate uncertainty in iur based on the credible intervales of IUR with fixed estimated tm and t0
        p = Params()
        init_params = p.sig_iur_params[:2]
        match_N = minimize(func_target_sigmoid_iur_ci,init_params,args=(dat_iur_fit,p,'ub'),method='nelder-mead',options={'xatol': 1e-5, 'disp': True}) #change argument between ub and lb
        opt_params_lb = np.array([0.05263581, 0.02736219])
        opt_params_ub = np.array([0.05502749, 0.03724505])
        fig_iur,ax_iur = plt.subplots()
        ax_iur.plot(dat_iur['year'],dat_iur['median'],color='steelblue',lw=2.,label='Historical estimation')
        ax_iur.fill_between(dat_iur['year'].astype(float),dat_iur['lb'].astype(float),dat_iur['ub'].astype(float),facecolor='skyblue',alpha=0.3)
        ax_iur.set_xlim([2007,2022])
        plt.show()
        arr_year_sim = np.arange(2007,2035) 
        arr_iur_sim = sigmoid(p.sig_iur_params[0],p.sig_iur_params[1],p.sig_iur_params[2],arr_year_sim-p.sig_iur_params[3])
        arr_iur_sim_lb = sigmoid(opt_params_lb[0],opt_params_lb[1],p.sig_iur_params[2],arr_year_sim-p.sig_iur_params[3])
        arr_iur_sim_ub = sigmoid(opt_params_ub[0],opt_params_ub[1],p.sig_iur_params[2],arr_year_sim-p.sig_iur_params[3])
        ax_iur.plot(arr_year_sim,arr_iur_sim,'r--',label='Sigmoid estimation')
        ax_iur.fill_between(arr_year_sim,arr_iur_sim_lb,arr_iur_sim_ub,facecolor='salmon',alpha=0.3)
        ax_iur.set_xlim([2007,2034])
        ax_iur.set_xticks(np.arange(2007,2035,3))
        ax_iur.set_ylim([0.02,0.06])
        ax_iur.set_yticks(np.arange(0.02,0.07,0.01))
        ax_iur.set_xlabel('Year',fontsize=14)
        ax_iur.set_ylabel('Incidence-to-unsuppressed ratio',fontsize=14)
        ax_iur.legend(loc='upper right',fontsize=12,fancybox=True)
        plt.subplots_adjust(top=0.95, bottom=0.12, left=0.14,right=0.95,hspace=0.3,wspace=0.25)
        fig_iur.savefig('my_path/results/fig_cali_iur_uncertain_today.png',dpi=300) #Figure S2

        #####load HIV status for the cohort every 6 months for 2008-2016 and estimate PLWH with pvl by the end of each year
        dat_hiv = pd.read_pickle('data/hiv_status.pkl') #load HIV status to build baseline status
        list_dates = list(dat_hiv)[4:]
        list_annual_dates = list_dates[::2]
        dat_pvl = pd.DataFrame({'year':np.arange(2007,2017)})
        dat_pvl['spvl'] = pd.Series([dat_hiv[dat_hiv[date_i]=='S'].shape[0] for date_i in list_annual_dates])
        dat_pvl['unsupp_diag'] = pd.Series([dat_hiv[dat_hiv[date_i].isin(['A','U_on','U_off'])].shape[0] for date_i in list_annual_dates])
        dat_all = dat_all.merge(dat_pvl,how='left',on='year')

        #####simulate HIV incident cases based on pvl, prev and iur200 over time
        arr_year = dat_all['year'].values
        list_inc_sim = [(dat_all['prev'][dat_all['year']==arr_year[i]].values[0]-dat_all['spvl'][dat_all['year']==arr_year[i]].values[0])*dat_all['iur200'][dat_all['year']==year_i].values[0] for i,year_i in enumerate(arr_year[1:])]
        list_inc_sim.insert(0,np.nan)
        dat_all['inc_sim'] = pd.Series(list_inc_sim)

        #####load characterisicts for the cohort to estimate annual HIV diagnosis
        arr_year = dat_all['year'].values
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        dat_diag = dat_cohort[dat_cohort['moh_id'].isin(dat_hiv['moh_id'])].reset_index(drop=True)
        dat_all['new_diag'] = pd.Series([dat_diag[dat_diag['earliest_HIV'].dt.year==year_i].shape[0] for year_i in dat_all['year']])
        dat_all['new_diag_cdc'] = pd.Series([dat_diag[(dat_diag['earliest_HIV'].dt.year==year_i)&(dat_diag['diag_source']=='CDC_POSTEST')].shape[0] for year_i in dat_all['year']])
        dat_all['new_migr'] = dat_all['new_diag']-dat_all['new_diag_cdc']
        list_undiag = [(dat_all['prev'][dat_all['year']==arr_year[i]].values[0]-dat_all['spvl'][dat_all['year']==arr_year[i]].values[0]-dat_all['unsupp_diag'][dat_all['year']==arr_year[i]].values[0]) for i,year_i in enumerate(arr_year[1:])]
        list_undiag.insert(0,np.nan)
        dat_all['undiag'] = pd.Series(list_undiag)
        dat_all['rdiag'] = dat_all['new_diag']/dat_all['undiag'] 
        dat_all['rdiag_cdc'] = dat_all['new_diag_cdc']/dat_all['undiag']

        #####use a sigmoid function to simulate time-varying diagnosis rate
        dat_diag_fit = dat_all[['year','rdiag_cdc']][dat_all['year'].isin(np.arange(2008,2017))].reset_index(drop=True).rename(columns={'rdiag_cdc':'rdiag'})
        init_params = [0.09253435, 0.28276672, 8.92294154, 2007.] 
        match_N = minimize(func_target_sigmoid_rdiag,init_params,args=(dat_diag_fit),method='nelder-mead',options={'xatol': 1e-5, 'disp': True})
        opt_params = match_N.x
        opt_params = np.array([1.05539568e-01, 2.00892319e-01, 2.88342681e+00, 2.01116228e+03]) 
        arr_year_sim = np.arange(2007,2035)
        arr_rdiag_sim = sigmoid(opt_params[0],opt_params[1],opt_params[2],arr_year_sim-opt_params[3])

        #####use a sigmoid function to simulate annual number of new cases from assumed-to-be migration
        dat_migr_fit = dat_all[['year','new_migr']][dat_all['year'].isin(np.arange(2008,2017))].reset_index(drop=True)
        init_params = [280,200,8.,2013.]
        match_N = minimize(func_target_sigmoid_general,init_params,args=(dat_migr_fit,'new_migr'),method='nelder-mead',options={'xatol': 1e-5, 'disp': True})
        opt_params = match_N.x
        opt_params = np.array([2.84370282e+02, 2.46489938e+02, 1.43629892e+00, 2.01282960e+03])
        arr_year_sim = np.arange(2007,2035)
        arr_migr_sim = sigmoid(opt_params[0],opt_params[1],opt_params[2],arr_year_sim-opt_params[3])

        #####Figure S1: output plots for diagnosis rate and annual migrations
        fig_nd,axs_nd = plt.subplots(1,2,figsize=(12,5))
        axs_nd[0].plot(dat_all['year'][(dat_all['year']>=2008)&(dat_all['year']<=2016)],dat_all['rdiag_cdc'][(dat_all['year']>=2008)&(dat_all['year']<=2016)],'-',color='steelblue',lw=2,label='Historical estimation')
        axs_nd[0].plot(arr_year_sim,arr_rdiag_sim,'--',color='salmon',lw=2,label='Sigmoid estimation')
        axs_nd[0].text(2004,0.25+(0.25-0.05)*0.05,'A)',fontsize=14)
        axs_nd[0].set_xticks(np.arange(2007,2035,3))
        axs_nd[0].set_xlim([2007,2034])
        axs_nd[0].set_yticks(np.arange(0.05,0.3,0.05))
        axs_nd[0].set_ylim([0.05,0.25])
        axs_nd[0].set_xlabel('Year',fontsize=14)
        axs_nd[0].set_ylabel('Ratio between new diagnosis\nand undiagnosed PLWH',fontsize=14)
        axs_nd[0].legend(loc='best',fontsize=12,fancybox=True)
        axs_nd[1].plot(dat_all['year'][(dat_all['year']>=2008)&(dat_all['year']<=2016)],dat_all['new_migr'][(dat_all['year']>=2008)&(dat_all['year']<=2016)],'-',color='steelblue',lw=2,label='Historical estimation')
        axs_nd[1].plot(arr_year_sim,arr_migr_sim,'-',color='salmon',lw=2,label='Sigmoid estimation')
        axs_nd[1].text(2004,320+100*0.05,'B)',fontsize=14)
        axs_nd[1].set_xticks(np.arange(2007,2035,3))
        axs_nd[1].set_xlim([2007,2034])
        axs_nd[1].set_ylim([220,320])
        axs_nd[1].set_xlabel('Year',fontsize=14)
        axs_nd[1].set_ylabel('Annual number of new diagnosedPLWH\nmoving to BC',fontsize=14)
        axs_nd[1].legend(loc='best',fontsize=12,fancybox=True)
        plt.subplots_adjust(top=0.9, bottom=0.12, left=0.1,right=0.96,hspace=0.3,wspace=0.3)
        plt.show()
        fig_nd.savefig('results/fig_cali_rdiag_migr_today.png',dpi=300) 

        #####estimate suppressed and unsuppressed diagnosed PLWH every 6 months for microsimulation
        dat_hiv = pd.read_pickle('data/hiv_status.pkl') #load HIV status to build baseline status
        list_dates = list(dat_hiv)[4:]
        dat_hist_diag = pd.DataFrame({'year':np.arange(2007,2016.5,0.5)})
        dat_hist_diag['spvl'] = pd.Series([dat_hiv[dat_hiv[date_i]=='S'].shape[0] for date_i in list_dates])
        dat_hist_diag['unsupp_diag'] = pd.Series([dat_hiv[dat_hiv[date_i].isin(['A','U_on','U_off'])].shape[0] for date_i in list_dates])

        #####use microsimulation to estimate undiagnosed PLWH and prevalence over time based on IUR and historical records of diagnosed PLWH
        p = Params()
        p.year0 = 2007. #t=0 at the end of 2007
        p.init_prev = dat_all['prev'][dat_all['year']==p.year0].values[0] #introduce initial hiv prevalence for microsimulation
        p.t0 = pd.Timestamp('2008-01-01') 
        p.t_end = pd.Timestamp('2017-01-01') 
        p.dt = 0.5 #use one year time step
        p.n_step = int((p.t_end.year-p.t0.year)/p.dt)
        p.n_pool = 25 
        p.num_sim = 1000 
        p.arr_seed = np.arange(p.num_sim) #use ascending order seed sequence from 0 to num
        pool = Pool(processes=p.n_pool,maxtasksperchild=1) 
        results = [pool.apply_async(mp_microsim_inc_undiag_test,args=(p,dat_hist_diag,p.arr_seed[i])) for i in range(p.num_sim)]
        list_results = [x.get() for x in results] 

        #####compare cumulative new HIV cases, prevalence and new diagnosis with historical records
        dat_sim_inc = func_micro_1outcome_ci(p,list_results,'new_inc','inc')
        dat_sim_prev = func_micro_1outcome_ci(p,list_results,'plwh','prev')
        dat_sim_undiag = func_micro_1outcome_ci(p,list_results,'undiag','prev')
        dat_sim_diag = func_micro_1outcome_ci(p,list_results,'new_diag','inc')
        dat_sim_migr = func_micro_1outcome_ci(p,list_results,'new_migr','inc')
        dat_hist = dat_all[dat_all['year'].isin(dat_sim_inc['year'].values)].reset_index(drop=True)
        
        #####Figure S3: output figure to compare simulated new diagnosed PLWH in BC with and without adjustment (p.coeff_rdiag=1 or opt_params)
        #####axs[1] should be updated after optimizing p.coeff_rdiag below
        fig,axs = plt.subplots(1,2,figsize=(12,5))
        axs[0].plot(dat_hist['year'][(dat_hist['year']>=2008)&(dat_hist['year']<=2016)],dat_hist['new_diag'][(dat_hist['year']>=2008)&(dat_hist['year']<=2016)],'o',color=dic_color_palette['blue'][0],label='STOP HIV/AIDS')
        axs[0].plot(dat_sim_migr['year'],dat_sim_migr['new_migr med'].values+dat_sim_diag['new_diag med'],color='steelblue',label='Simulations')
        axs[0].fill_between(dat_hist['year'].astype(float),dat_sim_diag['new_diag lb'].astype(float).values+dat_sim_migr['new_migr med'].astype(float).values,
                            dat_sim_diag['new_diag ub'].astype(float).values+dat_sim_migr['new_migr med'].astype(float).values,facecolor='skyblue',alpha=0.3)
        axs[0].text(2007,540+120*0.05,'A)',fontsize=14)
        axs[0].set_xlim([2008,2016])
        axs[0].set_ylim([420,540])
        axs[0].set_xlabel('Year',fontsize=14)
        axs[0].set_ylabel('Annual number of new diagnosed\nPLWH in BC',fontsize=14)
        axs[0].legend(loc='best',fontsize=12)
        axs[1].plot(dat_hist['year'][(dat_hist['year']>=2008)&(dat_hist['year']<=2016)],dat_hist['new_diag'][(dat_hist['year']>=2008)&(dat_hist['year']<=2016)],'o',color=dic_color_palette['blue'][0],label='STOP HIV/AIDS')
        axs[1].plot(dat_sim_migr['year'],dat_sim_migr['new_migr med'].values+dat_sim_diag['new_diag med'],color='steelblue',label='Simulations after adjustment')
        axs[1].fill_between(dat_hist['year'].astype(float),dat_sim_diag['new_diag lb'].astype(float).values+dat_sim_migr['new_migr med'].astype(float).values,
                            dat_sim_diag['new_diag ub'].astype(float).values+dat_sim_migr['new_migr med'].astype(float).values,facecolor='skyblue',alpha=0.3)
        axs[1].text(2007,540+120*0.05,'B)',fontsize=14)
        axs[1].set_xlim([2008,2016])
        axs[1].set_ylim([420,540])
        axs[1].set_xlabel('Year',fontsize=14)
        axs[1].legend(loc='best',fontsize=12)
        plt.subplots_adjust(top=0.92, bottom=0.12, left=0.09,right=0.98,hspace=0.3,wspace=0.25)
        plt.show()
        fig.savefig('my_path/results/fig_cali_newstop_today.png',dpi=300) 

        #####introduce coefficient to adjust rdiag, fitting to cumulative number of new STOP participants
        p = Params()
        p.year0 = 2007. 
        p.init_prev = dat_all['prev'][dat_all['year']==p.year0].values[0] #introduce initial hiv prevalence for microsimulation
        p.t0 = pd.Timestamp('2008-01-01') 
        p.t_end = pd.Timestamp('2017-01-01') 
        p.dt = 0.5 
        p.n_step = int((p.t_end.year-p.t0.year)/p.dt)
        p.n_pool = 10
        p.num_sim = 1000 
        p.arr_seed = np.arange(p.num_sim) #use ascending order seed sequence from 0 to num
        init_params = [p.coeff_rdiag]
        match_N = minimize(func_cali_target_micro_rdiag,init_params,args=(dat_hist_diag,dat_all,p),method='nelder-mead',options={'xatol': 1e-4}) 
        opt_params = np.array([1.02912445]) #keep the record of the optimized parameter

    if 'record_hiv_status_every6m' in runs:
        """determine HIV status for each eligible individual (age>=18, alive by 2008-01-01, diagnosed before 2017) in STOP cohort every 6 months from 2008-01-01 to 2016-12-31"""

        p = Params()

        #####load historical data and derived rebound/failure information
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        file_data = open('data/spvl_rebound_status.pkl','rb') #load derived spvl/rebound(status included) information
        dat_spvl_rebound = pickle.load(file_data)
        list_spvl_rebound = [[str(i)+'_spvl_dt',str(i)+'_rebound_dt'] for i in range(1,11)]
        list_columns = [item for sublist in list_spvl_rebound for item in sublist] #columns with spvl/rebound date
        dat_spvl_rebound[list_columns] = dat_spvl_rebound[list_columns].fillna(pd.Timestamp('2050-01-01')) #replace nan by dummy dates for comparison purpose
        list_status_columns = ['FARVDT']+list_columns #consider period from FARVDT to last spvl/rebound

        #####following PLWH since 2008, determine ART and spvl status every 6 months
        p.t0 = pd.Timestamp('2008-01-01') 
        p.t_end = pd.Timestamp('2017-01-01') 
        p.dt = 0.5
        p.n_step = int((p.t_end.year-p.t0.year)/p.dt)
        dat_study = dat_cohort[dat_cohort.end_fu_dt.dt.year>=2008].reset_index(drop=True)
        dat_study = dat_study[(pd.isnull(dat_study.sex_at_birth_dv)==False)&(pd.isnull(dat_study.DOB)==False)].reset_index(drop=True) 
        dat_study['age_diag'] = pd.Series((dat_study['earliest_HIV']-dat_study['DOB']).dt.days/p.def_year) 
        dat_study = dat_study[dat_study.age_diag>=18].reset_index(drop=True)
        dat_hist = dat_study[['moh_id','earliest_HIV','FARVDT','end_fu_dt']][(dat_study.end_fu_dt>=p.t0)&(dat_study.earliest_HIV<p.t0)].reset_index(drop=True)
        dat_hist[str(p.t0)[:10]] = dat_hist.FARVDT.apply(lambda x: 'T' if x<p.t0 else 'A')
        dat_hist[str(p.t0)[:10]] = dat_hist.apply(func_apply_spvl_status,args=(str(p.t0)[:10],dat_spvl_rebound,list_status_columns),axis=1)
        for i in range(1,p.n_step+1):
            t_start = p.t0+pd.to_timedelta(p.def_year/2*(i-1),unit='d')
            t_stop = p.t0+pd.to_timedelta(p.def_year/2*i,unit='d')
            dat_new = dat_study[['moh_id','earliest_HIV','FARVDT','end_fu_dt']][(dat_study.earliest_HIV>=t_start)&(dat_study.earliest_HIV<t_stop)].reset_index(drop=True) #only keep those diagnosed after t_start and alive by t_stop
            dat_hist = pd.merge(dat_hist,dat_new,how='outer',on=['moh_id','earliest_HIV','FARVDT','end_fu_dt',])
            dat_hist[str(t_stop)[:10]] = dat_hist.apply(lambda x: 'T' if (x['FARVDT']<t_stop and x['end_fu_dt']>=t_stop) else ('D' if x['end_fu_dt']<t_stop else 'A'),axis=1) 
            dat_hist[str(t_stop)[:10]] = dat_hist.apply(func_apply_spvl_status,args=(str(t_stop)[:10],dat_spvl_rebound,list_status_columns),axis=1) #update spvl/rebound status based on pre-determined status T
        file_data = open('data/hiv_status.pkl','wb') 
        pickle.dump(dat_hist, file_data)

    if 'record_hiv_status_every6m_adj' in runs:
        """Adjust unsuppressed HIV status based on hiv_status_pre.pkl, in which changes between AF/U_on and AI/U_off before suppression were not considered
        for both STOP and DTP data"""

        p = Params()

        #####load cohort information, ART dispensation records and HIV status data for those in DTP, used for dtp_hiv_status derivation
        dat_cohort = pd.read_sas('data/dtp.sas7bdat',encoding='latin-1') 
        dat_cohort = dat_cohort.rename(columns={'PSEUDO':'moh_id'})
        cutoff_dt = pd.Timestamp('2019-12-31')
        dat_cohort['end_fu_dt'] = dat_cohort.apply(lambda x: x['LASTCTDT'] if x['LASTCTDT']==x['DTHDT'] or (cutoff_dt-x['LASTCTDT']).days>365.25/2*3 else cutoff_dt, axis=1) #apply last_contact_dt similar to STOP, to leave enough time to make sure the person is truely 'lost to follow up', LASTCTDT should be conservative as not many contact resources are linked to DTP (unlike STOP)
        dat_art = pd.read_sas('data/art_records.sas7bdat',encoding='latin-1') #ART dispensation records from the DTP
        dat_art = dat_art.rename(columns={'PSEUDO':'moh_id'})
        dat_art = dat_art.sort_values(by=['moh_id','STARTDATE']) 
        file_data = open('data/dtp_hiv_status_pre.pkl','rb') #use status at each step without any adjustments
        dat_hiv = pickle.load(file_data) #be aware that state AF(ART failure)=U_on(unsuppressed and on ART) and AI(ART interruption)=U_off(unsuppressed and off ART)
        list_dates = list(dat_hiv)[3:] 
        dat_hiv = dat_hiv.replace({'T':'U_on','AF':'U_on','AI':'U_off'}) #replace status by S/U_off/U_on 

        #####load cohort information, ART dispensation records and HIV status data every 6 months, used for hiv_status derivation
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        dat_art = pd.read_sas('data/art_records.sas7bdat',encoding='latin-1') #art dispensation information 
        file_data = open('data/hiv_status_pre.pkl','rb') #HIV status without adjustment
        dat_hiv = pickle.load(file_data) #be aware that state AF(ART failure)=U_on(unsuppressed and on ART) and AI(ART interruption)=U_off(unsuppressed and off ART)
        list_dates = list(dat_hiv)[4:]
        dat_hiv = dat_hiv.replace({'T':'U_on','AF':'U_on','AI':'U_off'}) #replace status by S/U_off/U_on 

        #####check unsupressed status at each time step based on ART dispensation record
        dat_hiv_adj = dat_hiv.copy()
        for date_i in list_dates:
            arr_id_check = dat_hiv_adj['moh_id'][(pd.isnull(dat_hiv_adj[date_i])==False)&(dat_hiv_adj[date_i].str.contains('U'))].values
            
            ####only modify the status if ART dispensation record is available
            for id_j in arr_id_check:
                dat_art_j = dat_art[dat_art['moh_id']==id_j].reset_index(drop=True)
                if not dat_art_j.empty:
                    for index_j in dat_art_j.index:
                        if index_j<dat_art_j.index.max():
                            t_start = pd.Timestamp(dat_art_j['STARTDATE'][dat_art_j.index==index_j].values[0])
                            t_stop = pd.Timestamp(dat_art_j['STOPDATE'][dat_art_j.index==index_j].values[0])
                            t_start_new = pd.Timestamp(dat_art_j['STARTDATE'][dat_art_j.index==index_j+1].values[0])
                            if t_start<pd.Timestamp(date_i)<=t_stop or (t_stop<pd.Timestamp(date_i)<=t_start_new and (t_start_new-t_stop).days<90): #define ART interruption period by 90 days gap between stop date and next start date and check whether the time step fell in the period (PMID:29048508)
                                dat_hiv_adj.loc[dat_hiv_adj['moh_id']==id_j,date_i] = 'U_on'
                                break;
                            elif t_stop<pd.Timestamp(date_i)<=t_start_new and (t_start_new-t_stop).days>=90:
                                dat_hiv_adj.loc[dat_hiv_adj['moh_id']==id_j,date_i] = 'U_off'
                                break;
                        else:
                            #####need specify final t_start and t_stop separately as t_start_new does not exist
                            t_start = pd.Timestamp(dat_art_j['STARTDATE'][dat_art_j.index==index_j].values[0])
                            t_stop = pd.Timestamp(dat_art_j['STOPDATE'][dat_art_j.index==index_j].values[0])
                            if (pd.Timestamp(date_i)-t_stop).days<90:
                                dat_hiv_adj.loc[dat_hiv_adj['moh_id']==id_j,date_i] = 'U_on'
                            else:
                                dat_hiv_adj.loc[dat_hiv_adj['moh_id']==id_j,date_i] = 'U_off'

        #####output adjusted HIV status for STOP and DTP separately
        file_data = open('data/dtp_hiv_status.pkl','wb') #derived information for DTP cohort
        #file_data = open('data/hiv_status.pkl','wb') #derived information for STOP cohort
        pickle.dump(dat_hiv_adj, file_data)
        file_data.close()

    if 'record_prob_ART_init' in runs:
        """estimate probability from aware/undiagnosed to S/U_on/U_off after 6 months time step"""

        p = Params()

        #####load STOP cohort and derived HIV status every 6 months
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1')
        dat_hiv = pd.read_pickle('data/hiv_status.pkl') 
        list_dates = list(dat_hiv)[4:]

        #####estimate probability to different state (S/U_on/U_off) from A or new diagnosis (nan)
        dat_prob = pd.DataFrame([],columns=['date','A2S','A2U_on','A2U_off','U2S','U2U_on','U2U_off'])
        for i,date_i in enumerate(list_dates[1:]):
            dic_prob = {'date':date_i}
            for key_j in ['S','U_on','U_off']:
                dic_prob['A2'+key_j] = dat_hiv[(dat_hiv[list_dates[i]]=='A')&(dat_hiv[date_i]==key_j)].shape[0]
                dic_prob['U2'+key_j] = dat_hiv[(pd.isnull(dat_hiv[list_dates[i]]))&(dat_hiv[date_i]==key_j)].shape[0]
            dat_prob = pd.concat([dat_prob,pd.DataFrame(dic_prob,index=[0])],ignore_index=True)
        dat_prob['p_A2S_A'] = dat_prob['A2S']/dat_prob[['A2S','A2U_on','A2U_off']].sum(axis=1) 
        dat_prob['p_A2S_all'] = dat_prob[['A2S','U2S']].sum(axis=1)/dat_prob[['A2S','A2U_on','A2U_off','U2S','U2U_on','U2U_off']].sum(axis=1) 
        dat_prob['p_A2Uon_A'] = dat_prob['A2U_on']/dat_prob[['A2S','A2U_on','A2U_off']].sum(axis=1) 
        dat_prob['p_A2Uon_all'] = dat_prob[['A2U_on','U2U_on']].sum(axis=1)/dat_prob[['A2S','A2U_on','A2U_off','U2S','U2U_on','U2U_off']].sum(axis=1) 

        #####estimate overall distribution of S/U_on/U_off right after ART initiation (6m step)
        A2S_A = dat_prob['A2S'].sum()/dat_prob[['A2S','A2U_on','A2U_off']].sum(axis=1).sum() 
        A2S_all = dat_prob[['A2S','U2S']].sum(axis=1).sum()/dat_prob[['A2S','A2U_on','A2U_off','U2S','U2U_on','U2U_off']].sum(axis=1).sum()
        A2Uon_A = dat_prob['A2U_on'].sum()/dat_prob[['A2S','A2U_on','A2U_off']].sum(axis=1).sum()
        A2Uon_all = dat_prob[['A2U_on','U2U_on']].sum(axis=1).sum()/dat_prob[['A2S','A2U_on','A2U_off','U2S','U2U_on','U2U_off']].sum(axis=1).sum()
        prob_A2S = 0.501024823502619 #used in func_params_update
        prob_A2Uon = 0.4618537918469597

    if 'record_hiv_status_every6m_dtp' in runs:
        """determine HIV status for each eligible individual (age>=18, alive by 2008-01-01) in DTP every 6 months from 2008-01-01 to 2019-12-31"""

        p = Params()

        #####load historical data and derived rebound/failure information
        dat_cohort = pd.read_sas('data/dtp.sas7bdat',encoding='latin-1') 
        dat_cohort = dat_cohort.rename(columns={'PSEUDO':'moh_id'})
        cutoff_dt = pd.Timestamp('2019-12-31')
        dat_cohort['end_fu_dt'] = dat_cohort.apply(lambda x: x['LASTCTDT'] if x['LASTCTDT']==x['DTHDT'] or (cutoff_dt-x['LASTCTDT']).days>365.25/2*3 else cutoff_dt, axis=1) #apply last_contact_dt similar to STOP, to leave enough time to make sure the person is truely 'lost to follow up', LASTCTDT should be conservative as not many contact resources are linked to DTP (unlike STOP)
        file_data = open('data/dtp_spvl_rebound_status.pkl','rb') 
        dat_spvl_rebound = pickle.load(file_data)
        list_spvl_rebound = [[str(i)+'_spvl_dt',str(i)+'_rebound_dt'] for i in range(1,12)]  
        list_columns = [item for sublist in list_spvl_rebound for item in sublist] #columns with spvl/rebound date
        dat_spvl_rebound[list_columns] = dat_spvl_rebound[list_columns].fillna(pd.Timestamp('2050-01-01')) #replace nan by dummy dates for comparison purpose
        list_status_columns = ['FARVDT']+list_columns #consider period from FARVDT to last spvl/rebound

        #####set up time step and eligible population
        p.t0 = pd.Timestamp('2008-01-01') 
        p.t_end = pd.Timestamp('2020-01-01') 
        p.dt = 0.5
        p.n_step = int((p.t_end.year-p.t0.year)/p.dt)
        dat_study = dat_cohort[dat_cohort.end_fu_dt.dt.year>=2008].reset_index(drop=True) 
        dat_study = dat_study[(pd.isnull(dat_study.SEX_AT_BIRTH)==False)&(pd.isnull(dat_study.BIRTHDATE)==False)].reset_index(drop=True) 
        dat_study['age_art'] = pd.Series((dat_study['FARVDT']-dat_study['BIRTHDATE']).dt.days/p.def_year) 
        dat_study = dat_study[dat_study.age_art>=18].reset_index(drop=True) 

        #####determine status at every time step from 2008-01-01
        dat_hist = dat_study[['moh_id','FARVDT','end_fu_dt']][(dat_study.end_fu_dt>=p.t0)&(dat_study.FARVDT<p.t0)].reset_index(drop=True) #plwh with ART initiation before 2008-01-01
        dat_hist[str(p.t0)[:10]] = pd.Series(['T']*dat_hist.shape[0]) #all plwh in DTP have initiated ART
        dat_hist[str(p.t0)[:10]] = dat_hist.apply(func_apply_spvl_status,args=(str(p.t0)[:10],dat_spvl_rebound,list_status_columns),axis=1) 
        for i in range(1,p.n_step+1):
            t_start = p.t0+pd.to_timedelta(p.def_year/2*(i-1),unit='d')
            if i<p.n_step:
                t_stop = p.t0+pd.to_timedelta(p.def_year/2*i,unit='d')
            else:
                t_stop = cutoff_dt #the final t_stop needs to be defined separately as t_stop might be beyond cutoff_dt and we don't have data (LASTCTDT,STARTDATE) after cutoff_dt
            dat_new = dat_study[['moh_id','FARVDT','end_fu_dt']][(dat_study['FARVDT']>=t_start)&(dat_study['FARVDT']<t_stop)].reset_index(drop=True) #only keep those initiated ART between t_start and t_stop
            dat_hist = pd.merge(dat_hist,dat_new,how='outer',on=['moh_id','FARVDT','end_fu_dt',])
            dat_hist[str(t_stop)[:10]] = dat_hist.apply(lambda x: 'T' if (x['FARVDT']<t_stop and x['end_fu_dt']>=t_stop) else 'D',axis=1) 
            dat_hist[str(t_stop)[:10]] = dat_hist.apply(func_apply_spvl_status,args=(str(t_stop)[:10],dat_spvl_rebound,list_status_columns),axis=1) #update spvl/rebound status based on pre-determined status T
        file_data = open('data/dtp_hiv_status.pkl','wb') 
        pickle.dump(dat_hist, file_data)

    if 'microsim_full_hist_cali' in runs:
        """calibrate model simulation to fit historical targets after including health behaviour factors"""

        #####load initial setting of the parameter set and update the ones with calibrations
        p0 = Params()
        p0.rand_hb = int(input('Input 0/1 indicating whether to consider health behaviours: ')) #0/1 indicating whether to consider the impacts of health behaviours (physical activity, alcohol, smoking) on comorb/mortality incidence
        if p0.rand_hb==0:
            print ('Health behaviours not considered')
        elif p0.rand_hb==1:
            print ('Health behaviours considered')
        else:
            print ('Error: double-check coefficients for health behaviours')
        p = func_params_update(p0)

        #####set up time period for microsimulation
        p.year0 = 2007. #introduce year information at each time step to use sigmoid function for estimations of new infection, new diagnosis and undiagnosed PLWH
        p.t0 = pd.Timestamp('2008-01-01') #different from study_dt in Params() as it can be changed for simulation 
        p.t_end = pd.Timestamp('2017-01-01') #use January 1st instead of December 31st as the time is 00:00 in Timestamp
        p.dt = 0.5 
        p.n_step = int((p.t_end.year-p.t0.year)/p.dt)
        
        #####load STOP data to build a synthesized population of PLWH with necessary characteristics sampled, the number of diagnosis is the same as STOP
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1') #load individual characteristics from STOP
        dat_ncd4 = pd.read_pickle('data/cd4_nadir_baseline.pkl') #load CD4 nadir data at baseline
        dat_su = pd.read_pickle('data/substance_use.pkl') #load information of substance use for those with age and sex, alive in 2008, at least 18 at diagnosis 
        dat_comorb = pd.read_pickle('data/comorbidities.pkl') #load information of comorbidity diagnosis at comorb_baseline for those with age and sex, alive in 2008, at least 18 at diagnosis 
        list_col_comorb = list(dat_comorb)[-11:] #list of column names for comorbidity incidence date
        for col_i in list_col_comorb:
            dat_comorb[col_i] = dat_comorb[col_i].fillna(p.dummy_end_dt) #replace NaT by dummy_end_dt for comparison purpose
        dat_hiv = pd.read_pickle('data/hiv_status.pkl') #load HIV status to build baseline status
        dat_status_pvl = pd.read_pickle('data/pvl_status.pkl') #load viral suppression status for each 1 year period, needed for baseline characteristics 
        dic_status_reg = pd.read_pickle('data/art_reg_status.pkl') #regimen type by the end of each time step, needed for baseline regimen status if on ART
        dic_prop_reg = pd.read_pickle('data/prop_time_art_reg.pkl') #proportion of time on regimen while on ART, needed for baseline characteristics 
        for reg_i in dic_prop_reg.keys(): #replace reg_prop>1 by 1 for simplification
            list_reg_dates = list(dic_prop_reg[reg_i])[4:]
            for date_j in list_reg_dates:
                dic_prop_reg[reg_i].loc[dic_prop_reg[reg_i][date_j]>1.01,date_j] = 1.

        #####create population of PLWH starting from 2008-01-01 with necessary characteristics for microsimulation
        p.rand_pop = input('Input hist08/hist/random indicating the way to create cohort and characteristics: ') #use hist for calibration
        dic_raw_data = { #put cohort information together for easy loading in the function
            'cohort':dat_cohort, 'cd4':dat_ncd4,'su':dat_su,'comorb':dat_comorb,'hiv':dat_hiv,'pvl':dat_status_pvl,'reg_status':dic_status_reg,'reg_prop':dic_prop_reg,
        }
        dic_update_data = func_hist_prepare(p,dic_raw_data) #modify cohort information for consistency between dataframes

        #####create targets based on historical data 
        dic_hist_target = func_hist_cali_target(p,dic_update_data)

        #####initialize parameters for multiprocessing
        p.n_pool = int(input('Number of threads used for multiprocessing: '))
        p.num_sim = int(input('Number of microsimulation iterations run: '))
        p.arr_seed = np.arange(p.num_sim) #use ascending order seed sequence from 0 to num

        #####update coefficients for probability adjustment after run func_cali_target_micro_all for each module separately and combined
        dic_opt_prob_comorb_coeff = {
            'cvd':0.64019531, 
            'htn':0.95234375, 
            'dm':0.8453125, 
            'oa':0.975, 
            'copd':0.44560547, 
            'ckd':0.81875, 
            'cld':0.5840332, 
            'cancer':0.88125, 
            'manx':0.5109375, 
            'sczo':0.50019531, 
            'prsn':0.928125, 
            }
        for key_i in dic_opt_prob_comorb_coeff.keys():
            p.dic_prob_comorb_coeff[key_i] = dic_opt_prob_comorb_coeff[key_i]
        p.dt_split_prob_a2t = pd.Timestamp('2013-12-31') 
        dic_opt_coeff_adj_a2t = {
            'bf08':np.array([0.36818393, 0.25761546]), 
            '0811':np.array([0.69744916, 0.17756822]), 
            'sin12':np.array([0.69908445, 0.42727508]), 
            }
        p.dic_coeff_adj_a2t_pw['bf08'] = np.copy(dic_opt_coeff_adj_a2t['bf08'])
        p.dic_coeff_adj_a2t_pw['0811'] = np.copy(dic_opt_coeff_adj_a2t['0811']) 
        p.dic_coeff_adj_a2t_pw['sin12'] = np.copy(dic_opt_coeff_adj_a2t['sin12'])
        dic_opt_coeff_prob_art = {
            'opt':np.array([0.82516378, 0.45341352, 0.55834187, 1.53005645, 1.08743479, 1.27687444]), 
        }
        opt_art_key = 'opt' 
        arr_opt_coeff_prob_art = dic_opt_coeff_prob_art[opt_art_key]
        p.dic_coeff_prob_art = {k:arr_opt_coeff_prob_art[i] for i,k in enumerate(list(p.dic_coeff_prob_art))}
        dic_opt_coeff_prob_dead = {
            'opt':np.array([1.55372862, 1.18309458, 4.23182396, 1.18484426]), 
        }
        opt_dead_key = 'opt' 
        p.dt_split_prob_dead_noart = pd.Timestamp('2012-12-31') #time to split piece-wise coefficient for A2D 
        arr_opt_coeff_prob_dead = dic_opt_coeff_prob_dead[opt_dead_key]
        p.coeff_prob_dead_ltfu_art = dic_opt_coeff_prob_dead[opt_dead_key][0]
        p.coeff_prob_dead_ltfu_noart_pw = np.array([dic_opt_coeff_prob_dead[opt_dead_key][1],dic_opt_coeff_prob_dead[opt_dead_key][2]]) #piece-wise adjustment for A2D
        p.dic_prob_dead_art['pvl_dv'][0] = dic_opt_coeff_prob_dead[opt_dead_key][-1]
        
        #####optimization process to minimize residuals between simulated and historical outcomes
        dic_params_set = {
            'init':np.ones(4), #initial guess of coefficients to fit deaths
            #'init':np.ones(6), #initial guess of coefficients to fit new art at once or art transitions
            #'init':np.ones(2), #initial guess of coeffciients to fit new art for different periods
            #'init':np.ones(1), #initial guess of coefficients to fit one comorbidity at a time
            #'init':np.ones(11), #initial guess of coefficients to fit all comorbidities
            #'init':np.ones(21), #initial guess of coefficients to fit all targets
        }
        init_params = dic_params_set['init']
        args_cali = input('Calibration on specific targets: ') # calibration for different module: all, comorb (cvd/htn...), new_art (bf08, 0811, sin12), trans_art, dead
        match_N = minimize(func_cali_target_micro_all,init_params,args=(dic_hist_target,dic_update_data,p,args_cali),method='nelder-mead',options={'xatol': 1e-4}) #comorb to indicate calibration on comorbidity incidence only

        #####use Latin-hypercube sampling to randomize adjusted coefficients and filtered by residuals for comorb/new_art/trans_art/dead based on reasonable initial guess after initial calibrations
        coeff_var = 1. #variation for each adjusted coefficient, use smaller range for further optimization, e.g., 0.05
        p.dic_mc_range = { 
            'comorb': {key_i:[p.dic_prob_comorb_coeff[key_i]*(1-coeff_var), p.dic_prob_comorb_coeff[key_i]*(1+coeff_var)] for key_i in p.list_comorb},
            'new_art': {key_i:np.array([p.dic_coeff_adj_a2t_pw[key_i]*(1-coeff_var),p.dic_coeff_adj_a2t_pw[key_i]*(1+coeff_var)]) for key_i in ['bf08','0811','sin12']}, #row1 as lower bound for variation, col1 as coefficient 1 
            'trans_art':np.array([np.array(list(p.dic_coeff_prob_art.values()))*(1-coeff_var),np.array(list(p.dic_coeff_prob_art.values()))*(1+coeff_var)]), #dimention 2*6
            'dead':np.array([arr_opt_coeff_prob_dead*(1-coeff_var),arr_opt_coeff_prob_dead*(1+coeff_var)]), #dimension 2*3
        }
        n_lhs = 1000 #400 #100 #500 
        print ('Number of random samples for adjusted coefficients: ', n_lhs)
        n_coeff = len(p.dic_mc_range['comorb'])+6+6+4 #introduce piece-wise coefficients for new_art sin12 and coeff_prob_dead_ltfu_noart
        np.random.seed(1000000) #specify random generator for x_sample for reproductive purpose
        random.seed(1000000)
        x_sample = lhs(n_coeff,samples=n_lhs)
        list_rand_coeff = []
        list_rand_res = []
        for i in range(n_lhs):
            rand_results = func_mc_filter_micro_all(x_sample[i,:],dic_hist_target,dic_update_data,p)
            list_rand_coeff.append(rand_results[0])
            list_rand_res.append(rand_results[1])
        
        #####output list_rand_coeff and list_rand_res as .pkl file
        dic_output = {
            'path':'my_path/results/cali/',
            'name':'lhs_cali_',
            'date':'today', 
        }
        file_results = open(dic_output['path']+dic_output['name']+dic_output['date']+'.pkl','wb')
        pickle.dump([list_rand_coeff,list_rand_res,p.dic_mc_range],file_results) 
        file_results.close()

        #####load list_rand_coeff and list_rand_res from .pkl file
        file_results = open(dic_output['path']+dic_output['name']+dic_output['date']+'.pkl','rb')
        list_results = pickle.load(file_results)
        list_rand_coeff = list_results[0]
        list_rand_res = list_results[1]
        p.dic_mc_range = copy.deepcopy(list_results[2]) 

        #####update dic_mc_range by comparing the results above with targeted residual 
        dic_res_target = {
            #####update target residuals base on the residual for parameters before randomization using lhs
            'new_art bf08': 9.424899999999608,
            'new_art 0811': 19.980900000000243,
            'new_art sin12': 584.1889000000035,
            'S2Uoff': 17.55610000000046,
            'Uoff2S': 37.33210000000017,
            'S2Uon': 6.300099999999954,
            'Uon2S': 655.8721000000065,
            'Uon2Uoff': 61.62250000000036,
            'Uoff2Uon': 883.2784000000016,
            'A2D': 0.980100000000018,
            'S2D': 200.50560000000232,
            'Uon2D': 22.184100000000075,
            'Uoff2D': 386.9088999999995,
            'cvd': 0.6400000000000182,
            'htn': 596.336399999998,
            'dm': 295.83999999999963,
            'oa': 35.76040000000022,
            'copd': 293.43689999999987,
            'ckd': 159.2644000000001,
            'cld': 72.25,
            'cancer': 15.523599999999982,
            'manx': 398.0025000000018,
            'sczo': 15.76089999999999,
            'prsn': 5.198400000000005
        }
        dic_filter = func_mc_filter_res(list_rand_res,dic_res_target)
        list_sum_res = [sum(list(rand_res.values())) for rand_res in list_rand_res]
        dic_mc_range_update = func_mc_param_range_update(list_rand_coeff,dic_filter,p)
        p.dic_mc_range = copy.deepcopy(dic_mc_range_update)

        #####update parameter ranges by limiting to the ones close to smallest residual sum, stop when min_res cannot be improved further
        arr_res_sum = np.array([sum(list(rand_res.values())) for rand_res in list_rand_res])
        arr_sort_ind = np.argsort(arr_res_sum)
        dic_mc_range_update = func_mc_param_range_update_res(list_rand_coeff,list_rand_res,p,0.05)
        p.dic_mc_range = copy.deepcopy(dic_mc_range_update)

        #####update parameter based on the optimized set and run microsimulation for figures in SI
        ind_res_min = np.argmin(arr_res_sum)
        dic_opt_rand_coeff = copy.deepcopy(list_rand_coeff[ind_res_min])
        for key_i in dic_opt_rand_coeff['comorb'].keys():
            p.dic_prob_comorb_coeff[key_i] = dic_opt_rand_coeff['comorb'][key_i] 
        p.dic_coeff_adj_a2t_pw['bf08'] = np.copy(dic_opt_rand_coeff['new_art']['bf08'])
        p.dic_coeff_adj_a2t_pw['0811'] = np.copy(dic_opt_rand_coeff['new_art']['0811']) 
        p.dic_coeff_adj_a2t_pw['sin12'] = np.copy(dic_opt_rand_coeff['new_art']['sin12']) 
        p.dic_coeff_prob_art = {k:dic_opt_rand_coeff['trans_art'][i] for i,k in enumerate(list(p.dic_coeff_prob_art))}
        p.coeff_prob_dead_ltfu_art = dic_opt_rand_coeff['dead'][0]
        p.coeff_prob_dead_ltfu_noart_pw = np.array([dic_opt_rand_coeff['dead'][1],dic_opt_rand_coeff['dead'][2]])
        p.dic_prob_dead_art['pvl_dv'][0] = dic_opt_rand_coeff['dead'][-1] 

        #####re-run multiprocessing with updated coefficients for probabilities and derive simulated targets
        pool = Pool(processes=p.n_pool,maxtasksperchild=1)
        results = [pool.apply_async(mp_microsim_outcomes_cali,args=(p,dic_update_data,p.arr_seed[i])) for i in range(p.num_sim)]
        list_results = [x.get() for x in results]
        dic_summary = func_micro_outcomes_target_ci(p,list_results)

        #####save summarized results after re-calibration
        file_results = open(dic_output['path']+'results_cali_sim100_'+dic_output['date']+'.pkl','wb')
        pickle.dump(dic_summary,file_results) 
        file_results.close()

        #####re-load results from calibrated parameter set
        file_results = open(dic_output['path']+'results_cali_sim100_'+dic_output['date']+'.pkl','rb')
        dic_summary = pickle.load(file_results)

        #####change dic_output['date'] to update figures for SI
        dic_output['date'] = 'new_date' 

        #####Figures S4-S6: compare simulated outcomes with credible intervals with historical targets
        dic_char_trans = {
            'new_diag':{
                'color':[dic_color_palette['green'][1],dic_color_palette['green'][2]], #[0] for historical and [1] for simulated
                'ylabel':'Cumulative HIV diagnosis'}, 
            'new_art':{
                'color':[dic_color_palette['blue'][1],dic_color_palette['blue'][2]],
                'ylabel':['Annual ART initiation','Cumulative number of ART initiations'],
                },
            'dead':{
                'color':[dic_color_palette['red'][1],dic_color_palette['red'][2]],
                'ylabel':['Annual all-cause deaths','Cumulative all-cause deaths'],
                },
            'new_art bf08':{
                'color':[dic_color_palette['blue'][0],dic_color_palette['blue'][1]],
                'ylabel':['Annual ART initiation','Cumulative number of ART initiations'],
                'title':'Among PLWH diagnosed before 2008',
                },
            'new_art 0811':{
                'color':[dic_color_palette['blue'][0],dic_color_palette['blue'][1]],
                'ylabel':['Annual ART initiation','Cumulative number of ART initiations'],
                'title':'Among PLWH diagnosed in 2008-2011',
                },
            'new_art sin12':{
                'color':[dic_color_palette['blue'][0],dic_color_palette['blue'][1]],
                'ylabel':['Annual ART initiation','Cumulative number of ART initiations'],
                'title':'Among PLWH diagnosed since 2012',
                },
            'S2Uoff':{
                'color':[dic_color_palette['green'][0],dic_color_palette['green'][1]], 
                'label':[r'From $S$ to $U_{off}$ (STOP HIV/AIDS)',r'From $S$ to $U_{off}$ (Model simulation)'],}, 
            'Uoff2S':{
                'color':[dic_color_palette['purple'][0],dic_color_palette['purple'][2]], 
                'label':[r'From $U_{off}$ to $S$ (STOP HIV/AIDS)',r'From $U_{off}$ to $S$ (Model simulations)'],},
            'S2Uon':{
                'color':[dic_color_palette['green'][0],dic_color_palette['green'][1]], 
                'label':[r'From $S$ to $U_{on}$ (STOP HIV/AIDS)',r'From $S$ to $U_{on}$ (Model simulations)'],},
            'Uon2S':{
                'color':[dic_color_palette['purple'][0],dic_color_palette['purple'][2]], 
                'label':[r'From $U_{on}$ to $S$ (STOP HIV/AIDS)',r'From $U_{on}$ to $S$ (Model simulations)'],},
            'Uon2Uoff':{
                'color':[dic_color_palette['green'][0],dic_color_palette['green'][1]], 
                'label':[r'From $U_{on}$ to $U_{off}$ (STOP HIV/AIDS)',r'From $U_{on}$ to $U_{off}$ (microsimulation)'],},
            'Uoff2Uon':{
                'color':[dic_color_palette['purple'][0],dic_color_palette['purple'][2]], 
                'label':[r'From $U_{off}$ to $U_{on}$ (STOP HIV/AIDS)',r'From $U_{off}$ to $U_{on}$ (Model simulations)'],},
            'diag':{
                'color':[dic_color_palette['red'][1],dic_color_palette['red'][2]],
                'label':['Diagnosed PLWH (hist)','Diagnosed PLWH (sim)']},
            'art_ever':{
                'color':[dic_color_palette['purple'][1],dic_color_palette['purple'][2]],
                'label':['Ever on ART (hist)','Ever on ART (sim)']},
            'art':{
                'color':[dic_color_palette['green'][1],dic_color_palette['green'][2]],
                'label':['On ART (hist)','On ART (sim)']},
            'S':{
                'color':[dic_color_palette['blue'][1],dic_color_palette['blue'][2]],
                'label':['Suppressed (hist)','Suppressed (sim)']},
            'A2D':{
                'color':[dic_color_palette['blue'][0],dic_color_palette['blue'][1]], 
                'label':[r'Among $A$ (STOP HIV/AIDS)',r'Among $A$ (Model simulations)'],}, 
            'S2D':{
                'color':[dic_color_palette['blue'][0],dic_color_palette['blue'][1]], 
                'label':[r'Among $S$ (STOP HIV/AIDS)',r'Among $S$ (Model simulations)'],},
            'Uon2D':{
                'color':[dic_color_palette['blue'][0],dic_color_palette['blue'][1]], 
                'label':[r'Among $U_{on}$ (STOP HIV/AIDS)',r'Among $U_{on}$ (Model simulations)'],},
            'Uoff2D':{
                'color':[dic_color_palette['blue'][0],dic_color_palette['blue'][1]], 
                'label':[r'Among $U_{off}$ (STOP HIV/AIDS)',r'Among $U_{off}$ (Model simulations)'],},
        }
        list_panel = ['A)','B)','C)']
        fig_new3,axs_new3 = plt.subplots(1,3,figsize=(14,4)) #show cum cases for calibration up to 2015 in SI
        for i,key_i in enumerate(['new_art bf08','new_art 0811','new_art sin12']):
            axs_new3[i].plot(dic_hist_target['hiv']['year'].values[:-1],dic_hist_target['hiv'][key_i].cumsum()[:-1],'-',color=dic_char_trans[key_i]['color'][0],lw=1.5,label='STOP HIV/AIDS') 
            axs_new3[i].errorbar(dic_summary['hiv']['year'].values[:-1],dic_summary['hiv']['cum '+key_i+' med'].values[:-1],
                                yerr=[dic_summary['hiv']['cum '+key_i+' med'].values[:-1]-dic_summary['hiv']['cum '+key_i+' lb'].values[:-1],
                                      dic_summary['hiv']['cum '+key_i+' ub'].values[:-1]-dic_summary['hiv']['cum '+key_i+' med'].values[:-1]],
                                      fmt='o',markersize=4,color=dic_char_trans[key_i]['color'][1],ecolor=dic_char_trans[key_i]['color'][1],capsize=2.5,label='Model simulations')
            axs_new3[i].text(2007,1500*1.05,list_panel[i],fontsize=12)
            axs_new3[i].set_xlabel('Year',fontsize=12)
            axs_new3[i].set_xlim([2007.8,2015.2])
            axs_new3[i].set_ylim([0,1500])
            axs_new3[i].legend(loc='upper left',fontsize=10)
        axs_new3[0].set_ylabel(dic_char_trans[key_i]['ylabel'][1],fontsize=12)
        fig_new3.subplots_adjust(top=0.91, bottom=0.13, left=0.06,right=0.98,hspace=0.25,wspace=0.25)
        plt.show()
        fig_new3.savefig('my_path/results/fig_cali_a2t_sim100_'+dic_output['date']+'.png',dpi=300) #Figure S4
        list_dead_panel = [['A)','B)'],['C)','D)']]
        list_dead_ylim = [[[0,1000],[0,1200]],[[0,250],[0,200]]]
        fig_dead2,axs_dead2 = plt.subplots(2,2,figsize=(10,8)) #compare annual/cum deaths from each state
        for i,group_i in enumerate([['A2D','S2D'],['Uon2D','Uoff2D']]):
            for j,key_i in enumerate(group_i):
                axs_dead2[i,j].plot(dic_hist_target['hiv']['year'].values[:-1],dic_hist_target['hiv'][key_i].cumsum()[:-1],'-',color=dic_char_trans[key_i]['color'][0],lw=1.5,label=dic_char_trans[key_i]['label'][0])
                axs_dead2[i,j].errorbar(dic_summary['hiv']['year'].values[:-1],dic_summary['hiv']['cum '+key_i+' med'].values[:-1],
                                    yerr=[dic_summary['hiv']['cum '+key_i+' med'].values[:-1]-dic_summary['hiv']['cum '+key_i+' lb'].values[:-1],
                                        dic_summary['hiv']['cum '+key_i+' ub'].values[:-1]-dic_summary['hiv']['cum '+key_i+' med'].values[:-1]],
                                        fmt='o',markersize=4,color=dic_char_trans[key_i]['color'][1],ecolor=dic_char_trans[key_i]['color'][1],capsize=2.5,label=dic_char_trans[key_i]['label'][1])
                axs_dead2[i,j].text(2007,list_dead_ylim[i][j][1]*1.05,list_dead_panel[i][j],fontsize=12)
                axs_dead2[i,j].set_xlim([2007.8,2015.2])
                axs_dead2[1,j].set_xlabel('Year',fontsize=12)
                axs_dead2[i,j].set_ylim(list_dead_ylim[i][j])
                axs_dead2[i,j].legend(loc='upper left',fontsize=10,fancybox=True)
            axs_dead2[i,0].set_ylabel('Cumulative number of deaths',fontsize=12)
        fig_dead2.subplots_adjust(top=0.95, bottom=0.08, left=0.09,right=0.97,hspace=0.25,wspace=0.25)
        plt.show()
        fig_dead2.savefig('my_path/results/fig_cali_dead_sim100_'+dic_output['date']+'.png',dpi=300) #Figure S6
        list_trans_ylim = [[0,1000],[0,3500],[0,1400]]
        fig_trans2,axs_trans2 = plt.subplots(1,3,figsize=(16,4))
        for i,group_i in enumerate([['S2Uoff','Uoff2S'],['S2Uon','Uon2S'],['Uon2Uoff','Uoff2Uon']]):
            for j,key_j in enumerate(group_i):
                axs_trans2[i].plot(dic_hist_target['hiv']['year'].values[:-1],dic_hist_target['hiv'][key_j].cumsum()[:-1],'-',color=dic_char_trans[key_j]['color'][0],lw=2.,label=dic_char_trans[key_j]['label'][0])
                axs_trans2[i].errorbar(dic_summary['hiv']['year'].values[:-1],dic_summary['hiv']['cum '+key_j+' med'].values[:-1],
                                yerr=[dic_summary['hiv']['cum '+key_j+' med'].values[:-1]-dic_summary['hiv']['cum '+key_j+' lb'].values[:-1],
                                      dic_summary['hiv']['cum '+key_j+' ub'].values[:-1]-dic_summary['hiv']['cum '+key_j+' med'].values[:-1]],
                                      fmt='o',markersize=4,color=dic_char_trans[key_j]['color'][1],ecolor=dic_char_trans[key_j]['color'][1],capsize=2.5,label=dic_char_trans[key_j]['label'][1])
            axs_trans2[i].text(2007,list_trans_ylim[i][1]*1.05,list_panel[i],fontsize=12)
            axs_trans2[i].set_xlim([2007.8,2015.2])
            axs_trans2[i].set_xlabel('Year',fontsize=12)
            axs_trans2[i].set_ylim(list_trans_ylim[i])
            axs_trans2[i].legend(loc='upper left',fontsize=9,fancybox=True) 
        axs_trans2[0].set_ylabel('Cumulative number of transitions',fontsize=12)
        fig_trans2.subplots_adjust(top=0.92, bottom=0.12, left=0.05,right=0.98,hspace=0.3,wspace=0.25)
        plt.show()
        fig_trans2.savefig('my_path/results/fig_cali_trans_art_sim100_'+dic_output['date']+'.png',dpi=300) #Figure S5 

        #####Figure S7:plots to show historical and simulated comorbidity incidence (physical and mental)
        dic_comorb_label = {
            'cvd':'CVD','htn':'HTN','dm':'Diabetes','oa':'OA',
            'copd':'COPD','ckd':'CKD','cld':'CLD','cancer':'Cancers',
            'manx':'MANX','prsn':'PD','sczo':'SCZ',
        }
        list_phys_panel = [['A)','B)','C)','D)'],['E)','F)','G)','H)'],['I)','J)','K)','L)']]
        fig_phys,axs_phys = plt.subplots(3,4,figsize=(18,9))
        for i,row_i in enumerate([['htn','cvd','dm','copd'],['ckd','cld','oa','cancer'],['manx','sczo','prsn']]):
            for j,comorb_j in enumerate(row_i):
                axs_phys[i,j].plot(dic_hist_target['comorb']['year'].values,dic_hist_target['comorb'][comorb_j].cumsum(),'-',color=dic_color_palette['blue'][0],lw=1.5,label=dic_comorb_label[comorb_j]+' (STOP HIV/AIDS)') 
                axs_phys[i,j].errorbar(dic_summary['comorb']['year'].values,dic_summary['comorb']['cum '+comorb_j+' med'].values,
                                yerr=[dic_summary['comorb']['cum '+comorb_j+' med'].values-dic_summary['comorb']['cum '+comorb_j+' lb'].values,
                                      dic_summary['comorb']['cum '+comorb_j+' ub'].values-dic_summary['comorb']['cum '+comorb_j+' med'].values],
                                      fmt='o',markersize=4,color=dic_color_palette['blue'][1],ecolor=dic_color_palette['blue'][1],capsize=2.5,label=dic_comorb_label[comorb_j]+' (Model simulations)')
                axs_phys[i,j].text(2007,800*1.05,list_phys_panel[i][j],fontsize=12)
                axs_phys[i,j].set_xlim([2007.8,2016.2])
                axs_phys[i,j].set_ylim([0,800])
                if i==2 and j in [1,2]:
                    axs_phys[i,j].legend(loc='upper right',fontsize=10)
                else:
                    axs_phys[i,j].legend(loc='upper left',fontsize=10)
        fig_phys.text(0.51,0.01,'Year',ha='center',fontsize=12)
        fig_phys.text(0.01,0.5,'Cumulative number of new comorbidity cases',rotation=90,va='center',fontsize=12)
        axs_phys[2,3].set_visible(False)
        fig_phys.subplots_adjust(top=0.95, bottom=0.07, left=0.05,right=0.99,hspace=0.25,wspace=0.25)
        plt.show()
        fig_phys.savefig('my_path/results/fig_cali_both_sim100_'+dic_output['date']+'.png',dpi=300) #Figure S7

    if 'microsim_full_rand_vali' in runs:
        """Full microsimulation model with randomized new diagnosis/migration and randomized characteristics for model validation"""

        #####load initial setting of the parameter set and update the ones with calibrations
        p0 = Params()
        p0.rand_hb = int(input('Input 0/1 indicating whether to consider health behaviours: ')) #0/1 indicating whether to consider the impacts of health behaviours (physical activity, alcohol, smoking) on comorb/mortality incidence
        if p0.rand_hb==0:
            print ('Health behaviours not considered')
        elif p0.rand_hb==1:
            print ('Health behaviours considered')
        else:
            print ('Error: double-check coefficients for health behaviours')
        p = func_params_update(p0)

        #####set up time period for microsimulation
        p.year0 = 2007. 
        p.t0 = pd.Timestamp('2008-01-01') 
        p.t_end = pd.Timestamp('2020-01-01') #extend simulation peirod beyond the one for calibration
        p.dt = 0.5 
        p.n_step = int((p.t_end.year-p.t0.year)/p.dt)
        
        #####load STOP data to build a synthesized population of PLWH with necessary characteristics sampled, the number of diagnosis is the same as STOP
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1') #load individual characteristics from STOP
        dat_ncd4 = pd.read_pickle('data/cd4_nadir_baseline.pkl') #load CD4 nadir data at baseline
        dat_su = pd.read_pickle('data/substance_use.pkl') #load information of substance use for those with age and sex, alive in 2008, at least 18 at diagnosis 
        dat_comorb = pd.read_pickle('data/comorbidities.pkl') #load information of comorbidity diagnosis at comorb_baseline for those with age and sex, alive in 2008, at least 18 at diagnosis 
        list_col_comorb = list(dat_comorb)[-11:] #list of column names for comorbidity incidence date
        for col_i in list_col_comorb:
            dat_comorb[col_i] = dat_comorb[col_i].fillna(p.dummy_end_dt) #replace NaT by dummy_end_dt for comparison purpose
        dat_hiv = pd.read_pickle('data/hiv_status.pkl') #load HIV status to build baseline status
        dat_status_pvl = pd.read_pickle('data/pvl_status.pkl') #load viral suppression status for each 1 year period, needed for baseline characteristics 
        dic_status_reg = pd.read_pickle('data/art_reg_status.pkl') #regimen type by the end of each time step, needed for baseline regimen status if on ART
        dic_prop_reg = pd.read_pickle('data/prop_time_art_reg.pkl') #proportion of time on regimen while on ART, needed for baseline characteristics 
        for reg_i in dic_prop_reg.keys(): #replace reg_prop>1 by 1 for simplification
            list_reg_dates = list(dic_prop_reg[reg_i])[4:]
            for date_j in list_reg_dates:
                dic_prop_reg[reg_i].loc[dic_prop_reg[reg_i][date_j]>1.01,date_j] = 1.

        #####create population of PLWH starting from 2008-01-01 with necessary characteristics for microsimulation
        p.rand_pop = input('Input hist08/hist/random indicating the way to create cohort and characteristics: ') #use random
        dic_raw_data = { #put cohort information together for easy loading in the function
            'cohort':dat_cohort, 'cd4':dat_ncd4,'su':dat_su,'comorb':dat_comorb,'hiv':dat_hiv,'pvl':dat_status_pvl,'reg_status':dic_status_reg,'reg_prop':dic_prop_reg,
        }
        dic_update_data = func_hist_prepare(p,dic_raw_data) 

        #####create targets based on historical data 
        dic_hist_target = func_hist_cali_target(p,dic_update_data)

        #####initialize parameters for multiprocessing
        p.n_pool = int(input('Number of threads used for multiprocessing: '))
        p.num_sim = int(input('Number of microsimulation iterations run: '))
        p.arr_seed = np.arange(p.num_sim) #use ascending order seed sequence from 0 to num

        #####update parameters using the optimized parameter set
        dic_opt_record = {
            'path':'my_path/results/cali/', 
            'name':'lhs_cali',
            'date':'today', 
        }
        file_results = open(dic_opt_record['path']+dic_opt_record['name']+dic_opt_record['date']+'.pkl','rb')
        list_results = pickle.load(file_results)
        list_rand_coeff = list_results[0]
        list_rand_res = list_results[1]
        file_results.close() 
        arr_res_sum = np.array([sum(list(rand_res.values())) for rand_res in list_rand_res])
        ind_res_min = np.argmin(arr_res_sum)
        dic_opt_rand_coeff = copy.deepcopy(list_rand_coeff[ind_res_min])
        for key_i in dic_opt_rand_coeff['comorb'].keys():
            p.dic_prob_comorb_coeff[key_i] = dic_opt_rand_coeff['comorb'][key_i] 
        p.dic_coeff_adj_a2t_pw['bf08'] = np.copy(dic_opt_rand_coeff['new_art']['bf08'])
        p.dic_coeff_adj_a2t_pw['0811'] = np.copy(dic_opt_rand_coeff['new_art']['0811']) 
        p.dic_coeff_adj_a2t_pw['sin12'] = np.copy(dic_opt_rand_coeff['new_art']['sin12']) 
        p.dic_coeff_prob_art = {k:dic_opt_rand_coeff['trans_art'][i] for i,k in enumerate(list(p.dic_coeff_prob_art))}
        p.coeff_prob_dead_ltfu_art = dic_opt_rand_coeff['dead'][0]
        p.coeff_prob_dead_ltfu_noart_pw = np.array([dic_opt_rand_coeff['dead'][1],dic_opt_rand_coeff['dead'][2]]) 
        p.dic_prob_dead_art['pvl_dv'][0] = dic_opt_rand_coeff['dead'][-1] 

        #####run the microsimulation model with the optimized parameter set
        pool = Pool(processes=p.n_pool,maxtasksperchild=1)
        results = [pool.apply_async(mp_microsim_outcomes_sim,args=(p,dic_update_data,p.arr_seed[i])) for i in range(p.num_sim)] 
        list_results = [x.get() for x in results]

        #####derive age distribution among PLWH on ART based on model simulations
        list_hiv_all = [result['hiv'] for result in list_results]
        dat_age_summary = func_micro_outcomes_age_ci(p,list_hiv_all)

        #####derive HIV cascade of care with credible interval
        list_hiv = [result['hiv'][result['hiv']['sex']=='all'].reset_index(drop=True) for result in list_results]
        dat_hiv_summary = func_micro_outcomes_hiv_ci(p,list_hiv) 

        #####initiate path and date for output results
        dic_output = {
            'path':'results/', 
            'date':'today', 
        }
        
        #####save results for validation period
        file_results = open(dic_output['path']+'results_all_valid_sim1000_'+dic_output['date']+'.pkl','wb') #save all results instead of summarized age or hiv-related
        pickle.dump(list_results,file_results)
        file_results.close() 

        #####re-load results for validation
        file_results = open(dic_output['path']+'results_all_valid_sim1000_'+dic_output['date']+'.pkl','rb')
        list_results = pickle.load(file_results) 
        file_results.close() 
        
        #####load hiv information from DTP cohort and compare with microsimulation
        dat_dtp = pd.read_sas('/data/dtp.sas7bdat',encoding='latin-1') 
        dat_dtp = dat_dtp.rename(columns={'PSEUDO':'moh_id'})
        file_data = open('data/dtp_hiv_status.pkl','rb') 
        dat_hiv_dtp = pd.read_pickle(file_data) 
        file_data.close() 
        dat_hiv_dtp = dat_hiv_dtp.merge(dat_dtp[['moh_id','BIRTHDATE']],how='left',on='moh_id') 
        list_dates_dtp = list(dat_hiv_dtp)[3:][::2]
        dat_hist_target_dtp,dat_hist_art_age_dtp = func_hist_cali_target_dtp(p,dat_hiv_dtp)

        #####Figure S9: figure to compare age distribution among PLWH on ART between microsimulation and DTP record
        dic_age_label = {'<30':'<30','30-40':'30-39','40-50':'40-49','50-60':'50-59','60-70':'60-69','>=70':r'$\geq$70'} 
        dic_age_ylim = {'<30':[0,0.1],'30-40':[0.1,0.2],'40-50':[0.2,0.45],'50-60':[0.25,0.4],'60-70':[0.,0.25],'>=70':[0,0.1]}
        list_age_panel = [['A)','B)','C)'],['D)','E)','F)']]
        fig_age,axs_age = plt.subplots(2,3,figsize=(15,8))
        for i,row_i in enumerate([['<30','30-40','40-50'],['50-60','60-70','>=70']]):
            for j,age_j in enumerate(row_i):
                axs_age[i,j].plot(dat_hist_art_age_dtp['year'],dat_hist_art_age_dtp[age_j],'o',color=dic_color_palette['blue'][1],label=dic_age_label[age_j]+' (DTP)') 
                axs_age[i,j].fill_between(dat_age_summary['year'].astype(float),dat_age_summary[age_j+' lb'].astype(float),dat_age_summary[age_j+' ub'].astype(float),facecolor=dic_color_palette['blue'][2],alpha=0.3)
                axs_age[i,j].plot(dat_age_summary['year'],dat_age_summary[age_j+' med'],'-',color=dic_color_palette['blue'][0],label=dic_age_label[age_j]+' (Model simulations)') 
                axs_age[i,j].text(2005,dic_age_ylim[age_j][1]+(dic_age_ylim[age_j][1]-dic_age_ylim[age_j][0])*0.05,list_age_panel[i][j],fontsize=12)
                axs_age[i,j].set_xlim(2006.5,2019.5)
                axs_age[1,j].set_xlabel('Year',fontsize=12)
                axs_age[i,j].set_ylim(dic_age_ylim[age_j])
                if i==0:
                    axs_age[i,j].legend(loc='upper right',fontsize=12,fancybox=True)
                else:
                    axs_age[i,j].legend(loc='lower right',fontsize=12,fancybox=True)
        axs_age[0,0].set_ylabel('Age distribution among PLWH on ART',fontsize=12)
        axs_age[1,0].set_ylabel('Age distribution among PLWH on ART',fontsize=12)
        plt.subplots_adjust(top=0.95, bottom=0.07, left=0.06,right=0.98,hspace=0.25,wspace=0.25)
        plt.show()
        fig_age.savefig(dic_output['path']+'fig_vali_age_art_sim1000_'+dic_output['date']+'.png',dpi=300)

        #####Figure S8: figure to compare HIV cascade between microsimulation and DTP record
        dic_cas_ylabel = {
            'S':'Number of PLWH on ART\nwith viral suppression',
            'art':'Number of PLWH on ART',
            'art_ever':'Number of PLWH ever on ART'}
        list_cas_panel = ['A)','B)','C)']
        fig_cas,axs_cas = plt.subplots(1,3,figsize=(15,4))
        for i,key_i in enumerate(['S','art','art_ever']):
            axs_cas[i].plot(dat_hist_target_dtp['year'],dat_hist_target_dtp[key_i],'o',color=dic_color_palette['blue'][0],markersize=4,label='DTP') 
            axs_cas[i].errorbar(dat_hiv_summary['year'],dat_hiv_summary[key_i+' med'],
                yerr=[dat_hiv_summary[key_i+' med'].values-dat_hiv_summary[key_i+' lb'].values,dat_hiv_summary[key_i+' ub'].values-dat_hiv_summary[key_i+' med'].values],
                fmt='o',color=dic_color_palette['blue'][1],markersize=3,capsize=2.5,label='Model simulations')
            axs_cas[i].text(2005,9000+5500*0.05,list_cas_panel[i],fontsize=12)
            axs_cas[i].set_xlabel('Year',fontsize=12)
            axs_cas[i].set_ylim([3500,9000])
            axs_cas[i].set_ylabel(dic_cas_ylabel[key_i],fontsize=12)
            axs_cas[i].legend(loc='best',fontsize=10,fancybox=True)
        plt.subplots_adjust(top=0.92, bottom=0.12, left=0.07,right=0.98,hspace=0.3,wspace=0.25)
        plt.show()
        fig_cas.savefig(dic_output['path']+'fig_vali_cas_sim1000_'+dic_output['date']+'.png',dpi=300)

        #####load PHAC HIV incidence and prevalence for model validation
        excel_record = pd.ExcelFile('data/phac_inc_prev_uncertain.xlsx')
        excel_sheet = excel_record.sheet_names
        dat_inc = excel_record.parse(excel_sheet[0])
        dat_prev = excel_record.parse(excel_sheet[1])
        dat_pvl = excel_record.parse(excel_sheet[2])
        excel_record.close() 
        dat_all = dat_inc.merge(dat_prev,how='left',on='year').merge(dat_pvl[['year','200']].copy().rename(columns={'200':'pvl_R'}),how='left',on='year') #'200' reprepsents PLWH with pvl<200
        dat_pvl = dic_hist_target['hiv'][['year','S','new_diag']].copy().rename(columns={'S':'spvl'})
        dat_pvl['unsupp_diag'] = pd.Series(dic_hist_target['hiv'][['A','U_on','U_off']].sum(axis=1))
        dat_all = dat_all.merge(dat_pvl,how='left',on='year')
        dat_all['undiag'] = dat_all['prev']-dat_all['spvl']-dat_all['unsupp_diag']

        #####Figure S10: show simulated new infection, and HIV prevalence
        list_results_hiv = [result['hiv'][result['hiv']['sex']=='all'].reset_index(drop=True) for result in list_results] 
        for result_i in list_results_hiv:
            result_i['prev'] = result_i[['A','S','U_on','U_off','undiag']].sum(axis=1)
            result_i['dplwh'] = result_i[['A','S','U_on','U_off']].sum(axis=1)
        dat_sim_inc = func_micro_1outcome_ci(p,list_results_hiv,'new_inc','inc')
        dat_sim_prev = func_micro_1outcome_ci(p,list_results_hiv,'prev','prev') 
        dat_hist = dat_all[dat_all['year'].isin(dat_sim_inc['year'].values)].reset_index(drop=True)
        fig,axs = plt.subplots(1,2,figsize=(12,5))
        axs[0].errorbar(dat_hist['year'],dat_hist['inc'],yerr=[dat_hist['inc'].values-dat_hist['inc_lb'].values,dat_hist['inc_ub'].values-dat_hist['inc'].values],
                        fmt='o',color=dic_color_palette['red'][0],capsize=2.5,label='PHAC')
        axs[0].fill_between(dat_hist['year'].astype(float),dat_sim_inc['new_inc lb'].astype(float),dat_sim_inc['new_inc ub'].astype(float),facecolor=dic_color_palette['red'][2],alpha=0.3,label='Model simulations') 
        axs[0].text(2005,450+250*0.05,'A)',fontsize=14)
        axs[0].set_xlabel('Year',fontsize=14)
        axs[0].set_ylim([50,450])
        axs[0].set_ylabel('New HIV incident cases',fontsize=14)
        axs[0].legend(loc='best',fontsize=12,fancybox=True)
        axs[1].errorbar(dat_hist['year'],dat_hist['prev'],yerr=[dat_hist['prev'].values-dat_hist['prev_lb'].values,
                        dat_hist['prev_ub'].values-dat_hist['prev'].values],fmt='o',color='k',capsize=2.5,label='PHAC')
        axs[1].fill_between(dat_hist['year'].astype(float),dat_sim_prev['prev lb'].astype(float),dat_sim_prev['prev ub'].astype(float),facecolor='darkgray',alpha=0.3,label='Microsimulation')
        axs[1].text(2005,11500+3500*0.05,'B)',fontsize=14)
        axs[1].set_xlabel('Year',fontsize=14)
        axs[1].set_ylim([8000,11500])
        axs[1].set_ylabel('HIV prevalence',fontsize=14)
        axs[1].legend(loc='best',fontsize=12,fancybox=True)
        plt.subplots_adjust(top=0.92, bottom=0.12, left=0.07,right=0.97,hspace=0.3,wspace=0.25)
        plt.show()
        fig.savefig(dic_output['path']+'fig_vali_inc_prev_sim1000_'+dic_output['date']+'.png',dpi=300)

    if 'microsim_rand_sim_results' in runs:
        """run microsimulation using randomized initial population and new diagnosis"""

        #####set up path and version for output files
        dic_output = {
            'path':'my_path/results/', 
            'date':'today', 
        }

        #####load initial setting of the parameter set and update the ones with calibrations
        p0 = Params()
        p0.rand_hb = int(input('Input 0/1 indicating whether to consider health behaviours: ')) #0/1 indicating whether to consider the impacts of health behaviours (physical activity, alcohol, smoking) on comorb/mortality incidence
        if p0.rand_hb==0:
            print ('Health behaviours not considered')
        elif p0.rand_hb==1:
            print ('Health behaviours considered')
        else:
            print ('Error: double-check coefficients for health behaviours')
        p = func_params_update(p0)

        #####set up time period for microsimulation
        p.year0 = 2007. #introduce year information at each time step to use sigmoid function for estimations of new infection, new diagnosis and undiagnosed PLWH
        p.t0 = pd.Timestamp('2008-01-01') #different from study_dt in Params() 
        p.t_end = pd.Timestamp(input('Input the end date of simulation period (e.g., 2020/2035-01-01): '))
        p.dt = 0.5 
        p.n_step = int((p.t_end.year-p.t0.year)/p.dt)
        print ('Total time steps: ', p.n_step)
        
        #####load STOP data to build a synthesized population of PLWH with necessary characteristics sampled, the number of diagnosis is the same as STOP
        dat_cohort = pd.read_sas('data/stop.sas7bdat',encoding='latin-1') #load individual characteristics from STOP
        dat_ncd4 = pd.read_pickle('data/cd4_nadir_baseline.pkl') #load CD4 nadir data at baseline
        dat_su = pd.read_pickle('data/substance_use.pkl') #load information of substance use for those with age and sex, alive in 2008, at least 18 at diagnosis 
        dat_comorb = pd.read_pickle('data/comorbidities.pkl') #load information of comorbidity diagnosis at comorb_baseline for those with age and sex, alive in 2008, at least 18 at diagnosis 
        list_col_comorb = list(dat_comorb)[-11:] #list of column names for comorbidity incidence date
        for col_i in list_col_comorb:
            dat_comorb[col_i] = dat_comorb[col_i].fillna(p.dummy_end_dt) #replace NaT by dummy_end_dt for comparison purpose
        dat_hiv = pd.read_pickle('data/hiv_status.pkl') #load HIV status to build baseline status
        dat_status_pvl = pd.read_pickle('data/pvl_status.pkl') #load viral suppression status for each 1 year period, needed for baseline characteristics 
        dic_status_reg = pd.read_pickle('data/art_reg_status.pkl') #regimen type by the end of each time step, needed for baseline regimen status if on ART
        dic_prop_reg = pd.read_pickle('data/prop_time_art_reg.pkl') #proportion of time on regimen while on ART, needed for baseline characteristics 
        for reg_i in dic_prop_reg.keys(): #replace reg_prop>1 by 1 for simplification
            list_reg_dates = list(dic_prop_reg[reg_i])[4:]
            for date_j in list_reg_dates:
                dic_prop_reg[reg_i].loc[dic_prop_reg[reg_i][date_j]>1.01,date_j] = 1.

        #####create population of PLWH starting from 2008-01-01 with necessary characteristics for microsimulation
        p.rand_pop = input('Input hist08/hist/random indicating the way to create cohort and characteristics: ')
        dic_raw_data = { #put cohort information together for easy loading in the function
            'cohort':dat_cohort, 'cd4':dat_ncd4,'su':dat_su,'comorb':dat_comorb,'hiv':dat_hiv,'pvl':dat_status_pvl,'reg_status':dic_status_reg,'reg_prop':dic_prop_reg,
        }
        dic_update_data = func_hist_prepare(p,dic_raw_data) 

        #####create targets based on historical data 
        dic_hist_target = func_hist_cali_target(p,dic_update_data)

        #####initialize parameters for multiprocessing
        p.n_pool = int(input('Number of threads used for multiprocessing: '))
        p.num_sim = int(input('Number of microsimulation iterations run: '))
        p.arr_seed = np.arange(p.num_sim) #use ascending order seed sequence from 0 to num

        #####update parameters using calibrated values after multiple rounds of calibration
        dic_opt_record = {
            'path':'my_path/calibration/', 
            'name':'cali_',
            'date':'today', 
        }
        file_results = open(dic_opt_record['path']+dic_opt_record['name']+dic_opt_record['date']+'.pkl','rb')
        list_results = pickle.load(file_results)
        list_rand_coeff = list_results[0]
        list_rand_res = list_results[1]
        file_results.close() 
        arr_res_sum = np.array([sum(list(rand_res.values())) for rand_res in list_rand_res])
        ind_res_min = np.argmin(arr_res_sum)
        dic_opt_rand_coeff = copy.deepcopy(list_rand_coeff[ind_res_min]) #use the one with minimal sum of residual
        for key_i in dic_opt_rand_coeff['comorb'].keys():
            p.dic_prob_comorb_coeff[key_i] = dic_opt_rand_coeff['comorb'][key_i] 
        p.dic_coeff_adj_a2t_pw['bf08'] = np.copy(dic_opt_rand_coeff['new_art']['bf08'])
        p.dic_coeff_adj_a2t_pw['0811'] = np.copy(dic_opt_rand_coeff['new_art']['0811']) 
        p.dic_coeff_adj_a2t_pw['sin12'] = np.copy(dic_opt_rand_coeff['new_art']['sin12']) 
        p.dic_coeff_prob_art = {k:dic_opt_rand_coeff['trans_art'][i] for i,k in enumerate(list(p.dic_coeff_prob_art))}
        p.coeff_prob_dead_ltfu_art = dic_opt_rand_coeff['dead'][0]
        p.coeff_prob_dead_ltfu_noart_pw = np.array([dic_opt_rand_coeff['dead'][1],dic_opt_rand_coeff['dead'][2]]) 
        p.dic_prob_dead_art['pvl_dv'][0] = dic_opt_rand_coeff['dead'][-1] 

        #####run microsimulation with modified comorbidity incidence probability for sensitivity analysis
        for comorb_i in p.list_comorb:
            for sce_j in ['plus','minus']:

                ####update probability adjustment for plus/minus scenario for specific comorbidity
                p.dic_prob_comorb_sens_coeff[comorb_i] = p.dic_prob_comorb_sens_coeff_range[sce_j]

                ####run the model with modified parameter values
                pool = Pool(processes=p.n_pool,maxtasksperchild=1)
                results = [pool.apply_async(mp_microsim_outcomes_sim,args=(p,dic_update_data,p.arr_seed[i])) for i in range(p.num_sim)] 
                list_results = [x.get() for x in results]

                ####examine change in comorbidity incidence:
                list_comorb = [result['comorb_inc'] for result in list_results]
                arr_comorb_inc = np.array([data_comorb[comorb_i].sum() for data_comorb in list_comorb])

                ####examine change in comorbidity prevalence by 2034:
                list_comorb = [result['comorb_prev'] for result in list_results]
                year_i = 2034 
                arr_n_all = [data_comorb[['0_b','1_b','2_b','>=3_b']][(data_comorb['year']==year_i)].sum().sum() for data_comorb in list_comorb]
                arr_n_comorb_j = [data_comorb[comorb_i][(data_comorb['year']==year_i)].sum() for data_comorb in list_comorb]

                ####examine change in physical multimorbidity by 2034:
                year_i = 2034 
                arr_n_all = [data_comorb[['0_b','1_b','2_b','>=3_b']][(data_comorb['year']==year_i)].sum().sum() for data_comorb in list_comorb]
                arr_n_comorb_j = [data_comorb[['2_p','>=3_p']][(data_comorb['year']==year_i)].sum().sum() for data_comorb in list_comorb]

                ####examine change in >=3 comorbidities (both) by 2034:
                year_i = 2034 
                arr_n_all = [data_comorb[['0_b','1_b','2_b','>=3_b']][(data_comorb['year']==year_i)].sum().sum() for data_comorb in list_comorb]
                arr_n_comorb_j = [data_comorb['>=3_b'][(data_comorb['year']==year_i)].sum() for data_comorb in list_comorb]

                ####output projected outcomes for sensititivy analysis
                file_results = open(dic_output['path']+'microsim_comorb_sens_'+comorb_i+'_'+sce_j+'_'+str(p.num_sim)+'_'+dic_output['date']+'.pkl','wb') 
                pickle.dump(list_results,file_results) 
                file_results.close()

            ####turn the adjustment back to status quo for different comorbidity scenario
            p.dic_prob_comorb_sens_coeff[comorb_i] = 1.

        #####main projection: multi-processing to run microsimulation with randomized new diagnosis/migration, return to integrated outcomes at population level
        pool = Pool(processes=p.n_pool,maxtasksperchild=1)
        results = [pool.apply_async(mp_microsim_outcomes_sim,args=(p,dic_update_data,p.arr_seed[i])) for i in range(p.num_sim)] #microsimulation beyond the calibration period with derived final outcomes
        list_results = [x.get() for x in results]

        #####output aggregated outcomes of microsimulation as .pkl file
        file_results = open(dic_output['path']+'microsim_baseline2034_cc_s1000_'+dic_output['date']+'.pkl','wb') #only save additional results of ranking of combination of two comorbidities to save space
        #file_results = open(dic_output['path']+'microsim_baseline2034_s1000_'+dic_output['date']+'.pkl','wb')
        pickle.dump(list_results,file_results) 
        file_results.close()

    if 'figure_dplwh_age_comorb' in runs:
        """generate plots for diagnosed PLWH, age-stratified multimorbidity and comorbidity prevalence for the manuscript"""

        #####set up path and version for output files
        dic_output = {
            'path':'my_path/results/', 
            'date':'today',
        }

        #####load results of the microsimulation model
        file_name = dic_output['path']+'microsim_baseline2034_s1000_today.pkl' #simulated aggregated HIV cascade of care and comorbidity prevalence over time for 1000 simulations
        file_results = open(file_name,'rb')
        list_results = pickle.load(file_results)
        file_results.close()

        #####Figure 2: derive age distribution among diagnosed PLWH and show as bar plots
        dic_age_cat = {'<30':[0,30],'30-40':[30,40],'40-50':[40,50],'50-60':[50,60],'60-70':[60,70],'>=70':[70,200]}
        list_hiv = [result['hiv'] for result in list_results]
        dat_age_summary = pd.DataFrame([],columns=['year'])
        for year_i in list_hiv[0]['year'].unique():
            dic_age_summary = {'year':year_i}
            arr_n_all = [data_hiv[['A','S','U_on','U_off']][(data_hiv['year']==year_i)&(data_hiv['sex']=='all')].sum().sum() for data_hiv in list_hiv]
            print (year_i,np.percentile(arr_n_all,[50,2.5,97.5]))
            for age_j in dic_age_cat.keys():
                arr_n_age_j = [data_hiv[['A','S','U_on','U_off']][(data_hiv['year']==year_i)&(data_hiv['sex']!='all')&(data_hiv['age']==age_j)].sum().sum() for data_hiv in list_hiv]
                dic_age_summary['n '+age_j],dic_age_summary['n '+age_j+' lb'],dic_age_summary['n '+age_j+' ub'] = np.percentile(np.array(arr_n_age_j),[50,2.5,97.5])
                dic_age_summary['p '+age_j],dic_age_summary['p '+age_j+' lb'],dic_age_summary['p '+age_j+' ub'] = np.percentile(np.array(arr_n_age_j)/np.array(arr_n_all),[50,2.5,97.5])
            dat_age_summary = pd.concat([dat_age_summary,pd.DataFrame(dic_age_summary,index=[0])],ignore_index=True)
        dic_age_color = {
            '<30':dic_color_palette['blue'][0],'30-40':dic_color_palette['blue'][1],'40-50':dic_color_palette['blue'][2],
            '50-60':dic_color_palette['green'][2],'60-70':dic_color_palette['green'][1],'>=70':dic_color_palette['green'][0],
        }
        dic_age_label = {'<30':r'$<30$','30-40':r'$30-39$','40-50':r'$40-49$','50-60':r'$50-59$','60-70':r'$60-69$','>=70':r'$\geq$70'}
        fig_age,axs_age = plt.subplots(1,2,figsize=(12,4.5))
        arr_bottom0 = 0.
        arr_p_age0 = arr_p_age1 = 0.
        for age_i in dic_age_cat.keys():
            axs_age[0].bar(dat_age_summary['year'].values,dat_age_summary['n '+age_i].values,bottom=arr_bottom0,color=dic_age_color[age_i])
            arr_bottom0 += dat_age_summary['n '+age_i].values
            arr_p_age1 += dat_age_summary['p '+age_i].values
            axs_age[1].fill_between(dat_age_summary['year'].astype(float),arr_p_age0,arr_p_age1,facecolor=dic_age_color[age_i],label=dic_age_label[age_i])
            arr_p_age0 += dat_age_summary['p '+age_i].values
        axs_age[0].text(2001,10500,'A',fontsize=14)
        axs_age[1].text(2001,1.05,'B',fontsize=14)
        axs_age[0].set_ylim([0,10000])
        axs_age[0].set_xlabel('Year',fontsize=12)
        axs_age[0].set_ylabel('Number of diagnosed PLWH',fontsize=12)
        axs_age[1].set_ylim([0,1])
        vals = axs_age[1].get_yticks()
        axs_age[1].set_yticklabels(['{:3.0f}%'.format(x*100) for x in vals])
        axs_age[1].set_xlabel('Year',fontsize=12)
        axs_age[1].set_ylabel('Age distribution of diagnosed PLWH',fontsize=12)
        axs_age[1].legend(loc='upper left',fontsize=10,fancybox=True)
        plt.subplots_adjust(top=0.91, bottom=0.12, left=0.08,right=0.98,hspace=0.15,wspace=0.27)
        plt.show()
        fig_age.savefig(dic_output['path']+'fig_age_dplwh_sim2034_'+dic_output['date']+'.png',dpi=600)

        #####output .xlsx file to save numbers in the plot
        wb = Workbook()  
        ws = wb.active
        ws.cell(row=1,column=1).value = 'Year'
        for j,type_j in enumerate(['count','distribution']):
            ws.cell(row=1,column=2+j*6).value = type_j
            for k,age_k in enumerate(list(dic_age_label)):
                ws.cell(row=2,column=2+j*6+k).value = age_k
                if type_j=='distribution':
                    type_m = 'p '+age_k
                else:
                    type_m = 'n '+age_k
                for m,year_m in enumerate(dat_age_summary['year'].values):
                    ws.cell(row=3+m*2,column=1).value = year_m
                    v_m,v_lb,v_ub = dat_age_summary[[type_m,type_m+' lb',type_m+' ub']][(dat_age_summary['year']==year_m)].values[0]
                    if type_j=='count':
                        ws.cell(row=3+m*2,column=2+j*6+k).value = '{:.0f}'.format(v_m)
                        ws.cell(row=3+m*2+1,column=2+j*6+k).value = '({:.0f}, {:.0f})'.format(v_lb,v_ub)
                    else:
                        ws.cell(row=3+m*2,column=2+j*6+k).value = '{:.2f}'.format(round(v_m*100,2))
                        ws.cell(row=3+m*2+1,column=2+j*6+k).value = '({:.2f}, {:.2f})'.format(round(v_lb*100,2),round(v_ub*100,2))
        wb.save(dic_output['path']+'record_paper_fig_age_dplwh_sim2034_'+dic_output['date']+'.xlsx')

        #####Figure 3: derive comorbidity burden (prevalence) among diagnosed PLWH and show as bar plots
        list_comorb = [result['comorb_prev'] for result in list_results]
        dat_ncomorb_summary = pd.DataFrame([],columns=['year'])
        for year_i in list_comorb[0]['year'].unique():
            dic_ncomorb_summary = {'year':year_i}
            arr_n_all = [data_comorb[['0_b','1_b','2_b','>=3_b']][(data_comorb['year']==year_i)].sum().sum() for data_comorb in list_comorb]
            for comorb_j in ['0_b','1_b','2_b','>=3_b','0_p','1_p','2_p','>=3_p','0_m','1_m','2_m','3_m']:
                arr_n_comorb_j = [data_comorb[comorb_j][(data_comorb['year']==year_i)].sum() for data_comorb in list_comorb]
                dic_ncomorb_summary['n '+comorb_j],dic_ncomorb_summary['n '+comorb_j+' lb'],dic_ncomorb_summary['n '+comorb_j+' ub'] = np.percentile(np.array(arr_n_comorb_j),[50,2.5,97.5]) 
                dic_ncomorb_summary['p '+comorb_j],dic_ncomorb_summary['p '+comorb_j+' lb'],dic_ncomorb_summary['p '+comorb_j+' ub'] = np.percentile(np.array(arr_n_comorb_j)/np.array(arr_n_all),[50,2.5,97.5]) 
            dat_ncomorb_summary = pd.concat([dat_ncomorb_summary,pd.DataFrame(dic_ncomorb_summary,index=[0])],ignore_index=True)
        dic_comorb_char = {
            'key':{'both':['0_b','1_b','2_b','>=3_b'],'phys':['0_p','1_p','2_p','>=3_p'],'mental':['0_m','1_m','2_m','3_m']},
            'color':[list_color_add[-1],dic_color_palette['purple'][2],dic_color_palette['purple'][1],dic_color_palette['purple'][0]], 
            'legend':{
                'both':[r'$0$'+' physical/mental comorbidities',r'$1$'+' physical/mental comorbidity',r'$2$'+' physical/mental comorbidities',r'$\geq3$'+' physical/mental comorbidities'],
                'phys':[r'$0$'+' physical comorbidities',r'$1$'+' physical comorbidity',r'$2$'+' physical comorbidities',r'$\geq3$'+' physical comorbidities'],
                'mental':[r'$0$'+' mental comorbidities',r'$1$'+' mental comorbidity',r'$2$'+' mental comorbidities',r'$\geq3$'+' mental comorbidities']},
        }
        list_panel_label = [['A','B','C'],['D','E','F']]
        fig_com,axs_com = plt.subplots(2,3,figsize=(16,8))
        for i,key_i in enumerate(['both','phys','mental']):
            arr_bottom = 0.
            arr_p_comorb0 = arr_p_comorb1 = 0.
            for j,comorb_j in enumerate(dic_comorb_char['key'][key_i]):
                axs_com[0,i].bar(dat_ncomorb_summary['year'].values,dat_ncomorb_summary['n '+comorb_j].values,bottom=arr_bottom,color=dic_comorb_char['color'][j],label=dic_comorb_char['legend'][key_i][j])
                arr_bottom += dat_ncomorb_summary['n '+comorb_j].values
                arr_p_comorb1 += dat_ncomorb_summary['p '+comorb_j].values
                axs_com[1,i].fill_between(dat_ncomorb_summary['year'].astype(float),arr_p_comorb0,arr_p_comorb1,facecolor=dic_comorb_char['color'][j],label=dic_comorb_char['legend'][key_i][j])
                arr_p_comorb0 += dat_ncomorb_summary['p '+comorb_j].values
            axs_com[0,i].text(1999,10000*1.05,list_panel_label[0][i],fontsize=14)
            axs_com[1,i].text(1999,1.05,list_panel_label[1][i],fontsize=14)
            axs_com[0,i].set_ylim([0,10000])
            axs_com[0,i].set_ylabel('Number of diagnosed PLWH',fontsize=12)
            axs_com[1,i].set_ylim([0,1])
            vals = axs_com[1,i].get_yticks()
            axs_com[1,i].set_yticklabels(['{:3.0f}%'.format(x*100) for x in vals])
            axs_com[1,i].set_xlabel('Year',fontsize=12)
            axs_com[1,i].legend(loc='lower left',fontsize=8,fancybox=True)
            axs_com[0,i].legend(loc='lower left',fontsize=8,fancybox=True)
        axs_com[1,0].set_ylabel('Distribution of multimorbidity\n(physical/mental) among diagnosed PLWH',fontsize=12)
        axs_com[1,1].set_ylabel('Distribution of multimorbidity (physical)\namong diagnosed PLWH',fontsize=12)
        axs_com[1,2].set_ylabel('Distribution of multimorbidity (mental)\namong diagnosed PLWH',fontsize=12)
        plt.subplots_adjust(top=0.95, bottom=0.07, left=0.07,right=0.98,hspace=0.2,wspace=0.3)
        plt.show()
        fig_com.savefig(dic_output['path']+'fig_multicomorb_dplwh_sim2034_'+dic_output['date']+'.png',dpi=600)

        #####output .xlsx file to save numbers in the plot
        dic_key_multicomorb = {
            'both':['0_b','1_b','2_b','>=3_b'],
            'physical':['0_p','1_p','2_p','>=3_p'],
            'mental':['0_m','1_m','2_m','3_m'],
        }
        wb = Workbook()  
        ws = wb.active
        ws.cell(row=1,column=1).value = 'Year'
        for i,key_i in enumerate(dic_key_multicomorb.keys()):
            ws.cell(row=1,column=2+i*8).value = key_i
            for j,type_j in enumerate(['count','proportion']):
                ws.cell(row=2,column=2+i*8+j*4).value = type_j
                for k,comorb_k in enumerate(dic_key_multicomorb[key_i]):
                    ws.cell(row=3,column=2+i*8+j*4+k).value = comorb_k
                    if type_j=='proportion':
                        type_m = 'p '+comorb_k
                    else:
                        type_m = 'n '+comorb_k
                    for m,year_m in enumerate(dat_age_summary['year'].values):
                        ws.cell(row=4+m*2,column=1).value = year_m
                        v_m,v_lb,v_ub = dat_ncomorb_summary[[type_m,type_m+' lb',type_m+' ub']][(dat_ncomorb_summary['year']==year_m)].values[0]
                        if type_j=='count':
                            ws.cell(row=4+m*2,column=2+i*8+j*4+k).value = '{:.0f}'.format(v_m)
                            ws.cell(row=4+m*2+1,column=2+i*8+j*4+k).value = '({:.0f}, {:.0f})'.format(v_lb,v_ub)
                        else:
                            ws.cell(row=4+m*2,column=2+i*8+j*4+k).value = '{:.2f}'.format(round(v_m*100,2))
                            ws.cell(row=4+m*2+1,column=2+i*8+j*4+k).value = '({:.2f}, {:.2f})'.format(round(v_lb*100,2),round(v_ub*100,2))
        wb.save(dic_output['path']+'record_paper_fig_multicomorb_dplwh_sim2034_'+dic_output['date']+'.xlsx')

        #####Figure 4: derive single comorbidity prevalence (physical/mental) among diagnosed PLWH and show as one plot
        dic_comorb_label = {
            'htn':'Hypertension','cvd':'Cardiovascular disease','dm':'Diabetes','ckd':'Chronic kidney disease',
            'cld':'Chronic liver disease','copd':'Chronic obstructive\npulmonary disease','oa':'Osteoarthritis',
            'cancer':'Non-AIDS-related cancers','manx':'Mood and anxiety disorders','sczo':'Schizophrenia','prsn':'Personality disorders'} 
        list_comorb = [result['comorb_prev'] for result in list_results]
        dat_comorb_summary = pd.DataFrame([],columns=['year'])
        for year_i in list_comorb[0]['year'].unique():
            dic_comorb_summary = {'year':year_i}
            arr_n_all = [data_comorb[['0_b','1_b','2_b','>=3_b']][(data_comorb['year']==year_i)].sum().sum() for data_comorb in list_comorb]
            for comorb_j in dic_comorb_label.keys():
                arr_n_comorb_j = [data_comorb[comorb_j][(data_comorb['year']==year_i)].sum() for data_comorb in list_comorb]
                dic_comorb_summary['n '+comorb_j],dic_comorb_summary['n '+comorb_j+' lb'],dic_comorb_summary['n '+comorb_j+' ub'] = np.percentile(np.array(arr_n_comorb_j),[50,2.5,97.5]) 
                dic_comorb_summary['p '+comorb_j],dic_comorb_summary['p '+comorb_j+' lb'],dic_comorb_summary['p '+comorb_j+' ub'] = np.percentile(np.array(arr_n_comorb_j)/np.array(arr_n_all),[50,2.5,97.5]) 
            dat_comorb_summary = pd.concat([dat_comorb_summary,pd.DataFrame(dic_comorb_summary,index=[0])],ignore_index=True)
        list_panel_label = [['A','B','C','D'],['E','F','G','H'],['I','J','K']]
        fig_both,axs_both = plt.subplots(3,4,figsize=(15,10))
        for i,row_i in enumerate([['htn','cvd','dm','ckd'],['cld','copd','oa','cancer'],['manx','sczo','prsn']]):
            for j,key_j in enumerate(row_i):
                arr_val = dat_comorb_summary['p '+key_j].values
                arr_lb = dat_comorb_summary['p '+key_j+' lb'].values
                arr_ub = dat_comorb_summary['p '+key_j+' ub'].values
                axs_both[i,j].errorbar(dat_comorb_summary['year'].values,arr_val, yerr=[arr_val-arr_lb,arr_ub-arr_val],fmt='o',
                                      color=dic_color_palette['red'][1],ecolor=dic_color_palette['red'][1],markersize=4,elinewidth=1.5,
                                      capsize=2.5,label=dic_comorb_label[key_j])
                if i<2:
                    axs_both[i,j].set_ylim([0.,0.2])
                    axs_both[i,j].set_yticks(np.arange(0,0.24,0.04))
                    axs_both[i,j].text(2001.5,0.2*1.05,list_panel_label[i][j],fontsize=14)
                else:
                    axs_both[i,j].set_ylim([0.,0.55])
                    axs_both[i,j].text(2001.5,0.55*1.05,list_panel_label[i][j],fontsize=14)
                vals = axs_both[i,j].get_yticks()
                axs_both[i,j].set_yticklabels(['{:3.0f}%'.format(x*100) for x in vals])
                if key_j in ['sczo','prsn']:
                    axs_both[i,j].legend(loc='upper right',fontsize=10)
                else:
                    axs_both[i,j].legend(loc='lower right',fontsize=10)
        fig_both.text(0.51,0.01,'Year',ha='center',fontsize=12)
        axs_both[1,0].set_ylabel('Prevalence',fontsize=12)
        axs_both[2,3].set_visible(False)
        plt.subplots_adjust(top=0.95, bottom=0.06, left=0.06,right=0.98,hspace=0.25,wspace=0.25)
        plt.show()
        fig_both.savefig(dic_output['path']+'fig_comorb_prev_dplwh_sim2034_'+dic_output['date']+'.png',dpi=600)

        #####output .xlsx file to save numbers in the plot
        wb = Workbook()  
        ws = wb.active
        ws.cell(row=1,column=1).value = 'Year'
        for j,comorb_j in enumerate(dic_comorb_label.keys()):
            ws.cell(row=1,column=2+j).value = dic_comorb_label[comorb_j]
            type_m = 'p '+comorb_j
            for m,year_m in enumerate(dat_comorb_summary['year'].values):
                ws.cell(row=2+m*2,column=1).value = year_m
                v_m,v_lb,v_ub = dat_comorb_summary[[type_m,type_m+' lb',type_m+' ub']][(dat_comorb_summary['year']==year_m)].values[0]
                ws.cell(row=2+m*2,column=2+j).value = '{:.2f}'.format(round(v_m*100,2))
                ws.cell(row=2+m*2+1,column=2+j).value = '({:.2f}, {:.2f})'.format(round(v_lb*100,2),round(v_ub*100,2))
        wb.save(dic_output['path']+'record_paper_fig_comorb_prev_dplwh_sim2034_'+dic_output['date']+'.xlsx')

        #####Figure 5: derive comorbidity burden stratified by age among diagnosed PLWH and show as bar plots
        list_comorb = [result['comorb_prev'] for result in list_results]
        dat_comage_summary = pd.DataFrame([],columns=['year','age'])
        for year_i in list_comorb[0]['year'].unique():
            for age_j in list(dic_age_cat):
                dic_comage_summary = {'year':year_i,'age':age_j}
                for comorb_k in ['0_b','1_b','2_b','>=3_b','0_p','1_p','2_p','>=3_p','0_m','1_m','2_m','3_m']:
                    arr_n_comorb_ij = [data_comorb[comorb_k][(data_comorb['year']==year_i)&(data_comorb['age']==age_j)].sum() for data_comorb in list_comorb]
                    dic_comage_summary['n '+comorb_k],dic_comage_summary['n '+comorb_k+' lb'],dic_comage_summary['n '+comorb_k+' ub'] = np.percentile(np.array(arr_n_comorb_ij),[50,2.5,97.5]) 
                dat_comage_summary = pd.concat([dat_comage_summary,pd.DataFrame(dic_comage_summary,index=[0])],ignore_index=True)
        dic_comage_char = {
            'panel':[['A','B','C'],['D','E','F']],
            'label':[['<30','30-40','40-50'],['50-60','60-70','>=70']],
            'age_label':{'<30':r'$<30$','30-40':r'$30-39$','40-50':r'$40-49$','50-60':r'$50-59$','60-70':r'$60-69$','>=70':r'$\geq70$'},
            'key':['0_p','1_p','2_p','>=3_p'],
            'legend':[r'$0$'+' physical comorbidities',r'$1$'+' physical comorbidity',r'$2$'+' physical comorbidities',r'$\geq3$'+' physical comorbidities'],
            'color':[list_color_add[-1],dic_color_palette['purple'][2],dic_color_palette['purple'][1],dic_color_palette['purple'][0]], 
        }
        fig_comage,axs_comage = plt.subplots(2,3,figsize=(16,8))
        for i,row_i in enumerate(dic_comage_char['label']):
            for j,age_j in enumerate(row_i):
                arr_bottom = 0.
                for k,key_k in enumerate(dic_comage_char['key']):
                    axs_comage[i,j].bar(dat_comage_summary['year'].unique(),dat_comage_summary['n '+key_k][(dat_comage_summary['age']==age_j)].values,bottom=arr_bottom,color=dic_comage_char['color'][k],label=dic_comage_char['legend'][k])
                    arr_bottom += dat_comage_summary['n '+key_k][dat_comage_summary['age']==age_j].values
                axs_comage[i,j].text(1998,3200*1.05,dic_comage_char['panel'][i][j],fontsize=14)
                axs_comage[i,j].set_ylim([0,3200])
                axs_comage[i,j].set_ylabel('Number of diagnosed PLWH\n(aged '+dic_comage_char['age_label'][age_j]+')',fontsize=12)
                axs_comage[1,j].set_xlabel('Year',fontsize=12)
        axs_comage[0,0].legend(loc='upper left',fontsize=10,fancybox=True)
        plt.subplots_adjust(top=0.95, bottom=0.07, left=0.07,right=0.98,hspace=0.2,wspace=0.35)
        plt.show()
        fig_comage.savefig(dic_output['path']+'fig_comage_dplwh_sim2034_'+dic_output['date']+'.png',dpi=600)
        ####Figure in SI:add age-specific trends in mental multimorbidity among diagnosed PLWH
        dic_comage_char['panel'] = [['A)','B)','C)'],['D)','E)','F)']]
        dic_comage_char['key'] = ['0_m','1_m','2_m','3_m']
        dic_comage_char['legend'] = [r'$0$'+' mental comorbidities',r'$1$'+' mental comorbidity',r'$2$'+' mental comorbidities',r'$3$'+' mental comorbidities']
        fig_comage_m,axs_comage_m = plt.subplots(2,3,figsize=(16,8))
        for i,row_i in enumerate(dic_comage_char['label']):
            for j,age_j in enumerate(row_i):
                arr_bottom = 0.
                for k,key_k in enumerate(dic_comage_char['key']):
                    axs_comage_m[i,j].bar(dat_comage_summary['year'].unique(),dat_comage_summary['n '+key_k][(dat_comage_summary['age']==age_j)].values,bottom=arr_bottom,color=dic_comage_char['color'][k],label=dic_comage_char['legend'][k])
                    arr_bottom += dat_comage_summary['n '+key_k][dat_comage_summary['age']==age_j].values
                axs_comage_m[i,j].text(1998,3200*1.05,dic_comage_char['panel'][i][j],fontsize=14)
                axs_comage_m[i,j].set_ylim([0,3200])
                axs_comage_m[i,j].set_ylabel('Number of diagnosed PLWH\n(aged '+dic_comage_char['age_label'][age_j]+')',fontsize=12)
                axs_comage_m[1,j].set_xlabel('Year',fontsize=12)
        axs_comage_m[0,0].legend(loc='upper left',fontsize=10,fancybox=True)
        plt.subplots_adjust(top=0.95, bottom=0.07, left=0.07,right=0.98,hspace=0.2,wspace=0.35)
        plt.show()
        fig_comage_m.savefig(dic_output['path']+'fig_comage_m_dplwh_sim2034_'+dic_output['date']+'.png',dpi=300)

        #####output .xlsx file to save numbers in the plot
        dic_key_multicomorb = {
            'both':['0_b','1_b','2_b','>=3_b'],
            'physical':['0_p','1_p','2_p','>=3_p'],
            'mental':['0_m','1_m','2_m','3_m'],
        }
        wb = Workbook()  
        ws = wb.active
        ws.cell(row=1,column=1).value = 'Year'
        for i,age_i in enumerate(dic_age_label.keys()):
            ws.cell(row=1,column=2+i*12).value = age_i
            for j,type_j in enumerate(dic_key_multicomorb.keys()):
                ws.cell(row=2,column=2+i*12+j*4).value = type_j
                for k,comorb_k in enumerate(dic_key_multicomorb[type_j]):
                    ws.cell(row=3,column=2+i*12+j*4+k).value = comorb_k
                    type_m = 'n '+comorb_k
                    for m,year_m in enumerate(dat_comage_summary['year'].unique()):
                        ws.cell(row=4+m*2,column=1).value = year_m
                        v_m,v_lb,v_ub = dat_comage_summary[[type_m,type_m+' lb',type_m+' ub']][(dat_comage_summary['year']==year_m)&(dat_comage_summary['age']==age_i)].values[0]
                        ws.cell(row=4+m*2,column=2+i*12+j*4+k).value = '{:.0f}'.format(np.round(v_m,0))
                        ws.cell(row=4+m*2+1,column=2+i*12+j*4+k).value = '({:.0f}, {:.0f})'.format(np.round(v_lb,0),np.round(v_ub))
        wb.save(dic_output['path']+'record_paper_fig_comage_dplwh_sim2034_'+dic_output['date']+'.xlsx')

        #####convert .png file to jpg file
        fig_path = dic_output['path']+figure_name+dic_output['date']
        im = Image.open(fig_path+'.png')
        im.load()
        background = Image.new("RGB", im.size, (255, 255, 255))
        background.paste(im, mask=im.split()[3]) 
        background.save(fig_path+'.jpg', 'JPEG', quality=100)

    if 'figure_ranking_comorbidities' in runs:
        """generate plots/tables showing the ranking of comorbidities for diagnosed PLWH, stratified by age/sex"""

        #####initiate output path and version by date
        dic_output = {
            'path':'my_path/results/', 
            'date':'today',
        }

        #####load results from microsimulations
        file_name = 'microsim_baseline2034_s1000_today.pkl'
        file_results = open(dic_output['path']+file_name,'rb') 
        list_results = pickle.load(file_results)
        file_results.close()

        list_comorb = ['cvd','htn','dm','oa','copd','ckd','cld','cancer','manx','sczo','prsn']
        list_comorb_phys = ['cvd','htn','dm','oa','copd','ckd','cld','cancer'] 

        #####derive ranking of comorbidities overall and by sex and age over time
        list_comorb_all = [result['comb1_both'] for result in list_results]
        list_comorb_agesex = [result['comb1_both_agesex'] for result in list_results]
        dat_agesex_summary = pd.DataFrame([],columns=['year']) #for each year, determine rank of each comorbidity over 1000 simulations and prevalence range
        for year_i in [2024,2034]: 
            dat_agesex_summary_i = pd.DataFrame([],columns=['year'])
            for comorb_j in list_comorb:
                if comorb_j not in ['manx','sczo','prsn']:
                    arr_c_rankp = np.array([data['rank_phys'][(data['year']==year_i)&(data['c']==comorb_j)].values[0] for data in list_comorb_all])
                arr_c_rankb = np.array([data['rank_both'][(data['year']==year_i)&(data['c']==comorb_j)].values[0] for data in list_comorb_all])
                arr_c_prev = np.array([data['n_c'][(data['year']==year_i)&(data['c']==comorb_j)].values[0]/data['n_dplwh'][(data['year']==year_i)&(data['c']==comorb_j)].values[0] for data in list_comorb_all])
                prev,prev_lb,prev_ub = np.percentile(arr_c_prev,[50,2.5,97.5])
                if comorb_j not in ['manx','sczo','prsn']:
                    dic_summary = {
                        'year':year_i,'sex':'all','age':'all','c':comorb_j,
                        'rankb':sum(arr_c_rankb)/1000,'rankp':sum(arr_c_rankp)/1000,
                        'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                    }
                else:
                    dic_summary = {
                        'year':year_i,'sex':'all','age':'all','c':comorb_j,'rankb':sum(arr_c_rankb)/1000,
                        'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                    }
                dat_agesex_summary_i = pd.concat([dat_agesex_summary_i,pd.DataFrame(dic_summary,index=[0])],ignore_index=True)
            dat_agesex_summary_i['rank_both'] = dat_agesex_summary_i['rankb'].rank(method='min') #min has highest rank in the group
            dat_agesex_summary_i['rank_phys'] = dat_agesex_summary_i['rankp'].rank(method='min')
            dat_agesex_summary = pd.concat([dat_agesex_summary,dat_agesex_summary_i],ignore_index=True)
            for sex_k in ['M','F']:
                for age_m in ['<50','50-60','60-70','>=70']:
                    dat_agesex_summary_i = pd.DataFrame([],columns=['year'])
                    for comorb_j in list_comorb:
                        if comorb_j not in ['manx','sczo','prsn']:
                            arr_c_rankp = np.array([data['rank_phys'][(data['year']==year_i)&(data['sex']==sex_k)&(data['age']==age_m)&(data['c']==comorb_j)].values[0] for data in list_comorb_agesex])
                        arr_c_rankb = np.array([data['rank_both'][(data['year']==year_i)&(data['sex']==sex_k)&(data['age']==age_m)&(data['c']==comorb_j)].values[0] for data in list_comorb_agesex])
                        arr_c_prev = np.array([data['n_c'][(data['year']==year_i)&(data['sex']==sex_k)&(data['age']==age_m)&(data['c']==comorb_j)].values[0]/data['n_dplwh'][(data['year']==year_i)&(data['sex']==sex_k)&(data['age']==age_m)&(data['c']==comorb_j)].values[0] for data in list_comorb_agesex])
                        prev,prev_lb,prev_ub = np.percentile(arr_c_prev,[50,2.5,97.5])
                        if comorb_j not in ['manx','sczo','prsn']:
                            dic_summary = {
                                'year':year_i,'sex':sex_k,'age':age_m,'c':comorb_j,
                                'rankb':sum(arr_c_rankb)/1000,'rankp':sum(arr_c_rankp)/1000,
                                'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                            }
                        else:
                            dic_summary = {
                                'year':year_i,'sex':sex_k,'age':age_m,'c':comorb_j,'rankb':sum(arr_c_rankb)/1000,
                                'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                            }
                        dat_agesex_summary_i = pd.concat([dat_agesex_summary_i,pd.DataFrame(dic_summary,index=[0])],ignore_index=True)
                    dat_agesex_summary_i['rank_both'] = dat_agesex_summary_i['rankb'].rank(method='min') #min has highest rank in the group
                    dat_agesex_summary_i['rank_phys'] = dat_agesex_summary_i['rankp'].rank(method='min')
                    dat_agesex_summary = pd.concat([dat_agesex_summary,dat_agesex_summary_i],ignore_index=True)

        #####derive ranking of comorbidities by age or sex
        dat_sex_summary = pd.DataFrame([],columns=['year'])
        for year_i in [2024,2034]: 
            arr_data = [data[data['year']==year_i].groupby(by=['sex','c'])['n_c'].sum().reset_index() for data in list_comorb_agesex]
            arr_n = [data[data['year']==year_i].groupby(by=['sex','c'])['n_dplwh'].sum().reset_index() for data in list_comorb_agesex]
            for sex_k in ['M','F']:
                arr_data_k = [data[data['sex']==sex_k].reset_index(drop=True) for data in arr_data]
                arr_data_k_mod = []
                for data_m in arr_data_k:
                    data_m['rank_both'] = data_m['n_c'].rank(method='min',ascending=False)
                    data_m_phys = data_m[data_m['c'].isin(['manx','prsn','sczo'])==False].reset_index(drop=True)
                    data_m_phys['rank_phys'] = data_m_phys['n_c'].rank(method='min',ascending=False)
                    data_m = data_m.merge(data_m_phys[['c','rank_phys']],how='left',on='c')
                    arr_data_k_mod.append(data_m)
                dat_sex_summary_ik = pd.DataFrame([],columns=['year'])
                for comorb_j in list_comorb:
                    if comorb_j not in ['manx','sczo','prsn']:
                        arr_c_rankp = np.array([data['rank_phys'][(data['c']==comorb_j)].values[0] for data in arr_data_k_mod])
                    arr_c_rankb = np.array([data['rank_both'][(data['c']==comorb_j)].values[0] for data in arr_data_k_mod])
                    arr_c_prev = np.array([data['n_c'][(data['c']==comorb_j)].values[0]/data_n['n_dplwh'][(data_n['sex']==sex_k)&(data_n['c']==comorb_j)].values[0] for data,data_n in zip(arr_data_k_mod,arr_n)])
                    prev,prev_lb,prev_ub = np.percentile(arr_c_prev,[50,2.5,97.5])
                    if comorb_j not in ['manx','sczo','prsn']:
                        dic_summary = {
                            'year':year_i,'sex':sex_k,'c':comorb_j,
                            'rankb':sum(arr_c_rankb)/1000,'rankp':sum(arr_c_rankp)/1000,
                            'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                        }
                    else:
                        dic_summary = {
                            'year':year_i,'sex':sex_k,'c':comorb_j,'rankb':sum(arr_c_rankb)/1000,
                            'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                        }
                    dat_sex_summary_ik = pd.concat([dat_sex_summary_ik,pd.DataFrame(dic_summary,index=[0])],ignore_index=True)
                dat_sex_summary_ik['rank_both'] = dat_sex_summary_ik['rankb'].rank(method='min') #min has highest rank in the group
                dat_sex_summary_ik['rank_phys'] = dat_sex_summary_ik['rankp'].rank(method='min')
                dat_sex_summary = pd.concat([dat_sex_summary,dat_sex_summary_ik],ignore_index=True)
        dat_age_summary = pd.DataFrame([],columns=['year'])
        for year_i in [2024,2034]: 
            arr_data = [data[data['year']==year_i].groupby(by=['age','c'])['n_c'].sum().reset_index() for data in list_comorb_agesex]
            arr_n = [data[data['year']==year_i].groupby(by=['age','c'])['n_dplwh'].sum().reset_index() for data in list_comorb_agesex]
            for age_k in ['<50','50-60','60-70','>=70']:
                arr_data_k = [data[data['age']==age_k].reset_index(drop=True) for data in arr_data]
                arr_data_k_mod = []
                for data_m in arr_data_k:
                    data_m['rank_both'] = data_m['n_c'].rank(method='min',ascending=False)
                    data_m_phys = data_m[data_m['c'].isin(['manx','prsn','sczo'])==False].reset_index(drop=True)
                    data_m_phys['rank_phys'] = data_m_phys['n_c'].rank(method='min',ascending=False)
                    data_m = data_m.merge(data_m_phys[['c','rank_phys']],how='left',on='c')
                    arr_data_k_mod.append(data_m)
                dat_age_summary_ik = pd.DataFrame([],columns=['year'])
                for comorb_j in list_comorb:
                    if comorb_j not in ['manx','sczo','prsn']:
                        arr_c_rankp = np.array([data['rank_phys'][(data['c']==comorb_j)].values[0] for data in arr_data_k_mod])
                    arr_c_rankb = np.array([data['rank_both'][(data['c']==comorb_j)].values[0] for data in arr_data_k_mod])
                    arr_c_prev = np.array([data['n_c'][(data['c']==comorb_j)].values[0]/data_n['n_dplwh'][(data_n['age']==age_k)&(data_n['c']==comorb_j)].values[0] for data,data_n in zip(arr_data_k_mod,arr_n)])
                    prev,prev_lb,prev_ub = np.percentile(arr_c_prev,[50,2.5,97.5])
                    if comorb_j not in ['manx','sczo','prsn']:
                        dic_summary = {
                            'year':year_i,'age':age_k,'c':comorb_j,
                            'rankb':sum(arr_c_rankb)/1000,'rankp':sum(arr_c_rankp)/1000,
                            'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                        }
                    else:
                        dic_summary = {
                            'year':year_i,'age':age_k,'c':comorb_j,'rankb':sum(arr_c_rankb)/1000,
                            'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                        }
                    dat_age_summary_ik = pd.concat([dat_age_summary_ik,pd.DataFrame(dic_summary,index=[0])],ignore_index=True)
                dat_age_summary_ik['rank_both'] = dat_age_summary_ik['rankb'].rank(method='min') #min has highest rank in the group
                dat_age_summary_ik['rank_phys'] = dat_age_summary_ik['rankp'].rank(method='min')
                dat_age_summary = pd.concat([dat_age_summary,dat_age_summary_ik],ignore_index=True)

        #####Figure 6: plot 2*3 pyramid bar plots to show comorbidity burden by sex in 2024 vs 2034, with one legend 
        dic_fig_char = {
            'color':{
                'htn':dic_color_palette['blue'][2],'cvd':dic_color_palette['blue'][1],'dm':dic_color_palette['blue'][0],
                'ckd':dic_color_palette['purple'][2],'cld':dic_color_palette['purple'][1],'copd':dic_color_palette['purple'][0],
                'oa':dic_color_palette['green'][2],'cancer':dic_color_palette['green'][1],
                'manx':dic_color_palette['red'][2],'prsn':dic_color_palette['red'][1],'sczo':dic_color_palette['red'][0]},
            'panel_label':[['A','B','C'],['D','E','F']],
            'comorb_label':{
                'cvd':'CVD','htn':'HTN','dm':'Diabetes','oa':'OA','copd':'COPD',
                'ckd':'CKD','cld':'CLD','cancer':'Cancers','manx':'MANX','sczo':'SCZ','prsn':'PD'},
            'sex_label':{'All':'All','M':'Male','F':'Female'},
        }
        fig_sex,axs_sex = plt.subplots(2,3,figsize=(12,9))
        for i,sex_i in enumerate(['All','M','F']):
            for k,type_k in enumerate(['both','phys']):
                if type_k=='both':
                    j_range = np.arange(1,12)
                    rank_type = 'rank_both'
                else:
                    j_range = np.arange(1,9)
                    rank_type = 'rank_phys'
                for j in j_range:
                    if sex_i=='All':
                        comorb_2020,val_2020,lb_2020,ub_2020 = dat_agesex_summary[['c','prev','prev_lb','prev_ub']][(dat_agesex_summary['year']==2024)&(dat_agesex_summary['sex']=='all')&(dat_agesex_summary[rank_type]==j)].values[0]
                    else:
                        comorb_2020,val_2020,lb_2020,ub_2020 = dat_sex_summary[['c','prev','prev_lb','prev_ub']][(dat_sex_summary['year']==2024)&(dat_sex_summary['sex']==sex_i)&(dat_sex_summary[rank_type]==j)].values[0]
                    axs_sex[k,i].barh(j*(-1),val_2020*(-1),color=dic_fig_char['color'][comorb_2020])
                    axs_sex[k,i].errorbar(val_2020*(-1),j*(-1),xerr=np.array([(val_2020-lb_2020),(ub_2020-val_2020)]).reshape(2,1),color='k',capsize=1.5,elinewidth=0.8) 
                    if j==1:
                        if k==0:
                            label_xloc_2020 = ub_2020*(-1)-0.5
                        else:
                            label_xloc_2020 = -0.4*0.8
                    axs_sex[k,i].text(label_xloc_2020,j*(-1),dic_fig_char['comorb_label'][comorb_2020]+': '+str(np.round(val_2020*100,1))+'\n('+str(np.round(lb_2020*100,1))+', '+str(np.round(ub_2020*100,1))+')',fontsize=8,ha='center',va='center')
                    if sex_i=='All':
                        comorb_2030,val_2030,lb_2030,ub_2030 = dat_agesex_summary[['c','prev','prev_lb','prev_ub']][(dat_agesex_summary['year']==2034)&(dat_agesex_summary['sex']=='all')&(dat_agesex_summary[rank_type]==j)].values[0]
                    else:
                        comorb_2030,val_2030,lb_2030,ub_2030 = dat_sex_summary[['c','prev','prev_lb','prev_ub']][(dat_sex_summary['year']==2034)&(dat_sex_summary['sex']==sex_i)&(dat_sex_summary[rank_type]==j)].values[0]
                    axs_sex[k,i].barh(j*(-1),val_2030,color=dic_fig_char['color'][comorb_2030])
                    axs_sex[k,i].errorbar(val_2030,j*(-1),xerr=np.array([val_2030-lb_2030,ub_2030-val_2030]).reshape(2,1),color='k',capsize=1.5,elinewidth=0.8)
                    if j==1:
                        if k==0:
                            label_xloc_2030 = ub_2030+0.5
                        else:
                            label_xloc_2030 = 0.4*0.8
                    axs_sex[k,i].text(label_xloc_2030,j*(-1),dic_fig_char['comorb_label'][comorb_2030]+': '+str(np.round(val_2030*100,1))+'\n('+str(np.round(lb_2030*100,1))+', '+str(np.round(ub_2030*100,1))+')',fontsize=8,ha='center',va='center')
                    if i==0:
                        if k==0:
                            axs_sex[k,i].text(-1.8,j*(-1),str(j),fontsize=12,ha='center',va='center')
                        else:
                            axs_sex[k,i].text(-0.48,j*(-1),str(j),fontsize=12,ha='center',va='center')
                axs_sex[k,i].plot(np.zeros(20),np.linspace(-12,0,20),'-',color='w',lw=0.6)
                if k==0:
                    axs_sex[k,i].text(-1.6,-0.1,dic_fig_char['panel_label'][k][i],fontsize=16)
                    axs_sex[k,i].text(-0.75,-12,'2024',ha='center',va='center',fontsize=14) 
                    axs_sex[k,i].text(0.75,-12,'2034',ha='center',va='center',fontsize=14)
                    axs_sex[k,i].text(0.,-12.5,dic_fig_char['sex_label'][sex_i],ha='center',va='center',fontsize=14)
                    axs_sex[k,i].set_xlim([-1.5,1.5])
                    axs_sex[k,i].set_xticks([])
                    axs_sex[k,i].set_ylim([-11.5,-0.5])
                else:
                    axs_sex[k,i].text(-0.42,-0.1,dic_fig_char['panel_label'][k][i],fontsize=16)
                    axs_sex[k,i].text(-0.2,-9,'2024',ha='center',va='center',fontsize=14) 
                    axs_sex[k,i].text(0.2,-9,'2034',ha='center',va='center',fontsize=14)
                    axs_sex[k,i].text(0.,-9.5,dic_fig_char['sex_label'][sex_i],ha='center',va='center',fontsize=14)
                    axs_sex[k,i].set_xlim([-0.4,0.4])
                    axs_sex[k,i].set_xticks([])
                    axs_sex[k,i].set_ylim([-8.5,-0.5])
                axs_sex[k,i].get_yaxis().set_visible(False)
                axs_sex[k,i].spines[['left','right']].set_visible(False)
        plt.tight_layout() #should be before subplots_adjust to avoid overriding
        plt.subplots_adjust(top=0.95, bottom=0.15, left=0.05,right=0.98,hspace=0.25,wspace=0.12)
        legend_handles = [Patch(facecolor=dic_fig_char['color'][k],label=dic_fig_char['comorb_label'][k]) for k in list_comorb]
        fig_sex.legend(handles=legend_handles,loc='lower center',ncol=6,fontsize=12,fancybox=True,bbox_to_anchor=(0.5, 0.01))
        fig_sex.savefig(dic_output['path']+'fig_comorb_both_sex_2024_2034_'+dic_output['date']+'.png',dpi=600) 

        #####Figure S12:plot 2*4 pyramid bar plots to show comorbidity burden by age in 2024 vs 2034
        dic_fig_char = {
            'color':{
                'htn':dic_color_palette['blue'][2],'cvd':dic_color_palette['blue'][1],'dm':dic_color_palette['blue'][0],
                'ckd':dic_color_palette['purple'][2],'cld':dic_color_palette['purple'][1],'copd':dic_color_palette['purple'][0],
                'oa':dic_color_palette['green'][2],'cancer':dic_color_palette['green'][1],
                'manx':dic_color_palette['red'][2],'prsn':dic_color_palette['red'][1],'sczo':dic_color_palette['red'][0]},
            'panel_label':[['A)','B)','C)','D)'],['E)','F)','G)','H)']],
            'comorb_label':{
                'cvd':'CVD','htn':'HTN','dm':'Diabetes','oa':'OA','copd':'COPD',
                'ckd':'CKD','cld':'CLD','cancer':'Cancers','manx':'MANX','sczo':'SCZ','prsn':'PD'},
            'age_label':{'<50':r'$<50$','50-60':r'$50-59$','60-70':r'$60-69$','>=70':r'$\geq70$'},
            'xlim':{'<50':[-0.9,0.9],'50-60':[-0.9,0.9],'60-70':[-0.9,0.9],'>=70':[-0.9,0.9]},
        }
        fig_age,axs_age = plt.subplots(2,4,figsize=(15,9))
        for i,age_i in enumerate(dic_fig_char['age_label'].keys()):
            for k,type_k in enumerate(['both','phys']):
                if type_k=='both':
                    j_range = np.arange(1,12)
                    rank_type = 'rank_both'
                else:
                    j_range = np.arange(1,9)
                    rank_type = 'rank_phys'
                for j in j_range:
                    comorb_2020,val_2020,lb_2020,ub_2020 = dat_age_summary[['c','prev','prev_lb','prev_ub']][(dat_age_summary['year']==2024)&(dat_age_summary['age']==age_i)&(dat_age_summary[rank_type]==j)].values[0]
                    axs_age[k,i].barh(j*(-1),val_2020*(-1),color=dic_fig_char['color'][comorb_2020])
                    axs_age[k,i].errorbar(val_2020*(-1),j*(-1),xerr=np.array([(val_2020-lb_2020),(ub_2020-val_2020)]).reshape(2,1),color='k',capsize=1.5,elinewidth=0.8)
                    if j==1:
                        if k==0:
                            label_xloc_2020 = ub_2020*(-1)-0.5
                        else:
                            label_xloc_2020 = -dic_fig_char['xlim'][age_i][1]*0.8
                    axs_age[k,i].text(label_xloc_2020,j*(-1),dic_fig_char['comorb_label'][comorb_2020]+': '+str(np.round(val_2020*100,1))+'\n('+str(np.round(lb_2020*100,1))+', '+str(np.round(ub_2020*100,1))+')',fontsize=8,ha='center',va='center')
                    comorb_2030,val_2030,lb_2030,ub_2030 = dat_age_summary[['c','prev','prev_lb','prev_ub']][(dat_age_summary['year']==2034)&(dat_age_summary['age']==age_i)&(dat_age_summary[rank_type]==j)].values[0]
                    axs_age[k,i].barh(j*(-1),val_2030,color=dic_fig_char['color'][comorb_2030])
                    axs_age[k,i].errorbar(val_2030,j*(-1),xerr=np.array([val_2030-lb_2030,ub_2030-val_2030]).reshape(2,1),color='k',capsize=1.5,elinewidth=0.8)
                    if j==1:
                        if k==0:
                            label_xloc_2030 = ub_2030+0.5
                        else:
                            label_xloc_2030 = dic_fig_char['xlim'][age_i][1]*0.8
                    axs_age[k,i].text(label_xloc_2030,j*(-1),dic_fig_char['comorb_label'][comorb_2030]+': '+str(np.round(val_2030*100,1))+'\n('+str(np.round(lb_2030*100,1))+', '+str(np.round(ub_2030*100,1))+')',fontsize=8,ha='center',va='center')
                    if i==0:
                        if k==0:
                            axs_age[k,i].text(-1.8,j*(-1),str(j),fontsize=12,ha='center',va='center')
                        else:
                            axs_age[k,i].text(-1.1,j*(-1),str(j),fontsize=12,ha='center',va='center')
                axs_age[k,i].plot(np.zeros(20),np.linspace(-12,0,20),'-',color='w',lw=0.6)
                if k==0:
                    axs_age[k,i].text(-1.6,-0.1,dic_fig_char['panel_label'][k][i],fontsize=16)
                    axs_age[k,i].text(-0.75,-12,'2024',ha='center',va='center',fontsize=14) 
                    axs_age[k,i].text(0.,-12.5,dic_fig_char['age_label'][age_i],ha='center',va='center',fontsize=14)
                    axs_age[k,i].set_xlim([-1.5,1.5])
                    axs_age[k,i].set_xticks([])
                    axs_age[k,i].set_ylim([-11.5,-0.5])
                else:
                    axs_age[k,i].text(-1,-0.1,dic_fig_char['panel_label'][k][i],fontsize=16)
                    axs_age[k,i].text(-0.45,-9,'2024',ha='center',va='center',fontsize=14) 
                    axs_age[k,i].text(0.45,-9,'2034',ha='center',va='center',fontsize=14)
                    axs_age[k,i].text(0.,-9.5,dic_fig_char['age_label'][age_i],ha='center',va='center',fontsize=14)
                    axs_age[k,i].set_xlim(dic_fig_char['xlim'][age_i])
                    axs_age[k,i].set_xticks([])
                    axs_age[k,i].set_ylim([-8.5,-0.5])
                axs_age[k,i].get_yaxis().set_visible(False)
                axs_age[k,i].spines[['left','right']].set_visible(False)
        plt.tight_layout() #should be before subplots_adjust to avoid overriding
        plt.subplots_adjust(top=0.95, bottom=0.15, left=0.05,right=0.98,hspace=0.25,wspace=0.12)
        legend_handles = [Patch(facecolor=dic_fig_char['color'][k],label=dic_fig_char['comorb_label'][k]) for k in list_comorb]
        fig_age.legend(handles=legend_handles,loc='lower center',ncol=6,fontsize=12,fancybox=True,bbox_to_anchor=(0.5, 0.01))
        fig_age.savefig(dic_output['path']+'fig_comorb_both_age_2024_2034_'+dic_output['date']+'.png',dpi=300)

        #####convert .png file to jpg file
        fig_path = dic_output['path']+'fig_comorb_both_sex_2024_2034_'+dic_output['date']
        im = Image.open(fig_path+'.png')
        im.load()
        background = Image.new("RGB", im.size, (255, 255, 255))
        background.paste(im, mask=im.split()[3]) # 3 is the alpha channel
        background.save(fig_path+'.jpg', 'JPEG', quality=100)

        #####Figure S13: 4*1 bar plots to show top 5 most prevalent comorbidities stratified by age and sex 
        dic_fig_char['xlabel'] = {'M':'Male','F':'Female'}
        dic_fig_char['panel_label'] = ['A)','B)','C)','D)']
        fig_as,axs_as = plt.subplots(4,1,figsize=(15,8))
        for i,sex_i in enumerate(['M','F']):
            for j,year_j in enumerate([2024,2034]):
                for k in np.arange(1,6):
                    comorb,val,lb,ub = dat_agesex_summary[['c','prev','prev_lb','prev_ub']][(dat_agesex_summary['year']==year_j)&(dat_agesex_summary['sex']=='all')&(dat_agesex_summary['rank_both']==k)].values[0]
                    axs_as[i*2+j].bar(k,val,color=dic_fig_char['color'][comorb],width=1.)
                    axs_as[i*2+j].errorbar(k,val,yerr=np.array([(val-lb),(ub-val)]).reshape(2,1),color='k',capsize=1.5,elinewidth=0.8)
                    axs_as[i*2+j].text(k,ub+0.06,dic_fig_char['comorb_label'][comorb],fontsize=6,ha='center',va='center') 
                    comorb,val,lb,ub = dat_sex_summary[['c','prev','prev_lb','prev_ub']][(dat_sex_summary['year']==year_j)&(dat_sex_summary['sex']==sex_i)&(dat_sex_summary['rank_both']==k)].values[0]
                    axs_as[i*2+j].bar(6+k,val,color=dic_fig_char['color'][comorb],width=1.)
                    axs_as[i*2+j].errorbar(6+k,val,yerr=np.array([(val-lb),(ub-val)]).reshape(2,1),color='k',capsize=1.5,elinewidth=0.8)
                    axs_as[i*2+j].text(6+k,ub+0.06,dic_fig_char['comorb_label'][comorb],fontsize=6,ha='center',va='center') 
                for m,age_m in enumerate(['<50','50-60','60-70','>=70']):
                    for k in np.arange(1,6):
                        comorb,val,lb,ub = dat_agesex_summary[['c','prev','prev_lb','prev_ub']][(dat_agesex_summary['year']==year_j)&(dat_agesex_summary['sex']==sex_i)&(dat_agesex_summary['age']==age_m)&(dat_agesex_summary['rank_both']==k)].values[0]
                        axs_as[i*2+j].bar(12+m*6+k,val,color=dic_fig_char['color'][comorb],width=1.)
                        axs_as[i*2+j].errorbar(12+m*6+k,val,yerr=np.array([(val-lb),(ub-val)]).reshape(2,1),color='k',capsize=1.5,elinewidth=0.8)
                        axs_as[i*2+j].text(12+m*6+k,ub+0.06,dic_fig_char['comorb_label'][comorb],fontsize=6,ha='center',va='center')
                axs_as[i*2+j].text(-1,0.8*1.05,dic_fig_char['panel_label'][i*2+j],fontsize=14)
                axs_as[i*2+j].set_xlim([0,36])
                axs_as[i*2+j].set_xticks(np.arange(3,35,6))
                axs_as[i*2+j].set_xticklabels(['All, '+str(year_j),dic_fig_char['xlabel'][sex_i]+', '+str(year_j)]+[dic_fig_char['xlabel'][sex_i]+', '+dic_fig_char['age_label'][age_m] for age_m in ['<50','50-60','60-70','>=70']])
                axs_as[i*2+j].set_ylim([0,0.8])
                axs_as[i*2+j].set_yticks([0,0.25,0.5,0.75])
                axs_as[i*2+j].set_yticklabels(['0%','25%','50%','75%'])
        fig_as.text(0.01,0.5,'Comorbidity prevalence',va='center',rotation=90,fontsize=12)
        plt.subplots_adjust(top=0.96, bottom=0.06, left=0.06,right=0.97,hspace=0.4,wspace=0.15)
        plt.show()
        fig_as.savefig(dic_output['path']+'fig_comorb_both5_agesex_2024_2034_'+dic_output['date']+'.png',dpi=300)

    if 'figure_ranking_comb_comorbidities' in runs:
        """generate plots/tables showing prevalence of combination of comorbidities"""

        #####initialize output path and version 
        dic_output = {
            'path':'my_path/results/', 
            'date':'today',  
        }

        #####load results from microsimulations
        file_name = 'microsim_baseline2034_s1000_today.pkl'
        file_results = open(dic_output['path']+file_name,'rb') 
        list_results = pickle.load(file_results)
        file_results.close()

        list_comorb = ['htn','cvd','dm','ckd','cld','copd','oa','cancer','manx','sczo','prsn']
        list_phys_comorb = ['htn','cvd','dm','ckd','cld','copd','oa','cancer']

        #####load simulated results regarding comorbidity prevalence and prevalence of combined comorbidities
        list_comorb_all = [result['comb1_both'] for result in list_results]
        list_comorb_agesex = [result['comb1_both_agesex'] for result in list_results]
        list_cc_all = [result['comb2_both'] for result in list_results]
        list_cc_agesex = [result['comb2_both_agesex'] for result in list_results]

        #####derive median/lb/ub of combination of 2 comorbidities (both/physical) for diagnosed PLWH over time
        dat_agesex_summary = pd.DataFrame([],columns=['year'])
        for year_i in [2019, 2020, 2024, 2030, 2034]: #create records for multiple years
            dat_agesex_summary_i = pd.DataFrame([],columns=['year'])
            for cc_j in list_cc_all[0]['cc'].unique():
                if 'manx' not in cc_j or 'prsn' not in cc_j or 'sczo' not in cc_j:
                    arr_c_rankp = np.array([data['rank_phys'][(data['year']==year_i)&(data['cc']==cc_j)].values[0] for data in list_cc_all])
                arr_c_rankb = np.array([data['rank_both'][(data['year']==year_i)&(data['cc']==cc_j)].values[0] for data in list_cc_all])
                arr_c_prev = np.array([data['n_cc'][(data['year']==year_i)&(data['cc']==cc_j)].values[0]/data['n_dplwh'][(data['year']==year_i)&(data['cc']==cc_j)].values[0] for data in list_cc_all])
                prev,prev_lb,prev_ub = np.percentile(arr_c_prev,[50,2.5,97.5])
                if 'manx' not in cc_j or 'prsn' not in cc_j or 'sczo' not in cc_j:
                    dic_summary = {
                        'year':year_i,'sex':'all','age':'all','cc':cc_j,
                        'rankb':sum(arr_c_rankb)/1000,'rankp':sum(arr_c_rankp)/1000,
                        'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                    }
                else:
                    dic_summary = {
                        'year':year_i,'sex':'all','age':'all','cc':cc_j,'rankb':sum(arr_c_rankb)/1000,
                        'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                    }
                dat_agesex_summary_i = pd.concat([dat_agesex_summary_i,pd.DataFrame(dic_summary,index=[0])],ignore_index=True)
            dat_agesex_summary_i['rank_both'] = dat_agesex_summary_i['rankb'].rank(method='min') #min has highest rank in the group
            dat_agesex_summary_i['rank_phys'] = dat_agesex_summary_i['rankp'].rank(method='min')
            dat_agesex_summary = pd.concat([dat_agesex_summary,dat_agesex_summary_i],ignore_index=True)
        dat_overall_summary = dat_agesex_summary.copy() #outcomes for all diagnosed PLWH

        #####derive median/lb/ub of combination of 2 comorbidities for diagnosed PLWH by age or sex over time
        list_age_cat = ['<50','50-60','60-70','>=70']
        a = time.time()
        dic_age_summary = {}
        for age_j in list_age_cat:
            dic_age_summary[age_j] = pd.DataFrame([],columns=['year'])
            for year_k in [2019, 2020, 2024, 2030, 2034]: 
                list_cc_age_j = []
                for n in range(1000):
                    dat_cc_age_jn_m = list_cc_agesex[n]['M'][age_j][['year','n_dplwh','cc','n_cc']][list_cc_agesex[n]['M'][age_j]['year']==year_k].reset_index(drop=True).rename(columns={'n_dplwh':'n_dplwh_m','n_cc':'n_cc_m'})
                    dat_cc_age_jn_f = list_cc_agesex[n]['F'][age_j][['n_dplwh','cc','n_cc']][list_cc_agesex[n]['F'][age_j]['year']==year_k].reset_index(drop=True).rename(columns={'n_dplwh':'n_dplwh_f','n_cc':'n_cc_f'})
                    dat_cc_age_jn = dat_cc_age_jn_m.merge(dat_cc_age_jn_f,how='left',on='cc')
                    dat_cc_age_jn['n_dplwh'] = dat_cc_age_jn['n_dplwh_m']+dat_cc_age_jn['n_dplwh_f']
                    dat_cc_age_jn['n_cc'] = dat_cc_age_jn['n_cc_m']+dat_cc_age_jn['n_cc_f']
                    dat_cc_age_jn['rank_both'] = dat_cc_age_jn['n_cc'].rank(method='min',ascending=False)
                    dat_cc_age_jn_phys = dat_cc_age_jn[['cc','n_cc']][dat_cc_age_jn['cc'].str.contains('manx|prsn|sczo')==False].reset_index(drop=True)
                    dat_cc_age_jn_phys['rank_phys'] = dat_cc_age_jn_phys['n_cc'].rank(method='min',ascending=False)
                    dat_cc_age_jn = dat_cc_age_jn.merge(dat_cc_age_jn_phys[['cc','rank_phys']],how='left',on='cc')
                    list_cc_age_j.append(dat_cc_age_jn)
                dat_age_summary_k = pd.DataFrame([],columns=['year'])
                for cc_m in list_cc_all[0]['cc'].unique():
                    if 'manx' not in cc_m or 'prsn' not in cc_m or 'sczo' not in cc_m:
                        arr_c_rankp = np.array([data['rank_phys'][data['cc']==cc_m].values[0] for data in list_cc_age_j])
                    arr_c_rankb = np.array([data['rank_both'][data['cc']==cc_m].values[0] for data in list_cc_age_j])
                    arr_c_prev = np.array([data['n_cc'][data['cc']==cc_m].values[0]/data['n_dplwh'][data['cc']==cc_m].values[0] for data in list_cc_age_j])
                    prev,prev_lb,prev_ub = np.percentile(arr_c_prev,[50,2.5,97.5])
                    if 'manx' not in cc_m or 'prsn' not in cc_m or 'sczo' not in cc_m:
                        dic_summary = {
                            'year':year_k,'age':age_j,'cc':cc_m,
                            'rankb':sum(arr_c_rankb)/1000,'rankp':sum(arr_c_rankp)/1000,
                            'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                        }
                    else:
                        dic_summary = {
                            'year':year_k,'age':age_j,'cc':cc_m,'rankb':sum(arr_c_rankb)/1000,
                            'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                        }
                    dat_age_summary_k = pd.concat([dat_age_summary_k,pd.DataFrame(dic_summary,index=[0])],ignore_index=True)
                dat_age_summary_k['rank_both'] = dat_age_summary_k['rankb'].rank(method='min') #min has highest rank in the group
                dat_age_summary_k['rank_phys'] = dat_age_summary_k['rankp'].rank(method='min')
                dic_age_summary[age_j] = pd.concat([dic_age_summary[age_j],dat_age_summary_k],ignore_index=True)
            print (age_j, dat_age_summary_k[dat_age_summary_k['rank_phys']<=5])
        dic_sex_summary = {}
        for sex_i in ['M','F']:
            dic_sex_summary[sex_i] = pd.DataFrame([],columns=['year'])
            for year_k in [2019, 2020, 2024, 2030, 2034]: 
                list_cc_sex_i = []
                for n in range(1000):
                    dat_cc_sex_in = list_cc_agesex[n][sex_i]['<50'][['year','n_dplwh','cc','n_cc']][list_cc_agesex[n][sex_i]['<50']['year']==year_k].reset_index(drop=True).rename(columns={'n_dplwh':'n_dplwh_0','n_cc':'n_cc_0'})
                    for j,age_j in enumerate(list_age_cat[1:]):
                        dat_cc_sex_in_j = list_cc_agesex[n][sex_i][age_j][['n_dplwh','cc','n_cc']][list_cc_agesex[n][sex_i][age_j]['year']==year_k].reset_index(drop=True).rename(columns={'n_dplwh':'n_dplwh_'+str(j+1),'n_cc':'n_cc_'+str(j+1)})
                        dat_cc_sex_in = dat_cc_sex_in.merge(dat_cc_sex_in_j,how='left',on='cc')
                    list_key_dplwh = ['n_dplwh_'+str(j) for j in range(len(list_age_cat))]
                    list_key_cc = ['n_cc_'+str(j) for j in range(len(list_age_cat))]
                    dat_cc_sex_in['n_dplwh'] = dat_cc_sex_in[list_key_dplwh].sum(axis=1)
                    dat_cc_sex_in['n_cc'] = dat_cc_sex_in[list_key_cc].sum(axis=1)
                    dat_cc_sex_in['rank_both'] = dat_cc_sex_in['n_cc'].rank(method='min',ascending=False)
                    dat_cc_sex_in_phys = dat_cc_sex_in[['cc','n_cc']][dat_cc_sex_in['cc'].str.contains('manx|prsn|sczo')==False].reset_index(drop=True)
                    dat_cc_sex_in_phys['rank_phys'] = dat_cc_sex_in_phys['n_cc'].rank(method='min',ascending=False)
                    dat_cc_sex_in = dat_cc_sex_in.merge(dat_cc_sex_in_phys[['cc','rank_phys']],how='left',on='cc')
                    list_cc_sex_i.append(dat_cc_sex_in)
                dat_sex_summary_k = pd.DataFrame([],columns=['year'])
                for cc_m in list_cc_all[0]['cc'].unique():
                    if 'manx' not in cc_m or 'prsn' not in cc_m or 'sczo' not in cc_m:
                        arr_c_rankp = np.array([data['rank_phys'][data['cc']==cc_m].values[0] for data in list_cc_sex_i])
                    arr_c_rankb = np.array([data['rank_both'][data['cc']==cc_m].values[0] for data in list_cc_sex_i])
                    arr_c_prev = np.array([data['n_cc'][data['cc']==cc_m].values[0]/data['n_dplwh'][data['cc']==cc_m].values[0] for data in list_cc_sex_i])
                    prev,prev_lb,prev_ub = np.percentile(arr_c_prev,[50,2.5,97.5])
                    if 'manx' not in cc_m or 'prsn' not in cc_m or 'sczo' not in cc_m:
                        dic_summary = {
                            'year':year_k,'sex':sex_i,'cc':cc_m,
                            'rankb':sum(arr_c_rankb)/1000,'rankp':sum(arr_c_rankp)/1000,
                            'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                        }
                    else:
                        dic_summary = {
                            'year':year_k,'sex':sex_i,'cc':cc_m,'rankb':sum(arr_c_rankb)/1000,
                            'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                        }
                    dat_sex_summary_k = pd.concat([dat_sex_summary_k,pd.DataFrame(dic_summary,index=[0])],ignore_index=True)
                dat_sex_summary_k['rank_both'] = dat_sex_summary_k['rankb'].rank(method='min') #min has highest rank in the group
                dat_sex_summary_k['rank_phys'] = dat_sex_summary_k['rankp'].rank(method='min')
                dic_sex_summary[sex_i] = pd.concat([dic_sex_summary[sex_i],dat_sex_summary_k],ignore_index=True)

        #####derive median/lb/ub of combination of 2 comorbidities (both/physical) for diagnosed PLWH by age and sex over time
        list_age_cat = ['<50','50-60','60-70','>=70']
        dic_agesex_summary = {}
        for sex_i in ['M','F']:
            dic_agesex_summary[sex_i] = {}
            for age_j in list_age_cat:
                dic_agesex_summary[sex_i][age_j] = pd.DataFrame([],columns=['year'])
                for year_k in [2019, 2020, 2024, 2030, 2034]: 
                    dat_agesex_summary_k = pd.DataFrame([],columns=['year'])
                    for cc_m in list_cc_all[0]['cc'].unique():
                        if 'manx' not in cc_m or 'prsn' not in cc_m or 'sczo' not in cc_m:
                            arr_c_rankp = np.array([data[sex_i][age_j]['rank_phys'][(data[sex_i][age_j]['year']==year_k)&(data[sex_i][age_j]['cc']==cc_m)].values[0] for data in list_cc_agesex])
                        arr_c_rankb = np.array([data[sex_i][age_j]['rank_both'][(data[sex_i][age_j]['year']==year_k)&(data[sex_i][age_j]['cc']==cc_m)].values[0] for data in list_cc_agesex])
                        arr_c_prev = np.array([data[sex_i][age_j]['n_cc'][(data[sex_i][age_j]['year']==year_k)&(data[sex_i][age_j]['cc']==cc_m)].values[0]/data[sex_i][age_j]['n_dplwh'][(data[sex_i][age_j]['year']==year_k)&(data[sex_i][age_j]['cc']==cc_m)].values[0] for data in list_cc_agesex])
                        prev,prev_lb,prev_ub = np.percentile(arr_c_prev,[50,2.5,97.5])
                        if 'manx' not in cc_m or 'prsn' not in cc_m or 'sczo' not in cc_m:
                            dic_summary = {
                                'year':year_k,'sex':sex_i,'age':age_j,'cc':cc_m,
                                'rankb':sum(arr_c_rankb)/1000,'rankp':sum(arr_c_rankp)/1000,
                                'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                            }
                        else:
                            dic_summary = {
                                'year':year_k,'sex':sex_i,'age':age_j,'cc':cc_m,'rankb':sum(arr_c_rankb)/1000,
                                'prev':prev,'prev_lb':prev_lb,'prev_ub':prev_ub
                            }
                        dat_agesex_summary_k = pd.concat([dat_agesex_summary_k,pd.DataFrame(dic_summary,index=[0])],ignore_index=True)
                    dat_agesex_summary_k['rank_both'] = dat_agesex_summary_k['rankb'].rank(method='min') #min has highest rank in the group
                    dat_agesex_summary_k['rank_phys'] = dat_agesex_summary_k['rankp'].rank(method='min')
                    dic_agesex_summary[sex_i][age_j] = pd.concat([dic_agesex_summary[sex_i][age_j],dat_agesex_summary_k],ignore_index=True)

        #####output derived records as .pkl file
        dic_record = {'all':dat_overall_summary,'sex':dic_sex_summary,'age':dic_age_summary,'agesex':dic_agesex_summary}
        file_data = open(dic_output['path']+'record_ranking_cc_'+dic_output['date']+'.pkl','wb') 
        pickle.dump(dic_record, file_data)
        file_data.close()

        #####load derived records from .pkl file
        file_data = open(dic_output['path']+'record_ranking_cc_'+dic_output['date']+'.pkl','rb') 
        dic_record = pickle.load(file_data)
        file_data.close()
        dat_overall_summary = dic_record['all']
        dic_sex_summary = dic_record['sex']
        dic_age_summary = dic_record['age']
        dic_agesex_summary = dic_record['agesex']

        #####Figures 7-8: create 2*3/2*4 heat maps for prevalence of combined comorbidities (both) by age or sex in two different years
        dic_sex_record = {'all':dat_overall_summary,'M':dic_sex_summary['M'],'F':dic_sex_summary['F']}
        list_panel = [['A','B','C'],['D','E','F']] 
        dic_sex_label = {'all':'All','M':'Male','F':'Female'}
        dic_comorb_label = {
            'htn':'HTN','cvd':'CVD','dm':'Diabetes','ckd':'CKD','cld':'CLD',
            'copd':'COPD','oa':'OA','cancer':'Cancers','manx':'MANX','sczo':'SCZ','prsn':'PD'}
        fig_cc_both,axs_cc_both = plt.subplots(2,3,figsize=(12,8))
        cbar_ax1 = fig_cc_both.add_axes([0.94,0.558,.01,.38]) 
        cbar_ax2 = fig_cc_both.add_axes([0.94,0.08,.01,.38])
        for m,sex_m in enumerate(dic_sex_record.keys()):
            dat_cc = dic_sex_record[sex_m]
            for k,year_k in enumerate([2024,2034]):
                matr_cc_prev = np.zeros(shape=(len(list_comorb),len(list_comorb)))
                for i,c1_i in enumerate(list_comorb):
                    for j,c2_j in enumerate(list_comorb):
                        if c1_i==c2_j:
                            matr_cc_prev[i,j] = np.nan
                        else:
                            matr_cc_prev[i,j] = dat_cc['prev'][(dat_cc['year']==year_k)&(dat_cc['cc'].str.contains(c1_i))&(dat_cc['cc'].str.contains(c2_j))].values[0]
                mask = np.zeros_like(matr_cc_prev)
                mask[np.triu_indices_from(mask)] = True
                sns.heatmap(matr_cc_prev*100,mask=mask,square=True,vmin=0.,vmax=14.,cmap='Reds',annot=True,fmt='.1f',annot_kws={'size':8},cbar_ax=cbar_ax1 if k==0 else cbar_ax2,ax=axs_cc_both[k,m])
                cbar = axs_cc_both[k,m].collections[0].colorbar
                cbar.set_ticks([0.,2.,4.,6.,8.,10.,12.,14.])
                cbar.set_ticklabels(['0%','2%','4%','6%','8%','10%','12%','14%'],fontsize=8)
                axs_cc_both[k,m].set_xticklabels([dic_comorb_label[comorb] for comorb in list_comorb],rotation=75,fontsize=8) 
                axs_cc_both[k,m].set_yticklabels([dic_comorb_label[comorb] for comorb in list_comorb],rotation='horizontal',fontsize=8)
                axs_cc_both[k,m].xaxis.set_tick_params(length=0) 
                axs_cc_both[k,m].yaxis.set_tick_params(length=0)
                axs_cc_both[k,m].text(-2.,-0.5,list_panel[k][m],fontsize=14)
                axs_cc_both[k,m].set_xlabel(dic_sex_label[sex_m]+', '+str(year_k),fontsize=12) 
        plt.subplots_adjust(top=0.95, bottom=0.11, left=0.03,right=0.95,hspace=0.4,wspace=0.04) 
        plt.show()
        fig_cc_both.savefig(dic_output['path']+'fig_cc_both_sex_2024_2034_'+dic_output['date']+'.png',bbox_inches='tight',dpi=600)
        dic_age_label = {'<50':r'$<50$','50-60':r'$50-59$','60-70':r'$60-69$','>=70':r'$\geq70$'}
        list_panel = [['A','B','C','D'],['E','F','G','H']] 
        fig_cc_both,axs_cc_both = plt.subplots(2,4,figsize=(16,8))
        cbar_ax1 = fig_cc_both.add_axes([0.95,0.558,.01,.38]) 
        cbar_ax2 = fig_cc_both.add_axes([0.95,0.08,.01,.38])
        for m,age_m in enumerate(dic_age_label.keys()):
            dat_cc = dic_age_summary[age_m]
            for k,year_k in enumerate([2024,2034]):
                matr_cc_prev = np.zeros(shape=(len(list_comorb),len(list_comorb)))
                for i,c1_i in enumerate(list_comorb):
                    for j,c2_j in enumerate(list_comorb):
                        if c1_i==c2_j:
                            matr_cc_prev[i,j] = np.nan
                        else:
                            matr_cc_prev[i,j] = dat_cc['prev'][(dat_cc['year']==year_k)&(dat_cc['cc'].str.contains(c1_i))&(dat_cc['cc'].str.contains(c2_j))].values[0]
                mask = np.zeros_like(matr_cc_prev)
                mask[np.triu_indices_from(mask)] = True
                sns.heatmap(matr_cc_prev*100,mask=mask,square=True,vmin=0.,vmax=18.,cmap='Reds',annot=True,fmt='.1f',annot_kws={'size':8},cbar_ax=cbar_ax1 if k==0 else cbar_ax2,ax=axs_cc_both[k,m])
                cbar = axs_cc_both[k,m].collections[0].colorbar
                cbar.set_ticks([0.,3.,6.,9.,12.,15.,18.])
                cbar.set_ticklabels(['0%','3%','6%','9%','12%','15%','18%'],fontsize=8)
                axs_cc_both[k,m].set_xticklabels([dic_comorb_label[comorb] for comorb in list_comorb],rotation=75,fontsize=8) 
                axs_cc_both[k,m].set_yticklabels([dic_comorb_label[comorb] for comorb in list_comorb],rotation='horizontal',fontsize=8)
                axs_cc_both[k,m].xaxis.set_tick_params(length=0) 
                axs_cc_both[k,m].yaxis.set_tick_params(length=0)
                axs_cc_both[k,m].text(-2.,-0.5,list_panel[k][m],fontsize=14)
                axs_cc_both[k,m].set_xlabel('Aged '+dic_age_label[age_m]+', '+str(year_k),fontsize=12) 
        plt.subplots_adjust(top=0.95, bottom=0.11, left=0.02,right=0.96,hspace=0.4,wspace=0.03) 
        plt.show()
        fig_cc_both.savefig(dic_output['path']+'fig_cc_both_age_2024_2034_'+dic_output['date']+'.png',bbox_inches='tight',dpi=600) 

        #####convert .png file to jpg file
        fig_path = dic_output['path']+'fig_cc_both_age_2024_2034_'+dic_output['date']
        #fig_path = dic_output['path']+'fig_cc_both_sex_2024_2034_'+dic_output['date']
        im = Image.open(fig_path+'.png')
        im.load()
        background = Image.new("RGB", im.size, (255, 255, 255))
        background.paste(im, mask=im.split()[3]) 
        background.save(fig_path+'.jpg', 'JPEG', quality=100)

        #####Figure S14: create 2*4 heat map for prevalence of combined comorbidities (both) by age and sex in 2034
        dic_comorb_label = {
            'htn':'HTN','cvd':'CVD','dm':'Diabetes','ckd':'CKD','cld':'CLD',
            'copd':'COPD','oa':'OA','cancer':'Cancers','manx':'MANX','sczo':'SCZ','prsn':'PD'}
        list_panel = [['A)','B)','C)','D)'],['E)','F)','G)','H)']]
        fig_cc_both,axs_cc_both = plt.subplots(2,4,figsize=(16,8))
        cbar_ax1 = fig_cc_both.add_axes([0.96,0.558,.01,.38]) 
        cbar_ax2 = fig_cc_both.add_axes([0.96,0.08,.01,.38])
        for m,sex_m in enumerate(['M','F']):
            for n,age_n in enumerate(list_age_cat):
                dat_cc = dic_agesex_summary[sex_m][age_n]
                year_k = 2034
                matr_cc_prev = np.zeros(shape=(len(list_comorb),len(list_comorb)))
                for i,c1_i in enumerate(list_comorb):
                    for j,c2_j in enumerate(list_comorb):
                        if c1_i==c2_j:
                            matr_cc_prev[i,j] = np.nan
                        else:
                            matr_cc_prev[i,j] = dat_cc['prev'][(dat_cc['year']==year_k)&(dat_cc['cc'].str.contains(c1_i))&(dat_cc['cc'].str.contains(c2_j))].values[0]
                mask = np.zeros_like(matr_cc_prev)
                mask[np.triu_indices_from(mask)] = True
                sns.heatmap(matr_cc_prev*100,mask=mask,square=True,vmin=0.,vmax=19.,cmap='Reds',annot=True,fmt='.1f',annot_kws={'size':8},cbar_ax=cbar_ax1 if m==0 else cbar_ax2,ax=axs_cc_both[m,n])
                cbar = axs_cc_both[m,n].collections[0].colorbar
                cbar.set_ticks([0.,3.,6.,9.,12.,15.,18.])
                cbar.set_ticklabels(['0%','3%','6%','9%','12%','15%','18%'],fontsize=8)
                axs_cc_both[m,n].set_xticklabels([dic_comorb_label[com] for com in list_comorb],rotation=75,fontsize=8)
                axs_cc_both[m,n].set_yticklabels([dic_comorb_label[com] for com in list_comorb],rotation='horizontal',fontsize=8)
                axs_cc_both[m,n].xaxis.set_tick_params(length=0) 
                axs_cc_both[m,n].yaxis.set_tick_params(length=0)
                axs_cc_both[m,n].text(-2.,-0.5,list_panel[m][n],fontsize=14)
        plt.subplots_adjust(top=0.94, bottom=0.08, left=0.03,right=0.95,hspace=0.25,wspace=0.05)
        plt.show()
        fig_cc_both.savefig(dic_output['path']+'fig_cc_both_agesex_2034_'+dic_output['date']+'.png',bbox_inches='tight',dpi=300)

    if 'figure_sens_comorb_prob' in runs:
        """generate forest plot to show the relative difference in the proportion with >=3 comorbidities (both physical and mental)
        under scenarios with +/-25% comorbidity incidence, similar to the PEARL study (PMID:38215160)"""

        #####initialize output path and version as date
        dic_output = {
            'path':'my_path/results/', 
            'date':'today', 
        }

        #####load results of the microsimulation model
        file_name = dic_output['path']+'microsim_baseline2034_s1000_today.pkl' #simulated aggregated HIV cascade of care and comorbidity prevalence over time for 1000 simulations
        file_results = open(file_name,'rb')
        list_results = pickle.load(file_results)
        file_results.close()

        #####initialize dataframe to save prevalence of >=3 multimorbidity (both physical and mental) under different scenarios
        list_comorb = [result['comorb_prev'] for result in list_results]
        year_i = 2034 
        arr_n_all = np.array([data_comorb[['0_b','1_b','2_b','>=3_b']][(data_comorb['year']==year_i)].sum().sum() for data_comorb in list_comorb])
        arr_n_comorb_j = [data_comorb['>=3_b'][(data_comorb['year']==year_i)].sum() for data_comorb in list_comorb]
        med,lb,ub = np.percentile(arr_n_comorb_j/arr_n_all,[50,2.5,97.5])
        dat_sens = pd.DataFrame([],columns=['comorb','type','median','lb','ub'])
        dic_sens_sq = {'comorb':'sq','type':'--','median':med,'lb':lb,'ub':ub}
        dat_sens = pd.concat([dat_sens,pd.DataFrame(dic_sens_sq,index=[0])],ignore_index=True)

        #####load results from scenarios for sensitivity analysis and add to dataframe
        list_comorb = ['htn','cvd','dm','ckd','cld','copd','oa','cancer','manx','sczo','prsn']
        for comorb_i in list_comorb:
            for type_j in ['plus','minus']:

                ####load simulation results
                file_name = dic_output['path']+'microsim_comorb_sens_'+comorb_i+'_'+type_j+'_1000_today.pkl' #simulated aggregated HIV cascade of care and comorbidity prevalence over time for 1000 simulations
                file_results = open(file_name,'rb')
                list_results = pickle.load(file_results)
                file_results.close()

                ####find >=3 multimorbidity prevalence
                list_comorb = [result['comorb_prev'] for result in list_results]
                year_i = 2034
                arr_n_all = np.array([data_comorb[['0_b','1_b','2_b','>=3_b']][(data_comorb['year']==year_i)].sum().sum() for data_comorb in list_comorb])
                arr_n_comorb_j = [data_comorb['>=3_b'][(data_comorb['year']==year_i)].sum() for data_comorb in list_comorb]
                med,lb,ub = np.percentile(arr_n_comorb_j/arr_n_all,[50,2.5,97.5])
                dic_sens_ij = {'comorb':comorb_i,'type':type_j,'median':med,'lb':lb,'ub':ub}
                dat_sens = pd.concat([dat_sens,pd.DataFrame(dic_sens_ij,index=[0])],ignore_index=True)

        #####sort results by absolute difference between sensitivity scenarios and status quo
        val_sq,lb_sq,ub_sq = dat_sens[['median','lb','ub']][dat_sens['comorb']=='sq'].values[0]
        dat_sens['diff'] = (dat_sens['median']-val_sq)/val_sq*100
        dat_sens['diff_lb'] = (dat_sens['lb']-val_sq)/val_sq*100
        dat_sens['diff_ub'] = (dat_sens['ub']-val_sq)/val_sq*100
        dat_sens['diff_abs'] = abs(dat_sens['diff'])
        dat_sens_sort = dat_sens.sort_values(by='diff_abs')

        #####Figure S15: create forest plot to show relative difference under sens. scenarios compared with status quo
        dic_comorb_label = {
            'htn':'HTN','cvd':'CVD','dm':'Diabetes','ckd':'CKD','cld':'CLD',
            'copd':'COPD','oa':'OA','cancer':'Cancers','manx':'MANX','sczo':'SCZ','prsn':'PD'}
        dic_type_label = {'plus':'+25%','minus':'-25%'}
        df_markersize = 6
        fig_fore,ax_fore = plt.subplots(figsize=(8,10))
        row_n = 0
        diff_sq,diff_sq_lb,diff_sq_ub = dat_sens[['diff','diff_lb','diff_ub']][dat_sens['comorb']=='sq'].values[0]
        ax_fore.errorbar(diff_sq,row_n,xerr=np.array([diff_sq-diff_sq_lb,diff_sq_ub-diff_sq]).reshape(-1,1),fmt='o',
                                        ms=df_markersize,color='k',ecolor='k',elinewidth=1.5,capsize=5.)
        ax_fore.text(-12,row_n,'Baseline simulation',fontsize=12,verticalalignment='center',horizontalalignment='center')
        row_n += 1
        for i,comorb_i in enumerate(dat_sens_sort['comorb'].unique()[1:]):
            for j,type_j in enumerate(['plus','minus']):
                
                diff,diff_lb,diff_ub = dat_sens[['diff','diff_lb','diff_ub']][(dat_sens['comorb']==comorb_i)&(dat_sens['type']==type_j)].values[0]
                ax_fore.errorbar(diff,row_n,xerr=np.array([diff-diff_lb,diff_ub-diff]).reshape(-1,1),fmt='o',
                                        ms=df_markersize,color='darkgray',ecolor='k',elinewidth=1.5,capsize=5.)
                ax_fore.text(-12,row_n,dic_comorb_label[comorb_i]+' '+dic_type_label[type_j],fontsize=12,verticalalignment='center',horizontalalignment='center')
                row_n += 1
        ax_fore.plot(np.zeros((row_n+4)*4),np.linspace(-1,row_n+4,(row_n+4)*4),'--',color='k',markersize=0.4,alpha=0.6)
        ax_fore.set_xlim([-8,8])
        ax_fore.set_ylim([-1,23])
        ax_fore.spines['top'].set_visible(False)
        ax_fore.spines['right'].set_visible(False)
        ax_fore.spines['left'].set_visible(False)
        ax_fore.yaxis.set_visible(False)
        ax_fore.axes.get_yaxis().set_visible(False)
        ax_fore.set_xticks([-8,-4,0,4,8])
        ax_fore.set_xticklabels([-8,-4,0,4,8],fontsize=10)
        ax_fore.set_xlabel('Relative difference in the proportion of PLWH\nwith at least three comorbidities (including both\nphysical and mental conditions) in 2034',fontsize=12)
        plt.subplots_adjust(top=0.98, bottom=0.1, left=0.3,right=0.95,hspace=0.3,wspace=0.2)
        plt.show()
        fig_fore.savefig(dic_output['path']+'fig_comorb_sens_2034_'+dic_output['date']+'.png',dpi=600)
